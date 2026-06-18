# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com/
https://qualcoder.org/
"""

from copy import copy, deepcopy
import datetime
import logging
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline
import os
import qtawesome as qta

from PyQt6 import QtGui, QtWidgets, QtCore
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QBrush

from .code_in_all_files import DialogCodeInAllFiles
from .color_selector import TextColor
from .GUI.ui_dialog_report_comparisons import Ui_Dialog_reportComparisons
from .GUI.ui_dialog_report_code_frequencies import Ui_Dialog_reportCodeFrequencies
from .helpers import Message, ExportDirectoryPathDialog, init_persistent_tree_header, restore_persistent_tree_widths
from .information import DialogInformation
from .report_attributes import DialogSelectAttributeParameters
from .select_items import DialogSelectItems

path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogReportCodeFrequencies(QtWidgets.QDialog):
    """ Show code and category frequencies, overall and for each coder in tree widget.
    This is for text, image and av coding. """

    def __init__(self, app, parent_textedit):

        self.app = app
        self.parent_textEdit = parent_textedit
        self.attributes = []
        self.coders, self.codes, self.categories, = [], [], []
        self.coded = []  # Used to refactor name
        self.truncated_code_names = True
        self.contains_long_names = False
        self.file_ids = []
        self.get_data()
        self.calculate_code_frequencies()
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_reportCodeFrequencies()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_file_attributes.pressed.connect(self.get_files_from_attributes)
        self.ui.pushButton_exporttext.pressed.connect(self.export_text_file)
        self.ui.pushButton_exporttext.setIcon(qta.icon('mdi6.export-variant', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_export_excel.pressed.connect(self.export_excel_file)
        self.ui.pushButton_export_excel.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_select_files.pressed.connect(self.select_files_button)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        font = f'font: {self.app.settings["treefontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.treeWidget.setStyleSheet(font)
        self.ui.treeWidget.setSelectionMode(QtWidgets.QTreeWidget.SelectionMode.ExtendedSelection)
        init_persistent_tree_header(self.ui.treeWidget, self.app, 'dialogreportcodefrequencies_tree_widths')
        self.fill_tree()
        # These signals after the tree is filled the first time
        self.ui.treeWidget.itemCollapsed.connect(self.get_collapsed)
        self.ui.treeWidget.itemExpanded.connect(self.get_collapsed)
        self.ui.treeWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.treeWidget.customContextMenuRequested.connect(self.tree_menu)
        self.ui.radioButton.clicked.connect(self.sort_by_alphabet)
        self.ui.radioButton_2.clicked.connect(self.sort_by_totals)
        self.ui.checkBox_source_breakdown.stateChanged.connect(self._toggle_source_columns)
        tree_header = self.ui.treeWidget.header()
        tree_header.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        tree_header.customContextMenuRequested.connect(self._header_menu)
        self.app.project_events.project_data_changed.connect(self._on_project_data_changed)

    def _toggle_source_columns(self):
        """ Show or hide the per-source (text / image / A-V) breakdown columns (2, 3, 4).
        The data is always calculated; only the visibility changes. """
        breakdown = self.ui.checkBox_source_breakdown.isChecked()
        for col in (2, 3, 4):
            self.ui.treeWidget.setColumnHidden(col, not breakdown)

    def _header_menu(self, position):
        """ Right click on a tree header: hide that column, or re-show hidden columns.
        Column 0 (the code tree itself) cannot be hidden. Exports follow visibility. """
        tree_header = self.ui.treeWidget.header()
        col = tree_header.logicalIndexAt(position)
        menu = QtWidgets.QMenu()
        menu.setStyleSheet(f"QMenu {{font-size:{self.app.settings['fontsize']}pt}} ")
        action_hide = None
        if col > 0 and not self.ui.treeWidget.isColumnHidden(col):
            action_hide = menu.addAction(_("Hide column: ") + self.ui.treeWidget.headerItem().text(col))
        show_actions = {}
        for i in range(1, self.ui.treeWidget.columnCount()):
            if self.ui.treeWidget.isColumnHidden(i):
                act = menu.addAction(_("Show column: ") + self.ui.treeWidget.headerItem().text(i))
                show_actions[act] = i
        if action_hide is None and not show_actions:
            return
        action = menu.exec(tree_header.mapToGlobal(position))
        if action is None:
            return
        if action == action_hide:
            self.ui.treeWidget.setColumnHidden(col, True)
        elif action in show_actions:
            self.ui.treeWidget.setColumnHidden(show_actions[action], False)

    def select_files_button(self):
        """ Report code frequencies for all files or selected files.
        Set:
            self.file_ids
        """

        filenames = self.app.get_filenames()
        if len(filenames) == 0:
            return
        ui = DialogSelectItems(self.app, filenames, _("Select files to view"), "many")
        ok = ui.exec()
        tooltip = _("Files selected: ")
        self.attributes = []
        self.ui.pushButton_file_attributes.setToolTip(_("File Attributes"))
        self.file_ids = []
        if ok:
            selected_files = ui.get_selected()  # List of dictionaries
            files_text = ""
            for row in selected_files:
                self.file_ids.append(row['id'])
                files_text += f"\n{row['name']}"
            tooltip += files_text
            if len(self.file_ids) > 0:
                self.ui.pushButton_select_files.setToolTip(tooltip)
        self.get_data()
        self.calculate_code_frequencies()
        self.fill_tree()

    def get_files_from_attributes(self):
        """ Select files based on attribute selections.
        Attribute results are a dictionary of:
        first item is a Boolean AND or OR list item
        Followed by each attribute list item
        Set:
            self.file_ids
        """

        # Clear ui
        self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
        ui = DialogSelectAttributeParameters(self.app)

        ui.fill_parameters(self.attributes)
        temp_attributes = deepcopy(self.attributes)
        self.attributes = []

        ok = ui.exec()
        if not ok:
            self.attributes = temp_attributes
            self.ui.pushButton_file_attributes.setIcon(
                qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
            if self.attributes:
                self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable-box', options=[{'scale_factor': 1.3}]))
            return
        self.attributes = ui.parameters
        if len(self.attributes) == 1:  # Boolean parameter, no attributes selected
            self.ui.pushButton_file_attributes.setIcon(
                qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
            self.attributes = []
            return
        if not ui.result_file_ids:
            Message(self.app, _("Nothing found") + " " * 20, _("No matching files found")).exec()
            self.ui.pushButton_file_attributes.setIcon(
                qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
            return
        self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable-box', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_file_attributes.setToolTip(ui.tooltip_msg)
        self.file_ids = ui.result_file_ids
        self.ui.pushButton_select_files.setToolTip(_("Select files"))
        msg = ""
        filenames = self.app.get_filenames()
        for i, f in enumerate(filenames):
            if f['id'] in ui.result_file_ids:
                if i < 20:
                    msg += f"\n{f['name']}"
        if len(ui.result_file_ids) > 20:
            msg += f"\nand more. Total files: {len(ui.result_file_ids)}"
        Message(self.app, _("Files selected by attributes"), msg).exec()
        self.get_data()
        self.calculate_code_frequencies()
        self.fill_tree()

    def get_data(self):
        """ Called from init.
        Calls calculate_code_frequency - for each code.
        Adds a list item that is ready to be used by the treeWidget to display multiple
        columns with the coder frequencies.
        Not using  app.get_codes_categories method as this adds extra columns for each end user
        """

        cur = self.app.conn.cursor()
        self.categories = []
        cur.execute("select name, catid, owner, date, ifnull(memo,''), supercatid from code_cat order by lower(name)")
        result = cur.fetchall()
        for row in result:
            self.categories.append({'name': row[0], 'catid': row[1], 'owner': row[2],
                                    'date': row[3], 'memo': row[4], 'supercatid': row[5],
                                    'display_list': [row[0], 'catid:' + str(row[1])]})
        self.codes = []
        cur.execute("select name, ifnull(memo,''), owner, date, cid, catid, color, supercid from code_name order by lower(name)")
        result = cur.fetchall()
        for row in result:
            self.codes.append({'name': row[0], 'memo': row[1], 'owner': row[2], 'date': row[3],
                               'cid': row[4], 'catid': row[5], 'color': row[6], 'supercid': row[7],
                               'display_list': [row[0], 'cid:' + str(row[4])]})
        self.coders = []
        cur.execute("select distinct owner from code_text union select distinct owner from code_image union "
                    "select distinct owner from code_av")
        result = cur.fetchall()
        self.coders = []
        for row in result:
            self.coders.append(row[0])
        self.coded = []
        if True:
            cur.execute("select cid, owner, fid from code_text")
            result = cur.fetchall()
            for row in result:
                if row[2] in self.file_ids or self.file_ids == []:
                    self.coded.append((row[0], row[1], row[2], 'text'))
            cur.execute("select cid, owner, id from code_image")
            result = cur.fetchall()
            for row in result:
                if row[2] in self.file_ids or self.file_ids == []:
                    self.coded.append((row[0], row[1], row[2], 'image'))
            cur.execute("select cid, owner, id from code_av")
            result = cur.fetchall()
            for row in result:
                if row[2] in self.file_ids or self.file_ids == []:
                    self.coded.append((row[0], row[1], row[2], 'av'))

    def calculate_code_frequencies(self):
        """ Calculate the frequency of each code for all coders and the total.
        Add a list item to each code that can be used to display in treeWidget.
        For codings in code_image, code_text.
        """

        for c in self.codes:
            # Per-source breakdown first: these columns sit between Id and the coder columns
            n_text = n_image = n_av = 0
            for cit in self.coded:
                if cit[0] == c['cid']:
                    if cit[3] == 'text':
                        n_text += 1
                    elif cit[3] == 'image':
                        n_image += 1
                    elif cit[3] == 'av':
                        n_av += 1
            c['display_list'] += [n_text, n_image, n_av]
            total = 0
            for cn in self.coders:
                count = 0
                for cit in self.coded:
                    if cit[1] == cn and cit[0] == c['cid']:
                        count += 1
                        total += 1
                c['display_list'].append(count)
            c['display_list'].append(total)

        # Map each code to the category of its top ancestor code, so a sub-code's codings
        # are counted under the category that its parent code belongs to.
        def effective_catid(code):
            seen = set()
            cur = code
            while cur is not None and cur['catid'] is None and cur.get('supercid') \
                    and cur['cid'] not in seen:
                seen.add(cur['cid'])
                parent = None
                for cc in self.codes:
                    if cc['cid'] == cur['supercid']:
                        parent = cc
                        break
                cur = parent
            return cur['catid'] if cur else None

        # Add the number of codings of each code (including nested sub-codes) to its category
        for cat in self.categories:
            # 6 = cat name, cat id, total, and the three per-source columns
            cat_list = [0] * (len(self.coders) + 6)
            for c in self.codes:
                if effective_catid(c) == cat['catid']:
                    for i in range(2, len(c['display_list'])):
                        cat_list[i] += c['display_list'][i]
            cat_list = cat_list[2:]
            for count in cat_list:
                cat['display_list'].append(count)

        # find leaf categories, add to above categories, and gradually remove leaves
        # until only top categories are left
        sub_cats = copy(self.categories)
        counter = 0
        while len(sub_cats) > 0 or counter < 10000:
            leaf_list = []
            branch_list = []
            for c in sub_cats:
                for c2 in sub_cats:
                    if c['catid'] == c2['supercatid']:
                        branch_list.append(c)
            for cat in sub_cats:
                if cat not in branch_list:
                    leaf_list.append(cat)
            # Add totals for each coder and overall total to higher category
            for leaf_cat in leaf_list:
                for cat in self.categories:
                    if cat['catid'] == leaf_cat['supercatid']:
                        for i in range(2, len(cat['display_list'])):
                            cat['display_list'][i] += leaf_cat['display_list'][i]
                sub_cats.remove(leaf_cat)
            counter += 1

        header = ["Code Tree", "Id"]
        for coder in self.coders:
            header.append(coder)
        header.append("Total")

    def sort_by_totals(self, ):
        """ Sort by totals descending. """

        self.get_data()
        self.calculate_code_frequencies()
        self.categories = sorted(self.categories, key=lambda i: (i['display_list'][-1]), reverse=True)
        self.codes = sorted(self.codes, key=lambda i: (i['display_list'][-1]), reverse=True)
        self.fill_tree()

    def sort_by_alphabet(self, ):
        """ Sort alphabetically ascending. """

        self.get_data()
        self.calculate_code_frequencies()
        self.categories = sorted(self.categories, key=lambda i: (i['display_list'][0]))
        self.codes = sorted(self.codes, key=lambda i: (i['display_list'][0]))
        self.fill_tree()

    def _on_project_data_changed(self, tables, source):
        """Handle project change events from other dialogs.

        Args:
            tables: Changed database table names.
            source: Event emitter, ignored when it is this dialog.
        """

        if source is self or not isinstance(tables, list):
            return
        tables = set(tables)
        watched_tables = {"code_cat", "code_name", "code_text", "code_av", "code_image"}
        if watched_tables.isdisjoint(tables):
            return
        if self.ui.radioButton_2.isChecked():
            self.sort_by_totals()
            return
        self.sort_by_alphabet()

    def depthgauge(self, item):
        """ Get depth for treewidget item. """

        depth = 0
        while item.parent() is not None:
            item = item.parent()
            depth += 1
        return depth

    def export_text_file(self):
        """ Export coding frequencies to a structured text file, mirroring the tree view:
        full hierarchy (indented), full names, and exactly the columns currently visible
        (breakdown, coders, total, ids). """

        filename = "Code_frequencies.txt"
        e = ExportDirectoryPathDialog(self.app, filename)
        filepath = e.filepath
        if filepath is None:
            return
        text_ = _("Code frequencies") + "\n"
        text_ += _("Project: ") + f"{self.app.project_name}\n"
        text_ += _("Date: ") + datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        if self.file_ids:
            text_ += _("Files: ") + f"{len(self.file_ids)} " + _("selected") + "\n"
        else:
            text_ += _("Files: ") + _("All files") + "\n"
        text_ += "\n"
        headers = self.ui.treeWidget.headerItem()
        visible_cols = [i for i in range(1, self.ui.treeWidget.columnCount())
                        if not self.ui.treeWidget.isColumnHidden(i)]
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        while item:
            prefix = "  " * self.depthgauge(item)
            is_cat = item.text(1).split(':')[0] == "catid"
            label = _("Category: ") if is_cat else _("Code: ")
            # Tooltip holds the full (untruncated) name
            name = item.toolTip(0) if item.toolTip(0) != "" else item.text(0)
            line = f"{prefix}{label}{name}"
            for i in visible_cols:
                line += f" | {headers.text(i)}: {item.text(i)}"
            text_ += line + "\n"
            it += 1
            item = it.value()
        with open(filepath, 'w', encoding='utf-8-sig') as file_:
            file_.write(text_)
        msg = _("Coding frequencies text file exported to: ") + filepath
        Message(self.app, _('Text file Export'), msg).exec()
        self.parent_textEdit.append(msg)

    def export_excel_file(self):
        """ Export to a three-sheet Excel workbook, mirroring the current view (only the
        visible columns are exported):
        1. Report  - the indented hierarchy, exactly as displayed.
        2. Outline - the same report with native Excel row grouping, so each branch can
                     be collapsed/expanded with the +/- margin buttons.
        3. Data    - a flat 'tidy' sheet (one column per level, plus type and depth),
                     suited to pivot tables, filters and formulas. """

        headers_item = self.ui.treeWidget.headerItem()
        visible_cols = [i for i in range(1, self.ui.treeWidget.columnCount())
                        if not self.ui.treeWidget.isColumnHidden(i)]
        # Walk the tree once, collecting every row with its depth, ancestry and values
        rows = []
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        while item:
            is_code = item.text(1)[:3] == "cid"
            # Tooltip holds the full (untruncated) name
            name = item.toolTip(0) if item.toolTip(0) != "" else item.text(0)
            color = None
            if is_code:
                cid = int(item.text(1)[4:])
                for code_ in self.codes:
                    if cid == code_['cid']:
                        color = code_['color']
                        break
            ancestors = []  # (name, is_code) pairs, topmost first
            parent = item.parent()
            while parent is not None:
                p_name = parent.toolTip(0) if parent.toolTip(0) != "" else parent.text(0)
                ancestors.insert(0, (p_name, parent.text(1)[:3] == "cid"))
                parent = parent.parent()
            chain = ancestors + [(name, is_code)]
            cat_chain = [n for n, code_flag in chain if not code_flag]
            code_chain = [n for n, code_flag in chain if code_flag]
            values = []
            for col in visible_cols:
                try:
                    values.append(int(item.text(col)))
                except ValueError:
                    values.append(item.text(col))
            rows.append({'depth': len(ancestors), 'name': name, 'is_code': is_code,
                         'color': color, 'cat_chain': cat_chain, 'code_chain': code_chain,
                         'values': values})
            it += 1
            item = it.value()

        date_str = datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S")
        if self.file_ids:
            files_txt = f"{len(self.file_ids)} " + _("selected files")
        else:
            files_txt = _("All files")
        header = [_("Code Tree")] + [headers_item.text(i) for i in visible_cols]
        wb = openpyxl.Workbook()

        def write_report_sheet(ws, outline):
            """ Indented report; with outline=True adds native Excel row grouping. """
            ws.cell(column=1, row=1,
                    value=_("Code frequencies") + " - " + self.app.project_name).font = Font(b=True, size=12)
            ws.cell(column=1, row=2, value=_("Date: ") + date_str)
            ws.cell(column=1, row=3, value=_("Files: ") + files_txt)
            header_row = 5
            for col, head in enumerate(header):
                ws.cell(column=col + 1, row=header_row, value=head).font = Font(b=True)
            ws.freeze_panes = f"A{header_row + 1}"
            if outline:
                # +/- buttons at the parent (summary) row, above each branch
                ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)
            row_n = header_row
            for r in rows:
                row_n += 1
                # Native cell indentation: the hierarchy stays visible but the cell value
                # is the clean name, without padding spaces
                name_cell = ws.cell(column=1, row=row_n, value=r['name'])
                if r['depth'] > 0:
                    name_cell.alignment = Alignment(horizontal='left', indent=r['depth'])
                if r['is_code'] and r['color']:
                    name_cell.fill = PatternFill(start_color=r['color'][1:],
                                                 end_color=r['color'][1:], fill_type="solid")
                    name_cell.font = Font(color=TextColor(r['color']).recommendation[1:])
                elif not r['is_code']:
                    name_cell.font = Font(b=True)
                for out_col, value in enumerate(r['values'], start=2):
                    ws.cell(column=out_col, row=row_n, value=value)
                if outline and r['depth'] > 0:
                    # Excel supports at most 7 grouping levels
                    ws.row_dimensions[row_n].outline_level = min(r['depth'], 7)
            ws.column_dimensions['A'].width = 45

        ws1 = wb.active
        ws1.title = _("Report")[:31]
        write_report_sheet(ws1, outline=False)
        ws2 = wb.create_sheet(_("Outline")[:31])
        write_report_sheet(ws2, outline=True)

        # Tidy data sheet: only code and sub-code rows (category rows are aggregates and
        # would add blank cells and double counting when pivoting). Each row repeats its
        # ancestry, split into typed columns: the category chain (Category level 1..N,
        # topmost first) and the code chain (Code level 1, then Sub-code level 2..N).
        data_rows = [r for r in rows if r['is_code']]
        max_cat = max((len(r['cat_chain']) for r in data_rows), default=0)
        max_code = max((len(r['code_chain']) for r in data_rows), default=0)
        ws3 = wb.create_sheet(_("Data")[:31])
        tidy_header = [_("Category level") + f" {i + 1}" for i in range(max_cat)]
        for i in range(max_code):
            tidy_header.append(_("Code level 1") if i == 0 else _("Sub-code level") + f" {i + 1}")
        tidy_header += [_("Depth")] + [headers_item.text(i) for i in visible_cols]
        for col, head in enumerate(tidy_header):
            ws3.cell(column=col + 1, row=1, value=head).font = Font(b=True)
        ws3.freeze_panes = "A2"
        for row_i, r in enumerate(data_rows, start=2):
            for col, level_name in enumerate(r['cat_chain']):
                ws3.cell(column=col + 1, row=row_i, value=level_name)
            for j, level_name in enumerate(r['code_chain']):
                ws3.cell(column=max_cat + 1 + j, row=row_i, value=level_name)
            base = max_cat + max_code + 1  # 1-based column of Depth
            ws3.cell(column=base, row=row_i, value=r['depth'] + 1)
            for off, value in enumerate(r['values']):
                ws3.cell(column=base + 1 + off, row=row_i, value=value)
        for i in range(max_cat + max_code):
            ws3.column_dimensions[get_column_letter(i + 1)].width = 28

        filename = "Code_frequencies.xlsx"
        export_dir = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_dir.filepath
        if filepath is None:
            return
        wb.save(filepath)
        msg = _("Coding frequencies exported to: ") + filepath
        Message(self.app, _('File export'), msg).exec()
        self.parent_textEdit.append(msg)

    def get_collapsed(self, item):
        """ On category collapse or expansion signal, find the collapsed parent category items.
        This will fill the self.app.collapsed_categories and is the expanded/collapsed tree is then replicated across
        other areas of the app. """

        if item.text(1)[:3] == "cid":
            return
        if not item.isExpanded() and item.text(1) not in self.app.collapsed_categories:
            self.app.collapsed_categories.append(item.text(1))
        if item.isExpanded() and item.text(1) in self.app.collapsed_categories:
            self.app.collapsed_categories.remove(item.text(1))

    def _nest_subcodes_in_tree(self):
        """ Re-parent code tree items so sub-codes (supercid) nest under their parent
        code. Runs after fill_tree has placed every code. Preserves item flags,
        checkboxes, colour and count because the existing item is moved, not rebuilt.
        No-op for projects without sub-codes. """
        tree = getattr(getattr(self, 'ui', None), 'treeWidget', None) or getattr(self, 'tree', None)
        if tree is None:
            return
        code_list = getattr(self, 'code_names', None)
        if code_list is None:
            code_list = getattr(self, 'codes', [])
        supercid_of = {c['cid']: c.get('supercid') for c in code_list}
        if not any(supercid_of.values()):
            return
        guard = 0
        moved = True
        while moved and guard < 10000:
            moved = False
            guard += 1
            cid_item = {}
            it = QtWidgets.QTreeWidgetItemIterator(tree)
            while it.value():
                node = it.value()
                t = node.text(1)
                if t.startswith('cid:'):
                    try:
                        cid_item[int(t[4:])] = node
                    except ValueError:
                        pass
                it += 1
            for cid_, node in cid_item.items():
                sup = supercid_of.get(cid_)
                if sup is None:
                    continue
                parent_node = cid_item.get(sup)
                if parent_node is None or node.parent() is parent_node:
                    continue
                cur_parent = node.parent()
                if cur_parent is None:
                    idx = tree.indexOfTopLevelItem(node)
                    taken = tree.takeTopLevelItem(idx)
                else:
                    taken = cur_parent.takeChild(cur_parent.indexOfChild(node))
                parent_node.addChild(taken)
                parent_node.setExpanded(True)  # show the nested sub-code from the start <- L
                taken.setExpanded(True)
                moved = True
                break

    def fill_tree(self):
        """ Fill tree widget, top level items are main categories and unlinked codes. """

        self.ui.treeWidget.clear()
        cats = copy(self.categories)
        codes = copy(self.codes)
        self.ui.treeWidget.clear()
        header = [_("Code Tree"), "Id", _("Text"), _("Image"), "A/V"]
        for coder in self.coders:
            header.append(coder)
        header.append("Total")
        self.ui.treeWidget.setColumnCount(len(header))
        self.ui.treeWidget.setHeaderLabels(header)
        if not self.app.settings['showids']:
            self.ui.treeWidget.setColumnHidden(1, True)
        else:
            self.ui.treeWidget.setColumnHidden(1, False)
        # The per-source columns (2, 3, 4) are shown only when the breakdown checkbox is ticked
        breakdown = self.ui.checkBox_source_breakdown.isChecked()
        for col in (2, 3, 4):
            self.ui.treeWidget.setColumnHidden(col, not breakdown)
        # Add top level categories
        remove_list = []
        for c in cats:
            if c['supercatid'] is None:
                display_list = []
                for i in c['display_list']:
                    display_list.append(str(i))
                if self.truncated_code_names and len(display_list[0]) > 62:
                    display_list[0] = display_list[0][:30] + '..' + display_list[0][-30:]
                    self.contains_long_names = True
                top_item = QtWidgets.QTreeWidgetItem(display_list)
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                if f"catid:{c['catid']}" in self.app.collapsed_categories:
                    top_item.setExpanded(False)
                else:
                    top_item.setExpanded(True)
                remove_list.append(c)
        for item in remove_list:
            cats.remove(item)
        ''' Add child categories. Look at each unmatched category, iterate through tree to
        add as child then remove matched categories from the list. '''
        count = 0
        while not (len(cats) < 1 or count > 10000):
            remove_list = []
            for c in cats:
                it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
                item = it.value()
                while item:  # while there is an item in the list
                    if item.text(1) == f'catid:{c["supercatid"]}':
                        display_list = []
                        for i in c['display_list']:
                            display_list.append(str(i))
                        if self.truncated_code_names and  len(display_list[0]) > 62:
                            display_list[0] = display_list[0][:30] + '..' + display_list[0][-30:]
                            self.contains_long_names = True
                        child = QtWidgets.QTreeWidgetItem(display_list)
                        child.setToolTip(0, c['name'])
                        item.addChild(child)
                        if f"catid:{c['catid']}" in self.app.collapsed_categories:
                            child.setExpanded(False)
                        else:
                            child.setExpanded(True)
                        remove_list.append(c)
                    it += 1
                    item = it.value()
            for item in remove_list:
                cats.remove(item)
            count += 1

        # Add unlinked codes as top level items
        remove_items = []
        for c in codes:
            if c['catid'] is None:
                display_list = []
                for i in c['display_list']:
                    display_list.append(str(i))
                if self.truncated_code_names and len(display_list[0]) > 62:
                    display_list[0] = f"{display_list[0][:30]}..{display_list[0][-30:]}"
                    self.contains_long_names = True
                top_item = QtWidgets.QTreeWidgetItem(display_list)
                top_item.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                color = TextColor(c['color']).recommendation
                top_item.setForeground(0, QBrush(QtGui.QColor(color)))
                top_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_items.append(c)
        for item in remove_items:
            codes.remove(item)

        # Add codes as children
        for c in codes:
            it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
            item = it.value()
            while item:
                if item.text(1) == f'catid:{c["catid"]}':
                    display_list = []
                    for i in c['display_list']:
                        display_list.append(str(i))
                    if self.truncated_code_names and len(display_list[0]) > 62:
                        display_list[0] = f"{display_list[0][:30]}..{display_list[0][-30:]}"
                        self.contains_long_names = True
                    child = QtWidgets.QTreeWidgetItem(display_list)
                    child.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                    color = TextColor(c['color']).recommendation
                    child.setForeground(0, QBrush(QtGui.QColor(color)))
                    child.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                    child.setToolTip(0, c['name'])
                    item.addChild(child)
                it += 1
                item = it.value()
        self._nest_subcodes_in_tree()
        self.ui.treeWidget.sortByColumn(0, QtCore.Qt.SortOrder.AscendingOrder)
        restore_persistent_tree_widths(self.ui.treeWidget)
        # Show the full hierarchy expanded (categories, sub-categories, codes and
        # sub-codes), without altering the app-wide collapsed-categories memory.
        self.ui.treeWidget.blockSignals(True)
        self.ui.treeWidget.expandAll()
        self.ui.treeWidget.blockSignals(False)

    def tree_menu(self, position):
        menu = QtWidgets.QMenu()
        menu.setStyleSheet(f"QMenu {{font-size:{self.app.settings['fontsize']}pt}} ")
        selected = self.ui.treeWidget.currentItem()
        action_expand_names = None
        if self.contains_long_names:
            action_expand_names = menu.addAction(_("Expand names"))
        action_truncate_names = None
        if self.contains_long_names and self.truncated_code_names is False:
            action_truncate_names = menu.addAction(_("Truncate names"))
        action_show_coded_media = None
        if selected is not None and selected.text(1)[0:3] == 'cid':
            action_show_coded_media = menu.addAction(_("Show coded files"))
        action_expand_collapse = None
        action_cat_show_coded_files = None
        if selected is not None and selected.text(1)[0:3] == 'cat':
            action_cat_show_coded_files = menu.addAction(_("Show coded files"))
            action_expand_collapse = menu.addAction(_("Expand or collapse branch"))
        action = menu.exec(self.ui.treeWidget.mapToGlobal(position))
        if action is None:
            return
        if action == action_show_coded_media:
            found_code = None
            tofind = int(selected.text(1)[4:])
            for code in self.codes:
                if code['cid'] == tofind:
                    found_code = code
                    break
            if found_code:
                DialogCodeInAllFiles(self.app, found_code)
            return
        if action == action_cat_show_coded_files:
            branch_codes = self.recursive_get_branch_codes(selected, [])
            DialogCodeInAllFiles(self.app, branch_codes, "File", selected.text(0))
            return
        if action == action_expand_names:
            self.truncated_code_names = False
            self.fill_tree()
            return
        if action == action_truncate_names:
            self.truncated_code_names = True
            self.fill_tree()
            return
        if action == action_expand_collapse:
            expand_toggle = not selected.isExpanded()
            self.recursive_expand_collapse_branch(selected, expand_toggle)
            return

    def recursive_get_branch_codes(self, item, branch_codes):
        """ Set all children of this item to be expanded or collapsed.
        Recurse through all child categories. """

        child_count = item.childCount()
        for i in range(child_count):
            if item.child(i).text(1)[0:3] == "cid":
                cid = int(item.child(i).text(1)[4:])
                for code_ in self.codes:
                    if cid == code_['cid']:
                        branch_codes.append(code_)
                        break
            if item.child(i).text(1)[0:3] == "cat":
                self.recursive_get_branch_codes(item.child(i), branch_codes)
        return branch_codes

    def recursive_expand_collapse_branch(self, item, expand_toggle):
        """ Set all children of this item to be expanded or collapsed.
        Recurse through all child categories. """

        child_count = item.childCount()
        for i in range(child_count):
            item.setExpanded(expand_toggle)
            self.recursive_expand_collapse_branch(item.child(i), expand_toggle)


class DialogReportCoderComparisons(QtWidgets.QDialog):
    """ Compare coded text sequences between coders using Cohen's Kappa. """

    def __init__(self, app, parent_textedit):

        self.app = app
        self.parent_textEdit = parent_textedit
        self.text_data = ""
        self.excel_data = []
        self.coders = []
        self.selected_coders = []
        self.file_summaries = []
        self.codes = []
        self.categories = []
        self.truncated_code_names = True
        self.contains_long_names = False
        self.get_data()
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_reportComparisons()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        self.ui.pushButton_run.setEnabled(False)
        self.ui.pushButton_run.pressed.connect(self.calculate_statistics)
        self.ui.pushButton_run.setIcon(qta.icon('mdi6.play', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_clear.pressed.connect(self.clear_selection)
        self.ui.pushButton_clear.setIcon(qta.icon('mdi6.refresh', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export.pressed.connect(self.export_excel)  # export_text_file)
        self.ui.pushButton_export.setToolTip(_("Export Excel"))
        self.ui.pushButton_export.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_help1.setIcon(qta.icon('mdi6.help'))
        self.ui.pushButton_help1.pressed.connect(self.information)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        font = f'font: {self.app.settings["treefontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.treeWidget.setStyleSheet(font)
        self.ui.treeWidget.setSelectionMode(QtWidgets.QTreeWidget.SelectionMode.ExtendedSelection)
        self.ui.treeWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.treeWidget.customContextMenuRequested.connect(self.tree_menu)
        init_persistent_tree_header(self.ui.treeWidget, self.app, 'dialogreportcodercomparisons_tree_widths')
        self.ui.comboBox_coders.insertItems(0, self.coders)
        self.ui.comboBox_coders.currentTextChanged.connect(self.coder_selected)
        if len(self.coders) == 3:  # includes empty slot
            self.ui.comboBox_coders.setCurrentIndex(1)
            self.ui.comboBox_coders.setCurrentIndex(2)
        self.fill_tree()
        # These signals after the tree is filled the first time
        self.ui.treeWidget.itemCollapsed.connect(self.get_collapsed)
        self.ui.treeWidget.itemExpanded.connect(self.get_collapsed)
        self.app.project_events.project_data_changed.connect(self._on_project_data_changed)

    def get_data(self):
        """ Called from init. gets coders, codes, categories, file_summaries.
        Images and A/V files are not used. """

        self.codes, self.categories = self.app.get_codes_categories()
        cur = self.app.conn.cursor()
        cur.execute("select id, length(fulltext) from source where fulltext is not null")
        self.file_summaries = cur.fetchall()
        sql = "select owner from  code_image union select owner from code_text union select owner from code_av"
        cur.execute(sql)
        result = cur.fetchall()
        self.coders = [""]
        for row in result:
            self.coders.append(row[0])

    def _on_project_data_changed(self, tables, source):
        """Handle project change events from other dialogs.

        Args:
            tables: Changed database table names.
            source: Event emitter, ignored when it is this dialog.
        """

        if source is self or not isinstance(tables, list):
            return
        tables = set(tables)
        watched_tables = {"code_cat", "code_name", "code_text"}
        if watched_tables.isdisjoint(tables):
            return
        if "code_text" in tables:
            self.get_data()
        else:
            self.codes, self.categories = self.app.get_codes_categories()
        self.fill_tree()

    def coder_selected(self):
        """ Select coders for comparison - only two coders can be selected. """

        coder = self.ui.comboBox_coders.currentText()
        if coder == "":
            return
        if len(self.selected_coders) == 0:
            self.selected_coders.append(coder)
            self.ui.label_selections.setText(coder)
        if len(self.selected_coders) == 1 and self.selected_coders[0] != coder:
            self.selected_coders.append(coder)
            coder1 = self.ui.label_selections.text()
            self.ui.label_selections.setText(f"{coder1} , {coder}")
        if len(self.selected_coders) == 2:
            self.ui.pushButton_run.setEnabled(True)

    def clear_selection(self):
        """ Clear the coder selection and tree widget statistics.
        text(1) Catid/cid, text(2) Agree%, text(3) A and B %, text(4) Not A Not B %
        text (5) Disagree %, text(6) AgreeCodedOnly%, text(7) Kappa
        """

        self.selected_coders = []
        self.ui.pushButton_run.setEnabled(False)
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        while item:  # while there is an item in the list
            if item.text(1)[0:4] == 'cid:':
                item.setText(2, "")
                item.setText(3, "")
                item.setText(4, "")
                item.setText(5, "")
                item.setText(6, "")
                item.setText(7, "")
            it += 1
            item = it.value()
        self.ui.label_selections.setText(_("No coders selected"))

    def export_excel(self):
        """ Export to Excel. """

        filename = "Coder_comparison.xlsx"
        export_path = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_path.filepath
        if filepath is None:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(column=1, row=1, value=f"Coder Comparison: {self.selected_coders[0]}, {self.selected_coders[1]}")
        headings = ["Code tree", "Agree %", "A and B %", "Not A Not B %", "Disagree %", "Agree coded only %", "Kappa"]
        for col, heading in enumerate(headings):
            ws.cell(column=col + 1, row=2, value=heading)
            ws.cell(column=col + 1, row=2).font = Font(b=True)
        for row in range(len(self.excel_data)):
            for col in range(7):
                ws.cell(column=col + 1, row=3 + row, value=self.excel_data[row][col])
            if self.excel_data[row][7] != "Category":
                pf = PatternFill(start_color=self.excel_data[row][7], end_color=self.excel_data[row][7], fill_type="solid")
                ws.cell(column=1, row=3 + row).fill = pf

        wb.save(filepath)
        msg = _('Coder comparisons report exported: ') + filepath
        Message(self.app, _('Report exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    '''def export_text_file(self):
        """ OLD Export coding comparison statistics to text file. """

        filename = "Coder_comparison.txt"
        export_path = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_path.filepath
        if filepath is None:
            return
        with open(filepath, 'w', encoding="'utf-8-sig'") as file_:
            file_.write(f"{self.app.project_name}\n")
            file_.write(_("Date: ") + datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S") + "\n")
            file_.write(self.text_data)
        msg = _("Coder comparison text file exported to: ") + filepath
        Message(self.app, _('Text file export'), msg, "information").exec()
        self.parent_textEdit.append(msg)'''

    def calculate_statistics(self):
        """ Iterate through tree widget, for all cids
        For each code calculate the two-coder comparison statistics. """

        self.text_data = "====" + _("CODER COMPARISON") + "====\n" + _("Selected coders: ")
        self.text_data += self.selected_coders[0] + ", " + self.selected_coders[1] + "\n"
        self.excel_data = []
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        row = 0
        while item:
            if item.text(1)[0:4] == 'cid:':
                excel_row = []
                agreement = self.calculate_agreement_for_code(int(item.text(1)[4:]))
                item.setText(2, f"{agreement['agreement']}%")
                item.setText(3, f"{agreement['dual_percent']}%")
                item.setText(4, f"{agreement['uncoded_percent']}%")
                item.setText(5, f"{agreement['disagreement']}%")
                item.setText(6, f"{agreement['agree_coded_only']}%")
                item.setText(7, f"{agreement['kappa']}")
                self.text_data += f"\n{item.text(0)} ({item.text(1)})\n"
                excel_row.append(item.text(0))
                self.text_data += _("agreement: ") + f"{agreement['agreement']}%"
                excel_row.append(f"{agreement['agreement']}%")
                self.text_data += _(", dual coded: ") + f"{agreement['dual_percent']}%"
                excel_row.append(f"{agreement['dual_percent']}%")
                self.text_data += _(", uncoded: ") + f"{agreement['uncoded_percent']}%"
                excel_row.append(f"{agreement['uncoded_percent']}%")
                self.text_data += _(", disagreement: ") + f"{agreement['disagreement']}%"
                excel_row.append(f"{agreement['disagreement']}%")
                self.text_data += _(", agree coded only: ") + f"{agreement['agree_coded_only']}%"
                excel_row.append(f"{agreement['agree_coded_only']}%")
                self.text_data += f", Kappa: {agreement['kappa']}"
                excel_row.append(agreement['kappa'])
                # Fix codename, add color
                cid = int(item.text(1)[4:])
                for code_ in self.codes:
                    if cid == code_['cid']:
                        excel_row.append(code_['color'][1:])
                self.excel_data.append(excel_row)
            else:  # Category
                self.excel_data.append([item.text(0), "", "", "", "", "", "", "Category"])
            it += 1
            item = it.value()
            row += 1

    def calculate_agreement_for_code(self, cid):
        """ Calculate the two-coder statistics for this cid
        Percentage agreement.
        Get the start and end positions in all files (source table) for this cid.

        self.file_summaries item [0] = id, [1] = full text length
        Look at each file separately to ge the commonly coded text.
        Each character that is coded by coder 1 or coder 2 is incremented, resulting in a list of 0, 1, 2
        where 0 is no codings at all, 1 is coded by only one coder and 2 is coded by both coders.
        'Disagree%':'','A not B':'','B not A':'', coded only:'' ,'K':''

        param:
            cid : integer source file id

        """

        # coded0 and coded1 are the total characters coded by coder 0 and coder 1
        total = {'dual_coded': 0, 'single_coded': 0, 'uncoded': 0, 'characters': 0, 'coded0': 0, 'coded1': 0,
                 'agree_coded_only': 0.0, 'agreement': 0.0, 'disagreement': 0.0, 'uncoded_percent': 0.0}
        # Loop through each source file
        cur = self.app.conn.cursor()
        sql = "select pos0,pos1,fid from code_text where fid=? and cid=? and owner=?"
        for f in self.file_summaries:
            cur.execute(sql, [f[0], cid, self.selected_coders[0]])
            result0 = cur.fetchall()
            cur.execute(sql, [f[0], cid, self.selected_coders[1]])
            result1 = cur.fetchall()
            # Determine the same characters coded by both coders, by adding 1 to each coded character
            char_list = [0] * f[1]
            for coded in result0:
                for char in range(coded[0], coded[1]):
                    try:
                        char_list[char] += 1
                        total['coded0'] += 1
                    except IndexError as e_:
                        msg = "DialogReportCoderComparisons.calculate_agreement_for_code "
                        msg += f"{e_} fid:{f[0]} len_text:{f[1]} pos1:{coded[1]}"
                        msg += f" cid:{cid} coder:{self.selected_coders[0]}"
                        print(msg)
                        logger.error(msg)
                        self.parent_textEdit.append(msg)
            for coded in result1:
                for char in range(coded[0], coded[1]):
                    try:
                        char_list[char] += 1
                        total['coded1'] += 1
                    except IndexError as e_:
                        msg = "DialogReportCoderComparisons.calculate_agreement_for_code "
                        msg += f"{e_} fid:{f[0]} len_text:{f[1]} pos1:{coded[1]}"
                        msg += " cid:" + str(cid) + " coder:" + self.selected_coders[0]
                        print(msg)
                        logger.error(msg)
                        self.parent_textEdit.append(msg)
            uncoded = 0
            single_coded = 0
            dual_coded = 0
            for char in char_list:
                if char == 0:
                    uncoded += 1
                if char == 1:
                    single_coded += 1
                if char == 2:
                    dual_coded += 1
            total['dual_coded'] += dual_coded
            total['single_coded'] += single_coded
            total['uncoded'] += uncoded
            total['characters'] += f[1]

        if total['characters'] != 0:
            total['agreement'] = round(100 * (total['dual_coded'] + total['uncoded']) / total['characters'], 2)
            total['dual_percent'] = round(100 * total['dual_coded'] / total['characters'], 2)
            total['uncoded_percent'] = round(100 * total['uncoded'] / total['characters'], 2)
            total['disagreement'] = round(100 - total['agreement'], 2)
            try:
                total['agree_coded_only'] = round(100 * total['dual_coded'] / (total['dual_coded'] + total['single_coded']), 2)
            except ZeroDivisionError:
                total['agree_coded_only'] = "zero div"
        else:
            total['agreement'] = "zero div"
            total['dual_percent'] = "zero div"
            total['uncoded_percent'] = "zero div"
            total['disagreement'] = "zero div"
            total['agree_coded_only'] = "zero div"
        # Cohen's Kappa
        '''
        https://en.wikipedia.org/wiki/Cohen%27s_kappa

        k = Po - Pe     Po is proportionate agreement (both coders coded this text / all coded text))
            -------     Pe is probability of random agreement
            1  - Pe

            Pe = Pyes + Pno
            Pyes = proportion Yes by A multiplied by proportion Yes by B
                 = total['coded0']/total_coded * total['coded1]/total_coded

            Pno = proportion No by A multiplied by proportion No by B
                = (total_coded - total['coded0']) / total_coded * (total_coded - total['coded1]) / total_coded

        IMMEDIATE BELOW IS INCORRECT - RESULTS IN THE TOTAL AGREEMENT SCORE
        Po = total['agreement'] / 100
        Pyes = total['coded0'] / total['characters'] * total['coded1'] / total['characters']
        Pno = (total['characters'] - total['coded0']) / total['characters'] * (total['characters'] - total['coded1']) / 
            total['characters']

        BELOW IS BETTER - ONLY LOOKS AT PROPORTIONS OF CODED CHARACTERS
        NEED TO CONFIRM THIS IS THE CORRECT APPROACH
        '''
        total['kappa'] = "zerodiv"
        try:
            unique_codings = total['coded0'] + total['coded1'] - total['dual_coded']
            Po = total['dual_coded'] / unique_codings
            Pyes = total['coded0'] / unique_codings * total['coded1'] / unique_codings
            Pno = (unique_codings - total['coded0']) / unique_codings * (
                        unique_codings - total['coded1']) / unique_codings
            Pe = Pyes * Pno
            kappa = round((Po - Pe) / (1 - Pe), 4)
            total['kappa'] = kappa
        except ZeroDivisionError:
            pass
        return total

    def get_collapsed(self, item):
        """ On category collapse or expansion signal, find the collapsed parent category items.
        This will fill the self.app.collapsed_categories and is the expanded/collapsed tree is then replicated across
        other areas of the app. """

        if item.text(1)[:3] == "cid":
            return
        if not item.isExpanded() and item.text(1) not in self.app.collapsed_categories:
            self.app.collapsed_categories.append(item.text(1))
        if item.isExpanded() and item.text(1) in self.app.collapsed_categories:
            self.app.collapsed_categories.remove(item.text(1))

    def _nest_subcodes_in_tree(self):
        """ Re-parent code tree items so sub-codes (supercid) nest under their parent
        code. Runs after fill_tree has placed every code. Preserves item flags,
        checkboxes, colour and count because the existing item is moved, not rebuilt.
        No-op for projects without sub-codes. """
        tree = getattr(getattr(self, 'ui', None), 'treeWidget', None) or getattr(self, 'tree', None)
        if tree is None:
            return
        code_list = getattr(self, 'code_names', None)
        if code_list is None:
            code_list = getattr(self, 'codes', [])
        supercid_of = {c['cid']: c.get('supercid') for c in code_list}
        if not any(supercid_of.values()):
            return
        guard = 0
        moved = True
        while moved and guard < 10000:
            moved = False
            guard += 1
            cid_item = {}
            it = QtWidgets.QTreeWidgetItemIterator(tree)
            while it.value():
                node = it.value()
                t = node.text(1)
                if t.startswith('cid:'):
                    try:
                        cid_item[int(t[4:])] = node
                    except ValueError:
                        pass
                it += 1
            for cid_, node in cid_item.items():
                sup = supercid_of.get(cid_)
                if sup is None:
                    continue
                parent_node = cid_item.get(sup)
                if parent_node is None or node.parent() is parent_node:
                    continue
                cur_parent = node.parent()
                if cur_parent is None:
                    idx = tree.indexOfTopLevelItem(node)
                    taken = tree.takeTopLevelItem(idx)
                else:
                    taken = cur_parent.takeChild(cur_parent.indexOfChild(node))
                parent_node.addChild(taken)
                parent_node.setExpanded(True)  # show the nested sub-code from the start <- L
                taken.setExpanded(True)
                moved = True
                break

    def fill_tree(self):
        """ Fill tree widget, top level items are main categories and unlinked codes. """

        cats = copy(self.categories)
        codes = copy(self.codes)
        self.ui.treeWidget.clear()
        self.ui.treeWidget.setColumnCount(7)
        self.ui.treeWidget.setHeaderLabels(
            [_("Code Tree"), "Id", "Agree %", "A and B %", "Not A Not B %", "Disagree %", "Agree coded only %", "Kappa"])
        self.ui.treeWidget.hideColumn(1)
        if self.app.settings['showids']:
            self.ui.treeWidget.showColumn(1)
        # Add top level categories
        remove_list = []
        for c in cats:
            if c['supercatid'] is None:
                top_item = QtWidgets.QTreeWidgetItem([c['name'], f'catid:{c["catid"]}'])
                if self.truncated_code_names and len(c['name']) > 62:
                    top_item.setText(0, f"{c['name'][:30]}..{c['name'][-30:]}")
                    self.contains_long_names = True
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                if f"catid:{c['catid']}" in self.app.collapsed_categories:
                    top_item.setExpanded(False)
                else:
                    top_item.setExpanded(True)
                remove_list.append(c)
        for item in remove_list:
            cats.remove(item)
        ''' Add child categories. Look at each unmatched category, iterate through tree to
        add as child then remove matched categories from the list. '''
        count = 0
        while len(cats) > 0 or count < 10000:
            remove_list = []
            for c in cats:
                it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
                item = it.value()
                while item:  # while there is an item in the list
                    if item.text(1) == f'catid:{c["supercatid"]}':
                        child = QtWidgets.QTreeWidgetItem([c['name'], f'catid:{c["catid"]}'])
                        if self.truncated_code_names and len(c['name']) > 62:
                            child.setText(0, f"{c['name'][:30]}..{c['name'][-30:]}")
                            self.contains_long_names = True
                        child.setToolTip(0, c['name'])
                        item.addChild(child)
                        if f"catid:{c['catid']}" in self.app.collapsed_categories:
                            child.setExpanded(False)
                        else:
                            child.setExpanded(True)
                        remove_list.append(c)
                    it += 1
                    item = it.value()
            for item in remove_list:
                cats.remove(item)
            count += 1
        # Add unlinked codes as top level items
        remove_items = []
        for c in codes:
            if c['catid'] is None:
                top_item = QtWidgets.QTreeWidgetItem([c['name'], f'cid:{c["cid"]}'])
                if self.truncated_code_names and len(c['name']) > 62:
                    top_item.setText(0, f"{c['name'][:30]}..{c['name'][-30:]}")
                    self.contains_long_names = True
                top_item.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                color = TextColor(c['color']).recommendation
                top_item.setForeground(0, QBrush(QtGui.QColor(color)))
                top_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_items.append(c)
        for item in remove_items:
            codes.remove(item)

        # Add codes as children
        for c in codes:
            it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
            item = it.value()
            while item:
                if item.text(1) == f'catid:{c["catid"]}':
                    child = QtWidgets.QTreeWidgetItem([c['name'], f'cid:{c["cid"]}'])
                    if self.truncated_code_names and len(c['name']) > 62:
                        child.setText(0, f"{c['name'][:30]}..{c['name'][-30:]}")
                        self.contains_long_names = True
                    child.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                    color = TextColor(c['color']).recommendation
                    child.setForeground(0, QBrush(QtGui.QColor(color)))
                    child.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                    child.setToolTip(0, c['name'])
                    item.addChild(child)
                it += 1
                item = it.value()
        self._nest_subcodes_in_tree()
        self.ui.treeWidget.sortByColumn(0, QtCore.Qt.SortOrder.AscendingOrder)
        restore_persistent_tree_widths(self.ui.treeWidget)

    def tree_menu(self, position):
        menu = QtWidgets.QMenu()
        menu.setStyleSheet(f"QMenu {{font-size:{self.app.settings['fontsize']}pt}} ")
        selected = self.ui.treeWidget.currentItem()
        action_expand_names = None
        if self.contains_long_names:
            action_expand_names = menu.addAction(_("Expand names"))
        action_truncate_names = None
        if self.contains_long_names and self.truncated_code_names is False:
            action_truncate_names = menu.addAction(_("Truncate names"))
        action_show_coded_media = None
        if selected is not None and selected.text(1)[0:3] == 'cid':
            action_show_coded_media = menu.addAction(_("Show coded files"))
        action_cat_show_coded_files = None
        action_expand_collapse = None
        if selected is not None and selected.text(1)[0:3] == 'cat':
            action_cat_show_coded_files = menu.addAction(_("Show coded files"))
            action_expand_collapse = menu.addAction(_("Expand or collapse branch"))
        action = menu.exec(self.ui.treeWidget.mapToGlobal(position))
        if action is None:
            return
        if action == action_show_coded_media:
            found_code = None
            tofind = int(selected.text(1)[4:])
            for code in self.codes:
                if code['cid'] == tofind:
                    found_code = code
                    break
            if found_code:
                DialogCodeInAllFiles(self.app, found_code)
        if action == action_cat_show_coded_files:
            branch_codes = self.recursive_get_branch_codes(selected, [])
            DialogCodeInAllFiles(self.app, branch_codes, "File", selected.text(0))
            return
        if action == action_expand_names:
            self.truncated_code_names = False
            self.fill_tree()
        if action == action_truncate_names:
            self.truncated_code_names = True
            self.fill_tree()
        if action == action_expand_collapse:
            expand_toggle = not selected.isExpanded()
            self.recursive_expand_collapse_branch(selected, expand_toggle)

    def recursive_get_branch_codes(self, item, branch_codes):
        """ Set all children of this item to be expanded or collapsed.
        Recurse through all child categories. """

        child_count = item.childCount()
        for i in range(child_count):
            if item.child(i).text(1)[0:3] == "cid":
                cid = int(item.child(i).text(1)[4:])
                for code_ in self.codes:
                    if cid == code_['cid']:
                        branch_codes.append(code_)
                        break
            if item.child(i).text(1)[0:3] == "cat":
                self.recursive_get_branch_codes(item.child(i), branch_codes)
        return branch_codes

    def recursive_expand_collapse_branch(self, item, expand_toggle):
        """ Set all children of this item to be expanded or collapsed.
        Recurse through all child categories. """

        child_count = item.childCount()
        for i in range(child_count):
            item.setExpanded(expand_toggle)
            self.recursive_expand_collapse_branch(item.child(i), expand_toggle)

    def information(self):
        """ Provide statistical help information. """

        ui = DialogInformation(self.app, "Statistics information", "")
        ui.setHtml(info)
        ui.exec()


info = "<b>Agree %</b>" \
       "<p>Calculated across all text files as the (total dual coded plus the total uncoded) / total characters</p>" \
       "<b>A and B %</b><p>Calculated as the total dual coded characters / total characters</p>" \
       "<b>Not A Not B %</b><p>The characters not coded by either coder / total characters</p>" \
       "<b>Disagree %</b><p>Is 100% minus the total agreement percent.</p>" \
       "<b>Agree coded only %</b><p>Is the dual coded characters divided by the dual coded and single coded characters" \
       "</p>" \
       "<b>Kappa</b><p>Used to measure inter-rater reliability. " \
       "Calculations are based on this site https://en.wikipedia.org/wiki/Cohen%27s_kappa</p>"
