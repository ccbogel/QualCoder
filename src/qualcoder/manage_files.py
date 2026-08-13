# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain C, Kai Dröge, Justin Missaghieh--Poncet, Lorenzo Salomón
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com/
https://qualcoder-org.github.io
https://qualcoder.org/
"""
from PyQt6.QtWidgets import QProgressDialog
import datetime
import pymupdf
import json
import openpyxl
import pandas as pd
import platform
from pathlib import Path
import PIL
from PIL import Image
from PyQt6 import QtCore, QtGui, QtWidgets
import qtawesome as qta  # see: https://pictogrammers.com/library/mdi/
from random import randint
import sqlite3
import time
from typing import Any
from shutil import copyfile, move
from striprtf.striprtf import rtf_to_text
import webbrowser
import zipfile

#from .__main__ import App
from .add_attribute import DialogAddAttribute
from .add_item_name import DialogAddItemName
from .code_pdf import DialogCodePdf, extract_pdf_fulltext, extract_pdf_highlights, \
    closest_qualcoder_color, pdf_annotations_to_file_memo, \
    code_pdf_highlights as code_pdf_highlights_shared  # Same extractor and word map as the PDF viewer
from .code_text import DialogCodeText  # for isinstance()
from .color_selector import colour_ranges, colors
from .confirm_delete import DialogConfirmDelete
from .docx import opendocx, getdocumenttext
from .edit_textfile import DialogEditTextFile
from .GUI.ui_dialog_manage_files import Ui_Dialog_manage_files
from .helpers import ExportDirectoryPathDialog, Message, msecs_to_hours_mins_secs, \
    extract_epub_fulltext
from .html_parser import *
from .latex_import import LatexImportError, tex_file_to_plain_text
from .memo import DialogMemo
from .pdf_preview import DialogPdfPagesToImages, DialogPdfPreview
from .pseudonyms import Pseudonyms
from .report_codes import DialogReportCodes  # for isInstance()
from .ris import Ris
from .select_items import DialogSelectItems
from .view_av import DialogViewAV
from .view_av_waveform import waveform_backend_available, generate_waveform_png_async, waveform_colour
from .code_av import DialogCodeAV  # for isinstance update files
from .view_image import DialogViewImage, DialogCodeImage  # for isinstance update files
from .text_decoding import decode_text_with_best_encoding as decode_text_with_best_encoding_helper

# If VLC not installed, it will not crash
vlc = None
try:
    import vlc
    # if VLC plugins stale: Open folder: Program FilesVideoLAN/VLC/plugins delete plugins.dat to force refresh
except Exception as e:  # python-vlc missing: Qt backend takes over, no console noise
    import logging as _logging
    _logging.getLogger(__name__).debug(f"python-vlc unavailable: {e}")

path = Path(__file__).resolve().parent

logger = logging.getLogger(__name__)



class FilterHeaderView(QtWidgets.QHeaderView):
    """
    Horizontal table header with filter funnel per column.
    Clicking the funnel emits filter_clicked with the column index.
    """

    filter_clicked = QtCore.pyqtSignal(int)
    ICON_SIZE = 15
    ICON_MARGIN = 4

    def __init__(self, parent=None):
        super().__init__(QtCore.Qt.Orientation.Horizontal, parent)
        self.filtered_sections = set()
        self._pressed_section = None
        self._icon_active = None  # Lazy: qtawesome needs a running QApplication
        self._icon_inactive = None
        self.setSectionsClickable(True)
        self.setHighlightSections(True)

    def set_filtered_sections(self, sections):
        """
        Columns with an active filter show a blue funnel.
        """

        self.filtered_sections = set(sections)
        self.viewport().update()

    def sectionSizeFromContents(self, logical_index):
        # Reserve room for the funnel so resizeColumnsToContents does not clip it
        size = super().sectionSizeFromContents(logical_index)
        return QtCore.QSize(size.width() + self.ICON_SIZE + self.ICON_MARGIN * 2, size.height())

    def _icon_rect(self, rect):
        return QtCore.QRect(rect.right() - self.ICON_SIZE - self.ICON_MARGIN,
                            rect.top() + (rect.height() - self.ICON_SIZE) // 2,
                            self.ICON_SIZE, self.ICON_SIZE)

    def paintSection(self, painter, rect, logical_index):
        painter.save()
        super().paintSection(painter, rect, logical_index)
        painter.restore()
        if rect.width() < self.ICON_SIZE * 2 + self.ICON_MARGIN * 2:
            return
        if self._icon_active is None:
            self._icon_active = qta.icon('mdi6.filter', color='#1e90ff')
            self._icon_inactive = qta.icon('mdi6.filter-outline', color='#909090')
        if logical_index in self.filtered_sections:
            self._icon_active.paint(painter, self._icon_rect(rect))
        else:
            self._icon_inactive.paint(painter, self._icon_rect(rect))

    def _section_rect(self, logical_index):
        return QtCore.QRect(self.sectionViewportPosition(logical_index), 0,
                            self.sectionSize(logical_index), self.height())

    def _icon_hit(self, pos):
        idx = self.logicalIndexAt(pos)
        if idx >= 0 and self._icon_rect(self._section_rect(idx)).contains(pos):
            return idx
        return None

    def mousePressEvent(self, event):
        if event.button() == QtCore.Qt.MouseButton.LeftButton:
            idx = self._icon_hit(event.position().toPoint())
            if idx is not None:
                # Swallow so the click does not select the column
                self._pressed_section = idx
                return
        self._pressed_section = None
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == QtCore.Qt.MouseButton.LeftButton and self._pressed_section is not None:
            idx = self._icon_hit(event.position().toPoint())
            if idx == self._pressed_section:
                self._pressed_section = None
                self.filter_clicked.emit(idx)
                return
        self._pressed_section = None
        super().mouseReleaseEvent(event)


class HeaderFilterPopup(QtWidgets.QFrame):
    """
    Value picker for one table column: search box, select all
    and checkable values. Changes apply immediately via apply_callback(excluded_keys).
    """

    def __init__(self, parent, keys_labels, excluded, apply_callback):
        super().__init__(parent, QtCore.Qt.WindowType.Popup)
        self.apply_callback = apply_callback
        self._updating = False
        self.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(4)
        self.search = QtWidgets.QLineEdit()
        self.search.setPlaceholderText(_("Search values"))
        self.search.setClearButtonEnabled(True)
        self.search.textChanged.connect(self.search_values)
        layout.addWidget(self.search)
        self.checkbox_select_all = QtWidgets.QCheckBox(_("(Select all)"))
        self.checkbox_select_all.clicked.connect(self.select_all_clicked)
        layout.addWidget(self.checkbox_select_all)
        self.list_widget = QtWidgets.QListWidget()
        for key, label in keys_labels:
            item = QtWidgets.QListWidgetItem(label)
            item.setData(QtCore.Qt.ItemDataRole.UserRole, key)
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled | QtCore.Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.CheckState.Unchecked if key in excluded
                               else QtCore.Qt.CheckState.Checked)
            self.list_widget.addItem(item)
        self.list_widget.itemChanged.connect(self.item_changed)
        layout.addWidget(self.list_widget)
        self.update_select_all_state()
        self.resize(260, min(420, 92 + self.list_widget.count() * 24))

    def visible_items(self):
        return [self.list_widget.item(r) for r in range(self.list_widget.count())
                if not self.list_widget.item(r).isHidden()]

    def search_values(self, text):
        text = text.lower()
        for r in range(self.list_widget.count()):
            item = self.list_widget.item(r)
            item.setHidden(text != "" and text not in item.text().lower())
        self.update_select_all_state()

    def select_all_clicked(self, checked):
        # Applies to the values matching the popup search only
        state = QtCore.Qt.CheckState.Checked if checked else QtCore.Qt.CheckState.Unchecked
        self._updating = True
        for item in self.visible_items():
            item.setCheckState(state)
        self._updating = False
        self.apply_excluded()

    def item_changed(self, item):
        if self._updating:
            return
        self.apply_excluded()

    def update_select_all_state(self):
        items = self.visible_items()
        all_checked = bool(items) and all(
            item.checkState() == QtCore.Qt.CheckState.Checked for item in items)
        self.checkbox_select_all.blockSignals(True)
        self.checkbox_select_all.setChecked(all_checked)
        self.checkbox_select_all.blockSignals(False)

    def apply_excluded(self):
        self.update_select_all_state()
        excluded = {self.list_widget.item(r).data(QtCore.Qt.ItemDataRole.UserRole)
                    for r in range(self.list_widget.count())
                    if self.list_widget.item(r).checkState() != QtCore.Qt.CheckState.Checked}
        self.apply_callback(excluded)


class DialogManageFiles(QtWidgets.QDialog):
    """ View, import, export, rename and delete text files.
    Files are normally imported into the qda project folder.
    Option to link to external files.
    """

    NAME_COLUMN = 0
    MEMO_COLUMN = 1
    DATE_COLUMN = 2
    ID_COLUMN = 3
    CASE_COLUMN = 4
    ATTRIBUTE_START_COLUMN = 5

    def __init__(self, app, parent_text_edit, tab_coding, tab_reports, main_window: QtWidgets.QMainWindow):

        self.app = app
        self.parent_text_edit = parent_text_edit
        self.tab_coding = tab_coding  # Tab widget coding for updates
        self.tab_reports = tab_reports  # Tab widget reports for updates
        self.main_window = main_window
        self.rows_hidden = []  # For save display profile, as column_name \t operator \t value
        self.source = []  # Dictionaries of source files
        self.header_labels = []
        self.default_import_directory = str(Path.home())
        self.attribute_names = []  # list of dictionary name:value for AddAttribute dialog
        self.attribute_labels_ordered = []  # helps with filling table data
        self.files_renamed = []  # list of dictionaries of old and new names and fid
        self.clipboard_text = ""  # Used to copy text into another cell
        self.pdf_import_code_highlights = None  # Per-batch tri-state: code PDF highlight annotations as codings
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_manage_files()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        self.default_import_directory = self.app.settings['directory']
        self.attribute_labels_ordered = []
        self.av_dialog_open = None  # Used for opened AV dialog
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        self.ui.pushButton_pseudonyms.setIcon(qta.icon('mdi6.account-cancel-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_pseudonyms.clicked.connect(self.pseudonyms)
        self.ui.pushButton_create.setIcon(qta.icon('mdi6.pencil-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_create.clicked.connect(self.create_text_file)
        self.ui.pushButton_view.setIcon(qta.icon('mdi6.magnify', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_view.clicked.connect(self.view)
        self.ui.pushButton_delete.setIcon(qta.icon('mdi6.delete-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_delete.clicked.connect(self.delete_button_multiple_files)
        self.ui.pushButton_import.setIcon(qta.icon('mdi6.file-document-plus-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_import.clicked.connect(self.import_files)
        self.ui.pushButton_import_survey.setIcon(qta.icon('mdi6.clipboard-text-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_import_survey.clicked.connect(self.import_survey) 
        self.ui.pushButton_link.setIcon(qta.icon('mdi6.link-variant', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_link.clicked.connect(self.link_files)
        self.ui.pushButton_import_from_linked.setIcon(
            qta.icon('mdi6.link-variant-minus', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_import_from_linked.clicked.connect(self.button_import_linked_file)
        self.ui.pushButton_export_to_linked.setIcon(qta.icon('mdi6.link-variant-plus', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export_to_linked.clicked.connect(self.button_export_file_as_linked_file)
        self.ui.pushButton_export.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export.clicked.connect(self.export)
        self.ui.pushButton_add_attribute.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_add_attribute.clicked.connect(self.add_attribute)
        self.ui.pushButton_export_attributes.setIcon(
            qta.icon('mdi6.file-export-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export_attributes.clicked.connect(self.export_attributes)
        self.ui.pushButton_undo.setIcon(qta.icon('mdi6.undo', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_undo.clicked.connect(self.undo_file_rename)
        self.ui.pushButton_mark_speakers.setIcon(qta.icon('mdi6.pin-outline', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_mark_speakers.pressed.connect(self.mark_speakers)

        self.ui.pushButton_display_save.setIcon(qta.icon('mdi6.table-plus', options=[{'scale_factor': 1.2}]))
        self.ui.pushButton_display_save.clicked.connect(self.table_display_save)
        self.ui.pushButton_display_load.setIcon(qta.icon('mdi6.text-account', options=[{'scale_factor': 1.2}]))
        self.ui.pushButton_display_load.clicked.connect(self.table_display_load)
        self.ui.pushButton_display_delete.setIcon(qta.icon('mdi6.table-minus', options=[{'scale_factor': 1.2}]))
        self.ui.pushButton_display_delete.clicked.connect(self.table_display_delete)

        self.ui.pushButton_bulk_rename.setIcon(qta.icon('mdi6.file-multiple-outline', options=[{'scale_factor': 1.2}]))
        self.ui.pushButton_bulk_rename.clicked.connect(self.bulk_rename_database_entry)
        self.ui.pushButton_help.setIcon(qta.icon('mdi6.help'))
        self.ui.pushButton_help.pressed.connect(self.help)
        self.ui.pushButton_clear_filter.setIcon(qta.icon('mdi6.filter-off-outline', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_clear_filter.pressed.connect(self.clear_file_filter)
        self.ui.pushButton_clear_filter.setToolTip(_("Clear filter"))
        self.ui.pushButton_clear_filter.setVisible(False)  # hidden until a filter is active
        self.ui.lineEdit_search_files.setToolTip(_("File name filter"))
        self.ui.lineEdit_search_files.setPlaceholderText(_("Search files"))
        self.ui.lineEdit_search_files.textChanged.connect(self.apply_file_filter)
        self.setup_header_filters()
        self.ui.tableWidget.setTabKeyNavigation(False)
        self.ui.tableWidget.itemChanged.connect(self.cell_modified)
        self.ui.tableWidget.cellClicked.connect(self.cell_selected)
        self.ui.tableWidget.cellDoubleClicked.connect(self.cell_double_clicked)
        self.ui.tableWidget.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget.customContextMenuRequested.connect(self.table_menu)
        # self.ui.tableWidget.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection) OLD
        self.ui.tableWidget.installEventFilter(self)
        self.ui.tableWidget.horizontalHeader().setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget.horizontalHeader().customContextMenuRequested.connect(self.table_header_menu)
        self.ui.tableWidget.horizontalHeader().setToolTip(_("Right click header row to hide columns"))
        self.load_file_data()
        if getattr(self.app, "project_events", None) is not None:
            self.app.project_events.project_data_changed.connect(self._on_project_data_changed)

    def _emit_project_table_changes(self, tables):
        """Notify other open dialogs about changed project tables."""

        if getattr(self.app, "project_events", None) is not None:
            self.app.project_events.emit_table_changes(tables, source=self)

    def _current_source_id(self):
        """Return the currently focused source id, if any."""

        row = self.ui.tableWidget.currentRow()
        if row < 0:
            return None
        item = self.ui.tableWidget.item(row, self.ID_COLUMN)
        if item is None:
            return None
        try:
            return int(item.text())
        except (TypeError, ValueError):
            return None

    def _restore_current_source(self, source_id:int|None):
        """ Restore the current table selection for one source id when possible."""

        if source_id is None:
            return
        for row, source_item in enumerate(self.source):
            if int(source_item.get("id", -1)) == int(source_id):
                self.ui.tableWidget.setCurrentCell(row, self.NAME_COLUMN)
                return

    def _reload_after_attribute_change(self):
        """Reload file data after an external attribute update."""

        current_source_id = self._current_source_id()
        self.load_file_data()
        self._restore_current_source(current_source_id)

    def _on_project_data_changed(self, tables, source):
        """Refresh the file table when attributes change elsewhere."""

        if source is self or not isinstance(tables, list):
            return
        changed_tables = set(tables)
        if not changed_tables.intersection({"attribute", "attribute_type"}):
            return
        self._reload_after_attribute_change()

    def help(self):
        """ Open help for transcribe section in browser. """
        self.app.help_wiki("3.2.-Files")

    FILE_TYPE_KEYS = ("text", "pdf", "image", "audio", "video")
    PLAIN_TEXT_KEY = "__plain__"
    NO_DATE_KEY = "__nodate__"
    NO_CASE_KEY = "__nocase__"

    def setup_header_filters(self):
        """
        Install the filter header on the file table.
        """

        self.header_filters = {}  # Header label -> set of excluded value keys
        self.header_filter_popup = None
        self.filter_header = FilterHeaderView(self.ui.tableWidget)
        self.ui.tableWidget.setHorizontalHeader(self.filter_header)
        self.filter_header.filter_clicked.connect(self.show_header_filter_popup)
        self.filter_header.setVisible(True)

    BASE_COLUMN_KEYS = {0: "name", 1: "memo", 2: "date", 3: "id", 4: "case"}

    def column_filter_key(self, col):
        """
        Stable key of one column for the header_filters dict.
        Base columns use fixed keys, attribute columns use 'attr:' + name,
        so a translated UI or an attribute named like a base column cannot clash.
        """

        if col in self.BASE_COLUMN_KEYS:
            return self.BASE_COLUMN_KEYS[col]
        att_pos = col - self.ATTRIBUTE_START_COLUMN
        if 0 <= att_pos < len(self.attribute_labels_ordered):
            return "attr:" + self.attribute_labels_ordered[att_pos]
        return None

    def column_of_filter_key(self, key):
        """
        Column index of one filter key, or None when the column is gone.
        """

        for col, base_key in self.BASE_COLUMN_KEYS.items():
            if key == base_key:
                return col
        if key.startswith("attr:"):
            name = key[5:]
            if name in self.attribute_labels_ordered:
                return self.ATTRIBUTE_START_COLUMN + self.attribute_labels_ordered.index(name)
        return None

    def header_filter_keys_of(self, data, col):
        """
        Filter keys of one source for one column.
        The name column filters by file type and extension, matching its icon.
        """

        if col == self.NAME_COLUMN:
            return [f"type\t{self.file_type_of(data)}", f"ext\t{self.file_ext_of(data)}"]
        if col == self.MEMO_COLUMN:
            return ["memo" if data['memo'] != "" else ""]
        if col == self.DATE_COLUMN:
            return [self.file_date_of(data)]
        if col == self.ID_COLUMN:
            return [str(data['id'])]
        if col == self.CASE_COLUMN:
            return sorted(self.file_cases_of(data))
        att_pos = col - self.ATTRIBUTE_START_COLUMN
        if 0 <= att_pos < len(data['attributes']):
            return [data['attributes'][att_pos]]
        return []

    def row_matches_column(self, data, col, excluded):
        """
        True when the header filter of one column keeps this file visible.
        """

        keys = self.header_filter_keys_of(data, col)
        if not keys:
            return True
        if col == self.CASE_COLUMN:
            # A file in several cases stays visible while any of them is checked
            return not set(keys).issubset(excluded)
        return not any(key in excluded for key in keys)

    def row_passes_filters(self, row, data, skip_col=None):
        """
        True when the search text, the header filters (except skip_col) and
        the context menu / loaded display criteria keep this row visible.
        """

        search = self.ui.lineEdit_search_files.text().lower()
        if search and search not in data['name'].lower():
            return False
        for key, excluded in self.header_filters.items():
            if not excluded:
                continue
            col = self.column_of_filter_key(key)
            if col is None or col == skip_col:
                continue
            if not self.row_matches_column(data, col, excluded):
                return False
        return not self.row_hidden_by_criteria(row)

    def header_filter_options(self, col):
        """
        (key, label) options for one column popup: distinct
        values from the rows that pass the search and every other column filter.
        """

        sources = [data for row, data in enumerate(self.source)
                   if self.row_passes_filters(row, data, skip_col=col)]
        if col == self.NAME_COLUMN:
            type_labels = {'text': _("Text"), 'pdf': _("PDF"), 'image': _("Image"),
                           'audio': _("Audio"), 'video': _("Video")}
            present = {self.file_type_of(s) for s in sources}
            options = [(f"type\t{key}", type_labels[key]) for key in self.FILE_TYPE_KEYS
                       if key in present]
            exts = {self.file_ext_of(s) for s in sources}
            if self.PLAIN_TEXT_KEY in exts:
                options.append((f"ext\t{self.PLAIN_TEXT_KEY}", _("Plain text")))
            options += [(f"ext\t{ext}", ext) for ext in sorted(exts - {self.PLAIN_TEXT_KEY})]
            return options
        if col == self.MEMO_COLUMN:
            keys = {self.header_filter_keys_of(s, col)[0] for s in sources}
            options = []
            if "memo" in keys:
                options.append(("memo", _("Memo")))
            if "" in keys:
                options.append(("", _("No memo")))
            return options
        if col == self.DATE_COLUMN:
            dates = {self.file_date_of(s) for s in sources}
            options = [(d, d) for d in sorted(dates - {self.NO_DATE_KEY}, reverse=True)]
            if self.NO_DATE_KEY in dates:
                options.append((self.NO_DATE_KEY, _("No date")))
            return options
        if col == self.ID_COLUMN:
            ids = {str(s['id']) for s in sources}
            return [(id_, id_) for id_ in sorted(ids, key=lambda v: int(v) if v.isdigit() else 0)]
        if col == self.CASE_COLUMN:
            names = {name for s in sources for name in self.file_cases_of(s)}
            options = [(name, name) for name in sorted(names - {self.NO_CASE_KEY})]
            if self.NO_CASE_KEY in names:
                options.append((self.NO_CASE_KEY, _("No case")))
            return options
        att_pos = col - self.ATTRIBUTE_START_COLUMN
        values = sorted({s['attributes'][att_pos] for s in sources
                         if 0 <= att_pos < len(s['attributes'])})
        options = []
        for value in values:
            label = value.replace("\n", "; ") if value != "" else _("(no value)")
            if len(label) > 40:
                label = label[:40] + "..."
            options.append((value, label))
        return options

    def show_header_filter_popup(self, col):
        """
        Open the value picker under the clicked header section.
        """

        filter_key = self.column_filter_key(col)
        if filter_key is None:
            return
        options = self.header_filter_options(col)
        excluded = self.header_filters.get(filter_key, set())
        self.header_filter_popup = HeaderFilterPopup(
            self, options, excluded, lambda excl, key=filter_key: self.set_header_filter(key, excl))
        x = self.filter_header.sectionViewportPosition(col)
        pos = self.filter_header.viewport().mapToGlobal(QtCore.QPoint(x, self.filter_header.height()))
        screen = self.screen().availableGeometry()
        pos.setX(max(screen.left(), min(pos.x(), screen.right() - self.header_filter_popup.width())))
        self.header_filter_popup.move(pos)
        self.header_filter_popup.show()

    def set_header_filter(self, filter_key, excluded):
        """
        Store the excluded value keys of one column and re-apply the filter.
        """

        if excluded:
            self.header_filters[filter_key] = set(excluded)
        else:
            self.header_filters.pop(filter_key, None)
        self.update_header_filter_icons()
        self.apply_file_filter()

    def update_header_filter_icons(self):
        """
        Blue funnel on the columns with an active filter.
        """

        sections = {self.column_of_filter_key(key) for key in self.header_filters}
        self.filter_header.set_filtered_sections(sections - {None})

    def prune_header_filters(self):
        """ After a reload, drop filters of removed columns and of values no
        longer present, so the funnel icons stay truthful. """

        pruned = {}
        for filter_key, excluded in self.header_filters.items():
            col = self.column_of_filter_key(filter_key)
            if col is None:
                continue
            all_keys = {key for data in self.source
                        for key in self.header_filter_keys_of(data, col)}
            keep = excluded & all_keys
            if keep:
                pruned[filter_key] = keep
        self.header_filters = pruned
        self.update_header_filter_icons()

    def file_type_of(self, data):
        """
        Derive the file type key for one source diccionario
        """

        mediapath = data['mediapath'] or ""
        if mediapath[:7] == "/audio/" or mediapath[:6] == "audio:":
            return "audio"
        if mediapath[:7] == "/video/" or mediapath[:6] == "video:":
            return "video"
        if mediapath[:8] == "/images/" or mediapath[:7] == "images:":
            return "image"
        if data['name'].lower().endswith(".pdf"):
            return "pdf"
        return "text"

    def file_ext_of(self, data):
        """
        Lowercased name extension, or the plain text key when there is none.
        """

        suffix = Path(data['name']).suffix.lower()
        return suffix if suffix else self.PLAIN_TEXT_KEY

    def file_date_of(self, data):
        """
        Date portion (yyyy-mm-dd) of the source date, or the no date key.
        """

        date_ = (data['date'] or "")[:10]
        return date_ if date_ else self.NO_DATE_KEY

    def file_cases_of(self, data):
        """
        Set of case names the file belongs to, or the no case key.
        """

        names = {name for name in (data['case'] or "").split(";") if name}
        return names if names else {self.NO_CASE_KEY}


    def file_filter_active(self):
        """
        True when the search text or a header column filter restricts the table.
        """

        if self.ui.lineEdit_search_files.text() != "":
            return True
        return any(self.header_filters.values())

    def row_hidden_by_criteria(self, row):
        """
        True when a rows_hidden criterion (context menu or loaded display) hides this row.
        Criteria are 'colname \t operator \t value'; unknown columns are skipped.
        """

        for criterion in self.rows_hidden:
            try:
                colname, operator, value = criterion.split("\t")
            except ValueError:
                continue
            col_idx = None
            for c in range(self.ui.tableWidget.columnCount()):
                if self.ui.tableWidget.horizontalHeaderItem(c).text() == colname:
                    col_idx = c
                    break
            if col_idx is None:
                continue
            item = self.ui.tableWidget.item(row, col_idx)
            text_ = item.text() if item is not None else ""
            if operator == "like" and text_.find(value) == -1:
                return True
            if operator == "=" and text_ != value:
                return True
            if operator == "hide" and text_.find(value) != -1:
                return True
        return False

    def apply_file_filter(self):
        """
        Hide rows not matching the search text, the header column filters and
        any context menu / loaded display criteria. Hidden rows are deselected
        so selection based actions only act on visible rows.
        """

        active = self.file_filter_active()
        selection_model = self.ui.tableWidget.selectionModel()
        for row, data in enumerate(self.source):
            if row >= self.ui.tableWidget.rowCount():
                break
            self.ui.tableWidget.setRowHidden(row, not self.row_passes_filters(row, data))
        if selection_model is not None:
            for index in selection_model.selectedIndexes():
                if self.ui.tableWidget.isRowHidden(index.row()):
                    selection_model.select(index, QtCore.QItemSelectionModel.SelectionFlag.Deselect)
        if active:
            self.ui.pushButton_clear_filter.setVisible(True)
            self.ui.pushButton_clear_filter.setStyleSheet("background-color: #1e90ff; color: white;")
        else:
            self.ui.pushButton_clear_filter.setVisible(False)
            self.ui.pushButton_clear_filter.setStyleSheet("")
        self.update_label_file_count()

    def visible_selected_rows(self):
        """
        Selected row numbers excluding rows hidden by any filter
        """

        rows = {index.row() for index in self.ui.tableWidget.selectionModel().selectedIndexes()}
        return sorted(r for r in rows if not self.ui.tableWidget.isRowHidden(r))

    def clear_file_filter(self):
        """
        Reset the search text and every header column filter. Context menu
        criteria in rows_hidden stay applied; Ctrl+A or Show all rows clears those
        """

        self.ui.lineEdit_search_files.blockSignals(True)
        self.ui.lineEdit_search_files.clear()
        self.ui.lineEdit_search_files.blockSignals(False)
        self.header_filters = {}
        self.update_header_filter_icons()
        self.apply_file_filter()

    def pseudonyms(self):
        """ Pseudonymisation, data de-identification.
        User created list of pseudonyms to replace sensitive wording, e.g. participant names.
        The list must be created before any text files or survey qualitative columns are imported.
        The pseudonyms are case-sensitive.
        The original to pseudonym list is stored in the qda folder as pseudonyms.json
        This file can be removed and stored securely after the data is imported.
        Pseudonyms does not apply to PDF imports. Instead import plain text of the PDF. """

        ui_pseudomyms = Pseudonyms(self.app)
        ui_pseudomyms.exec()

    def table_display_save(self):
        """ Save rows and column settings for replicating table display.
        Table columns as name, width.
        Table rows as colname tab operator tab value
        operator is either: =, like, hide
        """

        row_txt = "\t\t".join(self.rows_hidden)
        row_msg = " AND ".join(self.rows_hidden)
        if len(row_msg) > 0:
            row_msg = _("Rows: ") + row_msg.replace("\t", " ")
        col_txt = ""
        col_msg = ""
        for c in range(0, self.ui.tableWidget.columnCount()):
            header_text = self.ui.tableWidget.horizontalHeaderItem(c).text()
            col_txt += f'{header_text}\t{self.ui.tableWidget.columnWidth(c)}\t\t'
            if self.ui.tableWidget.isColumnHidden(c):
                col_msg += f"{header_text}; "
        if len(col_msg) > 0:
            col_msg = _("Hidden columns: ") + col_msg
        if col_msg == "" and row_msg == "":
            Message(self.app, _("No special settings"),
                    _("No table display settings or rows or columns selected")).exec()
            return
        msg = f"{col_msg}\n{row_msg}\n" + _("Save as:")
        display_name, ok = QtWidgets.QInputDialog.getText(self, _("Save Table Display"), msg,
                                                          QtWidgets.QLineEdit.EchoMode.Normal)
        if not ok or display_name == "":
            return
        cur = self.app.conn.cursor()
        cur.execute("select name from manage_files_display where tblrows=? and tblcolumns=?", [row_txt, col_txt])
        res = cur.fetchone()
        if res:
            Message(self.app, _("Table display exists"),
                    _("This table display setting already exists: ") + res[0]).exec()
            return
        sql = "insert into manage_files_display (name, tblrows, tblcolumns, owner) values (?,?,?,?)"
        cur.execute(sql, [display_name, row_txt, col_txt, self.app.settings['codername']])
        self.app.conn.commit()

    def table_display_load(self):
        """ Load rows and column settings for replicating a table display.  """

        cur = self.app.conn.cursor()
        cur.execute("select name, tblrows,tblcolumns, owner from manage_files_display order by upper(name)")
        res = cur.fetchall()
        if not res:
            Message(self.app, _("Nothing saved"), _("No saved table displays")).exec()
            return
        keys = 'name', 'tblrows', 'tblcolumns', 'owner'
        displays = []
        for row in res:
            displays.append(dict(zip(keys, row)))
        ui = DialogSelectItems(self.app, displays, _("Select table display"), "single")
        ok = ui.exec()
        if not ok:
            return
        selection = ui.get_selected()
        column_name_width_list = selection['tblcolumns'].split("\t\t")  # Will contain a '' at the end
        column_name_width_list.pop()
        msg = _("Load table display settings") + f"\nProfile: {selection['name']}\n"

        # Reset columns and rows
        for col in range(0, self.ui.tableWidget.columnCount()):
            self.ui.tableWidget.setColumnHidden(col, False)
            # Must check for each column name, as some columns might be renamed. Renamed columns will be ignored
            for col_name_width in column_name_width_list:
                colname, width = col_name_width.split("\t")
                # print(colname, width)
                if self.ui.tableWidget.horizontalHeaderItem(col).text() == colname and int(width) > 0:
                    self.ui.tableWidget.setColumnWidth(col, int(width))
                    break
                if self.ui.tableWidget.horizontalHeaderItem(col).text() == colname and int(width) == 0:
                    self.ui.tableWidget.setColumnHidden(col, True)  # Sets different Qt flags
                    msg += _("Hidden column: ") + colname + "\n"
                    break
        if selection['tblrows'] == "":
            return
        row_parameters_list = selection['tblrows'].split("\t\t")
        # Need to re-create self.rows_hidden variable , for menu options and for any further display saving.
        self.rows_hidden = []
        for rpl in row_parameters_list:
            self.rows_hidden.append(rpl)
        # Warn about criteria whose column no longer exists
        if row_parameters_list:
            msg += _("Row settings:") + "\n"
        warning = ""
        for rpl in row_parameters_list:
            colname, operator, value = rpl.split("\t")
            msg += f"{colname} {operator} {value}\n"
            found = False
            for c in range(0, self.ui.tableWidget.columnCount()):
                if self.ui.tableWidget.horizontalHeaderItem(c).text() == colname:
                    found = True
                    break
            if not found:
                warning += f"Column name not found: {colname}\n"
        if warning:
            Message(self.app, _("Table column not present"), warning).exec()

        # Now hide the rows, combined with any active toolbar filter
        self.apply_file_filter()
        self.ui.pushButton_display_load.setToolTip(msg)

    def table_display_delete(self):
        """ Delete stored table displays """

        cur = self.app.conn.cursor()
        cur.execute("select name, tblrows,tblcolumns, owner from manage_files_display order by upper(name)")
        res = cur.fetchall()
        if not res:
            Message(self.app, _("Nothing saved"), _("No saved table displays")).exec()
            return
        keys = 'name', 'tblrows', 'tblcolumns', 'owner'
        displays = []
        for row in res:
            displays.append(dict(zip(keys, row)))
        ui = DialogSelectItems(self.app, displays, _("Delete table display"), "single")
        ok = ui.exec()
        if not ok:
            return
        selection = ui.get_selected()
        cur.execute("delete from manage_files_display where name=?", [selection['name']])
        self.app.conn.commit()
        Message(self.app, _("Deleted display"), selection['name']).exec()

    def keyPressEvent(self, event):
        """ Used to activate buttons.
        Ctrl 0 to 9
        """
        key = event.key()
        mods = QtWidgets.QApplication.keyboardModifiers()

        # Ctrl 0 to 4
        if mods & QtCore.Qt.KeyboardModifier.ControlModifier:
            if key == QtCore.Qt.Key.Key_1:
                self.view()
                return
            if key == QtCore.Qt.Key.Key_2:
                self.import_files()
                return
            if key == QtCore.Qt.Key.Key_3:
                self.link_files()
                return
            if key == QtCore.Qt.Key.Key_4:
                self.create_text_file()
                return
            if key == QtCore.Qt.Key.Key_5:
                self.button_import_linked_file()
                return
            if key == QtCore.Qt.Key.Key_6:
                self.button_export_file_as_linked_file()
                return
            if key == QtCore.Qt.Key.Key_7:
                self.add_attribute()
                return
            if key == QtCore.Qt.Key.Key_8:
                self.export_attributes()
                return
            if key == QtCore.Qt.Key.Key_9:
                self.export()
                return
            if key == QtCore.Qt.Key.Key_0:
                self.help()
                return
            if key == QtCore.Qt.Key.Key_C:
                x = self.ui.tableWidget.currentRow()
                y = self.ui.tableWidget.currentColumn()
                self.clipboard_text = self.ui.tableWidget.item(x, y).text()
                return
            if key == QtCore.Qt.Key.Key_V:
                x = self.ui.tableWidget.currentRow()
                y = self.ui.tableWidget.currentColumn()
                self.ui.tableWidget.item(x, y).setText(self.clipboard_text)
                return

    def eventFilter(self, object_, event):
        """ Using this event filter to
        Ctrl + A to show all rows
        Ctrl + Z Undo the last deletion.
        def
        """

        if type(event) == QtGui.QKeyEvent:
            key = event.key()
            mod = event.modifiers()
            if key == QtCore.Qt.Key.Key_A and mod == QtCore.Qt.KeyboardModifier.ControlModifier:
                self.rows_hidden = []
                self.clear_file_filter()
                return True
            if key == QtCore.Qt.Key.Key_Delete and self.ui.tableWidget.currentColumn() == 0:
                self.delete()
                return True
        return False

    def table_header_menu(self, position):
        """ Used to show and hide columns """

        index_at = self.ui.tableWidget.indexAt(position)
        header_index = int(index_at.column())
        menu = QtWidgets.QMenu(self)
        action_show_all_columns = menu.addAction(_("Show all columns"))
        action_hide_column = None
        if header_index > 0:
            action_hide_column = menu.addAction(_("Hide column"))
        action_hide_columns_starting = menu.addAction(_("Hide columns starting with"))
        action_show_columns_starting = menu.addAction(_("Show columns starting with"))
        action = menu.exec(self.ui.tableWidget.mapToGlobal(position))
        if action == action_show_all_columns:
            for c in range(0, self.ui.tableWidget.columnCount()):
                self.ui.tableWidget.setColumnHidden(c, False)
            if not self.app.settings['showids']:
                self.ui.tableWidget.setColumnHidden(self.ID_COLUMN, True)
            return
        if action == action_hide_column:
            self.ui.tableWidget.setColumnHidden(header_index, True)
            return
        if action == action_hide_columns_starting:
            msg = _("Hide columns starting with:")
            hide_filter, ok = QtWidgets.QInputDialog.getText(self, _("Hide Columns"), msg,
                                                             QtWidgets.QLineEdit.EchoMode.Normal)
            for c in range(1, self.ui.tableWidget.columnCount()):
                h_text = self.ui.tableWidget.horizontalHeaderItem(c).text()
                if len(h_text) >= len(hide_filter) and hide_filter == h_text[:len(hide_filter)]:
                    self.ui.tableWidget.setColumnHidden(c, True)
        if action == action_show_columns_starting:
            msg = _("Show columns starting with:")
            show_filter, ok = QtWidgets.QInputDialog.getText(self, _("Show Columns"), msg,
                                                             QtWidgets.QLineEdit.EchoMode.Normal)
            for c in range(4, self.ui.tableWidget.columnCount()):
                h_text = self.ui.tableWidget.horizontalHeaderItem(c).text()
                if len(h_text) >= len(show_filter) and show_filter == h_text[:len(show_filter)]:
                    self.ui.tableWidget.setColumnHidden(c, False)
                else:
                    self.ui.tableWidget.setColumnHidden(c, True)

    def table_menu(self, position):
        """ Context menu for displaying table rows in differing order,
        hiding table rows, assigning case to file, file rename, export import from linked. """

        row = self.ui.tableWidget.currentRow()
        col = self.ui.tableWidget.currentColumn()
        cell = self.ui.tableWidget.item(row, col)
        if cell is None:
            item_text = ""  # used for opening URLs action
        else:
            item_text = cell.text()
        # Use these next few lines to use for moving a linked file into or an internal file out of the project folder
        mediapath:str|None = None
        risid:int|None = None
        try:
            id_ = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())
        except AttributeError:
            # Occurs if a table cell is not clicked, but click occurs elsewhere in container
            return
        for s in self.source:
            if s['id'] == id_:
                mediapath = s['mediapath']
                risid = s['risid']
        # Check and get all selected indexes
        selected_indexes = self.ui.tableWidget.selectionModel().selectedIndexes()
        # Action cannot be None otherwise may default to one of the actions below depending on column clicked
        menu = QtWidgets.QMenu()
        menu.setStyleSheet(f"QMenu {{font-size:{self.app.settings['fontsize']}pt}} ")
        action_view = menu.addAction(_("View"))
        # Import a transcription from an external file (e.g. a noScribe .txt or .html export)
        # into the empty .txt transcription auto-created for an audio/video file.
        action_import_transcription = None
        if mediapath is not None and len(mediapath) > 5 and \
                mediapath[:6] in ("/audio", "audio:", "/video", "video:"):
            action_import_transcription = menu.addAction(_("Import transcription from file"))
        action_view_original_text = None
        action_pdf_to_images = None
        action_extract_pdf_text = None
        if mediapath is not None and len(mediapath) > 6 and (mediapath[:6] == '/docs/' or mediapath[:5] == 'docs:'):
            action_view_original_text = menu.addAction(_("View original text file"))
        if mediapath is not None and len(mediapath) > 6 and (mediapath[:6] == '/docs/' or mediapath[:5] == 'docs:') \
                and mediapath[-4:].lower() == ".pdf":
            action_pdf_to_images = menu.addAction(_("Pdf pages to images"))
            # Editable copy of the PDF text as a new text source: the stored fulltext
            # of a PDF is not editable, this copy is the way to edit its text.
            action_extract_pdf_text = menu.addAction(_("Extract pdf text to new file"))
        action_filename_asc = None
        action_filename_desc = None
        action_type = None
        if col == self.NAME_COLUMN:
            action_filename_asc = menu.addAction(_("Order ascending"))
            action_filename_desc = menu.addAction(_("Order descending"))
            action_type = menu.addAction(_("File type order"))
        action_date_asc = None
        action_date_desc = None
        if col == self.DATE_COLUMN:
            action_date_asc = menu.addAction(_("Order ascending"))
            action_date_desc = menu.addAction(_("Order descending"))
        action_casename_asc = None
        action_casename_desc = None
        action_assign_case = None
        if col == self.CASE_COLUMN:
            action_casename_asc = menu.addAction(_("Order ascending"))
            action_casename_desc = menu.addAction(_("Order descending"))
            action_assign_case = menu.addAction(_("Assign files to case"))
        action_show_values_like = None
        action_hide_values_like = None
        if col != self.MEMO_COLUMN:
            action_show_values_like = menu.addAction(_("Show values like"))
            action_hide_values_like = menu.addAction(_("Hide values like"))
        action_equals_value = menu.addAction(_("Show this value"))
        action_order_by_value_asc = None
        action_order_by_value_desc = None
        action_date_picker = None
        action_ref_apa = None
        action_ref_vancouver = None
        action_multiple_cells_value = None
        if col > self.CASE_COLUMN:
            action_order_by_value_asc = menu.addAction(_("Order ascending"))
            action_order_by_value_desc = menu.addAction(_("Order descending"))
            if "date" in self.header_labels[col].lower():
                # Check that a character date can be entered
                cur = self.app.conn.cursor()
                cur.execute("select valuetype from attribute_type where caseOrFile='file' and name=?",
                            [self.header_labels[col], ])
                result = cur.fetchone()
                if result is not None and result[0] == "character":
                    action_date_picker = menu.addAction(_("Enter date"))
            if self.header_labels[col] in ("Ref_Authors", "Ref_Title", "Ref_Journal", "Ref_Type", "Ref_Year"):
                action_ref_apa = menu.addAction(_("Copy reference to clipboard. APA"))
                action_ref_vancouver = menu.addAction(_("Copy reference to clipboard. Vancouver"))
            if len(selected_indexes) > 1:
                action_multiple_cells_value = menu.addAction(_("Set value of selected cells"))
        action_rename = None
        action_export = None
        action_delete = None
        action_export_to_linked = None
        action_import_linked = None
        action_mark_speakers = None
        if col == self.NAME_COLUMN:
            action_rename = menu.addAction(_("Rename database entry"))
            action_export = menu.addAction(_("Export"))
            action_delete = menu.addAction(_("Delete"))
            if mediapath is None or mediapath == "" or (mediapath is not None and mediapath[0] == "/"):
                action_export_to_linked = menu.addAction(_("Move file to externally linked file"))
            if mediapath is not None and mediapath != "" and mediapath[0] != "/":
                action_import_linked = menu.addAction(_("Import linked file"))
            if mediapath is None or \
                    (mediapath is not None and len(mediapath) > 6 and (mediapath[:6] == '/docs/' or
                                                                       mediapath[:5] == 'docs:')):
                action_mark_speakers = menu.addAction(_('Mark speakers'))
        action_show_all = None
        if self.rows_hidden:
            action_show_all = menu.addAction(_("Show all rows Ctrl A"))
        action_url = None
        # Regex HTTP HTTPS protocol
        regex_http = QtCore.QRegularExpression(
            r"^https?:\/\/(?:www\.)?[-a-zA-Z0-9@:%._\+~#=]{1,256}\.[a-zA-Z0-9()]{1,63}\b(?:[-a-zA-Z0-9()@:%_\+.~#?&\/=]*)$")
        # Regex Protocol optional
        regex_no_protocol = QtCore.QRegularExpression(r"^www\.[a-zA-Z0-9()]{1,63}\b(?:[-a-zA-Z0-9()@:%_\+.~#?&\/=]*)$")
        if bool(regex_no_protocol.match(item_text)) or bool(regex_http.match(item_text)):
            action_url = menu.addAction(_("Open URL"))

        action = menu.exec(self.ui.tableWidget.mapToGlobal(position))
        if action is None:
            return
        if action == action_view:
            self.view()
            return
        if action == action_import_transcription:
            self.import_transcription_from_file(id_)
            return
        if action == action_view_original_text:
            bad_link = bad_link = self.app.check_bad_file_links(id_)  # List is returned
            if bad_link:
                bad_link = bad_link[0]
                bad_path = bad_link['mediapath'].split(':', 1)[1]
                Message(self.app, _("Original file not found"), _("Update the file path") + "\n" + bad_path).exec()
                self.update_file_path(id_, bad_link)
                return
            self.view_original_text_file(mediapath)
            return
        if action == action_pdf_to_images:
            self.pdf_to_images(mediapath)
            return
        if action == action_extract_pdf_text:
            self.extract_pdf_text_copy(row)
            return
        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
            return
        if action == action_import_linked:
            self.import_linked_file(id_, mediapath)
        if action == action_export_to_linked:
            self.export_file_as_linked_file(id_, mediapath)
        if action == action_export:
            self.export()
        if action == action_delete:
            self.delete()
        if action == action_rename:
            self.rename_database_entry()
        if action == action_assign_case:
            self.assign_cases_to_file()
        if action == action_filename_asc:
            self.load_file_data()
        if action == action_filename_desc:
            self.load_file_data("filename desc")
        if action == action_date_asc:
            self.load_file_data("date asc")
        if action == action_date_desc:
            self.load_file_data("date desc")
        if action == action_type:
            self.load_file_data("filetype")
        if action == action_casename_asc:
            self.load_file_data("casename asc")
        if action == action_casename_desc:
            self.load_file_data("casename desc")
        if action == action_order_by_value_asc:
            self.load_file_data("attribute asc:" + self.header_labels[col])
        if action == action_order_by_value_desc:
            self.load_file_data("attribute desc:" + self.header_labels[col])
        if action == action_equals_value:
            # Hide rows that do not match this value
            item_to_compare = self.ui.tableWidget.item(row, col)
            compare_text = item_to_compare.text()
            self.rows_hidden.append(f'{self.ui.tableWidget.horizontalHeaderItem(col).text()}\t=\t{compare_text}')
            self.apply_file_filter()
            return
        if action == action_show_values_like:
            text_value, ok = QtWidgets.QInputDialog.getText(self, _("Text filter"), _("Show values like:"),
                                                            QtWidgets.QLineEdit.EchoMode.Normal)
            if ok and text_value != '':
                self.rows_hidden.append(f'{self.ui.tableWidget.horizontalHeaderItem(col).text()}\tlike\t{text_value}')
                self.apply_file_filter()
            return
        if action == action_hide_values_like:
            text_value, ok = QtWidgets.QInputDialog.getText(self, _("Text filter"), _("Hide values like:"),
                                                            QtWidgets.QLineEdit.EchoMode.Normal)
            if ok and text_value != '':
                self.rows_hidden.append(f'{self.ui.tableWidget.horizontalHeaderItem(col).text()}\thide\t{text_value}')
                self.apply_file_filter()
            return
        if action == action_show_all:
            self.rows_hidden = []
            self.apply_file_filter()
            self.ui.pushButton_display_load.setToolTip(_("Load table display settings"))
            return
        if action == action_url:
            print("URL open", item_text)
            webbrowser.open(item_text)
            return
        if action == action_date_picker:
            ui_memo = DialogMemo(self.app, "Date selector", "", "hide")
            ui_memo.ui.textEdit.hide()
            calendar = QtWidgets.QCalendarWidget()
            ui_memo.ui.gridLayout.addWidget(calendar, 0, 0, 1, 1)
            ok = ui_memo.exec()
            if ok:
                selected_date = calendar.selectedDate().toString("yyyy-MM-dd")
                self.ui.tableWidget.setItem(row, col, QtWidgets.QTableWidgetItem(selected_date))
            return
        if action == action_ref_apa:
            ris_obj = Ris(self.app)
            ris_obj.get_references(selected_ris=risid)
            apa = ris_obj.refs
            if not apa:
                return
            cb = QtWidgets.QApplication.clipboard()
            cb.setText(apa[0]['apa'].replace("\n", " "))
        if action == action_ref_vancouver:
            ris_obj = Ris(self.app)
            ris_obj.get_references(selected_ris=risid)
            vancouver = ris_obj.refs
            if not vancouver:
                return
            cb = QtWidgets.QApplication.clipboard()
            cb.setText(vancouver[0]['vancouver'].replace("\n", " "))
        if action == action_mark_speakers:
            self.mark_speakers()
        if action == action_multiple_cells_value:
            self.multiple_cells_value()

    def multiple_cells_value(self):
        """ Assign a value to all selected cells.
         If column > CASE_COLUMN. """

        value, ok = QtWidgets.QInputDialog.getText(self, _("Selected cells"), _("Set value:"),
                                                        QtWidgets.QLineEdit.EchoMode.Normal)
        if not ok: return
        msg = ""
        value = value.strip()
        selected_indexes = self.ui.tableWidget.selectionModel().selectedIndexes()
        for i in selected_indexes:
            col, row =  i.column(), i.row()
            if self.ui.tableWidget.isRowHidden(row):
                continue
            # Check if cell it an editable attribute
            if col > self.CASE_COLUMN and self.header_labels[col] not in ("Ref_Authors", "Ref_Journal","Ref_Title","Ref_Type","Ref_Year"):
                try:
                    prev_value = str(self.ui.tableWidget.item(row, col).text()).strip()
                except AttributeError:
                    prev_value = ""
                attribute_name = self.header_labels[col]
                cur = self.app.conn.cursor()
                # Check numeric for numeric attributes, clear "" if it cannot be cast
                cur.execute("select valuetype from attribute_type where caseOrFile='file' and name=?",
                            (attribute_name,))
                result = cur.fetchone()
                if result is None:
                    return
                if result[0] == "numeric" and value != "":
                    try:
                        float(value)
                    except ValueError:
                        value = prev_value
                        msg = _("Value must be numeric")
                cur.execute("update attribute set value=? where id=? and name=? and attr_type='file'",
                            (value, self.source[row]['id'], attribute_name))
                item = QtWidgets.QTableWidgetItem(value)
                self.ui.tableWidget.blockSignals(True)  # Otherwise, cell_modified() is called
                self.ui.tableWidget.setItem(row,col, item)
                self.ui.tableWidget.blockSignals(False)
                # Keep the loaded sources in sync for the header filters
                att_pos = col - self.ATTRIBUTE_START_COLUMN
                if 0 <= att_pos < len(self.source[row]['attributes']):
                    self.source[row]['attributes'][att_pos] = value
            self.app.conn.commit()
        self._emit_project_table_changes(["attribute"])
        if self.file_filter_active():
            # An edited value may now be excluded by a header filter
            self.apply_file_filter()
        if msg != "":
            Message(self.app, _("Value error"), msg).exec()

    def update_file_path(self, id_: int, bad_link: dict[str, Any]):
        """ Update the File Not Found file path to another path.
         Args:
             id_ : Integer source.id
             bad_link : Dictionary of name, mediapath (with prefix. e.g. docs: ...
        """

        extended_mediapath = bad_link['mediapath']
        file_typer, mediapath = extended_mediapath.split(':', 1)
        file_extension = Path(mediapath).suffix
        folder = self.app.settings['directory']
        new_file_path, filter_ = QtWidgets.QFileDialog.getOpenFileName(self, _("Update file path"), folder,
                                                              f"*{file_extension}")
        if new_file_path:
            new_file_path_extended = f"{file_typer}:{new_file_path}"
            cur = self.app.conn.cursor()
            sql = "update source set mediapath=? where id=?"
            cur.execute(sql, [new_file_path_extended, id_])
            self.app.conn.commit()
            self.parent_text_edit.append(_("Fixed file path: ") + f"{mediapath} -> {new_file_path}")
            # Reload data
            self.load_file_data()
            self.app.delete_backup = False
            self.update_files_in_dialogs()

    def extract_pdf_text_copy(self, row:int):
        """ Creates a new TEXT source with a copy of the PDF's fulltext (the one already
        extracted and stored at import, no re-extraction). The stored fulltext of a
        PDF is not editable, so this copy is the way to work with and edit its text,
        without touching the PDF's fulltext or breaking the page mapping.
        Args:
            row: table row of the pdf source, Integer
        """

        cur = self.app.conn.cursor()
        cur.execute("select name, fulltext from source where id=?", [self.source[row]['id']])
        res = cur.fetchone()
        if res is None:
            return
        fulltext = res[1] if res[1] is not None else ""
        if fulltext.strip() == "":
            Message(self.app, _("Extract pdf text"),
                    _("This PDF has no stored text (scanned PDF?)."), "warning").exec()
            return
        # Unique name: name.pdf.txt, with _n suffix if it already exists.
        base_name = res[0] + ".txt"
        new_name = base_name
        existing_names = {s['name'] for s in self.source}
        n = 1
        while new_name in existing_names:
            new_name = f"{res[0]}_{n}.txt"
            n += 1
        now_ = datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S")
        entry = {'name': new_name, 'id': -1, 'fulltext': fulltext, 'mediapath': None,
                 'memo': _("Text extracted from pdf: ") + res[0],
                 'owner': self.app.settings['codername'], 'date': now_, 'risid': None}
        cur.execute("insert into source(name,fulltext,mediapath,memo,owner,date) values(?,?,?,?,?,?)",
                    (entry['name'], entry['fulltext'], entry['mediapath'], entry['memo'],
                     entry['owner'], entry['date']))
        self.app.conn.commit()
        cur.execute("select last_insert_rowid()")
        id_ = cur.fetchone()[0]
        entry['id'] = id_
        # File attribute placeholders, as in create_text_file.
        cur.execute('select name from attribute_type where caseOrFile ="file"')
        attr_types = cur.fetchall()
        insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
        for a in attr_types:
            cur.execute(insert_sql, [a[0], id_, now_, self.app.settings['codername']])
        self.app.conn.commit()
        if self.app.settings['ai_enable'] == 'True':
            self.app.ai.sources_vectorstore.import_document(id_, entry['name'], entry['fulltext'])
        self.parent_text_edit.append(_("Text extracted from pdf to new file: ") + new_name)
        self.load_file_data()
        self.fill_table()
        self.app.delete_backup = False
        self.update_files_in_dialogs()

    def pdf_to_images(self, mediapath:str):
        """ Turn pdf pages into an image for each page.
        With a page range dialog, preview and resolution; skips duplicate names
        without per-page modal warnings, shows progress and closes the document
        (the handle previously stayed open).
        Args:
            mediapath : String
        """

        filepath = ""
        filename = ""
        if mediapath[:6] == '/docs/':
            filepath = Path(self.app.project_path) / "documents" / mediapath[6:]
            filename = mediapath[6:]
        if mediapath[:5] == 'docs:':
            filepath = mediapath[5:]
            filename = Path(filepath).name
        if filepath == "" or filename == "":
            return
        ui = DialogPdfPagesToImages(self.app, filepath, filename, self)
        if ui.total_pages == 0:
            Message(self.app, _("Image Error"), _("Cannot open: ") + filepath, "warning").exec()
            return
        if ui.exec() != QtWidgets.QDialog.DialogCode.Accepted:
            return
        page_from, page_to, dpi = ui.get_range_and_dpi()
        existing_names = {s['name'] for s in self.source}
        matrix = pymupdf.Matrix(dpi / 72.0, dpi / 72.0)
        progress_ = QtWidgets.QProgressDialog(_("Converting pages"), None, page_from,
                                              page_to + 1, self)
        progress_.setWindowModality(QtCore.Qt.WindowModality.WindowModal)
        try:
            pymu_pdf = pymupdf.open(filepath)
            try:
                for i in range(page_from, page_to + 1):
                    progress_.setValue(i)
                    QtCore.QCoreApplication.processEvents()
                    image_filename = filename + f"_p{i + 1}.jpg"
                    if image_filename in existing_names:
                        # No QMessageBox for each repeated page (large ranges).
                        self.parent_text_edit.append(_("Skipped duplicate image: ") + image_filename)
                        continue
                    pymu_page = pymu_pdf.load_page(i)
                    pymypdf_pixmap = pymu_page.get_pixmap(matrix=matrix)
                    # Other methods 'might' look for the forward slash. CC - ?
                    destination = Path(self.app.project_path) / "images" / image_filename
                    pymypdf_pixmap.save(destination)
                    self.load_media_reference(f"/images/{image_filename}")
                    self.parent_text_edit.append(_("Image loaded from pdf: ") + image_filename)
            finally:
                pymu_pdf.close()
        except Exception as err:
            logger.warning(f"pdf_to_images: {filepath} {err}")
            Message(self.app, _("Image Error"), _("Cannot open: ") + f"{filepath}\n{err}",
                    "warning").exec()
        progress_.setValue(page_to + 1)
        self.load_file_data()
        self.fill_table()
        self.app.delete_backup = False
        self.update_files_in_dialogs()

    def view_original_text_file(self, mediapath: str):
        """ View original text file.
         Args:
            mediapath: String '/docs/' for internal 'docs:/' for external """

        if mediapath[:6] == "/docs/":
            #media_path = self.app.project_path + "/documents/" + mediapath[6:]
            media_path = str(Path(self.app.project_path) / "documents" / mediapath[6:])
            webbrowser.open(media_path)
            return
        if mediapath[:5] == "docs:":
            media_path = mediapath[5:]
            webbrowser.open(media_path)
            return
        logger.error("Cannot open text file in browser " + mediapath)
        print(f"manage_files.view_original_text_file. Cannot open text file in browser {mediapath}")

    def assign_cases_to_file(self):
        """ Assign one or more cases to file. """

        '''row = self.ui.tableWidget.currentRow()
        fid = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())'''
        file_ids = []
        for row in self.visible_selected_rows():
            file_ids.append([int(self.ui.tableWidget.item(row, self.ID_COLUMN).text()), row])
        if not file_ids:
            return
        casenames = self.app.get_casenames()
        ui = DialogSelectItems(self.app, casenames, _("Assign files"), "multi")
        ok = ui.exec()
        if not ok:
            return
        selection = ui.get_selected()
        if not selection:
            return
        cur = self.app.conn.cursor()
        for item in file_ids:
            fid, row = item[0], item[1]
            cur.execute("select fulltext from source where id=?", [fid])
            res = cur.fetchone()
            len_text = 0
            if res is not None and res[0] is not None:
                len_text = len(res[0])
            for case_ in selection:
                # Check if already linked file to case
                cur.execute("select * from case_text where caseid = ? and fid=? and pos0=? and pos1=?",
                            (case_['id'], fid, 0, len_text))
                result = cur.fetchall()
                if len(result) == 0:
                    sql = "insert into case_text (caseid, fid, pos0, pos1, owner, date, memo) values(?,?,?,?,?,?,?)"
                    cur.execute(sql, (case_['id'], fid, 0, len_text, self.app.settings['codername'],
                                      datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"), ""))
                    self.app.conn.commit()
                # Visual feedback, keeping the loaded sources in sync for the header filters
                cases_text = self.get_cases_by_filename(self.ui.tableWidget.item(row, self.NAME_COLUMN).text())
                self.ui.tableWidget.item(row, self.CASE_COLUMN).setText(cases_text)
                self.source[row]['case'] = cases_text
        if self.file_filter_active():
            self.apply_file_filter()

    def rename_database_entry(self):
        """ Rename the database entry of the file. """

        row = self.ui.tableWidget.currentRow()
        if row == -1:
            return
        existing_name = self.ui.tableWidget.item(row, self.NAME_COLUMN).text()
        filenames = []
        for s in self.source:
            filenames.append({'name': s['name']})
        ui = DialogAddItemName(self.app, filenames, _("Rename database entry"), existing_name)
        ui.ui.lineEdit.setText(existing_name)
        ui.exec()
        new_name = ui.get_new_name()
        if new_name is None:
            return
        cur = self.app.conn.cursor()
        cur.execute("update source set name=? where name=?", [new_name, existing_name])
        self.app.conn.commit()
        self.parent_text_edit.append(_("Renamed database file entry: ") + f"{existing_name} -> {new_name}")
        entry = {'old_name': existing_name, 'name': new_name,
                 'fid': int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())}
        self.files_renamed.append(entry)
        self.ui.pushButton_undo.setEnabled(True)
        self.load_file_data()
        self.app.delete_backup = False
        self.update_files_in_dialogs()
        # update doc in vectorstore
        id_ = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())
        if self.app.settings['ai_enable'] == 'True':
            docs = self.app.get_file_texts(file_ids=[id_])
            self.app.ai.sources_vectorstore.import_document(docs[0]['id'], docs[0]['name'], docs[0]['fulltext'])

    def undo_file_rename(self):
        """ Undo file name rename. """

        if len(self.files_renamed) == 0:
            self.ui.pushButton_undo.setEnabled(False)
            # Could occur when file deleted
            return
        ui = DialogSelectItems(self.app, self.files_renamed, _("Undo file rename"), "single")
        ok = ui.exec()
        if not ok:
            return
        selection = ui.get_selected()
        if not selection:
            return
        filenames = self.app.get_filenames()
        for f in filenames:
            if f['name'] == selection['old_name']:
                Message(self.app, _("Cannot undo"), _("Another file has this name"), "warning").exec()
                self.files_renamed = [x for x in self.files_renamed if not (selection['fid'] == x.get('fid'))]
                if len(self.files_renamed) == 0:
                    self.ui.pushButton_undo.setEnabled(False)
                return
        cur = self.app.conn.cursor()
        cur.execute("update source set name=? where name=?", [selection['old_name'], selection['name']])
        self.app.conn.commit()
        self.parent_text_edit.append(_("Reversed renamed database file entry: ") +
                                     f"{selection['name']} -> {selection['old_name']}")
        self.load_file_data()
        if self.app.settings['ai_enable'] == 'True':
            self.app.ai.sources_vectorstore.update_vectorstore()
        self.files_renamed = [x for x in self.files_renamed if not (selection['fid'] == x.get('fid'))]
        if len(self.files_renamed) == 0:
            self.ui.pushButton_undo.setEnabled(False)

    def bulk_rename_database_entry(self):
        """ Bulk Rename source name database entries of the selected files. """

        rows = self.ui.tableWidget.rowCount()
        selected_rows = []
        for row in range(0, rows):
            if not self.ui.tableWidget.isRowHidden(row):
                selected_rows.append([int(self.ui.tableWidget.item(row, self.ID_COLUMN).text()),
                                      self.ui.tableWidget.item(row, self.NAME_COLUMN).text()])
        if not selected_rows:
            return
        # Sort selected rows by their id (order of entry) to ensure sequential renaming
        selected_rows.sort()
        # Display the rename dialog and ask for a base name
        additem = DialogAddItemName(self.app, [], _("Bulk Rename of database file name entries"),
                                    "Give a prefix for the names for all the displayed rows.\n"
                                    "e.g. prefix_001, prefix_002 ...")
        additem.ui.lineEdit.setText("prefix")
        ok = additem.exec()
        if not ok:
            return
        prefix_name = additem.get_new_name()
        if not prefix_name:
            return
        # Perform renaming for all visible rows
        err_msg = ""
        msg = ""
        cur = self.app.conn.cursor()
        for index, row in enumerate(selected_rows):
            fid = row[0]
            existing_name = row[1]
            new_name = f"{prefix_name}_{str(index + 1).zfill(3)}"  # Zero-padded to 3 digits
            # Update the database with the new name
            msg = ""
            try:
                cur.execute("update source set name=? where name=?", [new_name, existing_name])
                self.app.conn.commit()
                msg += f'{_("Renamed database file entry:")} {existing_name} -> {new_name}\n'
            except sqlite3.IntegrityError:
                err_msg += f'_("Bulk Rename. Not renamed in use:") {existing_name}\n'

            # Logging and tracking the renamed entry
            entry = {'old_name': existing_name, 'name': new_name, 'fid': fid}
            self.files_renamed.append(entry)
        self.parent_text_edit.append(msg + err_msg)
        # Updating vectorstore
        if self.app.settings['ai_enable'] == 'True':
            self.app.ai.sources_vectorstore.update_vectorstore()

        self.ui.pushButton_undo.setEnabled(True)
        self.load_file_data()
        self.app.delete_backup = False
        self.update_files_in_dialogs()

    def button_export_file_as_linked_file(self):
        """ User presses button to export current row's file.
         Only to work with an exportable file. """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        row = self.ui.tableWidget.currentRow()
        if row == -1:
            return
        mediapath = None
        id_ = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())
        for s in self.source:
            if s['id'] == id_:
                mediapath = s['mediapath']
        if id_ is None or mediapath is None:
            return
        if mediapath is None or (mediapath is not None and mediapath[0] == "/"):
            self.export_file_as_linked_file(id_, mediapath)

    def export_file_as_linked_file(self, id_:int, mediapath:str):
        """ Move an internal project file into an external location as a linked file.
        # Do not export text files as linked files. e.g. internally created in database, or
        docx, txt, md, odt files.

        Args:
            id_ : the file id, Integer
            mediapath: stored path to media, will be None for text files, or String
        """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        options = QtWidgets.QFileDialog.Option.DontResolveSymlinks | QtWidgets.QFileDialog.Option.ShowDirsOnly
        directory = QtWidgets.QFileDialog.getExistingDirectory(None,
                                                               _("Select directory to save file"),
                                                               self.app.last_export_directory, options)
        if directory == "":
            return
        if directory != self.app.last_export_directory:
            self.app.last_export_directory = directory
        destination = self.app.last_export_directory
        file_directory = ""
        if mediapath is not None and mediapath[:6] == "/docs/":
            mediapath = "/documents/" + mediapath[6:]
        if mediapath is not None:
            file_directory = mediapath.split('/')[1]  # as [0] will be blank
            destination = directory + "/" + mediapath.split('/')[-1]
        if mediapath is None:
            # Some older text files, and QC internally created text Db entries have None as mediapath
            cur = self.app.conn.cursor()
            cur.execute("select name from source where id=?", [id_, ])
            name = cur.fetchone()[0]
            file_directory = "documents"
            mediapath = "/documents/" + name
            destination = Path(directory) / name
        msg = f'{_("Export to")} {destination}\n'
        try:
            move(self.app.project_path + mediapath, destination)
        except Exception as err:
            logger.warning(str(err))
            Message(self.app, _("Cannot export"), _("Cannot export as linked file\n") + str(err), "warning").exec()
            return
        new_mediapath = ""
        if file_directory == "documents":
            new_mediapath = "docs:" + destination
        if file_directory == "images":
            new_mediapath = "images:" + destination
        if file_directory == "audio":
            new_mediapath = "audio:" + destination
        if file_directory == "video":
            new_mediapath = "video:" + destination
        cur = self.app.conn.cursor()
        cur.execute("update source set mediapath=? where id=?", [new_mediapath, id_])
        self.parent_text_edit.append(msg)
        self.app.conn.commit()
        self.update_files_in_dialogs()
        self.load_file_data()
        self.app.delete_backup = False

    def button_import_linked_file(self):
        """ User presses button to import a linked file into the project folder.
        Only to work with an importable file. """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        row = self.ui.tableWidget.currentRow()
        if row == -1:
            return
        mediapath = None
        id_ = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())
        for s in self.source:
            if s['id'] == id_:
                mediapath = s['mediapath']
        if id_ is None or mediapath is None:
            return
        if mediapath is not None and mediapath[0] != "/":
            self.import_linked_file(id_, mediapath)

    def import_linked_file(self, id_:int, mediapath:str):
        """ Import a linked file into the project folder, and change mediapath details.
        Args:
            id_ : Integer
            mediapath : String
        """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        name_split1 = mediapath.split(":")[1]
        filename = name_split1.split('/')[-1]
        if mediapath[0:6] == "audio:":
            copyfile(mediapath[6:], Path(self.app.project_path)/ "audio" / filename)
            mediapath = '/audio/' + filename
        if mediapath[0:6] == "video:":
            copyfile(mediapath[6:], Path(self.app.project_path) / "video" / filename)
            mediapath = '/video/' + filename
        if mediapath[0:7] == "images:":
            copyfile(mediapath[7:], Path(self.app.project_path) / "images" / filename)
            mediapath = '/images/' + filename
        # This must be the last if statement as mediapath can be None
        if mediapath[0:5] == "docs:":
            copyfile(mediapath[5:], Path(self.app.project_path) / "documents" / filename)
            mediapath = None
        cur = self.app.conn.cursor()
        cur.execute("update source set mediapath=? where id=?", [mediapath, id_])
        self.app.conn.commit()
        self.update_files_in_dialogs()
        self.load_file_data()
        self.app.delete_backup = False

    def mark_speakers(self):
        """ Mark the speakers in text files.
         Note: User generated files (not loaded files) have mediapath None.
         Preselection rules (Van's feedback):
         - Every selected AND VISIBLE text file carries over to the speakers dialog,
           not just the last selected one.
         - If nothing usable is selected (no selection, or the selection is hidden by
           the current filter), all VISIBLE text files become the preselection, which
           also carries over an active filter.
         When text files are found the text coding pane opens. """

        text_ids = {item['id'] for item in self.source if item['fulltext']}

        def row_text_file_id(row:int):
            """ id of the text file at a visible row, else None. """
            if self.ui.tableWidget.isRowHidden(row):
                return None  # hidden by the current filter: not usable
            id_item = self.ui.tableWidget.item(row, self.ID_COLUMN)
            if id_item is None:
                return None
            try:
                id_ = int(id_item.text())
            except ValueError:
                return None
            return id_ if id_ in text_ids else None

        ids = []
        selected_rows = sorted({index.row() for index in self.ui.tableWidget.selectedIndexes()})
        for row in selected_rows:
            id_ = row_text_file_id(row)
            if id_ is not None and id_ not in ids:
                ids.append(id_)
        if not ids:
            # Fallback: every visible text file (respects the active filter)
            for row in range(self.ui.tableWidget.rowCount()):
                id_ = row_text_file_id(row)
                if id_ is not None and id_ not in ids:
                    ids.append(id_)
        if not ids:
            Message(self.app, _('Mark speakers'), _('No text file selected.'), 'critical').exec()
            return
        self.main_window.text_coding(task='mark_speakers', doc_id=ids[0], doc_ids=ids)

    def check_attribute_placeholders(self):
        """ Files can be added after attributes are in the project.
         Need to add placeholder attribute values for these, if missing.
         Also,if a file is deleted, check and remove any isolated attribute values. """

        cur = self.app.conn.cursor()
        sql = "select id from source "
        cur.execute(sql)
        sources = cur.fetchall()
        sql = 'select name from attribute_type where caseOrFile ="file"'
        cur.execute(sql)
        attr_types = cur.fetchall()
        insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
        for source in sources:
            for attribute in attr_types:
                sql = "select value from attribute where id=? and name=?"
                cur.execute(sql, (source[0], attribute[0]))
                res = cur.fetchone()
                if res is None:
                    placeholders = [attribute[0], source[0], datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    self.app.settings['codername']]
                    cur.execute(insert_sql, placeholders)
        self.app.conn.commit()

        # Check and delete attribute values where file has been deleted
        attribute_to_del_sql = "SELECT distinct attribute.id FROM  attribute where \
        attribute.id not in (select source.id from source) order by attribute.id asc"
        cur.execute(attribute_to_del_sql)
        res = cur.fetchall()
        for r in res:
            cur.execute("delete from attribute where attr_type='file' and id=?", [r[0], ])
            self.app.conn.commit()

    def export_attributes(self):
        """ Export attributes from table to an Excel file. """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        shortname = self.app.project_name.split(".qda")[0]
        filename = f"{shortname}_file_attributes.xlsx"
        exp_dlg = ExportDirectoryPathDialog(self.app, filename)
        filepath = exp_dlg.filepath
        if filepath is None:
            return
        cols = self.ui.tableWidget.columnCount()
        rows = self.ui.tableWidget.rowCount()
        header = [self.ui.tableWidget.horizontalHeaderItem(i).text() for i in range(0, cols)]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "File Attributes"
        for col, col_name in enumerate(header):
            h_cell = ws.cell(row=1, column=col + 1)
            h_cell.value = col_name
        out_row = 2
        for row in range(rows):
            if self.ui.tableWidget.isRowHidden(row):
                continue
            for col in range(cols):
                cell = ws.cell(row=out_row, column=col + 1)
                data = ""
                try:
                    data = self.ui.tableWidget.item(row, col).text()
                except AttributeError:
                    pass
                cell.value = data
            out_row += 1
        wb.save(filepath)
        msg = _("File attributes exported to: ") + filepath
        Message(self.app, _('File Export'), msg).exec()
        self.parent_text_edit.append(msg)

    def load_file_data(self, order_by :str=""):
        """ Documents images and audio contain the filetype suffix.
        No suffix implies the 'file' was imported from a survey question or created internally.
        This also fills out the table header labels with file attribute names.
        Db versions < 5: Files with the '.transcribed' suffix mean they are associated with audio and
        video files.
        Db version 5+: av_text_id links the text file to the audio/video
        Obtain some file metadata to use in table tooltip.
        Fills table after data is loaded.
        Args:
            order_by: string ""= name asc, "filename desc" = name desc,
            "date asc" = date ascending, "date desc" = date descending, "filetype" = mediapath,
                "casename asc" = by alphabetic casename ascending
                "casename desc" = by alphabetic descending
                "attribute:attribute name" selected attribute - ascending
                "attribute desc: attribute name ttribute - descending
        """

        # Check a placeholder attribute is present for the file, add if missing
        self.check_attribute_placeholders()
        self.source = []
        cur = self.app.conn.cursor()
        placeholders = None
        # Default alphabetic order
        sql = "select name, id, fulltext, mediapath, ifnull(memo,''), owner, date, av_text_id, risid from source " \
              "order by upper(name)"
        if order_by == "filename desc":
            sql += " desc"
        if order_by == "date asc":
            sql = "select name, id, fulltext, mediapath, ifnull(memo,''), owner, date, av_text_id, risid from source " \
                  "order by date, upper(name)"
        if order_by == "date desc":
            sql = "select name, id, fulltext, mediapath, ifnull(memo,''), owner, date, av_text_id, risid from source " \
                  "order by date desc, upper(name) desc"
        if order_by == "filetype":
            sql = "select name, id, fulltext, mediapath, ifnull(memo,''), owner, date, av_text_id, risid from source " \
                  "order by mediapath"
        if order_by == "casename asc":
            sql = "select distinct source.name, source.id, source.fulltext, source.mediapath, ifnull(source.memo,''), "
            sql += "source.owner, source.date, av_text_id, risid "
            sql += "from source left join case_text on source.id=case_text.fid "
            sql += "left join cases on cases.caseid=case_text.caseid "
            sql += "order by cases.name, source.name "

        if order_by == "casename desc":
            sql = "select distinct source.name, source.id, source.fulltext, source.mediapath, ifnull(source.memo,''), "
            sql += "source.owner, source.date, av_text_id, risid "
            sql += "from source left join case_text on source.id=case_text.fid "
            sql += "left join cases on cases.caseid=case_text.caseid "
            sql += "order by cases.name desc, source.name desc"

        if order_by[:14] == "attribute asc:":
            attribute_name = order_by[14:]
            # Two types of ordering character or numeric
            cur.execute("select valuetype from attribute_type where name=?", [attribute_name])
            attr_type = cur.fetchone()[0]
            sql = "select source.name, source.id, fulltext, mediapath, ifnull(source.memo,''), source.owner, "
            sql += "source.date, av_text_id, risid from source join attribute on attribute.id = source.id "
            sql += " where attribute.attr_type = 'file' and attribute.name=? "
            if attr_type == "character":
                sql += "order by lower(attribute.value) asc "
            else:
                sql += "order by cast(attribute.value as numeric) asc"
            placeholders = [attribute_name]

        if order_by[:15] == "attribute desc:":
            attribute_name = order_by[15:]
            # two types of ordering character or numeric
            cur.execute("select valuetype from attribute_type where name=?", [attribute_name])
            attr_type = cur.fetchone()[0]
            sql = "select source.name, source.id, fulltext, mediapath, ifnull(source.memo,''), source.owner, "
            sql += "source.date, av_text_id, risid from source join attribute on attribute.id = source.id "
            sql += " where attribute.attr_type = 'file' and attribute.name=? "
            if attr_type == "character":
                sql += "order by lower(attribute.value) desc "
            else:
                sql += "order by cast(attribute.value as numeric) desc"
            placeholders = [attribute_name]

        if placeholders is not None:
            cur.execute(sql, placeholders)
        else:
            cur.execute(sql)
        result = cur.fetchall()
        for row in result:
            icon, metadata, err_ = self.get_icon_and_metadata(row[1])
            if err_:
                self.parent_text_edit.append(err_ + " : " + row[3])
            self.source.append({'name': row[0], 'id': row[1], 'fulltext': row[2],
                                'mediapath': row[3], 'memo': row[4], 'owner': row[5], 'date': row[6],
                                'av_text_id': row[7], 'risid': row[8], 'metadata': metadata, 'icon': icon,
                                'case': self.get_cases_by_filename(row[0]),
                                'attributes': []})
        # Auto-fix invalid filenames in the database
        cur2 = self.app.conn.cursor()
        for s in self.source:
            if s['name'].strip('.') == '' or s['name'].strip() == '':
                new_name = f"unnamed_file_{s['id']}"
                cur2.execute("update source set name=? where id=?", [new_name, s['id']])
                self.app.conn.commit()
                self.parent_text_edit.append(_("Auto-renamed invalid file: ") + f"'{s['name']}' -> {new_name}")
                s['name'] = new_name
        self.header_labels = [_("Name"), _("Memo"), _("Date"), _("Id"), _("Case")]
        # Attributes
        sql = "select name from attribute_type where caseOrFile='file' order by upper(name)"
        cur.execute(sql)
        attribute_names_res = cur.fetchall()
        self.attribute_names = []  # For AddAttribute dialog
        self.attribute_labels_ordered = []  # Help filling table more quickly
        for att_name in attribute_names_res:
            self.header_labels.append(att_name[0])
            self.attribute_labels_ordered.append(att_name[0])
            self.attribute_names.append({'name': att_name[0]})  # For AddAttribute dialog
        # Add list of attribute values to files, order matches header columns
        sql = "select ifnull(value, '') from attribute where attr_type='file' and attribute.name=? and id=?"
        for s in self.source:
            for att_name in self.attribute_labels_ordered:
                cur.execute(sql, [att_name, s['id']])
                res = cur.fetchone()
                if res:
                    tmp = res[0]
                    # Nicer display
                    if att_name == "Ref_authors":
                        tmp = tmp.replace(";", "\n")
                    s['attributes'].append(tmp)
        self.prune_header_filters()
        self.fill_table()

    def get_icon_and_metadata(self, id_:int):
        """ Get metadata used in table tooltip.
        Called by: create_text_file, load_file_data
        Args:
            id_  : integer source.id
        """

        cur = self.app.conn.cursor()
        cur.execute("select name, fulltext, mediapath from source where id=?", [id_])
        res = cur.fetchone()
        metadata = f"{res[0]}\n"
        icon = QtGui.QIcon(qta.icon('mdi6.text-box-outline', options=[{'scale_factor': 1.2}]))

        # Check if text file is a transcription and add details
        cur.execute("select name from source where av_text_id=?", [id_])
        transcript_res = cur.fetchone()
        if transcript_res is not None:
            metadata += _("Transcript for: ") + f"{transcript_res[0]}\n"
            icon = QtGui.QIcon(qta.icon('mdi6.text', options=[{'scale_factor': 1.2}]))
        if res[1] is not None and len(res[1]) > 0 and res[2] is None:
            metadata += _("Characters: ") + f"{len(res[1]):,}"
            return icon, metadata, ""
        if res[2] is None:
            logger.debug("empty media path error")
            return icon, metadata, ""
        if res[1] is not None and len(res[1]) > 5 and res[2][:6] == "/docs/":
            metadata += _("Characters: ") + f"{len(res[1]):,}"
            return icon, metadata, ""
        if res[1] is not None and len(res[1]) > 5 and res[2][:5] == "docs:":
            metadata += _("Characters: ") + f"{len(res[1]):,}"
            icon = QtGui.QIcon(qta.icon('mdi6.text-box-check-outline', options=[{'scale_factor': 1.2}]))
            return icon, metadata, ""

        abs_path = ""
        if 'audio:' == res[2][0:6]:
            abs_path = res[2][6:]
        elif 'video:' == res[2][0:6]:
            abs_path = res[2][6:]
        elif 'images:' == res[2][0:7]:
            abs_path = res[2][7:]
        else:
            abs_path = self.app.project_path + res[2]

        if res[2][:8] == "/images/":
            icon = QtGui.QIcon(qta.icon('mdi6.image-outline', options=[{'scale_factor': 1.2}]))
            try:
                image = Image.open(abs_path)
                w, h = image.size
            except (FileNotFoundError, PIL.UnidentifiedImageError):
                metadata += _("Cannot locate media. ") + abs_path
                return icon, metadata, "Not found error"
            except PIL.Image.DecompressionBombError:
                metadata += _("Image too large for PIL module. (DecompressionBombError): ") + abs_path
                return icon, metadata, "DecompressionError"
            metadata += f"W: {w:,} x H: {h:,}"
        if res[2][:7] == "images:":
            icon = QtGui.QIcon(qta.icon('mdi6.image-check-outline', options=[{'scale_factor': 1.2}]))
            try:
                image = Image.open(abs_path)
                w, h = image.size
            except (FileNotFoundError, PIL.UnidentifiedImageError, AttributeError):
                metadata += _("Cannot locate media. ") + abs_path
                return icon, metadata, "Other error"
            except PIL.Image.DecompressionBombError:
                metadata += _("Image too large for PIL module. (DecompressionBombError): ") + abs_path
                return icon, metadata, "DecompressionBombError"
            metadata += f"W: {w:,} x H: {h:,}"
        if res[2][:7] == "/video/":
            icon = QtGui.QIcon(qta.icon('mdi6.video-outline', options=[{'scale_factor': 1.2}]))
        if res[2][:6] == "video:":
            icon = QtGui.QIcon(qta.icon('mdi6.video-check-outline', options=[{'scale_factor': 1.2}]))
        if res[2][:7] == "/audio/":
            icon = QtGui.QIcon(qta.icon('mdi6.play', options=[{'scale_factor': 1.2}]))
        if res[2][:6] == "audio:":
            icon = QtGui.QIcon(qta.icon('mdi6.play-protected-content', options=[{'scale_factor': 1.2}]))
        if res[2][:6] in ("/audio", "audio:", "/video", "video:"):
            if not Path(abs_path).exists():
                metadata += _("Cannot locate media. ") + abs_path
                return icon, metadata, "Not found Error"
            if vlc:
                try:
                    try:
                        from .media_player_qt import metadata_vlc_instance
                        instance = metadata_vlc_instance(vlc)  # cached: metadata only
                    except NameError as name_err:
                        # NameError: no function 'libvlc_new'
                        logger.error(f"vlc.Instance: {name_err}")
                        return icon, metadata, f"Cannot use vlc. {name_err}"
                    media = instance.media_new(abs_path)
                    media.parse()
                    msecs = media.get_duration()
                    duration_txt = msecs_to_hours_mins_secs(msecs)
                    metadata += " " + _("Duration: ") + duration_txt
                    return icon, metadata, ""
                except AttributeError as err:
                    logger.warning(str(err))
                    metadata += _("Cannot locate media. ") + f"{abs_path}\n{err}"
                    return icon, metadata, "Not found error"
            else:
                metadata += _("Cannot get media duration.\nVLC not installed.")
                return icon, metadata, "Other error"
        bytes_ = 0
        try:
            bytes_ = Path(abs_path).stat().st_size
        except OSError as e_:
            logger.warning(str(e_))
        metadata += f"\nBytes: {bytes_}"
        if 1024 < bytes_ < 1024 * 1024:
            metadata += f"  {int(bytes_ / 1024)}KB"
        if bytes_ > 1024 * 1024:
            metadata += f"  {int(bytes_ / 1024 / 1024)}MB"
        # Get case names linked to the file
        txt = self.get_cases_by_filename(res[0])
        if txt != "":
            metadata += f'\n{_("Case linked:")}\n{txt}'
        return icon, metadata, ""

    def get_cases_by_filename(self, name: str):
        """ Called by get_icon_and_metadata, get_file_data
        Args:
            name String of filename """

        cur = self.app.conn.cursor()
        # Case_text is the table, but this also links av and images
        sql = "select distinct cases.name from cases join case_text on case_text.caseid=cases.caseid "
        sql += "join source on source.id=case_text.fid where source.name=? "
        text_ = ""
        cur.execute(sql, [name])
        res = cur.fetchall()
        if res:
            for r in res:
                text_ += f"{r[0]};"
            text_ = text_[:-1]
        return text_

    def add_attribute(self):
        """ When add button pressed, opens the AddAtribute dialog to get new attribute text.
        Then get the attribute type through a dialog.
        AddAttribute dialog checks for duplicate attribute name.
        New attribute is added to the model and database.
        Reserved attribute words - used for imported references:
        Ref_Type (Type of Reference) – character variable
        Ref_Author (authors list) – character
        Ref_Title – character
        Ref_Year (of publication) – numeric
        Ref_Journal - character
        """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        ui = DialogAddAttribute(self.app)
        ok = ui.exec()
        if not ok:
            return
        name = ui.new_name
        value_type = ui.value_type
        if name == "":
            return
        self.attribute_names.append({'name': name})
        # update attribute_type list and database
        now_date = str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        try:
            cur = self.app.conn.cursor()
            cur.execute("insert into attribute_type (name,date,owner,memo,caseOrFile, valuetype) values(?,?,?,?,?,?)",
                        (name, now_date, self.app.settings['codername'], "", 'file', value_type))
            sql = "select id from source"
            cur.execute(sql)
            ids = cur.fetchall()
            for id_ in ids:
                sql = "insert into attribute (name, value, id, attr_type, date, owner) values (?,?,?,?,?,?)"
                cur.execute(sql, (name, "", id_[0], 'file', now_date, self.app.settings['codername']))
            self.app.conn.commit()
            self.app.delete_backup = False
        except Exception as e_:
            print(e_)
            logger.debug(str(e_))
            self.app.conn.rollback()  # Revert all changes
            raise
        self._emit_project_table_changes(["attribute_type", "attribute"])
        self.load_file_data()
        self.fill_table()
        self.parent_text_edit.append(f'{_("Attribute added to files:")} {name}, {_("type")}: {value_type}')

    def cell_double_clicked(self):
        """ View file """

        y = self.ui.tableWidget.currentColumn()
        if y == self.NAME_COLUMN:
            self.view()

    def cell_selected(self):
        """ When the table widget memo cell is selected display the memo.
        Update memo text, or delete memo by clearing text.
        If a new memo, also show in table widget by displaying MEMO in the memo column. """

        x = self.ui.tableWidget.currentRow()
        y = self.ui.tableWidget.currentColumn()
        self.update_label_file_count()
        if y == self.MEMO_COLUMN:
            name = self.source[x]['name'].lower()
            cur = self.app.conn.cursor()
            # Need to dynamically get the memo text in case it has been changed in a coding dialog
            cur.execute('select memo from source where id=?', [self.source[x]['id']])
            self.source[x]['memo'] = cur.fetchone()[0]
            if name[-5:] == ".jpeg" or name[-4:] in ('.jpg', '.png', '.gif'):
                ui = DialogMemo(self.app, _("Memo for file ") + self.source[x]['name'],
                                self.source[x]['memo'], entity_type="file", entity_id=self.source[x]['id'])
                ui.exec()
                self.source[x]['memo'] = ui.memo
                cur.execute('update source set memo=? where id=?', (ui.memo, self.source[x]['id']))
                self.app.conn.commit()
            else:
                ui = DialogMemo(self.app, _("Memo for file ") + self.source[x]['name'],
                                self.source[x]['memo'], entity_type="file", entity_id=self.source[x]['id'])
                ui.exec()
                self.source[x]['memo'] = ui.memo
                cur = self.app.conn.cursor()
                cur.execute('update source set memo=? where id=?', (ui.memo, self.source[x]['id']))
                self.app.conn.commit()
            if self.source[x]['memo'] == "":
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem())
            else:
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem("Memo"))
            if self.file_filter_active():
                # The changed memo may now be excluded by a header filter
                self.apply_file_filter()

    def cell_modified(self):
        """ Attribute values can be changed.  """

        x = self.ui.tableWidget.currentRow()
        y = self.ui.tableWidget.currentColumn()
        # Update attribute value
        if y > self.CASE_COLUMN:
            value = str(self.ui.tableWidget.item(x, y).text()).strip()
            attribute_name = self.header_labels[y]
            cur = self.app.conn.cursor()
            # Check numeric for numeric attributes, clear "" if it cannot be cast
            cur.execute("select valuetype from attribute_type where caseOrFile='file' and name=?", (attribute_name,))
            result = cur.fetchone()
            if result is None:
                return
            if result[0] == "numeric" and value != "":
                try:
                    float(value)
                except ValueError:
                    self.ui.tableWidget.item(x, y).setText("")
                    value = ""
                    msg = _("This attribute is numeric")
                    Message(self.app, _("Warning"), msg, "warning").exec()
            cur.execute("update attribute set value=? where id=? and name=? and attr_type='file'",
                        (value, self.source[x]['id'], attribute_name))
            self.app.conn.commit()
            self._emit_project_table_changes(["attribute"])

            # Update self.source[attributes]
            # Add list of attribute values to files, order matches header columns
            sql = "select ifnull(value, '') from attribute where attr_type='file' and attribute.name=? and id=?"
            self.source[x]['attributes'] = []
            for att_name in self.attribute_labels_ordered:
                cur.execute(sql, [att_name, self.source[x]['id']])
                res = cur.fetchone()
                if res:
                    tmp = res[0]
                    # For nicer display
                    if att_name == "Ref_authors":
                        tmp = tmp.replace(";", "\n")
                    self.source[x]['attributes'].append(tmp)

            self.app.delete_backup = False
            self.ui.tableWidget.resizeColumnsToContents()
            if self.file_filter_active():
                # The edited value may now be excluded by a header filter
                self.apply_file_filter()

    def view(self):
        """ View and edit text file contents.
        Alternatively view an image, audio or video media. """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        x = self.ui.tableWidget.currentRow()
        self.ui.tableWidget.selectRow(x)
        if self.source[x]['mediapath'] is not None and 'docs:' != self.source[x]['mediapath'][0:5]:
            if len(self.source[x]['mediapath']) > 6 and self.source[x]['mediapath'][:7] in ("/images", "images:"):
                self.view_image(x)
                return
            if len(self.source[x]['mediapath']) > 5 and self.source[x]['mediapath'][:6] in ("/video", "video:"):
                self.view_av(x)
                return
            if len(self.source[x]['mediapath']) > 5 and self.source[x]['mediapath'][:6] in ("/audio", "audio:"):
                self.view_av(x)
                return
        # PDF fulltext is not editable (it must match the page extraction);
        # "Convert to txt" creates an editable copy as a new text source.
        mediapath_ = self.source[x]['mediapath']
        if mediapath_ is not None and mediapath_.lower().endswith(".pdf") and \
                (mediapath_[0:6] == '/docs/' or mediapath_[0:5] == 'docs:'):
            if mediapath_[0:6] == '/docs/':
                pdf_filepath = Path(self.app.project_path) / "documents" / mediapath_[6:]
            else:
                pdf_filepath = mediapath_[5:]
            preview = DialogPdfPreview(self.app, pdf_filepath, self.source[x]['name'], self)
            preview.exec()
            if preview.convert_txt_requested:
                # Reuses the context-menu extraction (name uniqueness, scanned
                # PDFs, attributes, vectorstore).
                self.extract_pdf_text_copy(x)
            return
        ui = DialogEditTextFile(self.app, self.source[x]['id'])
        result = ui.exec()
        # Get fulltext if changed (for metadata)
        cur = self.app.conn.cursor()
        cur.execute("select fulltext from source where id=?", [self.source[x]['id']])
        res = cur.fetchone()
        fulltext = ""
        if res is not None:
            fulltext = res[0]
        self.source[x]['fulltext'] = fulltext
        # The editor saved changes: notify the bus so open coding dialogs
        # (code_text, code_pdf) reload and do not keep stale positions.
        if result == QtWidgets.QDialog.DialogCode.Accepted and \
                getattr(self.app, "project_events", None) is not None:
            self.app.project_events.emit_table_changes(
                ['source', 'code_text', 'annotation', 'case_text'], source=self)

    def view_av(self, x: int):
        """ View an audio or video file. Edit the memo. Edit the transcript file.
        Added try block in case VLC bindings do not work.
        Uses a non-modal dialog.
        Args:
            x  :  row number Integer
        """

        if not vlc and self.app.settings.get('av_player', 'vlc') != 'qt':
            # Without python-vlc the Qt Multimedia backend still plays media
            self.app.settings['av_player'] = 'qt'
            self.app.write_config_ini(self.app.settings, self.app.ai_models)
        # Check media exists
        abs_path = ""
        if self.source[x]['mediapath'][0:6] in ('/audio', '/video'):
            abs_path = self.app.project_path + self.source[x]['mediapath']
        if self.source[x]['mediapath'][0:6] in ('audio:', 'video:'):
            abs_path = self.source[x]['mediapath'][6:]
        if not Path(abs_path).exists():
            self.parent_text_edit.append(_("Bad link or non-existent file ") + abs_path)
            return
        try:
            ui = DialogViewAV(self.app, self.source[x])
            # ui.exec()  # this dialog does not display well on Windows 10 so trying .show()
            # The vlc window becomes unmovable and not resizable
            self.av_dialog_open = ui
            ui.show()
            if self.source[x]['memo'] == "":
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem())
            else:
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem(_("Memo")))
        except Exception as err:
            logger.warning(str(err))
            Message(self.app, _('view AV error'), str(err), "warning").exec()
            self.av_dialog_open = None
            return

    def create_transcription_for_av(self, av_id, av_name):
        """ Create an empty transcription text source for an audio/video file and link it
        through av_text_id. Used when the original transcription was deleted (leaving a
        dangling av_text_id) or never existed. Mirrors the transcription-creation block
        in load_media_reference.

        param:
            av_id: source id of the audio/video file, Integer
            av_name: name of the audio/video file, String

        returns:
            (tr_id, tr_name) of the new transcription text source, or (None, None) on duplicate
        """

        tr_name = av_name + ".txt"
        if any(s['name'] == tr_name for s in self.source):
            QtWidgets.QMessageBox.warning(self, _('Duplicate file'),
                                          _("Duplicate filename.\nFile not created"))
            return None, None
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        entry = {'name': tr_name, 'id': -1, 'fulltext': "", 'mediapath': None, 'memo': "",
                 'owner': self.app.settings['codername'], 'date': now, 'av_text_id': None}
        cur = self.app.conn.cursor()
        cur.execute("insert into source(name,fulltext,mediapath,memo,owner,date) values(?,?,?,?,?,?)",
                    (entry['name'], entry['fulltext'], entry['mediapath'], entry['memo'], entry['owner'],
                     entry['date']))
        self.app.conn.commit()
        cur.execute("select last_insert_rowid()")
        tr_id = cur.fetchone()[0]
        entry['id'] = tr_id
        # Link the av file entry to this text file
        cur.execute("update source set av_text_id=? where id=?", [tr_id, av_id])
        self.app.conn.commit()

        # add doc to vectorstore
        if self.app.settings['ai_enable'] == 'True':
            self.app.ai.sources_vectorstore.import_document(entry['id'], entry['name'], entry['fulltext'])

        # Add file attribute placeholders
        cur.execute('select name from attribute_type where caseOrFile ="file"')
        attr_types = cur.fetchall()
        insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
        for a in attr_types:
            cur.execute(insert_sql, [a[0], tr_id, now, self.app.settings['codername']])
        self.app.conn.commit()
        self._emit_project_table_changes(["source", "attribute"])  #

        self.source.append(entry)
        self.parent_text_edit.append(tr_name + _(" created."))
        return tr_id, tr_name

    def import_transcription_from_file(self, id_):
        """ Load transcript text from an external file (e.g. a noScribe .txt or .html export)
        into the .txt transcription that QualCoder auto-creates for an audio or video file.

        The audio/video source row stores the linked transcription text source id in av_text_id
        (see load_media_reference). The content of the chosen file replaces that transcription's
        fulltext. Reading mirrors load_file_text so behaviour matches a normal text import:
        charset detection for plain text, html_to_text for html/htm, line-ending normalisation
        and pseudonym substitution. Warns before overwriting an existing transcription, codings
        or annotations.

        param:
            id_: source id of the selected audio/video file, Integer
        """

        # Release the media file if an AV dialog is open
        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None

        av_source = next((s for s in self.source if s['id'] == id_), None)
        if av_source is None:
            return

        # --- Locate the linked transcription text source via av_text_id
        cur = self.app.conn.cursor()
        cur.execute("select av_text_id from source where id=?", [id_])
        res = cur.fetchone()
        tr_id = res[0] if res is not None else None
        # The link may be stale: the transcription source could have been deleted while
        # av_text_id still points to its old id. Confirm the row actually exists.
        tr_name = None
        if tr_id is not None:
            cur.execute("select name from source where id=?", [tr_id])
            res = cur.fetchone()
            if res is None:
                tr_id = None  # dangling link -> treat as missing
            else:
                tr_name = res[0]
        # Fallback for older projects with no link: match a text source named "<av name>.txt"
        if tr_id is None:
            cur.execute("select id, name from source where name=? and mediapath is null",
                        [av_source['name'] + ".txt"])
            res = cur.fetchone()
            if res is not None:
                tr_id, tr_name = res[0], res[1]
                cur.execute("update source set av_text_id=? where id=?", [tr_id, id_])
                self.app.conn.commit()
        # Still nothing: recreate an empty transcription and link it (mirrors load_media_reference)
        if tr_id is None:
            tr_id, tr_name = self.create_transcription_for_av(id_, av_source['name'])
            if tr_id is None:
                return

        # --- Choose the transcription file
        start_dir = str(Path(self.app.project_path).parent)
        filepath, _filter = QtWidgets.QFileDialog.getOpenFileName(
            self, _("Select transcription file"), start_dir,
            _("Transcript files") + " (*.txt *.srt *.vtt *.md *.html *.htm);;" + _("All files") + " (*.*)")
        if not filepath:
            return

        # --- Read the file (same logic as load_file_text for txt and html)
        text_ = ""
        try:
            if filepath[-5:].lower() == ".html" or filepath[-4:].lower() == ".htm":
                with open(filepath, "r", encoding="utf-8", errors="surrogateescape") as sourcefile:
                    html_text = sourcefile.read()
                text_ = html_to_text(html_text)
            else:
                text_, detected_encoding = self.decode_text_with_best_encoding(filepath)
                logger.debug(f"Import transcription from {filepath} decoded as {detected_encoding}")
        except Exception as err:
            logger.warning(str(err))
            Message(self.app, _("Warning"), _("Cannot read file") + f"\n{filepath}\n{err}", "warning").exec()
            return
        if text_ is None:
            text_ = ""
        # Normalise line endings and strip BOM so stored positions match the editor (see load_file_text)
        text_ = text_.replace("\r\n", "\n").replace("\r", "\n")
        if text_ and text_[0] == "\ufeff":
            text_ = text_[1:]
        if text_.strip() == "":
            Message(self.app, _("Warning"),
                    _("The selected file has no readable text.") + f"\n{filepath}", "warning").exec()
            return

        # --- Apply pseudonyms, consistent with text import
        for pseudonym in self.load_pseudonyms():
            text_ = re.sub(rf"(?<!\w){re.escape(pseudonym['original'])}(?!\w)", pseudonym['pseudonym'], text_)

        # --- Warn before overwriting existing transcription, codings or annotations
        cur.execute("select length(ifnull(fulltext,'')) from source where id=?", [tr_id])
        existing_len = cur.fetchone()[0]
        cur.execute("select count(*) from code_text where fid=?", [tr_id])
        codings = cur.fetchone()[0]
        cur.execute("select count(*) from annotation where fid=?", [tr_id])
        annotations = cur.fetchone()[0]
        if existing_len > 0 or codings > 0 or annotations > 0:
            warn = _("The transcription already contains text.") + "\n"
            warn += _("Codings: ") + str(codings) + "    " + _("Annotations: ") + str(annotations) + "\n\n"
            warn += _("Replacing it may shift or remove existing codings and annotations. Continue?")
            reply = QtWidgets.QMessageBox.question(
                self, _("Overwrite transcription"), warn,
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                QtWidgets.QMessageBox.StandardButton.No)
            if reply == QtWidgets.QMessageBox.StandardButton.No:
                return

        # --- Store the transcription
        cur.execute("update source set fulltext=? where id=?", [text_, tr_id])
        self.app.conn.commit()
        for s in self.source:
            if s['id'] == tr_id:
                s['fulltext'] = text_
                break

        # Update the AI vectorstore if enabled
        if self.app.settings['ai_enable'] == 'True':
            self.app.ai.sources_vectorstore.import_document(tr_id, tr_name, text_)

        self.parent_text_edit.append(_("Transcription imported into ") + tr_name)
        self.load_file_data()
        self._emit_project_table_changes(["source"])  # notify other open dialogs
        Message(self.app, _("Transcription imported"),
                _("Transcription loaded into: ") + tr_name).exec()

    def view_image(self, x:int):
        """ View an image file and edit the image memo.
        Args:
            x  :  row number Integer
        """

        # Check image exists
        abs_path = ""
        if self.source[x]['mediapath'][:7] == "images:":
            abs_path = self.source[x]['mediapath'][7:]
        else:
            abs_path = self.app.project_path + self.source[x]['mediapath']
        if not Path(abs_path).exists():
            self.parent_text_edit.append(_("Bad link or non-existent file ") + abs_path)
            return
        ui = DialogViewImage(self.app, self.source[x])
        ui.exec()
        memo = ui.ui.textEdit.toPlainText()
        if self.source[x]['memo'] != memo:
            self.source[x]['memo'] = memo
            cur = self.app.conn.cursor()
            cur.execute('update source set memo=? where id=?', (self.source[x]['memo'],
                                                                self.source[x]['id']))
            self.app.conn.commit()
        if self.source[x]['memo'] == "":
            self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem())
        else:
            self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem(_("Memo")))

    def create_text_file(self):
        """ Create a new text file by entering text into the dialog.
        Implements the QtDesigner memo dialog. """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        ui = DialogAddItemName(self.app, self.source, _('New File'), _('Enter file name'))
        ui.exec()
        name = ui.get_new_name()
        # --- VALIDATE FILENAME
        if name is None:
            return
        if name.strip('.') == '' or name.strip() == '':
            Message(self.app, _("Warning"), _("Invalid file name."), "warning").exec()
            return

        # Create entry details to add to self.source and to database
        item = {'name': name, 'id': -1, 'fulltext': '', 'memo': "",
                'owner': self.app.settings['codername'], 'date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'mediapath': None, 'icon': None, 'metadata': '', 'case': ""}
        # Update database
        cur = self.app.conn.cursor()
        cur.execute("insert into source(name,fulltext,mediapath,memo,owner,date) values(?,?,?,?,?,?)",
                    (
                        item['name'], item['fulltext'], item['mediapath'], item['memo'], item['owner'],
                        item['date']))
        self.app.conn.commit()
        cur.execute("select last_insert_rowid()")
        id_ = cur.fetchone()[0]
        item['id'] = id_
        ui = DialogEditTextFile(self.app, id_)
        ui.ui.textEdit.setAcceptRichText(False)
        ui.exec()
        icon, metadata, err_ = self.get_icon_and_metadata(id_)
        item['icon'] = icon
        item['metadata'] = metadata
        item['attributes'] = []
        item['risid'] = None
        # Add file attribute placeholders
        att_sql = 'select name from attribute_type where caseOrFile ="file"'
        cur.execute(att_sql)
        attr_types = cur.fetchall()
        insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
        for a in attr_types:
            placeholders = [a[0], id_, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            self.app.settings['codername']]
            cur.execute(insert_sql, placeholders)
            self.app.conn.commit()
            item['attributes'].append('')
        self.update_files_in_dialogs()
        self.parent_text_edit.append(_("File created: ") + item['name'])
        self.source.append(item)
        self.fill_table()
        self.app.delete_backup = False

    def link_files(self):
        """ Trigger to link to file location. """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        self.import_files(True)

    def import_survey(self):
        """ Import from CSV/TSV/ODS/XLSX Header row to contain column headings.
        Process Qualitative Texts, Cases, Attributes and optional assign autocoding.
        Can assign attributes to either files or cases.
        The Case name can be absent. Or can be from one primary column, or can also collate values from additional columns.
        Qualitative texts from multiple columns are collated into one file.
        The header for each block of text is the column name from the survey.
        """

        filepath, filter_type = QtWidgets.QFileDialog.getOpenFileName(None, _("Select Survey"), "",
                                                                      _("Data files") + " (*.csv *.CSV *.tsv *.TSV *.ods *.ODS *.xlsx *.XLSX *.xls *.XLS)")
        if not filepath: return

        msg = _("Import from survey: ") + f"{filepath}"
        if filepath.lower().endswith('.csv') or filepath.lower().endswith('.tsv'):
            delimiter = ','
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    first_line = f.readline()
            except UnicodeDecodeError:
                with open(filepath, 'r', encoding='latin1') as f:
                    first_line = f.readline()
            # Potentially risky method to determine delimiter
            counts = {',': first_line.count(','), ';': first_line.count(';'), '\t': first_line.count('\t'), '|': first_line.count('|')}
            best_delimiter = max(counts, key=counts.get)
            if counts[best_delimiter] > 0:
                delimiter = best_delimiter
            msg += _("\nPresumed column delimiter for csv or tsv file: ") + delimiter
            if delimiter == "\t": msg += "tab"

            # Remove pandas trailing ".0" from numbers. Treat all columns as string
            try:
                df = pd.read_csv(filepath, sep=delimiter, encoding='utf-8', dtype=str)
            except UnicodeDecodeError:
                df = pd.read_csv(filepath, sep=delimiter, encoding='latin1', dtype=str)

        elif filepath.lower().endswith('.ods'):
            df = pd.read_excel(filepath, engine='odf', dtype=str)
        else:
            df = pd.read_excel(filepath, dtype=str)

        # User determines: Case, Attributes, Qual text, code texts, attributes as case or file.
        columns = [str(c) for c in df.columns]
        dialog = DialogSurveyImport(columns, self)
        if not dialog.exec(): return
        text_cols, case_cols, attr_cols = dialog.get_selections()
        #filename_col = dialog.get_filename_column()
        autocode_enabled = dialog.get_autocode_setting()
        attr_file_or_case = "case"
        if not dialog.get_case_setting() or not case_cols:
            attr_file_or_case = "file"
        popup_msg = ""
        if not text_cols:
            popup_msg = _("No columns assigned as qualitative.")
            msg += _("\nNo columns assigned as qualitative.")
            #msg += _("\nEmpty files will be created with attributes assigned to them.")
        cur = self.app.conn.cursor()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_attributes = {}  # Change from character to numeric attribute_type after checking when loading data
        for col in attr_cols:
            cur.execute("select name from attribute_type where name=?", [col])
            if not cur.fetchone():
                cur.execute("insert into attribute_type (name, date, owner, memo, caseOrFile, valuetype) values(?,?,?,?,?,?)",
                            (col, now, self.app.settings['codername'], "", attr_file_or_case, "character"))
                new_attributes[col] = 'numeric'

        def sanitize_name(name_str:str):
            return re.sub(r'[\\/:*?"<>|]', '-', str(name_str)).strip()

        # Get existing database filenames and source/ids
        # Needed for matching an updated data set to existing survey rows
        existing_files = self.app.get_text_filenames()

        count = 0
        for index, row in df.iterrows():
            fulltext = ""
            code_positions = []

            for t_col in text_cols:
                val = str(row[t_col]) if pd.notna(row[t_col]) else ""
                if val.strip():
                    if fulltext:
                        fulltext += "\n\n"
                    header = f"[{t_col}]:\n"
                    val_clean = val.strip()
                    current_len = len(fulltext)
                    fulltext += header + val_clean
                    start_pos = current_len + len(header)
                    end_pos = start_pos + len(val_clean)
                    code_positions.append((t_col, start_pos, end_pos, val_clean))
            # if not fulltext.strip(): continue

            case_name = ""
            if case_cols:
                c_vals = [str(row[c]) for c in case_cols if pd.notna(row[c])]
                case_name = sanitize_name("_".join(c_vals))
            '''if filename_col and pd.notna(row[filename_col]):
                base_filename = sanitize_name(row[filename_col])'''
            if case_name:
                base_filename = f"Survey_{case_name}"
            else:
                base_filename = f"Survey_Row_{index+1}"
            if not base_filename:
                base_filename = f"Survey_Row_{index+1}"

            filename = base_filename
            suffix = 1
            while True:
                cur.execute("select name from source where name=?", [filename])
                if not cur.fetchone():
                    break
                filename = f"{base_filename}_{suffix}"
                suffix += 1

            if text_cols:
                filename_txt = filename + ".txt"
                filepath_save = Path(self.app.project_path) / "documents" / filename_txt
                with open(filepath_save, 'w', encoding='utf-8') as f:
                    f.write(fulltext)
                cur.execute("insert into source(name, fulltext, mediapath, memo, owner, date) values(?,?,?,?,?,?)",
                            (filename, fulltext, None, "", self.app.settings['codername'], now))
                file_id = cur.lastrowid
            else:
                file_id = None

            if autocode_enabled:
                for col_name, start_pos, end_pos, text_chunk in code_positions:
                    cur.execute("select cid from code_name where name=?", [col_name])
                    res_code = cur.fetchone()
                    if res_code:
                        cid = res_code[0]
                    else:
                        grays = next((colr for colr in colour_ranges if colr['name'] == 'gray'), None)
                        color = colors[randint(grays['min'], grays['max'] - 1)]
                        cur.execute("insert into code_name (name, memo, owner, date, color) values(?,?,?,?,?)",
                                    (col_name, "", self.app.settings['codername'], now, color))
                        cid = cur.lastrowid
                    if text_cols:
                        cur.execute("insert into code_text (cid, fid, seltext, pos0, pos1, owner, date, memo) values(?,?,?,?,?,?,?,?)",
                                (cid, file_id, text_chunk, start_pos, end_pos, self.app.settings['codername'], now, ""))
            #  Correct auto-code positions, posiciones de autocodificación (aun sin resolver)
            case_id = -1
            if case_name:
                cur.execute("select caseid from cases where name=?", [case_name])
                res = cur.fetchone()
                if res:
                    case_id = res[0]
                else:
                    cur.execute("insert into cases (name, memo, owner, date) values(?,?,?,?)",
                                (case_name, "", self.app.settings['codername'], now))
                    case_id = cur.lastrowid
                if text_cols:
                    cur.execute("insert into case_text (caseid, fid, pos0, pos1, owner, date, memo) values(?,?,?,?,?,?,?)",
                            (case_id, file_id, 0, len(fulltext), self.app.settings['codername'], now, ""))

            # Insert file or case attributes from survey, and check if character or numeric
            updated_data = False
            for i, col in enumerate(attr_cols):
                val = str(row[col]) if pd.notna(row[col]) else ""
                if val != "":
                    try:
                        float(val)
                    except ValueError:
                        try:
                            new_attributes[col] = "character"
                        except KeyError:
                            pass
                file_or_case_id = file_id
                if attr_file_or_case == "case":
                    file_or_case_id = case_id
                # Check if fid is None, try getting from a file name match
                if attr_file_or_case == "file":
                    for item in existing_files:
                        if item['name'] == base_filename:
                            file_or_case_id = item['id']
                            break

                try:
                    # print(f"Insert name:{col}, value:{val}, Fid/Cid:{file_or_case_id}, F/C:{attr_file_or_case}")
                    cur.execute("insert into attribute (name, value, id, attr_type, date, owner) values(?,?,?,?,?,?)",
                                (col, val, file_or_case_id, attr_file_or_case, now, self.app.settings['codername']))
                except sqlite3.IntegrityError as err:
                    # Replace existing file or case attribute data with new data
                    try:
                        cur.execute("update attribute set value=?, date=? where name=? and id=? and attr_type=?",
                                    (val, now, col, file_or_case_id, attr_file_or_case))
                        updated_data = True
                    except Exception as update_err:
                        print("Update survey data:", update_err)
                        logger.error(f"update survey data: {update_err}")
    
            count += 1

        # Update attribute type for new attributes, if values were all numeric, default was character
        msg += "\n" + _("Attributes") + " → (" + attr_file_or_case + "):"
        for key, value in new_attributes.items():
            cur.execute("update attribute_type set valuetype=? where name=?", [value, key])
            msg += f"\n    {key} - {value}"
        self.app.conn.commit()
        changed_tables = {"source"}
        if attr_cols:
            changed_tables.update({"attribute_type", "attribute"})
        if case_cols:
            changed_tables.update({"cases", "case_text"})
        if autocode_enabled:
            changed_tables.update({"code_name", "code_text"})
        self._emit_project_table_changes(sorted(changed_tables))
        if updated_data:
            msg += "\n" + _("Some existing data updated.")
        msg += f"\n{count} " + _("rows imported.")
        msg += "\n" + "▔" * 20  # U2594
        self.app.delete_backup = False
        self.update_files_in_dialogs()
        self.load_file_data()
        self.parent_text_edit.append("<h2>" + _("Survey Import") + "</h2>")
        self.parent_text_edit.append(msg)
        dlg_msg = popup_msg + "\n" + _("{} rows imported.").format(count) + " " * 10
        if updated_data:
            dlg_msg += "\n" + _("Some existing data updated")
        Message(self.app, _("Import successful."), dlg_msg).exec()

    def import_files(self, link:bool=False):
        """ Import files and store into relevant directories (documents, images, audio, video).
        Convert documents to plain text and store this in data.qda
        Can import from plain text files, also import from html, odt, docx, rtf, tex, and md.
        md is text Markdown format.
        Note importing from html, odt, docx, rtf, tex all formatting is lost.
        Imports images as jpg, jpeg, png which are stored in an images directory.
        Imports audio as flac, mp3, wav, ogg, m4a which are stored in an audio directory.
        Imports video as mp4, mov, wmv, webm, m4v which are stored in a video directory.

        Args:
            link:   False - files are imported into project folder,
                    True - files are linked and not imported
        """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        response = QtWidgets.QFileDialog.getOpenFileNames(
            None, 
            _('Open file'),
            self.default_import_directory
        )
        imports = response[0]
        if not imports:
            return

        steps = len(imports)
        file_number = 0
        first_filename = Path(imports[0]).name
        progress = QtWidgets.QProgressDialog(first_filename, None, file_number, steps, self)
        progress.setWindowModality(QtCore.Qt.WindowModality.WindowModal)
        progress.setWindowTitle(_("Importing"))
        progress.setMinimumDuration(0)  # Show immediately
        # Without these, QProgressDialog auto-hides at maximum: with a single
        # file it closed before the page-by-page extraction started.
        progress.setAutoReset(False)
        progress.setAutoClose(False)
        progress.show()
        self.default_import_directory = str(Path(imports[0]).parent)
        pdf_msg = ""
        # Highlight coding option, asked ONCE per batch and only when the first PDF
        # with highlight annotations is detected (tri-state: None = not asked yet).
        self.pdf_import_code_highlights = None
        for import_path in imports:
            known_file_type = False
            link_path = ""
            if link:
                link_path = import_path
            # Check file size, any files over 2Gb are linked and not imported internally
            fileinfo = Path(import_path).stat()
            if fileinfo.st_size >= 2147483647:
                link_path = import_path
            # Need process events, if many large files are imported, leaves the FileDialog open and covering the screen.
            QtWidgets.QApplication.processEvents()
            filename = Path(import_path).name
            progress.setValue(file_number + 1)
            progress.setLabelText(filename)
            # Duplicate: check BEFORE copying and extracting. Previously the file was copied
            # and, for large PDFs, all the text was extracted only to be rejected at the end.
            if any(d['name'] == filename for d in self.source):
                QtWidgets.QMessageBox.warning(self, _('Duplicate file'),
                                              _("Duplicate filename.\nFile not imported") + f"\n{filename}")
                file_number += 1
                continue
            suffix = Path(import_path).suffix.lower()
            # Base destination path, extended per file type below
            destination = self.app.project_path
            if suffix in ('.docx', '.odt', '.rtf', '.tex', '.txt', '.htm', '.html', '.epub', '.md'):
                if suffix == '.tex':
                    try:
                        imported_ok = self.load_file_text(import_path, f"docs:{import_path}")
                    except LatexImportError as err:
                        logger.warning(f"LaTeX import error: {filename} {err}")
                        Message(self.app, _("Cannot import LaTeX file"),
                                _("Could not convert LaTeX to readable text") + f":\n{filename}",
                                "warning").exec()
                        continue
                    if not imported_ok:
                        continue
                    known_file_type = True
                    file_number += 1
                    continue
                destination += f"/documents/{filename}"
                if link_path == "":
                    try:
                        copyfile(import_path, destination)
                        imported_ok = self.load_file_text(import_path)
                    except PermissionError as e_:
                        msg = _("Cannot copy file: ") + f"{filename}\n" + _(
                            "Is the file open?\nIs there a permission restriction?") + f"\n{e_}"
                        Message(self.app, _("Copy file permission error"), msg).exec()
                        continue
                    if not imported_ok:
                        # Import failed (duplicate, empty, error): remove the copy
                        # so no residual files are left in /documents.
                        try:
                            Path(destination).unlink()
                        except FileNotFoundError as err:
                            logger.warning(_("Removing failed import copy: ") + str(err))
                        continue
                else:
                    self.load_file_text(import_path, f"docs:{link_path}")
                known_file_type = True
            if suffix == '.pdf':
                destination += f"/documents/{filename}"
                if link_path == "":
                    try:
                        copyfile(import_path, destination)
                        imported_ok = self.load_file_text(import_path, "", progress)
                    except PermissionError as e_:
                        msg = _("Cannot copy file: ") + f"{filename}\n" + _(
                            "Is the file open?\nIs there a permission restriction?") + f"\n{e_}"
                        Message(self.app, _("Copy file permission error"), msg).exec()
                        continue
                    if not imported_ok:
                        # Rejected PDF (protected, damaged, duplicate): remove the copy
                        # so no residual files are left in /documents.
                        try:
                            Path(destination).unlink()
                        except FileNotFoundError as err:
                            logger.warning(_("Removing failed import copy: ") + str(err))
                        continue
                else:
                    # Progress also applies to linked PDFs (extraction takes just as long).
                    self.load_file_text(import_path, f"docs:{link_path}", progress)
                known_file_type = True
            # Media files
            if Path(import_path).suffix.lower() in ('.jpg', '.jpeg', '.png'):
                if link_path == "":
                    destination += f"/images/{filename}"
                    try:
                        copyfile(import_path, destination)
                        self.load_media_reference(f"/images/{filename}")
                    except PermissionError as e_:
                        msg = _("Cannot copy file: ") + f"{filename}\n" + _(
                            "Is the file open?\nIs there a permission restriction?") + f"\n{e_}"
                        Message(self.app, _("Copy file permission error"), msg).exec()
                        continue
                else:
                    self.load_media_reference(f"images:{link_path}")
                known_file_type = True
            if Path(import_path).suffix.lower() in ('.flac', '.m4a', '.mp3',  '.ogg', '.wav'):
                if link_path == "":
                    destination += f"/audio/{filename}"
                    try:
                        copyfile(import_path, destination)
                        self.load_media_reference(f"/audio/{filename}")
                    except PermissionError as e_:
                        msg = _("Cannot copy file: ") + f"{filename}\n" + _(
                            "Is the file open?\nIs there a permission restriction?") + f"\n{e_}"
                        Message(self.app, _("Copy file permission error"), msg).exec()
                        continue
                else:
                    self.load_media_reference(f"audio:{link_path}")
                known_file_type = True
            if Path(import_path).suffix.lower() in ('.mkv', '.mov', '.mp4', '.m4v', '.wmv', '.webm'):
                if link_path == "":
                    destination += f"/video/{filename}"
                    try:
                        copyfile(import_path, destination)
                        self.load_media_reference(f"/video/{filename}")
                    except PermissionError as e_:
                        msg = _("Cannot copy file: ") + f"{filename}\n" + _(
                            "Is the file open?\nIs there a permission restriction?") + f"\n{e_}"
                        Message(self.app, _("Copy file permission error"), msg).exec()
                        continue
                else:
                    self.load_media_reference(f"video:{link_path}")
                known_file_type = True
            if not known_file_type:
                Message(self.app, _('Not supported file type'),
                        _("Cannot import file") + f":\n{import_path}", "warning").exec()
                continue

            file_number += 1

        progress.close()
        if pdf_msg != "":
            self.parent_text_edit.append(pdf_msg)
        self.load_file_data()
        self.fill_table()
        self.app.delete_backup = False
        self.update_files_in_dialogs()

    def update_files_in_dialogs(self):
        """ Update files list in any opened dialogs:
         DialogReportCodes, DialogCodeText, DialogCodeAV, DialogCodeImage """

        contents = self.tab_coding.layout()
        if contents is not None:
            for i in reversed(range(contents.count())):
                c = contents.itemAt(i).widget()
                if isinstance(c, DialogCodeImage):
                    c.get_files()
                if isinstance(c, DialogCodeAV):
                    c.get_files()
                if isinstance(c, DialogCodeText):
                    c.get_files()
                if isinstance(c, DialogCodePdf):
                    c.get_files()
        contents = self.tab_reports.layout()
        if contents is not None:
            # Examine widgets in layout
            for i in reversed(range(contents.count())):
                c = contents.itemAt(i).widget()
                if isinstance(c, DialogReportCodes):
                    c.get_files_and_cases()

    def create_waveform_png(self, file_id, mediapath):
        """ Pre-build the waveform image for an audio or video file, stored as
        audio/waveform_<id>.png, so the AV coding dialog can reuse it instead of regenerating
        it each time. For video the waveform is drawn from its audio track. Skipped on Linux
        (ffmpeg showwavespic can segfault on some distros). Requires ffmpeg installed. """

        if platform.system() == "Linux" or mediapath is None:
            return
        if not waveform_backend_available():
            # ffmpeg not installed: skip silently. Playback (VLC) and coding still work;
            # the coding dialog will show a "waveform unavailable" hint instead.
            return
        if mediapath[0:6] in ("/audio", "/video"):
            abs_path = self.app.project_path + mediapath
        elif mediapath[0:6] in ("audio:", "video:"):
            abs_path = mediapath[6:]
        else:
            return
        waveform_path = str(Path(self.app.project_path) / "audio" / f"waveform_{file_id}.png")
        # Fire-and-forget worker thread so importing long media does not freeze the UI.
        # Runs are serialised in the helper; the coding dialog waits if still pending.
        generate_waveform_png_async(abs_path, waveform_path,
                                    waveform_colour(self.app.settings['stylesheet']))

    def remove_waveform_png(self, file_id):
        """ Remove the cached waveform image for a deleted media file, to avoid residual files. """

        waveform_path = str(Path(self.app.project_path) / "audio" / f"waveform_{file_id}.png")
        try:
            if Path(waveform_path).exists():
                Path(waveform_path).unlink()
        except OSError as err:
            logger.warning(_("Deleting waveform error: ") + str(err))

    def load_media_reference(self, mediapath:str):
        """ Load media reference information for all file types.

        Args:
            mediapath: QualCoder project folder path OR external link path to file
                       External link path contains prefix 'docs:', 'images:, 'audio:', 'video:'
        """

        # Check for duplicated filename and update model, widget and database
        path_obj = Path(mediapath)
        head_path, filename = str(path_obj.parent), path_obj.name
        if any(d['name'] == filename for d in self.source):
            QtWidgets.QMessageBox.warning(self, _('Duplicate file'), _("Duplicate filename.\nFile not imported"))
            return
        entry = {'name': filename, 'id': -1, 'fulltext': None, 'memo': "", 'mediapath': mediapath,
                 'owner': self.app.settings['codername'], 'date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                 'av_text_id': None}
        cur = self.app.conn.cursor()
        cur.execute("insert into source(name,memo,owner,date, mediapath, fulltext) values(?,?,?,?,?,?)",
                    (
                        entry['name'], entry['memo'], entry['owner'], entry['date'], entry['mediapath'],
                        entry['fulltext']))
        self.app.conn.commit()
        cur.execute("select last_insert_rowid()")
        id_ = cur.fetchone()[0]
        entry['id'] = id_
        msg = entry['name']
        if ':' in mediapath:
            msg += _(" linked")
        else:
            msg += _(" imported.")
        self.parent_text_edit.append(msg)
        self.source.append(entry)

        # Pre-build the waveform image for audio and video files so the coding dialog loads faster
        if mediapath[:6] in ("/audio", "audio:", "/video", "video:"):
            self.create_waveform_png(id_, mediapath)

        # Create an empty transcription file for audio and video
        if mediapath[:6] in ("/audio", "audio:", "/video", "video:"):
            entry = {'name': filename + ".txt", 'id': -1, 'fulltext': "", 'mediapath': None, 'memo': "",
                     'owner': self.app.settings['codername'],
                     'date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     'av_text_id': None}
            cur = self.app.conn.cursor()
            cur.execute("insert into source(name,fulltext,mediapath,memo,owner,date) values(?,?,?,?,?,?)",
                        (entry['name'], entry['fulltext'], entry['mediapath'], entry['memo'], entry['owner'],
                         entry['date']))
            self.app.conn.commit()
            cur.execute("select last_insert_rowid()")
            tr_id = cur.fetchone()[0]
            entry['id'] = tr_id
            # Update av file entry with av_text_id link to this text file
            cur.execute("update source set av_text_id=? where id=?", [tr_id, id_])
            self.app.conn.commit()

            # add doc to vectorstore
            if self.app.settings['ai_enable'] == 'True':
                self.app.ai.sources_vectorstore.import_document(entry['id'], entry['name'], entry['fulltext'])

            # Add file attribute placeholders
            att_sql = 'select name from attribute_type where caseOrFile ="file"'
            cur.execute(att_sql)
            attr_types = cur.fetchall()
            insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
            for a in attr_types:
                placeholders = [a[0], tr_id, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                self.app.settings['codername']]
                cur.execute(insert_sql, placeholders)
                self.app.conn.commit()

            self.parent_text_edit.append(entry['name'] + _(" created."))
            self.source.append(entry)

    def load_pseudonyms(self):
        """ Pseudonyms stored in pseudonyms.json in qda data folder.
        Loads into list of dictionaries of 'original', ;pseudonym' keys.
        """

        pseudonyms = []
        pseudonyms_filepath = Path(self.app.project_path) / "pseudonyms.json"
        try:
            with open(pseudonyms_filepath, "r") as f:
                pseudonyms = json.load(f)
        except FileNotFoundError:
            pass
        return pseudonyms

    def decode_text_with_best_encoding(self, import_file:str):
        """ Decode text file bytes using robust encoding detection and fallbacks. """

        return decode_text_with_best_encoding_helper(import_file)

    def load_file_text(self, import_file:str, link_path:str="", progress_:QProgressDialog|None=None):
        """ Import from file types of odt, docx, rtf, pdf, epub, txt, html, htm.
        Implement character detection for txt imports.
        Loading pdf text. I have removed additional line breaks. See commented sections below.
        Removing these allows the pdf to be coded in Code_text and Code_pdf without positional shifting problems.
        If pseudonyms.json is present in the qda folder, apply the pseudonyms to the words / phrases on import.
        Args:
            import_file: filepath of file to be imported, String
            link_path:  filepath of file to be linked, String
            progress_: None or QProgressDialog
        """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        # Duplicate: check BEFORE extracting (extraction is expensive for large PDFs).
        # The check used to sit at the end, after all the extraction time was wasted.
        filename = Path(import_file).name
        if any(d['name'] == filename for d in self.source):
            QtWidgets.QMessageBox.warning(self, _('Duplicate file'),
                                          _("Duplicate filename.\nFile not imported"))
            return False
        text_ = ""
        suffix = Path(import_file).suffix.lower()
        # Import from odt
        if suffix == ".odt":
            text_ = self.convert_odt_to_text(import_file)
            text_ = text_.replace("\n", "\n\n")  # add line to paragraph spacing for visual format
        # Import from docx
        if suffix == ".docx":
            document = opendocx(import_file)
            list_ = getdocumenttext(document)
            text_ = "\n\n".join(list_)  # add line to paragraph spacing for visual format
        # Import from rtf
        if suffix == ".rtf":
            # text_ = rtf_to_text(import_file, encoding="latin-1", errors="replace")
            with open(import_file, "r", encoding="latin-1") as sourcefile:
                text_ = ""
                try:
                    rtf = sourcefile.read()
                    text_ = rtf_to_text(rtf)
                except Exception as err:
                    msg = "Importing rtf. Expecting characters encoded as latin-1. Import failed."
                    logger.debug(f"rtf_to_text error Not Latin-1: {err}")
                    Message(self.app, "rtf to text error", msg).exec()
        # Import from epub
        if import_file[-5:].lower() == ".epub":
            # Extraction shared with the reference attachment import.
            text_ = extract_epub_fulltext(import_file)
        # Import from html
        if suffix in (".html", ".htm"):
            import_errors = 0
            with open(import_file, "r", encoding="utf-8", errors="surrogateescape") as sourcefile:
                html_text = ""
                while 1:
                    line = sourcefile.readline()
                    if not line:
                        break
                    html_text += line
                text_ = html_to_text(html_text)
                if import_errors > 0:
                    Message(self.app, _("Warning"), str(import_errors) + _(" lines not imported"), "warning").exec()
        # Import PDF
        if suffix == '.pdf':
            # Extraction with the SAME extractor as the viewer (code_pdf.extract_pdf_fulltext),
            # otherwise coding positions do not map onto the PDF pages.
            def pdf_progress(current_page, total_pages):
                if progress_ is not None:
                    progress_.setLabelText(f"{Path(import_file).name} p:{current_page}/{total_pages}")
                QtCore.QCoreApplication.processEvents()
            try:
                # Paragraph layout; the viewer also verifies the classic line
                # layout, so older imports keep working.
                text_ = extract_pdf_fulltext(import_file, pdf_progress, join_lines=True)
            except ValueError as err:
                Message(self.app, _("Cannot import PDF"),
                        f"{Path(import_file).name}\n{err}", "warning").exec()
                return False
            except Exception as err:
                # Damaged or unreadable PDF
                logger.warning(f"PDF import error: {import_file} {err}")
                Message(self.app, _("Cannot import PDF"),
                        _("Damaged or unreadable PDF") + f":\n{Path(import_file).name}\n{err}",
                        "warning").exec()
                return False
            if text_.strip() == "":
                # No text layer (scanned PDF): imported anyway, with page markers,
                # so AREA coding works in the PDF viewer.
                msg = _("No text layer detected (scanned PDF?).") + "\n"
                msg += _("Area coding will be available in the PDF view, "
                         "but text coding and text search will not.")
                Message(self.app, _("PDF without text"),
                        f"{Path(import_file).name}\n{msg}", "warning").exec()
        if suffix == ".tex":
            try:
                text_ = tex_file_to_plain_text(import_file)
            except LatexImportError as err:
                logger.warning(f"LaTeX import error: {Path(import_file).name} {err}")
                Message(self.app, _("Cannot import LaTeX file"),
                        _("Could not convert LaTeX to readable text") + f":\n{Path(import_file).name}",
                        "warning").exec()
                return False
        # Try importing as a plain text file.
        # Never decode a PDF as plain text (it would produce unreadable binary).
        if text_ == "" and suffix not in ('.pdf', '.tex'):
            try:
                text_, detected_encoding = decode_text_with_best_encoding_helper(import_file)
                logger.debug(f"Importing plain text file: {import_file} decoded as {detected_encoding}")
                if text_ and text_[0] == "\ufeff":  # associated with notepad files
                    text_ = text_[1:]
            except Exception as err:
                logger.warning(str(err))
                Message(self.app, _("Warning"), _("Cannot import") + f"{import_file}\n{err}",
                        "warning").exec()
                return False
        # Import of text file did not work
        if text_ == "":
            Message(self.app, _("Warning"),
                    _("Cannot import ") + str(import_file) + "\nPlease check if the file is empty.", "warning").exec()
            return False
        # Normalise line endings and strip BOM: Qt converts \r\n/\r to \n on
        # setPlainText, so mismatches make stored positions drift.
        if suffix != '.pdf': # skip PDF
            text_ = text_.replace("\r\n", "\n").replace("\r", "\n")
            if text_ and text_[0] == "\ufeff":
                text_ = text_[1:]
        # Name and duplicate were already checked at the start, before extracting.

        # Apply pseudonym text replacement
        pseudonyms = self.load_pseudonyms()
        if suffix != '.pdf':
            for pseudonym in pseudonyms:
                pseudonymised = re.sub(rf"(?<!\w){re.escape(pseudonym['original'])}(?!\w)", pseudonym['pseudonym'], text_)
                text_ = pseudonymised

        # Internal storage
        mediapath = "/docs/" + filename
        if link_path != "":
            mediapath = link_path
        entry = {'name': filename, 'id': -1, 'fulltext': text_, 'mediapath': mediapath, 'memo': "",
                 'owner': self.app.settings['codername'], 'date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        cur = self.app.conn.cursor()
        cur.execute("insert into source(name,fulltext,mediapath,memo,owner,date) values(?,?,?,?,?,?)",
                    (
                        entry['name'], entry['fulltext'], entry['mediapath'], entry['memo'], entry['owner'],
                        entry['date']))
        self.app.conn.commit()
        cur.execute("select last_insert_rowid()")
        id_ = cur.fetchone()[0]
        entry['id'] = id_

        # Add file attribute placeholders
        att_sql = 'select name from attribute_type where caseOrFile ="file"'
        cur.execute(att_sql)
        attr_types = cur.fetchall()
        insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
        for a in attr_types:
            placeholders = [a[0], id_, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            self.app.settings['codername']]
            cur.execute(insert_sql, placeholders)
            self.app.conn.commit()

        # add doc to vectorstore
        if self.app.settings['ai_enable'] == 'True':
            self.app.ai.sources_vectorstore.import_document(entry['id'], entry['name'], entry['fulltext'])

        msg = entry['name']
        if link_path == "":
            msg += _(" imported")
        else:
            msg += _(" linked")
        self.parent_text_edit.append(msg)
        self.source.append(entry)
        # Offer (once per batch) to code highlight annotations; they are not
        # painted in the coding view (annots=False).
        if suffix == '.pdf':
            # Non-highlight annotations with text are appended to the file memo.
            entry['memo'] = pdf_annotations_to_file_memo(self.app, self.parent_text_edit, entry['id'],
                                                         import_file, entry['memo'])
            try:
                highlights = extract_pdf_highlights(import_file)
            except Exception as err:
                logger.warning(f"Highlight detection: {import_file} {err}")
                highlights = []
            if highlights:
                if self.pdf_import_code_highlights is None:
                    ask_msg = _("Highlighted segments were detected in the imported PDF(s).") + "\n\n"
                    ask_msg += _("Code those segments? A 'PDF Highlights' category will be "
                                 "created, with one code per highlight colour (named and "
                                 "coloured after the closest QualCoder colour).")
                    reply = QtWidgets.QMessageBox.question(
                        self, _("PDF highlights"), ask_msg,
                        QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                        QtWidgets.QMessageBox.StandardButton.Yes)
                    self.pdf_import_code_highlights = reply == QtWidgets.QMessageBox.StandardButton.Yes
                if self.pdf_import_code_highlights:
                    # Reuses the batch import dialog; the phase only shows after acceptance.
                    self.code_pdf_highlights(entry['id'], import_file, entry['fulltext'],
                                             highlights, progress_)
        return True  # Import completed; lets import_files clean up failed copies

    # why not use: color_selector.color_matcher() method
    @staticmethod
    def _closest_qualcoder_color(hex_color):
        """ Delegates to code_pdf.closest_qualcoder_color.
        Returns (hex of the palette colour, family name). """

        return closest_qualcoder_color(hex_color)

    def code_pdf_highlights(self, fid, filepath, fulltext, highlights, progress_=None):
        """ Delegates to code_pdf.code_pdf_highlights.
        Args:
            fid: source id, Integer
            filepath: PDF path
            fulltext: the imported fulltext (paragraph layout)
            highlights: output of extract_pdf_highlights TODO highlights variable - type()
            progress_: the batch QProgressDialog from import_files, or None
        """

        code_pdf_highlights_shared(self.app, self.parent_text_edit, fid, filepath, fulltext,
                                   highlights, progress_, self)

    def convert_odt_to_text(self, import_file:str):
        """ Convert odt to very rough equivalent with headings, list items and tables for
        html display in qTextEdits.
        Args:
            import_file: String """

        odt_file = zipfile.ZipFile(import_file)
        data = str(odt_file.read('content.xml'))  # bytes class to string
        # https://stackoverflow.com/questions/18488734/python3-unescaping-non-ascii-characters
        data = str(bytes([ord(char) for char in data.encode("utf_8").decode("unicode_escape")]), "utf_8")
        data_start = data.find("</text:sequence-decls>")
        data_end = data.find("</office:text>")
        if data_start == -1 or data_end == -1:
            logger.warning("ODT IMPORT ERROR")
            return ""
        data = data[data_start + 22: data_end]
        data = data.replace('</text:index-title-template>', '')
        data = data.replace('</text:index-entry-span>', '')
        data = data.replace('</text:table-of-content-entry-template>', '')
        data = data.replace('</text:index-title>', '')
        data = data.replace('</text:index-body>', '')
        data = data.replace('</text:table-of-contents>', '')
        data = data.replace('</text:table-of-content-source>', '')
        data = data.replace('<text:h', '\n<text:h')
        data = data.replace('</text:h>', '\n\n')
        data = data.replace('</text:list-item>', '\n')
        data = data.replace('</text:span>', '')
        data = data.replace('</text:p>', '\n')
        data = data.replace('</text:a>', ' ')
        data = data.replace('</text:list>', '')
        data = data.replace('</text:sequence>', '')
        data = data.replace('<text:list-item>', '')
        data = data.replace('<table:table table:name=', '\n=== TABLE ===\n<table:table table:name=')
        data = data.replace('</table:table>', '=== END TABLE ===\n')
        data = data.replace('</table:table-cell>', '\n')
        data = data.replace('</table:table-row>', '')
        data = data.replace('<draw:image', '\n=== IMG ===<draw:image')
        data = data.replace('</draw:frame>', '\n')
        text_ = ""
        tagged = False
        for i in range(0, len(data)):
            if data[i: i + 6] == "<text:" or data[i: i + 7] == "<table:" or data[i: i + 6] == "<draw:":
                tagged = True
            if not tagged:
                text_ += data[i]
            if data[i] == ">":
                tagged = False
        text_ = text_.replace("&apos;", "'")
        text_ = text_.replace("&quot;", '"')
        text_ = text_.replace("&gt;", '>')
        text_ = text_.replace("&lt;", '<')
        text_ = text_.replace("&amp;", '&')
        return text_

    def export(self):
        """ Export selected files to selected directory.
        If an imported file was from a docx, odt, pdf, html, epub then export the original file
        If the file was created within QualCoder (so only in the database), export as plain text.
        """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        rows = self.visible_selected_rows()
        if len(rows) == 0:
            return

        destination = self.app.settings['directory']

        export_msg = _("Export to") + f" {destination}\n"
        files_failed = 0
        for row in rows:
            filename = self.source[row]['name']
            mediapath = self.source[row]['mediapath']
            # Check for invalid filenames (e.g. ".", "..")
            if filename.strip('.') == '' or filename.strip() == '':
                msg = _("Invalid file name. Please rename this file before exporting.")
                msg += f"\n{filename}"
                Message(self.app, _("Warning"), msg, "warning").exec()
                return
            # Export text representation of linked files (e.g. odt, docx, txt, md, pdf)
            if mediapath is not None and ':' in mediapath and self.source[row]['fulltext'] != "":
                export_msg += f"{filename} - " + _("Linked file. Exported text representation.") + "\n"
                filedata = self.source[row]['fulltext']
                if not filename.endswith(".txt"):
                    filename += ".txt"
                path = Path(destination)/ filename
                with open(path, 'w', encoding='utf-8-sig') as textfile:
                    textfile.write(filedata)
                continue
            # Export audio, video, picture files
            if mediapath is not None and mediapath[0:6] != "/docs/":
                # Note: [1:] must remove leading slash
                file_path = Path(self.app.project_path) / mediapath[1:]
                try:
                    copyfile(file_path, Path(destination)/ filename)
                    if filename != mediapath[6:]:
                        export_msg += f"{filename} ({mediapath[6:]}) " + _("exported.") + "\n"
                    else:
                        export_msg += f"{filename} " + _("exported.") + "\n"
                except FileNotFoundError:
                    Message(self.app, _("Error"), _("File not found: ") + file_path).exec()
                    export_msg += f"{file_path} - " + _("Media file NOT exported.") + "\n"
                    files_failed += 1
                continue
            # Export qc-created transcribed files, user-created text files
            if mediapath is None:  # rest of this not needed as media done above and self.source[row]['fulltext'] != "":
                export_msg += f"{filename} - " + _("QC or user created file exported.") + "\n"
                filedata = self.source[row]['fulltext']
                if not filename.endswith(".txt"):
                    filename += ".txt"
                path = Path(destination) / filename
                with open(path, 'w', encoding='utf-8-sig') as textfile:
                    textfile.write(filedata)
                continue

            # Export md, pdf, docx, odt, epub, html files with text rep. if located in documents directory
            source_path = Path(self.app.project_path) / "documents" / mediapath[6:]  # 0-6 is /docs/
            document_exists = Path(source_path).exists()
            if document_exists:
                try:
                    # Remove '/docs/' from start of mediapath string
                    copyfile(source_path, Path(destination) / mediapath[6:])
                    filedata = self.source[row]['fulltext']
                    if not filename.endswith(".txt"):
                        filename += ".txt"
                    path = Path(destination) / filename
                    with open(path, 'w', encoding='utf-8-sig') as file_:
                        file_.write(filedata)
                    if filename != mediapath[6:]:
                        export_msg += f"{filename} ({mediapath[6:]}) " + _("exported.") + "\n"
                    else:
                        export_msg += f"{filename} " + _("exported.") + "\n"
                except (FileNotFoundError, PermissionError) as err:
                    files_failed += 1
                    msg = f"{err}\n{file_path} - " + _("File NOT exported.") + "\n"
                    logger.warning(msg)
                    print(msg)
                    Message(self.app, _("Error"), msg).exec()
                    continue

        msg = f"{len(rows) - files_failed} " + _(" files exported. ") + _("Exported to: ") + destination
        if files_failed > 0:
            msg += _("Files not exported: ") + len(files_failed)
        Message(self.app, _("Files exported"), msg).exec()
        export_msg += "\n" + msg
        self.parent_text_edit.append(export_msg)

    def release_files_in_coding_dialogs(self):
        """
        Releases file handles held by open coding dialogs. DialogCodePdf workers
        (render and text) keep the PDF open while coding; on Windows that blocks
        removing the file and the file would remain as a residual. Call ALWAYS before
        deleting project files.
        """

        contents = self.tab_coding.layout()
        if contents is None:
            return
        for i in reversed(range(contents.count())):
            c = contents.itemAt(i).widget()
            if isinstance(c, DialogCodePdf):
                try:
                    c.stop_workers()
                except (AttributeError, RuntimeError):
                    pass

    def vectorstore_delete_document_safe(self, fid:int):
        """
        Removes the document from the AI index without letting a vectorstore lock
        abort the project file deletion (e.g. "database is locked" when an embeddings
        worker is writing to search.sqlite in the background). The index is derived
        data: the next update_vectorstore prunes ids no longer in source, so failing
        soft here is safe and self-repairing.

        Args:
            fid: source id, Integer
        """

        if self.app.settings['ai_enable'] != 'True':
            return
        try:
            self.app.ai.sources_vectorstore.delete_document(fid)
        except Exception as err:
            logger.warning(f"vectorstore delete_document fid {fid}: {err}")
            self.parent_text_edit.append(
                _("AI index is busy; the deleted file will be removed from the index "
                  "on the next update."))

    def _unlink_media_with_retry(self, filepath):
        """ Players free the file handle asynchronously, so an immediate unlink
        can still hit WinError 32: retry briefly, then warn. """
        last_err = None
        for _attempt in range(12):
            try:
                Path(filepath).unlink()
                return True
            except FileNotFoundError as err:
                logger.warning(_("Deleting file error: ") + str(err))
                return True
            except PermissionError as err:
                last_err = err
                self._release_media_players_for(filepath)
                QtWidgets.QApplication.processEvents()
                time.sleep(0.15)
        logger.warning(f"Locked media file, could not delete: {filepath} {last_err}")
        Message(self.app, _("Cannot delete file"),
                _("The file is in use by another program and was not deleted:") +
                f"\n{filepath}", "warning").exec()
        return False

    def _release_media_players_for(self, filepath):
        """ Ask every live player holding this file to let go before unlink
        (an open handle raises WinError 32 on Windows). """
        try:
            target = str(Path(filepath).resolve())
        except Exception:
            target = str(filepath)
        for w in QtWidgets.QApplication.allWidgets():
            mp = getattr(w, 'mediaplayer', None)
            if mp is None:
                continue
            try:
                if type(mp).__module__.endswith('media_player_qt'):
                    src = mp.player.source().toLocalFile()
                    if src and str(Path(src).resolve()) == target:
                        mp.release()
                else:
                    med = mp.get_media()
                    if med is not None and target.replace("\\", "/") in \
                            str(med.get_mrl() or "").replace("%20", " "):
                        mp.stop()
                        mp.set_media(None)  # stop alone may keep the handle a moment
            except Exception:
                pass

    def delete_button_multiple_files(self):
        """ Delete files from database and update model and widget.
        Also, delete files from sub-directories, if not externally linked.

        Called by: delete button.
        """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
                if type(self.av_dialog_open.mediaplayer).__module__.endswith('media_player_qt'):
                    self.av_dialog_open.mediaplayer.release()  # free handle before unlink
            self.av_dialog_open.close()
            self.av_dialog_open = None
        # Respect active filters: only visible files are offered for deletion
        visible_sources = [s for r, s in enumerate(self.source)
                           if not self.ui.tableWidget.isRowHidden(r)]
        if not visible_sources:
            return
        ui = DialogSelectItems(self.app, visible_sources, _("Delete files"), "multi")
        ok = ui.exec()
        if not ok:
            return
        selection = ui.get_selected()
        if not selection:
            return
        names = ""
        for selected in selection:
            names = f"{names}{selected['name']}\n"
        ui = DialogConfirmDelete(self.app, names)
        ok = ui.exec()
        if not ok:
            return

        # Release PDFs open in coding tabs before deleting on disk.
        self.release_files_in_coding_dialogs()
        msg = ""
        cur = self.app.conn.cursor()
        for s in selection:
            msg += _("Deleted file: ") + s['name'] + "\n"
            self.files_renamed = [x for x in self.files_renamed if not (s['id'] == x.get('fid'))]
            # Delete text source
            if s['mediapath'] is None or s['mediapath'][0:5] == 'docs:' or s['mediapath'][0:6] == '/docs/':
                try:
                    if s['mediapath'] is None:
                        # Legacy for older < 3.4 QualCoder projects
                        p = Path(self.app.project_path) / "/documents/" / s['name']
                        p.unlink()
                    if s['mediapath'][0:6] == '/docs/':
                        # Previously sliced s['name'][6:] leaving the file as a residual.
                        p = Path(self.app.project_path) / "/documents/" / s['mediapath'][6:]
                        p.unlink()
                except FileNotFoundError as err:
                    logger.warning(_("Deleting file error: ") + str(err))
                # Delete stored coded sections and source details
                cur.execute("delete from source where id = ?", [s['id']])
                cur.execute("delete from code_text where fid = ?", [s['id']])
                # Coded areas over PDF pages (code_image with pdf_page); orphaned without this.
                cur.execute("delete from code_image where id = ?", [s['id']])
                cur.execute("delete from annotation where fid = ?", [s['id']])
                cur.execute("delete from case_text where fid = ?", [s['id']])
                cur.execute("delete from attribute where attr_type ='file' and id=?", [s['id']])
                # Clear stale transcript links: SQLite reuses row ids and a later
                # import could inherit this id, becoming a ghost transcript
                cur.execute("update source set av_text_id=null where av_text_id=?", [s['id']])
                self.app.conn.commit()
                # Delete from vectorstore
                self.vectorstore_delete_document_safe(s['id'])

                    # Delete image, audio or video source
            if s['mediapath'] is not None and s['mediapath'][0:5] != 'docs:' and s['mediapath'][0:6] != '/docs/':
                # Get linked transcript file id
                cur.execute("select av_text_id from source where id=?", [s['id']])
                res = cur.fetchone()
                av_text_id = res[0]
                # Remove avid links in code_text
                sql = "select avid from code_av where id=?"
                cur.execute(sql, [s['id']])
                avids = cur.fetchall()
                sql = "update code_text set avid=null where avid=?"
                for avid in avids:
                    cur.execute(sql, [avid[0]])
                self.app.conn.commit()
                # Remove project folder file, if internally stored
                if ':' not in s['mediapath']:
                    filepath = self.app.project_path + s['mediapath']
                    self._release_media_players_for(filepath)
                    self._unlink_media_with_retry(filepath)
                # Remove the cached waveform image, if any
                self.remove_waveform_png(s['id'])
                # Delete stored coded sections and source details
                cur.execute("delete from source where id = ?", [s['id']])
                cur.execute("delete from code_image where id = ?", [s['id']])
                cur.execute("delete from code_av where id = ?", [s['id']])
                cur.execute("delete from attribute where attr_type='file' and id=?", [s['id']])
                # Just in case, added this line
                cur.execute("delete from case_text where fid = ?", [s['id']])
                cur.execute("update source set av_text_id=null where av_text_id=?", [s['id']])
                self.app.conn.commit()

                # Delete linked transcription text file
                if av_text_id is not None:
                    cur.execute("delete from source where id = ?", [res[0]])
                    cur.execute("delete from code_text where fid = ?", [res[0]])
                    cur.execute("delete from annotation where fid = ?", [res[0]])
                    cur.execute("delete from case_text where fid = ?", [res[0]])
                    cur.execute("delete from attribute where attr_type ='file' and id=?", [res[0]])
                    self.app.conn.commit()
                    # Delete from vectorstore
                    if self.app.settings['ai_enable'] == 'True':
                        self.app.ai.sources_vectorstore.delete_document(res[0])

        self.update_files_in_dialogs()
        self.check_attribute_placeholders()
        self.parent_text_edit.append(msg)
        self.load_file_data()
        self.fill_table()
        self.app.delete_backup = False

    def delete(self):
        """ Delete files from database and update model and widget.
        Also, delete the files from subdirectories, if not externally linked.
        Called by: right-click table context menu.
        """

        if self.av_dialog_open is not None:
            # Guard: a dialog that failed mid-init can leave mediaplayer as None
            if getattr(self.av_dialog_open, 'mediaplayer', None) is not None:
                self.av_dialog_open.mediaplayer.stop()
                if type(self.av_dialog_open.mediaplayer).__module__.endswith('media_player_qt'):
                    self.av_dialog_open.mediaplayer.release()  # free handle before unlink
            self.av_dialog_open.close()
            self.av_dialog_open = None
        rows = self.visible_selected_rows()
        if len(rows) == 0:
            return
        filenames = ""
        for r in rows:
            filenames += f"\n{self.source[r]['name']}"
        ui = DialogConfirmDelete(self.app, filenames, _("Delete files"))
        ok = ui.exec()
        if not ok:
            return

        # Release PDFs open in coding tabs before deleting on disk.
        self.release_files_in_coding_dialogs()
        cur = self.app.conn.cursor()
        for row in rows:
            file_id = self.source[row]['id']
            # Delete text source
            if self.source[row]['mediapath'] is None or self.source[row]['mediapath'][0:5] == 'docs:' or \
                    self.source[row]['mediapath'][0:6] == '/docs/':
                try:
                    if self.source[row]['mediapath'] is None:
                        # Legacy for older QualCoder Projects < 3.3
                        # The condition was inverted (deleted when mediapath was present).
                        p = Path(self.app.project_path) /"/documents/" / self.source[row]['name']
                        p.unlink()
                    if self.source[row]['mediapath'] is not None and self.source[row]['mediapath'][0:6] == '/docs/':
                        p = Path(self.app.project_path) / "/documents/" / self.source[row]['mediapath'][6:]
                        p.unlink()
                except FileNotFoundError as err:
                    logger.warning(_("Deleting file error: ") + str(err))
                # Delete stored coded sections and source details
                cur.execute("delete from source where id = ?", [file_id])
                cur.execute("delete from code_text where fid = ?", [file_id])
                # Coded areas over PDF pages (code_image with pdf_page); orphaned without this.
                cur.execute("delete from code_image where id = ?", [file_id])
                cur.execute("delete from annotation where fid = ?", [file_id])
                cur.execute("delete from case_text where fid = ?", [file_id])
                cur.execute("delete from attribute where attr_type ='file' and id=?", [file_id])
                cur.execute("update source set av_text_id=null where av_text_id=?", [file_id])
                self.app.conn.commit()
                # Delete from vectorstore
                self.vectorstore_delete_document_safe(file_id)

            else:  # Delete image, audio or video source
                # Get linked transcript file id
                cur.execute("select av_text_id from source where id=?", [file_id])
                res = cur.fetchone()
                av_text_id = res[0]
                # Remove avid links in code_text
                sql = "select avid from code_av where id=?"
                cur.execute(sql, [file_id])
                avids = cur.fetchall()
                sql = "update code_text set avid=null where avid=?"
                for avid in avids:
                    cur.execute(sql, [avid[0]])
                self.app.conn.commit()
                # Remove folder file, if internally stored
                if ':' not in self.source[row]['mediapath']:
                    filepath = self.app.project_path + self.source[row]['mediapath']
                    self._release_media_players_for(filepath)
                    self._unlink_media_with_retry(filepath)
                # Remove the cached waveform image, if any
                self.remove_waveform_png(file_id)
                # Delete stored coded sections and source details
                cur.execute("delete from source where id = ?", [file_id])
                cur.execute("delete from code_image where id = ?", [file_id])
                cur.execute("delete from code_av where id = ?", [file_id])
                cur.execute("delete from attribute where attr_type='file' and id=?", [file_id])
                self.app.conn.commit()
                # Delete from vectorstore (this should not be necessary since it's not a text file, but just to be sure...)
                if self.app.settings['ai_enable'] == 'True':
                    self.app.ai.sources_vectorstore.delete_document(file_id)

                # Delete transcription text file
                if av_text_id is not None:
                    cur.execute("delete from source where id = ?", [res[0]])
                    cur.execute("delete from code_text where fid = ?", [res[0]])
                    cur.execute("delete from annotation where fid = ?", [res[0]])
                    cur.execute("delete from case_text where fid = ?", [res[0]])
                    cur.execute("delete from attribute where attr_type ='file' and id=?", [res[0]])
                    self.app.conn.commit()
                    # Delete from vectorstore
                    if self.app.settings['ai_enable'] == 'True':
                        self.app.ai.sources_vectorstore.delete_document(res[0])

            self.files_renamed = [x for x in self.files_renamed if not (file_id == x.get('fid'))]
        self.update_files_in_dialogs()
        self.check_attribute_placeholders()
        self.parent_text_edit.append(_("Deleted: ") + filenames)
        self.load_file_data()
        self.app.delete_backup = False

    def get_tooltip_values(self, attribute_name:str):
        """ Get values to display in tooltips for the value list column.
        Args:
            attribute_name : String """

        tt = ""
        cur = self.app.conn.cursor()
        sql_val_type = 'select valuetype from attribute_type where caseOrFile="file" and name=?'
        cur.execute(sql_val_type, [attribute_name])
        res_val_type = cur.fetchone()
        value_type = "character"
        if res_val_type is not None:
            value_type = res_val_type[0]
        if value_type == "numeric":
            sql = 'select min(cast(value as real)), max(cast(value as real)) from attribute where name=? and ' \
                  'attr_type="file"'
            cur.execute(sql, [attribute_name])
            res = cur.fetchone()
            tt = _("Minimum: ") + f"{res[0]}\n"
            tt += _("Maximum: ") + str(res[1])
        if value_type == "character":
            sql = 'select distinct value from attribute where name=? and attr_type="file" and length(value)>0 limit 10'
            cur.execute(sql, [attribute_name])
            res = cur.fetchall()
            for r in res:
                tt += f"\n{r[0]}"
            if len(tt) > 1:
                tt = tt[1:]
        return tt

    def update_label_file_count(self):
        """ Update label_file to show file count and current file name. """

        total = self.ui.tableWidget.rowCount()
        visible = 0
        for r in range(total):
            if not self.ui.tableWidget.isRowHidden(r):
                visible += 1
        row = self.ui.tableWidget.currentRow()
        if visible < total:
            count_text = f"{visible}/{total} " + _("Files")
        else:
            count_text = f"{total} " + _("Files")
        if 0 <= row < len(self.source):
            count_text += f". {self.source[row]['name']}"
        self.ui.label_file.setText(count_text)

    def fill_table(self):
        """ Fill the table widget with file details. """

        self.ui.tableWidget.blockSignals(True)
        self.ui.tableWidget.setColumnCount(len(self.header_labels))
        self.ui.tableWidget.setHorizontalHeaderLabels(self.header_labels)
        self.ui.tableWidget.horizontalHeader().setStretchLastSection(False)
        rows = self.ui.tableWidget.rowCount()
        for r in range(0, rows):
            self.ui.tableWidget.removeRow(0)
        for row, data in enumerate(self.source):
            self.ui.tableWidget.insertRow(row)
            icon = data['icon']
            name_item = QtWidgets.QTableWidgetItem(data['name'])
            name_item.setIcon(icon)
            # Having un-editable file names helps with assigning icons
            name_item.setFlags(name_item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            # Externally linked - add link details to tooltip
            name_tt = data['metadata']
            if data['mediapath'] is not None and ':' in data['mediapath']:
                name_tt += _("\nExternally linked file:\n")
                name_tt += data['mediapath'].split(':', 1)[1]
                badlink = self.app.check_bad_file_links(data['id'])
                if badlink:
                    name_tt += f"\nORIGINAL FILE NOT FOUND"
            name_item.setToolTip(name_tt)
            self.ui.tableWidget.setItem(row, self.NAME_COLUMN, name_item)
            trimmed_date = data['date'].split()[0]
            date_item = QtWidgets.QTableWidgetItem(trimmed_date)
            date_item.setFlags(date_item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, self.DATE_COLUMN, date_item)
            memo_string = ""
            if data['memo'] != "":
                memo_string = _("Memo")
            memo_item = QtWidgets.QTableWidgetItem(memo_string)
            if data['memo'] != "":
                memo_item.setToolTip(data['memo'])
            memo_item.setFlags(date_item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, self.MEMO_COLUMN, memo_item)
            fid = data['id']
            if fid is None:
                fid = ""
            iditem = QtWidgets.QTableWidgetItem(str(fid))
            iditem.setFlags(iditem.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, self.ID_COLUMN, iditem)
            case_item = QtWidgets.QTableWidgetItem(data['case'])
            case_item.setFlags(case_item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, self.CASE_COLUMN, case_item)
            # Add the attribute values
            # TODO consider using role type for numerics
            for offset, attribute in enumerate(data['attributes']):
                item = QtWidgets.QTableWidgetItem(attribute)
                self.ui.tableWidget.setItem(row, self.ATTRIBUTE_START_COLUMN + offset, item)
                if self.attribute_labels_ordered[offset] in (
                        "Ref_Authors", "Ref_Title", "Ref_Type", "Ref_Year", "Ref_Journal"):
                    item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
        # Resize columns and rows
        self.ui.tableWidget.hideColumn(self.ID_COLUMN)
        if self.app.settings['showids']:
            self.ui.tableWidget.showColumn(self.ID_COLUMN)
        self.ui.tableWidget.resizeColumnsToContents()
        for i in range(self.ui.tableWidget.columnCount()):
            if self.ui.tableWidget.columnWidth(i) > 500:
                self.ui.tableWidget.setColumnWidth(i, 500)
        self.ui.tableWidget.resizeRowsToContents()
        # self.ui.tableWidget.verticalHeader().setVisible(False)
        # Add statistics tooltips to table headers for attributes
        for i, attribute_name in enumerate(self.attribute_labels_ordered):
            tt = self.get_tooltip_values(attribute_name)
            self.ui.tableWidget.horizontalHeaderItem(self.ATTRIBUTE_START_COLUMN + i).setToolTip(
                _("Right click header row to hide columns") + "\n" + tt)

        self.update_label_file_count()
        if getattr(self, "header_filters", None) is not None and \
                (self.file_filter_active() or self.rows_hidden):
            self.apply_file_filter()
        self.ui.tableWidget.blockSignals(False)


class DialogSurveyImport(QtWidgets.QDialog):
    """ Survey import dialog. To assign cases, attributes and qualitative text. """

    def __init__(self, columns:list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle(_("Survey Import Assistant"))
        self.resize(780, 500)
        self.setMaximumWidth(850) 
        main_layout = QtWidgets.QVBoxLayout(self)
        
        # File Name Selector
        '''top_layout = QtWidgets.QHBoxLayout()
        # top_layout.addWidget(QtWidgets.QLabel(_("File Name Column (Optional):")))
        self.combo_filename = QtWidgets.QComboBox()
        self.combo_filename.addItem(_(" [Generate names automatically] "))
        self.combo_filename.addItems(columns)
        self.combo_filename.setMaximumWidth(450)
        # top_layout.addWidget(self.combo_filename)
        top_layout.addStretch()
        main_layout.addLayout(top_layout)'''
        layout = QtWidgets.QHBoxLayout()
        
        # Left Panel (Available Columns)
        left_layout = QtWidgets.QVBoxLayout()
        left_layout.addWidget(QtWidgets.QLabel(_("Columns:")))
        self.list_available = QtWidgets.QListWidget()
        self.list_available.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.list_available.setMinimumWidth(180)
        self.list_available.setMaximumWidth(220)
        self.list_available.addItems(columns)
        left_layout.addWidget(self.list_available)
        
        # Right Panel (Targets)
        right_layout = QtWidgets.QVBoxLayout()
        
        def create_target_block(label_text, list_widget, max_height=None, ttip=""):
            block_layout = QtWidgets.QVBoxLayout()
            label = QtWidgets.QLabel(label_text)
            label.setToolTip(ttip)
            block_layout.addWidget(label)
            
            h_layout = QtWidgets.QHBoxLayout()
            
            btn_layout = QtWidgets.QVBoxLayout()
            btn_add = QtWidgets.QPushButton(">")
            btn_remove = QtWidgets.QPushButton("<")
            btn_add.setFixedSize(30, 30) 
            btn_remove.setFixedSize(30, 30)
            
            btn_layout.addStretch()
            btn_layout.addWidget(btn_add)
            btn_layout.addWidget(btn_remove)
            btn_layout.addStretch()
            
            h_layout.addLayout(btn_layout)
            if max_height:
                list_widget.setMaximumHeight(max_height)
            
            list_widget.setMaximumWidth(450) 
            h_layout.addWidget(list_widget)
            
            block_layout.addLayout(h_layout)
            return block_layout, btn_add, btn_remove

        # 1. Cases
        self.list_case = QtWidgets.QListWidget()
        self.list_case.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        ttip = _("Adding columns to the Case will add those values to the case name") + "\n"
        ttip += _("Example: using columns id, country --> ID4_Fiji ")
        case_block, self.btn_case_add, self.btn_case_rem = create_target_block(_("1. Cases / Participants (e.g., ID, Name):"), self.list_case, 70, ttip)
        right_layout.addLayout(case_block)
        
        # 2. Attributes
        self.list_attr = QtWidgets.QListWidget()
        self.list_attr.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        attr_block, self.btn_attr_add, self.btn_attr_rem = create_target_block(_("2. Attributes (e.g., Age, Gender):"), self.list_attr, 70)
        right_layout.addLayout(attr_block)
        
        # 3. Texts
        self.list_text = QtWidgets.QListWidget()
        self.list_text.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        ttip2 = _("Multiple qualitative columns will be collated into one file.") + "\n"
        ttip2 += _("The column name is a header preceding each block of text.")
        text_block, self.btn_text_add, self.btn_text_rem = create_target_block(_("3. Qualitative Texts:"), self.list_text, 70, ttip2)
        right_layout.addLayout(text_block)
        
        layout.addLayout(left_layout)
        layout.addLayout(right_layout)
        main_layout.addLayout(layout)

        # Case or File Attributes Checkbox
        self.cb_case = QtWidgets.QCheckBox(_("Assign attributes to cases (check). Files (uncheck)"))
        self.cb_case.setChecked(False)
        self.cb_case.setEnabled(False)
        main_layout.addWidget(self.cb_case)

        # Autocode Checkbox
        self.cb_autocode = QtWidgets.QCheckBox(_("Autocode text segments using column name"))
        self.cb_autocode.setChecked(False) 
        main_layout.addWidget(self.cb_autocode)
        
        # OK / Cancel Buttons
        bbox = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Ok | QtWidgets.QDialogButtonBox.StandardButton.Cancel)
        bbox.accepted.connect(self.accept)
        bbox.rejected.connect(self.reject)
        main_layout.addWidget(bbox)
        
        # Button connections
        self.btn_case_add.clicked.connect(lambda: self.move_items(self.list_available, self.list_case))
        self.btn_case_rem.clicked.connect(lambda: self.move_items(self.list_case, self.list_available))

        self.btn_attr_add.clicked.connect(lambda: self.move_items(self.list_available, self.list_attr))
        self.btn_attr_rem.clicked.connect(lambda: self.move_items(self.list_attr, self.list_available))

        self.btn_text_add.clicked.connect(lambda: self.move_items(self.list_available, self.list_text))
        self.btn_text_rem.clicked.connect(lambda: self.move_items(self.list_text, self.list_available))

    def move_items(self, source, dest):
        for item in source.selectedItems():
            dest.addItem(item.text())
            source.takeItem(source.row(item))
        if self.list_case:
            self.cb_case.setEnabled(True)
        else:
            self.cb_case.setEnabled(False)
            self.cb_case.setChecked(False)

    def get_selections(self):
        texts = [self.list_text.item(i).text() for i in range(self.list_text.count())]
        cases = [self.list_case.item(i).text() for i in range(self.list_case.count())]
        attrs = [self.list_attr.item(i).text() for i in range(self.list_attr.count())]
        return texts, cases, attrs
        
    def get_autocode_setting(self):
        return self.cb_autocode.isChecked()

    def get_case_setting(self):
        return self.cb_case.isChecked()
