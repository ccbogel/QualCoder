# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
"""

from copy import deepcopy
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import os
import qtawesome as qta  # see: https://pictogrammers.com/library/mdi/

from PyQt6 import QtCore, QtWidgets, QtGui

from .GUI.ui_dialog_cooccurrence import Ui_Dialog_Coocurrence
from .helpers import ExportDirectoryPathDialog, Message
from .report_attributes import DialogSelectAttributeParameters
from .select_items import DialogSelectItems


path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogReportCooccurrence(QtWidgets.QDialog):
    """ Provide a co-occurrence report.
     This shows overlapping and edge-connected codes against codes.
    """

    app = None
    parent_tetEdit = None
    files = []
    attributes = []

    def __init__(self, app, parent_text_edit):
        self.app = app
        self.parent_textEdit = parent_text_edit
        self.attributes = []
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_Coocurrence()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        self.ui.pushButton_export.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export.pressed.connect(self.export_to_excel)
        self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file-outline', options=[{'scale_factor': 1.2}]))
        self.ui.pushButton_select_files.pressed.connect(self.select_files)
        self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_file_attributes.pressed.connect(self.get_files_from_attributes)

        self.ui.pushButton_select_codes.setIcon(qta.icon('mdi6.text', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_select_codes.pressed.connect(self.select_codes)
        self.ui.pushButton_select_categories.setIcon(qta.icon('mdi6.file-tree', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_select_categories.pressed.connect(self.select_categories)
        self.ui.checkBox_hide_blanks.stateChanged.connect(self.show_or_hide_empty_rows_and_cols)
        tablefont = f'font: 10pt "{self.app.settings["font"]}";'
        self.ui.tableWidget.setStyleSheet(tablefont)  # Should be smaller
        self.ui.tableWidget.cellClicked.connect(self.cell_selected)
        #self.ui.tableWidget.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        #self.ui.tableWidget.customContextMenuRequested.connect(self.table_menu)
        self.ui.splitter.setSizes([500, 0])

        self.codes, self.categories = self.app.get_codes_categories()
        self.code_names_list = []
        self.code_names_str = ""
        self.code_ids_str = ""
        for c in self.codes:
            self.code_names_list.append(c['name'])
            self.code_names_str += f"{c['name']}|"
            self.code_ids_str += f",{c['cid']}"
        self.code_ids_str = self.code_ids_str[1:]
        self.selected_codes = deepcopy(self.codes)
        self.result_relations = []
        self.max_count = 0
        self.data_counts = []
        self.data_colors = []
        self.data_details = []
        self.file_ids_names = self.app.get_text_filenames()
        self.process_data()

    def select_files(self):
        """ Select files, stored in self.file_ids_names, then re-load data. """

        selection_list = [{'id': -1, 'name': ''}]
        for file_name in self.app.get_text_filenames():
            selection_list.append(file_name)
        ui = DialogSelectItems(self.app, selection_list, _("Select files"), "multi")
        ok = ui.exec()
        if not ok:
            return
        selected = ui.get_selected()
        if not selected or selected[0]['name'] == '':
            self.file_ids_names = self.app.get_text_filenames()
            Message(self.app, _("Files selected"), _("All files selected")).exec()
            self.ui.pushButton_select_files.setToolTip(_("All files selected"))
            self.ui.pushButton_file_attributes.setToolTip(_("Select files by attributes"))
            self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file', options=[{'scale_factor': 1.4}]))
            self.attributes = []
        else:
            self.file_ids_names = selected
            msg = ""
            for item in self.file_ids_names:
                msg += f"{item['name']}\n"
            Message(self.app, _("Files selected"), msg).exec()
            self.ui.pushButton_select_files.setToolTip(msg)
            self.ui.pushButton_file_attributes.setToolTip(_("Select files by attributes"))
            self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file', options=[{'scale_factor': 1.4}]))
            self.attributes = []

        self.process_data()

    def get_files_from_attributes(self):
        """ Select text files based on attribute selections.
        Attribute results are a dictionary of:
        first item is a Boolean AND or OR list item
        Followed by each attribute list item
        Set:
            self.file_ids_names
        """

        # Clear ui
        self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
        ui = DialogSelectAttributeParameters(self.app)
        ui.fill_parameters(self.attributes)
        temp_attributes = deepcopy(self.attributes)
        self.attributes = []

        ok = ui.exec()
        if not ok:
            self.attributes = temp_attributes
            self.ui.pushButton_file_attributes.setIcon(
                qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
            if self.attributes:
                self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable-box', options=[{'scale_factor': 1.3}]))
                self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file-outline', options=[{'scale_factor': 1.4}]))
            return
        self.attributes = ui.parameters
        if len(self.attributes) == 1:  # Boolean parameter, no attributes selected
            self.ui.pushButton_file_attributes.setIcon(
                qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
            self.attributes = []
            return
        if not ui.result_file_ids:
            Message(self.app, _("Nothing found") + " " * 20, _("No matching files found")).exec()
            self.ui.pushButton_file_attributes.setIcon(
                qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
            return

        # Limit to text files
        text_files = self.app.get_text_filenames()
        self.file_ids_names = []
        msg = ui.tooltip_msg
        for i, file_ in enumerate(text_files):
            if file_['id'] in ui.result_file_ids:
                self.file_ids_names.append(file_)
                if i < 20:
                    msg += f"\n{file_['name']}"
        if len(ui.result_file_ids) > 20:
            msg += f"\nand more. Total files: {len(ui.result_file_ids)}"
        Message(self.app, _("Files selected by attributes"), msg).exec()
        self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable-box', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_file_attributes.setToolTip(msg)
        self.ui.pushButton_select_files.setToolTip(_("Select files"))
        self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file-outline', options=[{'scale_factor': 1.4}]))

        self.process_data()

    def select_codes(self):
        """ Select codes. """

        selection_list = [{'id': -1, 'name': ''}]
        for code_ in self.codes:
            selection_list.append(code_)
        ui = DialogSelectItems(self.app, selection_list, _("Select codes"), "multi")
        ok = ui.exec()
        if not ok:
            return
        self.selected_codes = ui.get_selected()
        if not self.selected_codes or self.selected_codes[0]['name'] == '':
            self.selected_codes = deepcopy(self.codes)
            Message(self.app, _("Codes selected"), _("All codes selected")).exec()
        else:
            msg = ""
            for s in self.selected_codes:
                msg += f"{s['name']}\n"
            Message(self.app, _("Codes selected"), msg).exec()

        self.code_ids_str = ""
        self.code_names_str = ""
        self.code_names_list = []
        for code_ in self.selected_codes:
            self.code_names_list.append(code_['name'])
            self.code_names_str += f"{code_['name']}|"
            self.code_ids_str += f",{code_['cid']}"
        self.code_ids_str = self.code_ids_str[1:]
        self.process_data()

    def select_categories(self):
        """ Select categories and their codes for table. """

        selection_list = [{'id': -1, 'name': ''}]
        for category in self.categories:
            selection_list.append(category)
        ui = DialogSelectItems(self.app, selection_list, _("Select categories"), "multi")
        ok = ui.exec()
        if not ok:
            return
        selected_categories = ui.get_selected()
        if not selected_categories or selected_categories[0]['name'] == '':
            selected_categories = deepcopy(self.categories)
            msg = _("All categories selected")
        else:
            msg = ""
            for category in selected_categories:
                msg += f"{category['name']}\n"

        self.selected_codes = []
        for category in selected_categories:
            codes = self.codes_of_category(category)
            for code_ in codes:
                if code_ not in self.selected_codes:
                    self.selected_codes.append(code_)
        Message(self.app, _("Categories selected"), msg).exec()

        self.code_names_list = []
        self.code_names_str = ""
        self.code_ids_str = ""
        for code_ in self.selected_codes:
            self.code_names_list.append(code_['name'])
            self.code_names_str += f"{code_['name']}|"
            self.code_ids_str += f",{code_['cid']}"
        self.code_ids_str = self.code_ids_str[1:]
        self.process_data()

    def codes_of_category(self, node):
        """ Get child codes of this category node.
        Only keep the category or code name.

        param: node : Dictionary of category

        return: selected_codes : List of Code Dictionaries
        """

        child_cat_names = []
        codes, categories = self.app.get_codes_categories()
        """ Create a list of this category (node) and all its category children.
        Maximum depth of 200. """
        selected_categories = [node]
        i = 0  # Ensure an exit from loop
        new_model_changed = True
        while categories != [] and new_model_changed and i < 200:
            new_model_changed = False
            append_list = []
            for sel_cat in selected_categories:
                for cat in categories:
                    if cat['supercatid'] == sel_cat['catid']:
                        append_list.append(cat)
                        child_cat_names.append({'name': cat['name'], 'catid': cat['catid']})
            for n in append_list:
                selected_categories.append(n)
                categories.remove(n)
                new_model_changed = True
            i += 1
        categories = selected_categories
        # Ignore codes that are not associated with these selected categories and sub-categories
        selected_codes = []
        for category in categories:
            for code in codes:
                if code['catid'] == category['catid']:
                    selected_codes.append(code)
        return selected_codes

    def process_data(self):
        """ Calculate the relations for selected codes for ALL coders (TODO only THIS coder).
        For text codings only. """

        self.ui.checkBox_hide_blanks.setChecked(False)
        self.ui.splitter.setSizes([500, 0])

        self.result_relations = []
        self.calculate_relations(self.code_ids_str)

        # Create data matrices zeroed, codes are ordered alphabetically by name
        self.data_counts = []
        self.data_colors = []
        self.data_details = []
        for row in self.selected_codes:
            self.data_counts.append([0] * len(self.selected_codes))
            self.data_colors.append([""] * len(self.selected_codes))
            self.data_details.append(["."] * len(self.selected_codes))

        self.max_count = 0
        for r in self.result_relations:
            row_pos = self.code_names_list.index(r['c0_name'])
            col_pos = self.code_names_list.index(r['c1_name'])
            self.data_counts[row_pos][col_pos] += 1
            if self.data_counts[row_pos][col_pos] > self.max_count:
                self.max_count = self.data_counts[row_pos][col_pos]
            '''print(r['cid0'], r['c0_name'], r['ctid0'], r['cid1'], r['c1_name'], r['ctid1'], r['fid'], r['file_name'],
                  r['owners'], r['c0_pos0'], r['c0_pos1'], r['c1_pos0'], r['c1_pos1'])'''
            res_list = [r['cid0'], r['c0_name'], r['ctid0'], r['cid1'], r['c1_name'], r['ctid1'], r['fid'],
                        r['file_name'], r['owners'], r['c0_pos0'], r['c0_pos1'], r['c1_pos0'], r['c1_pos1'],
                        r['text_before'], r['text_overlap'], r['text_after'], r['relation']]
            if self.data_details[row_pos][col_pos] == ".":
                self.data_details[row_pos][col_pos] = [res_list]
            else:
                self.data_details[row_pos][col_pos].append(res_list)

        # Color heat map for spread across 5 colours
        colors = ["#F8E0E0", "#F6CECE", "#F5A9A9", "#F78181", "#FA5858"]  # Light to dark red
        for row, row_data in enumerate(self.data_counts):
            for col, item_data in enumerate(row_data):
                if self.data_counts[row][col] > 0:
                    color_range_index = int(self.data_counts[row][col] / self.max_count * 5) - 1
                    if color_range_index < 0:
                        color_range_index = 0
                    self.data_colors[row][col] = colors[color_range_index]
        self.fill_table()

    def export_to_excel(self):
        """ Export to Excel file. """

        filename = "Code_cooccurrence.xlsx"
        export_dir = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_dir.filepath
        if filepath is None:
            return

        # Excel vertical and horizontal headers
        header = []
        for code_ in self.selected_codes:  # self.codes:
            name_split_50 = [code_['name'][y - 50:y] for y in range(50, len(code_['name']) + 50, 50)]
            # header_labels.append(code_['name'])  # OLD, need line separators
            header.append("\n".join(name_split_50))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Counts"
        wb.create_sheet("Details")
        ws2 = wb["Details"]
        for col in range(len(self.codes)):
            ws.column_dimensions[get_column_letter(col + 1)].width = 20
            ws2.column_dimensions[get_column_letter(col + 1)].width = 20
        for col, col_name in enumerate(header):
            h_cell = ws.cell(row=1, column=col + 2)
            h_cell.value = col_name
            h_cell2 = ws2.cell(row=1, column=col + 2)
            h_cell2.value = col_name
            v_cell = ws.cell(row=col + 2, column=1)
            v_cell.value = col_name
            v_cell2 = ws2.cell(row=col + 2, column=1)
            v_cell2.value = col_name
        # Co-occurrence counts
        for row, row_data in enumerate(self.data_counts):
            for col, col_data in enumerate(row_data):
                cell = ws.cell(row=row + 2, column=col + 2)
                cell.value = col_data
                if self.data_colors[row][col] != "":
                    cell.fill = PatternFill(start_color=self.data_colors[row][col][1:], end_color=self.data_colors[row][col][1:], fill_type="solid")
                # Details list
                if self.data_details[row][col] == ".":
                    continue
                details = ""
                for data in self.data_details[row][col]:
                    '''
                    0 - 5 [r['cid0'], r['c0_name'], r['ctid0'], r['cid1'], r['c1_name'], r['ctid1'], 
                    6 - 8 r['fid'], r['file_name'], r['owners'], 
                    9 - 12 r['c0_pos0'], r['c0_pos1'], r['c1_pos0'], r['c1_pos1'],
                    13 - 16 r['text_before'], r['text_overlap'], r['text_after'], r['relation']]
                    '''
                    details += f"Codes: {data[1]} ({data[9]} - {data[10]})| {data[4]} ({data[11]} - {data[12]})\n"
                    details += f"Coders: {data[8]}. (ctid0: {data[2]} | ctid1: {data[5]})\n"
                    details += f"File (fid {data[6]}): {data[7]}\n"
                    details += f"{data[13]}[[{data[14]}]]{data[15]}\n========\n"
                d_cell = ws2.cell(row=row + 2, column=col + 2)
                d_cell.value = details

        wb.save(filepath)
        msg = _('Co-occurrence exported: ') + filepath
        Message(self.app, _('Co-occurrence exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def show_or_hide_empty_rows_and_cols(self):
        """ Unchecked - show all rows and columns.
        Checked - hide rows and columns with no code co-occurrences. """

        if self.ui.checkBox_hide_blanks.isChecked():
            for row, row_data in enumerate(self.data_counts):
                if sum(row_data) == 0:
                    self.ui.tableWidget.hideRow(row)

            for col in range(len(self.data_counts)):
                col_sum = 0
                for row, row_data in enumerate(self.data_counts):
                    col_sum += row_data[col]
                if col_sum == 0:
                    self.ui.tableWidget.hideColumn(col)

        if not self.ui.checkBox_hide_blanks.isChecked():
            for row, row_data in enumerate(self.data_counts):
                self.ui.tableWidget.showRow(row)
                self.ui.tableWidget.showColumn(row)

    def cell_selected(self):
        """ When the table widget memo cell is selected display the memo.
        Update memo text, or delete memo by clearing text.
        If a new memo, also show in table widget by displaying MEMO in the memo column.
        """

        row = self.ui.tableWidget.currentRow()
        col = self.ui.tableWidget.currentColumn()
        text = self.ui.tableWidget.item(row, col).text()
        if text == "":
            return
        self.ui.textEdit.clear()
        data_list = self.data_details[row][col]
        '''
        0 - 5 [r['cid0'], r['c0_name'], r['ctid0'], r['cid1'], r['c1_name'], r['ctid1'], 
        6 - 8 r['fid'], r['file_name'], r['owners'], 
        9 - 12 r['c0_pos0'], r['c0_pos1'], r['c1_pos0'], r['c1_pos1'],
        13 - 16 r['text_before'], r['text_overlap'], r['text_after'], r['relation']
        '''
        # Colours for overlapping and non-overlapping text
        color_yellow = "#F4FA58"  # Coder0 - first and usually lowest pos0
        brush_yellow = QtGui.QBrush(QtGui.QColor(color_yellow))
        color_blue = "#81BEF7"  # Coder1 - second and usually highest pos1
        brush_blue = QtGui.QBrush(QtGui.QColor(color_blue))
        color_green = "#81F781"  # Overlap color
        brush_green = QtGui.QBrush(QtGui.QColor(color_green))

        for data in data_list:
            msg = f"Codes: {data[1]} ({data[9]} - {data[10]})| {data[4]} ({data[11]} - {data[12]})\n"
            msg += f"Coders: {data[8]}. (ctid0: {data[2]} | ctid1: {data[5]})\n"
            msg += f"File (fid {data[6]}): {data[7]}\n"
            # msg += f"\nrelation: {data[16]}\n"  # testing
            self.ui.textEdit.append(msg)

            # Coded text highlights - yellow code 0, green overlap, blue code 1
            start_pos_yellow = len(self.ui.textEdit.toPlainText())
            end_pos_yellow = start_pos_yellow + len(data[13]) + 1
            start_pos_green = end_pos_yellow
            end_pos_green = start_pos_green + len(data[14])
            start_pos_blue = end_pos_green
            end_pos_blue = start_pos_blue + len(data[15])

            msg = f"{data[13]}{data[14]}{data[15]}"
            self.ui.textEdit.append(msg)

            cursor = self.ui.textEdit.textCursor()
            fmt_before = QtGui.QTextCharFormat()
            cursor.setPosition(start_pos_yellow, QtGui.QTextCursor.MoveMode.MoveAnchor)
            cursor.setPosition(end_pos_yellow, QtGui.QTextCursor.MoveMode.KeepAnchor)
            fmt_before.setBackground(brush_yellow)
            cursor.mergeCharFormat(fmt_before)

            fmt_overlap = QtGui.QTextCharFormat()
            cursor.setPosition(start_pos_green, QtGui.QTextCursor.MoveMode.MoveAnchor)
            cursor.setPosition(end_pos_green, QtGui.QTextCursor.MoveMode.KeepAnchor)
            fmt_overlap.setBackground(brush_green)
            cursor.mergeCharFormat(fmt_overlap)

            # may need yellow or blue
            fmt_after = QtGui.QTextCharFormat()
            cursor.setPosition(start_pos_blue, QtGui.QTextCursor.MoveMode.MoveAnchor)
            cursor.setPosition(end_pos_blue, QtGui.QTextCursor.MoveMode.KeepAnchor)
            fmt_after.setBackground(brush_blue)
            if data[16] == "I":  # A code is Included inside another code
                fmt_after.setBackground(brush_yellow)
            cursor.mergeCharFormat(fmt_after)
            msg = "========"
            self.ui.textEdit.append(msg)

        self.ui.splitter.setSizes([300, 200])

    ''' TODO for future expansion maybe.
    def table_menu(self, position):
        """ Context menu for displaying table cell coding details.
        """

        row = self.ui.tableWidget.currentRow()
        col = self.ui.tableWidget.currentColumn()
        text = self.ui.tableWidget.item(row, col).text()
        print(row, col, text)

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_view = menu.addAction(_("View"))'''

    def fill_table(self):
        """ Fill table using code names alphabetically (case insensitive), using self.data """

        rows = self.ui.tableWidget.rowCount()
        for r in range(0, rows):
            self.ui.tableWidget.removeRow(0)
        cols = self.ui.tableWidget.columnCount()
        for c in range(0, cols):
            self.ui.tableWidget.removeColumn(0)

        header_labels = []
        # Wrong for selected codes
        for code_ in self.selected_codes:  # self.codes:
            name_split_50 = [code_['name'][y - 50:y] for y in range(50, len(code_['name']) + 50, 50)]
            # header_labels.append(code_['name'])  # OLD, needed line separators
            header_labels.append("\n".join(name_split_50))
        self.ui.tableWidget.setColumnCount(len(header_labels))
        self.ui.tableWidget.setHorizontalHeaderLabels(header_labels)
        self.ui.tableWidget.setRowCount(len(header_labels))
        self.ui.tableWidget.setVerticalHeaderLabels(header_labels)
        for row, row_data in enumerate(self.data_counts):
            for col, cell_data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem()
                if self.data_colors[row][col] != "":
                    item.setBackground(QtGui.QBrush(QtGui.QColor(self.data_colors[row][col])))
                    item.setForeground(QtGui.QBrush(QtGui.QColor("#000000")))
                if cell_data > 0:
                    item.setData(QtCore.Qt.ItemDataRole.DisplayRole, cell_data)
                item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
                self.ui.tableWidget.setItem(row, col, item)
        # self.ui.tableWidget.resizeColumnsToContents()  # Doesnt look great
        self.ui.tableWidget.resizeRowsToContents()

    def calculate_relations(self, code_ids_str):
        """ Calculate the relations for selected codes for all coders.
        For codings in code_text only.

        id1, id2, overlapindex, unionindex, distance, whichmin, whichmax, fid
        relation is 1 character: Inclusion, Overlap, Exact, (Proximity - not used)
        owners is a combination of: owner of ctid0 pipe owner of ctid1

        Args:
            code_ids_str (String): comma separated code ids
        """

        selected_relations = ['E', 'I', 'O']
        if self.file_ids_names is None:
            self.file_ids_names = self.app.get_text_filenames()

        # Get codings for each selected text file
        cur = self.app.conn.cursor()
        for fid in self.file_ids_names:
            sql = "select fid, code_text.cid, pos0, pos1, name, ctid,seltext, ifnull(code_text.memo,''), " \
                  "code_text.owner from code_text " \
                  "join code_name on code_name.cid=code_text.cid where fid=? " \
                  "and code_text.cid in (" + code_ids_str + ") order by code_text.cid"
            cur.execute(sql, [fid['id']])
            result = cur.fetchall()
            coded = [row for row in result if row[0] == fid['id']]

            # Look at each code again other codes, when done remove from list of codes
            cid = 1
            pos0 = 2
            pos1 = 3
            name = 4
            ctid = 5
            seltext = 6
            coded_memo = 7
            owner = 8
            while len(coded) > 0:
                c0 = coded.pop()
                for c1 in coded:
                    if c0[cid] != c1[cid]:
                        relation = self.relation(c0, c1)
                        # Add extra details for output
                        relation['c0_name'] = c0[name]
                        relation['c1_name'] = c1[name]
                        relation['fid'] = fid['id']
                        relation['file_name'] = fid['name']
                        relation['c0_pos0'] = c0[pos0]
                        relation['c0_pos1'] = c0[pos1]
                        relation['c1_pos0'] = c1[pos0]
                        relation['c1_pos1'] = c1[pos1]
                        relation['owners'] = f"{c0[owner]}|{c1[owner]}"
                        relation['ctid0'] = c0[ctid]
                        relation['ctid0_text'] = c0[seltext]
                        relation['ctid1'] = c1[ctid]
                        relation['ctid1_text'] = c1[seltext]
                        relation['coded_memo0'] = c0[coded_memo]
                        relation['coded_memo1'] = c1[coded_memo]
                        # Append relation based on comboBox selection
                        if relation['relation'] in selected_relations:
                            self.result_relations.append(relation)

    def relation(self, c0, c1):
        """ Relation function as in RQDA

        whichmin is the code with the lowest pos0, or None if equal
        whichmax is the code with the highest pos1 or None if equal
        operlapindex is the combined lowest to the highest positions. Only used for E, O, P
        unionindex is the lowest and highest positions of the union of overlap. Only used for E, O

        Called by:
            calculate_relations_for_coder_and_selected_codes

        Returns:
        id1, id2, overlapindex, unionindex, distance, whichmin, min, whichmax, max, fid
        relation is 1 character: Inclusion, Overlap, Exact, Proximity
        actual text as before, overlap, after
        """

        # fid = 0
        cid = 1
        pos0 = 2
        pos1 = 3
        result = {"cid0": c0[cid], "cid1": c1[cid], "relation": "", "whichmin": None, "min": 0,
                  "whichmax": None, "max": 0, "overlapindex": None, "unionindex": None, "distance": None,
                  "text_before": "", "text_overlap": "", "text_after": ""}

        cur = self.app.conn.cursor()

        # Which min
        if c0[pos0] < c1[pos0]:
            result['whichmin'] = c0[cid]
            result['min'] = c0[pos0]
        if c1[pos0] < c0[pos0]:
            result['whichmin'] = c1[cid]
            result['min'] = c1[pos0]

        # Which max
        if c0[pos1] > c1[pos1]:
            result['whichmax'] = c0[cid]
            result['max'] = c0[pos1]
        if c1[pos1] > c0[pos1]:
            result['whichmax'] = c1[cid]
            result['max'] = c1[pos1]

        # Check for Exact
        if c0[pos0] == c1[pos0] and c0[pos1] == c1[pos1]:
            result['relation'] = "E"
            result['overlapindex'] = [c0[pos0], c0[pos1]]
            result['unionindex'] = [c0[pos0], c0[pos1]]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c0[pos1] - c0[pos0], c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
                result['text_before'] = ""
                result['text_after'] = ""
                result['distance'] = 0
            return result

        # Check for Proximity
        if c0[pos1] < c1[pos0]:
            result['relation'] = "P"
            result['distance'] = c1[pos0] - c0[pos1]
            result['text_overlap'] = ""
            result['text_before'] = ""
            result['text_after'] = ""
            return result
        if c0[pos0] > c1[pos1]:
            result['relation'] = "P"
            result['distance'] = c0[pos0] - c1[pos1]
            result['text_overlap'] = ""
            result['text_before'] = ""
            result['text_after'] = ""
            return result

        # Check for Inclusion
        # Exact has been resolved above
        # c0 inside c1
        if c0[pos0] >= c1[pos0] and c0[pos1] <= c1[pos1]:
            result['relation'] = "I"
            result['overlapindex'] = [c0[pos0], c0[pos1]]
            result['unionindex'] = [c0[pos0], c0[pos1]]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c0[pos1] - c0[pos0], c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, c0[pos0] - c1[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos1] + 1, c1[pos1] - c0[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

        # c1 inside c0
        if c1[pos0] >= c0[pos0] and c1[pos1] <= c0[pos1]:
            result['relation'] = "I"
            result['overlapindex'] = [c1[pos0], c1[pos1]]
            result['unionindex'] = [c1[pos0], c1[pos1]]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, c1[pos1] - c1[pos0], c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c1[pos0] - c0[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos1] + 1, c0[pos1] - c1[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

        # Check for Overlap
        # Should be all that is remaining
        # c0 overlaps on the right side, left side is not overlapping
        if c0[pos0] < c1[pos0] and c0[pos1] < c1[pos1]:
            '''print("c0 overlaps on the right side, left side is not overlapping")
            print("c0", c0)
            print("C1", c1)'''
            result['relation'] = "O"
            # Reorder lowest to highest
            result['overlapindex'] = sorted([c0[pos0], c1[pos1]])
            result['unionindex'] = sorted([c0[pos1], c1[pos0]])
            overlap_length = result['unionindex'][1] - result['unionindex'][0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, overlap_length, c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c1[pos0] - c0[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos1] + 1, c1[pos1] - c0[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

        # c1 overlaps on the right side, left side is not overlapping
        if c1[pos0] < c0[pos0] and c1[pos1] < c0[pos1]:
            result['relation'] = "O"
            result['overlapindex'] = sorted([c1[pos0], c0[pos1]])
            result['unionindex'] = sorted([c1[pos1], c0[pos0]])
            overlap_length = result['unionindex'][1] - result['unionindex'][0]
            '''print("TODO c1 overlaps on the right, the left side is not overlapping")
            print("C0", c0)
            print("C1", c1)'''
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, overlap_length, c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, c0[pos0] - c1[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos1] + 1, c0[pos1] - c1[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result
