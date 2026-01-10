# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com/
https://qualcoder-org.github.io/
"""

from copy import copy, deepcopy
import csv
import logging
import openpyxl
import os
import pandas as pd
import plotly.express as px
import qtawesome as qta  # see: https://pictogrammers.com/library/mdi/
import statistics

from PyQt6 import QtGui, QtWidgets, QtCore
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QBrush

from .color_selector import TextColor
from .GUI.ui_dialog_code_relations import Ui_Dialog_CodeRelations
from .helpers import DialogCodeInText, ExportDirectoryPathDialog, Message
from .report_attributes import DialogSelectAttributeParameters
from .select_items import DialogSelectItems

path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogReportRelations(QtWidgets.QDialog):
    """ Show code relations/crossovers for one coder.
    This is for text only. """

    app = None
    parent_textEdit = None
    coder_names = []
    categories = []
    codes = []
    files = []
    result_relations = []
    result_summary = []
    dataframe = None
    attributes = []

    def __init__(self, app, parent_textedit):

        self.app = app
        self.parent_textEdit = parent_textedit
        self.get_code_data()
        self.result_relations = []
        self.result_summary = []
        self.attributes = []
        self.dataframe = None
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_CodeRelations()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        try:
            s0 = int(self.app.settings['dialogcodecrossovers_splitter0'])
            s1 = int(self.app.settings['dialogcodecrossovers_splitter1'])
            if s0 > 10 and s1 > 10:
                self.ui.splitter.setSizes([s0, s1])
        except KeyError:
            pass
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        font = f'font: {self.app.settings["treefontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.treeWidget.setStyleSheet(font)
        self.ui.label_codes.setStyleSheet(font)
        self.ui.pushButton_exportcsv.setIcon(qta.icon('mdi6.export'))
        self.ui.pushButton_exportcsv.pressed.connect(self.export_excel_file)  # changed from csv to excel export
        self.ui.pushButton_export_exact.setIcon(qta.icon('mdi6.export'))
        self.ui.pushButton_export_exact.pressed.connect(self.export_exact_excel_file)
        self.ui.pushButton_calculate.setIcon(qta.icon('mdi6.calculator', options=[{'scale_factor': 1.5}]))
        self.ui.pushButton_calculate.pressed.connect(self.calculate_code_relations)
        self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_select_files.pressed.connect(self.select_files)
        self.ui.pushButton_boxplots.setIcon(qta.icon('mdi6.chart-bar', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_boxplots.pressed.connect(self.create_boxplots)
        self.ui.pushButton_search_next.setIcon(qta.icon('mdi6.play'))
        self.ui.pushButton_search_next.clicked.connect(self.search_text)
        self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_file_attributes.pressed.connect(self.get_files_from_attributes)
        self.ui.tableWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget.customContextMenuRequested.connect(self.table_menu)
        self.ui.tableWidget.setTabKeyNavigation(False)
        self.ui.tableWidget_statistics.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget_statistics.customContextMenuRequested.connect(self.table_statistics_menu)
        self.ui.tableWidget_statistics.setTabKeyNavigation(False)
        self.ui.treeWidget.setSelectionMode(QtWidgets.QTreeWidget.SelectionMode.ExtendedSelection)
        self.ui.treeWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.treeWidget.customContextMenuRequested.connect(self.tree_menu)
        self.fill_tree()
        # These signals after the tree is filled the first time
        self.ui.treeWidget.itemCollapsed.connect(self.get_collapsed)
        self.ui.treeWidget.itemExpanded.connect(self.get_collapsed)

        # Default to select all files
        cur = self.app.conn.cursor()
        sql = "select distinct name, id from source where id in (select fid from code_text) order by name"
        cur.execute(sql)
        res = cur.fetchall()
        self.files = []
        for r in res:
            self.files.append({'name': r[0], 'fid': r[1]})

    def select_files(self):
        """ Select files for analysis. """

        cur = self.app.conn.cursor()
        sql = "select distinct name, id from source where id in (select fid from code_text) order by name"
        cur.execute(sql)
        res = cur.fetchall()
        all_files = [{'name': '', 'fid': -1}]
        for r in res:
            all_files.append({'name': r[0], 'fid': r[1]})
        ui = DialogSelectItems(self.app, all_files, _("Select files"), "multi")
        ok = ui.exec()
        if not ok:
            return
        self.files = []
        selected = ui.get_selected()
        for s in selected:
            if s['fid'] == -1:
                self.files = all_files[1:]
                self.ui.pushButton_select_files.setToolTip(_("All files selected"))
                self.ui.pushButton_file_attributes.setToolTip(_("Select files by attributes"))
                self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
                self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file', options=[{'scale_factor': 1.4}]))
                self.attributes = []
                return
        tt = _("Files selected: ")
        for s in selected:
            self.files.append(s)
            tt += "\n" + s['name']
        self.ui.pushButton_select_files.setToolTip(tt)
        self.ui.pushButton_file_attributes.setToolTip(_("Select files by attributes"))
        self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file', options=[{'scale_factor': 1.4}]))
        self.attributes = []

    def get_files_from_attributes(self):
        """ Select text files based on attribute selections.
        Attribute results are a dictionary of:
        first item is a Boolean AND or OR list item
        Followed by each attribute list item
        Set:
            self.file_ids
        """

        # Clear ui
        self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
        ui = DialogSelectAttributeParameters(self.app)

        ui.fill_parameters(self.attributes)
        temp_attributes = deepcopy(self.attributes)
        self.attributes = []

        ok = ui.exec()
        if not ok:
            self.attributes = temp_attributes
            self.ui.pushButton_file_attributes.setIcon(
                qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
            if self.attributes:
                self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable-box', options=[{'scale_factor': 1.3}]))
                self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file-outline', options=[{'scale_factor': 1.4}]))
            return
        self.attributes = ui.parameters
        if len(self.attributes) == 1:  # Boolean parameter, no attributes selected
            self.ui.pushButton_file_attributes.setIcon(
                qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
            self.attributes = []
            return
        if not ui.result_file_ids:
            Message(self.app, _("Nothing found") + " " * 20, _("No matching files found")).exec()
            self.ui.pushButton_file_attributes.setIcon(
                qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
            self.ui.pushButton_file_attributes.setToolTip(_("Attributes"))
            return

        # Limit to text files
        text_files = self.app.get_text_filenames()
        self.files = []
        msg = ui.tooltip_msg
        for i, file_ in enumerate(text_files):
            if file_['id'] in ui.result_file_ids:
                file_['fid'] = file_['id']  # Need fid key for calucte_code_relations
                self.files.append(file_)
                if i < 20:
                    msg += f"\n{file_['name']}"
        if len(ui.result_file_ids) > 20:
            msg += f"\nand more. Total files: {len(ui.result_file_ids)}"
        Message(self.app, _("Files selected by attributes"), msg).exec()

        self.ui.pushButton_file_attributes.setIcon(qta.icon('mdi6.variable-box', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_file_attributes.setToolTip(msg)
        self.ui.pushButton_select_files.setToolTip(_("Select files"))
        self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file-outline', options=[{'scale_factor': 1.4}]))

    def get_code_data(self):
        """ Called from init. gets code_names, categories and owner names.
        """

        self.coder_names = self.app.get_coder_names_in_project()
        self.codes, self.categories = self.app.get_codes_categories()

    def calculate_code_relations(self):
        """ Calculate the relations for selected codes for THIS coder or ALL coders.
        For codings in code_text only. """

        sel_codes = []
        codes_str = ""
        code_ids = ""
        items = self.ui.treeWidget.selectedItems()
        for i in items:
            if i.text(1)[:3] == "cid":
                sel_codes.append({"name": i.text(0), "cid": int(i.text(1)[4:])})
                codes_str += i.text(0) + "|"
                code_ids += "," + i.text(1)[4:]
        if len(sel_codes) < 2:
            msg = _("Select 2 or more codes\nUse Ctrl or Shift and mouse click")
            Message(self.app, _('Selection warning'), msg, "warning").exec()
            return
        code_ids = code_ids[1:]
        self.ui.label_codes.setText(_("Codes: ") + codes_str)
        self.ui.label_codes.setToolTip(_("Codes: ") + codes_str)
        self.result_relations = []
        if self.ui.radioButton_this.isChecked():
            self.calculate_relations_for_coder_and_selected_codes(self.app.settings['codername'], code_ids)
        else:
            for coder_name in self.coder_names:
                self.calculate_relations_for_coder_and_selected_codes(coder_name, code_ids)
        self.fill_table()
        self.summary_statistics()

    def calculate_relations_for_coder_and_selected_codes(self, coder_name, code_ids):
        """ Calculate the relations for selected codes for selected coder.
        For codings in code_text only.

        id1, id2, overlapindex, unionindex, distance, whichmin, whichmax, fid
        relation is 1 character: Inclusion, Overlap, Exact, Proximity
        Requires:
            self.files List of Dict
        """

        index = self.ui.comboBox_relation_type.currentIndex()
        selected_relations = ['E', 'I', 'O', 'P']
        if index == 1:  # Overlap
            selected_relations = ['O']
        if index == 2:  # Inclusion
            selected_relations = ['I']
        if index == 3:  # Exact
            selected_relations = ['E']
        if index == 4:  # Proximity
            selected_relations = ['P']
        if index == 5:  # Overlap Inclusion
            selected_relations = ['O', 'I']
        if index == 6:  # Overlap Inclusion Exact
            selected_relations = ['O', 'I', 'E']

        selected_fids = ""
        for f in self.files:
            selected_fids += "," + str(f['fid'])
        try:
            selected_fids = selected_fids[1:]
        except IndexError:
            return

        cur = self.app.conn.cursor()
        sql = "select distinct fid, name from code_text join source on source.id=code_text.fid " \
              "where code_text.owner=? and code_text.cid in (" + code_ids + ") and " \
                                                                            "fid in (" + selected_fids + ") order by fid"
        cur.execute(sql, [coder_name, ])
        result = cur.fetchall()
        file_ids_names = []
        for r in result:
            file_ids_names.append({'fid': r[0], 'filename': r[1]})
        if not file_ids_names:
            return

        # Get codings for each selected text file separately
        for fid_name in file_ids_names:
            sql = "select fid, code_text.cid, pos0, pos1, name, ctid,seltext, ifnull(code_text.memo,'') from code_text " \
                  "join code_name on code_name.cid=code_text.cid where code_text.owner=? and fid=? " \
                  "and code_text.cid in (" + code_ids + ") order by code_text.cid"
            cur.execute(sql, [coder_name, fid_name['fid']])
            result = cur.fetchall()
            coded = [row for row in result if row[0] == fid_name['fid']]
            '''for row in result:
                if row[0] == fid_name['fid']:
                    coded.append(row)'''

            # TODO later, find the closest Other code for relation analysis
            # Look at each code again other codes, when done remove from list of codes
            cid = 1
            pos0 = 2
            pos1 = 3
            name = 4
            ctid = 5
            seltext = 6
            coded_memo = 7
            while len(coded) > 0:
                c0 = coded.pop()
                for c1 in coded:
                    if c0[cid] != c1[cid]:
                        relation = self.relation(c0, c1)
                        # Add extra details for output
                        relation['c0_name'] = c0[name]
                        relation['c1_name'] = c1[name]
                        relation['fid'] = fid_name['fid']
                        relation['file_name'] = fid_name['filename']
                        relation['c0_pos0'] = c0[pos0]
                        relation['c0_pos1'] = c0[pos1]
                        relation['c1_pos0'] = c1[pos0]
                        relation['c1_pos1'] = c1[pos1]
                        relation['owner'] = coder_name
                        relation['ctid0'] = c0[ctid]
                        relation['ctid0_text'] = c0[seltext]
                        relation['ctid1'] = c1[ctid]
                        relation['ctid1_text'] = c1[seltext]
                        relation['coded_memo0'] = c0[coded_memo]
                        relation['coded_memo1'] = c1[coded_memo]
                        # Append relation based on comboBox selection
                        if relation['relation'] in selected_relations:
                            self.result_relations.append(relation)

    def relation(self, c0, c1):
        """ Relation function as in RQDA

        whichmin is the code with the lowest pos0, or None if equal
        whichmax is the code with the highest pos1 or None if equal
        operlapindex is the combined lowest to the highest positions. Only used for E, O, P
        unionindex is the lowest and highest positions of the union of overlap. Only used for E, O

        Called by:
            calculate_relations_for_coder_and_selected_codes

        Returns:
        id1, id2, overlapindex, unionindex, distance, whichmin, min, whichmax, max, fid
        relation is 1 character: Inclusion, Overlap, Exact, Proximity
        actual text as before, overlap, after
        """

        # fid = 0
        cid = 1
        pos0 = 2
        pos1 = 3
        result = {"cid0": c0[cid], "cid1": c1[cid], "relation": "", "whichmin": None, "min": 0,
                  "whichmax": None, "max": 0, "overlapindex": None, "unionindex": None, "distance": None,
                  "text_before": "", "text_overlap": "", "text_after": ""}

        cur = self.app.conn.cursor()

        # Which min
        if c0[pos0] < c1[pos0]:
            result['whichmin'] = c0[cid]
            result['min'] = c0[pos0]
        if c1[pos0] < c0[pos0]:
            result['whichmin'] = c1[cid]
            result['min'] = c1[pos0]

        # Which max
        if c0[pos1] > c1[pos1]:
            result['whichmax'] = c0[cid]
            result['max'] = c0[pos1]
        if c1[pos1] > c0[pos1]:
            result['whichmax'] = c1[cid]
            result['max'] = c1[pos1]

        # Check for Exact
        if c0[pos0] == c1[pos0] and c0[pos1] == c1[pos1]:
            result['relation'] = "E"
            result['overlapindex'] = [c0[pos0], c0[pos1]]
            result['unionindex'] = [c0[pos0], c0[pos1]]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c0[pos1] - c0[pos0], c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
                result['text_before'] = ""
                result['text_after'] = ""
                result['distance'] = 0
            return result

        # Check for Proximity
        if c0[pos1] < c1[pos0]:
            result['relation'] = "P"
            result['distance'] = c1[pos0] - c0[pos1]
            result['text_overlap'] = ""
            result['text_before'] = ""
            result['text_after'] = ""
            return result
        if c0[pos0] > c1[pos1]:
            result['relation'] = "P"
            result['distance'] = c0[pos0] - c1[pos1]
            result['text_overlap'] = ""
            result['text_before'] = ""
            result['text_after'] = ""
            return result

        # Check for Inclusion
        # Exact has been resolved above
        # c0 inside c1
        if c0[pos0] >= c1[pos0] and c0[pos1] <= c1[pos1]:
            result['relation'] = "I"
            result['overlapindex'] = [c0[pos0], c0[pos1]]
            result['unionindex'] = [c0[pos0], c0[pos1]]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c0[pos1] - c0[pos0], c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, c0[pos0] - c1[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos1] + 1, c1[pos1] - c0[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

        # c1 inside c0
        if c1[pos0] >= c0[pos0] and c1[pos1] <= c0[pos1]:
            result['relation'] = "I"
            result['overlapindex'] = [c1[pos0], c1[pos1]]
            result['unionindex'] = [c1[pos0], c1[pos1]]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, c1[pos1] - c1[pos0], c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c1[pos0] - c0[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos1] + 1, c0[pos1] - c1[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

        # Check for Overlap
        # Should be all that is remaining
        # c0 overlaps on the right side, left side is not overlapping
        if c0[pos0] < c1[pos0] and c0[pos1] < c1[pos1]:
            '''print("c0 overlaps on the right side, left side is not overlapping")
            print("c0", c0)
            print("C1", c1)'''
            result['relation'] = "O"
            # Reorder lowest to highest
            result['overlapindex'] = sorted([c0[pos0], c1[pos1]])
            result['unionindex'] = sorted([c0[pos1], c1[pos0]])
            overlap_length = result['unionindex'][1] - result['unionindex'][0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, overlap_length, c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c1[pos0] - c0[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos1] + 1, c1[pos1] - c0[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

        # c1 overlaps on the right side, left side is not overlapping
        if c1[pos0] < c0[pos0] and c1[pos1] < c0[pos1]:
            result['relation'] = "O"
            result['overlapindex'] = sorted([c1[pos0], c0[pos1]])
            result['unionindex'] = sorted([c1[pos1], c0[pos0]])
            overlap_length = result['unionindex'][1] - result['unionindex'][0]
            '''print("TODO c1 overlaps on the right, the left side is not overlapping")
            print("C0", c0)
            print("C1", c1)'''
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, overlap_length, c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, c0[pos0] - c1[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos1] + 1, c0[pos1] - c1[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

    def tree_menu(self, position):
        """ Context menu for treewidget code/category items.
        Select all codes. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        selected = self.ui.treeWidget.currentItem()
        action_clear_selected = menu.addAction(_("Clear all"))
        action_select_all = menu.addAction(_("Select all"))

        action = menu.exec(self.ui.treeWidget.mapToGlobal(position))
        if action == action_clear_selected:
            selected = self.ui.treeWidget.selectedItems()
            for tree_item in selected:
                tree_item.setSelected(False)
            return
        if action == action_select_all:
            self.ui.treeWidget.selectAll()

    def search_text(self):
        """ Search for text in the results. """

        search_text = self.ui.lineEdit_search_results.text()
        if search_text == "":
            return
        row_count = self.ui.tableWidget.rowCount()
        col_count = self.ui.tableWidget.columnCount()
        if row_count == 0:
            return
        current_row = 0
        current_col = 0
        try:
            current_row = self.ui.tableWidget.currentRow()
            current_col = self.ui.tableWidget.currentColumn()
        except AttributeError:
            # No table for table menu
            return
        cell_counter_pos = current_row * col_count + current_col
        found_row = -1
        found_col = -1
        for row in range(0, row_count):
            for col in range(0, col_count):
                try:
                    cell = self.ui.tableWidget.item(row, col).text()
                    if cell_counter_pos < row * col_count + col and found_row == -1 and search_text in cell:
                        found_row = row
                        found_col = col
                        break
                except AttributeError:
                    pass
                if found_row != -1:
                    break
        if found_row == -1:
            return
        self.ui.tableWidget.setCurrentCell(found_row, found_col)

    def table_menu(self, position):
        """ Context menu to show row text in original context, row ordering. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        cell_value = ""
        try:
            row = self.ui.tableWidget.currentRow()
            col = self.ui.tableWidget.currentColumn()
            cell_value = self.ui.tableWidget.item(row, col).text()
        except AttributeError:
            # No table for table menu
            return
        action_show_context = menu.addAction(_("View in context"))
        action_sort_ascending = menu.addAction(_("Sort ascending"))
        action_sort_descending = menu.addAction(_("Sort descending"))
        action_filter_equals = menu.addAction(_("Filter equals: ") + cell_value)
        action_filter_greater = menu.addAction(_("Filter greater or equals: ") + cell_value)
        action_filter_lower = menu.addAction(_("Filter lower or equals: ") + cell_value)
        action_clear_filter = menu.addAction(_("Clear filter"))
        action = menu.exec(self.ui.tableWidget.mapToGlobal(position))
        if action == action_show_context:
            self.show_context()
        if action == action_sort_ascending:
            self.ui.tableWidget.sortItems(col, QtCore.Qt.SortOrder.AscendingOrder)
        if action == action_sort_descending:
            self.ui.tableWidget.sortItems(col, QtCore.Qt.SortOrder.DescendingOrder)
        if action == action_clear_filter:
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
        if action == action_filter_equals:
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
                if self.ui.tableWidget.item(r, col).text() != cell_value:
                    self.ui.tableWidget.setRowHidden(r, True)
        if action == action_filter_greater:
            val_type = "str"
            if col in (0, 1, 2, 4, 5, 6, 7, 8, 9, 10, 15, 16):
                val_type = "int"
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
                if val_type == "str" and self.ui.tableWidget.item(r, col).text() < cell_value:
                    self.ui.tableWidget.setRowHidden(r, True)
                if val_type == "int" and int(self.ui.tableWidget.item(r, col).text()) < int(cell_value):
                    self.ui.tableWidget.setRowHidden(r, True)
        if action == action_filter_lower:
            val_type = "str"
            if col in (0, 1, 2, 4, 5, 6, 7, 8, 9, 10, 15, 16):
                val_type = "int"
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
                if val_type == "str" and self.ui.tableWidget.item(r, col).text() > cell_value:
                    self.ui.tableWidget.setRowHidden(r, True)
                if val_type == "int" and int(self.ui.tableWidget.item(r, col).text()) > int(cell_value):
                    self.ui.tableWidget.setRowHidden(r, True)

    def show_context(self):
        """ Show context of coding in dialog.
        Called by table_menu.
        """

        row = self.ui.tableWidget.currentRow()
        d = self.result_relations[row]
        codename0 = ""
        codename1 = ""
        color0 = ""
        color1 = ""
        for c in self.codes:
            if c['cid'] == d['cid0']:
                codename0 = c['name']
                color0 = c['color']
            if c['cid'] == d['cid1']:
                codename1 = c['name']
                color1 = c['color']
        # data: dictionary: codename, color, file_or_casename, pos0, pos1, text, coder, fid, file_or_case,
        # textedit_start, textedit_end
        data0 = {'codename': codename0, 'color': color0, 'file_or_casename': d['file_name'],
                 'pos0': d['c0_pos0'], 'pos1': d['c0_pos1'],
                 'text': '', 'coder': d['owner'], 'fid': d['fid'], 'file_or_case': 'File'}
        data1 = {'codename': codename1, 'color': color1, 'file_or_casename': d['file_name'],
                 'pos0': d['c1_pos0'], 'pos1': d['c1_pos1'],
                 'text': '', 'coder': d['owner'], 'fid': d['fid'], 'file_or_case': 'File'}
        ui = DialogCodeInText(self.app, data0)
        ui.add_coded_text(data1)
        ui.exec()
        return

    def fill_table(self):
        """ A table of:
        Tooltips with codenames on id1,id2, relation,fid - to minimise screen use
        id1, id2, overlapindex, unionindex, distance, whichmin, whichmax, fid
        relation is: inclusion, overlap, exact, proximity

        https://stackoverflow.com/questions/60512920/sorting-numbers-in-qtablewidget-work-doesnt-right-pyqt5
        """

        fid = 0
        code0 = 1
        code1 = 2
        relation_type = 3
        min_ = 4
        max_ = 5
        overlap0 = 6
        overlap1 = 7
        union0 = 8
        union1 = 9
        distance = 10
        text_before = 11
        text_overlap = 12
        text_after = 13
        owner = 14
        ctid0 = 15
        ctid1 = 16
        memo0 = 17
        memo1 = 18
        col_names = ["FID", _("Code") + " 0", _("Code") + " 1", "Rel", "Min", "Max", _("Overlap") + " 0",
                     _("Overlap") + " 1", _("Union") + " 0", _("Union") + " 1",
                     _("Distance"), _("Text before"), _("Overlap"), _("Text after"), _("Owner"), "ctid0", "ctid1",
                     _("Memo") + "0", _("Memo") + "1"]
        self.ui.tableWidget.setColumnCount(len(col_names))
        self.ui.tableWidget.setHorizontalHeaderLabels(col_names)
        rows = self.ui.tableWidget.rowCount()
        for r in range(0, rows):
            self.ui.tableWidget.removeRow(0)
        for r, i in enumerate(self.result_relations):
            self.ui.tableWidget.insertRow(r)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, i['fid'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            item.setToolTip(i['file_name'])
            self.ui.tableWidget.setItem(r, fid, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, i['c0_name'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            item.setToolTip(f"cid:{i['cid0']}\n{i['c0_pos0']} - {i['c0_pos1']}")
            self.ui.tableWidget.setItem(r, code0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole,i['c1_name'])  # i['cid1'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            item.setToolTip(f"cid:{i['cid1']}\n{i['c1_pos0']} - {i['c1_pos1']}")
            self.ui.tableWidget.setItem(r, code1, item)
            item = QtWidgets.QTableWidgetItem(str(i['relation']))
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            ttip = _("Proximity")
            if i['relation'] == "O":
                ttip = _("Overlap")
            if i['relation'] == "E":
                ttip = _("Exact")
            if i['relation'] == "I":
                ttip = _("Inclusion")
            item.setToolTip(ttip)
            self.ui.tableWidget.setItem(r, relation_type, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, i['whichmin'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            if i['whichmin'] is not None:
                ttip = i['c0_name']
                if i['whichmin'] == i['cid1']:
                    ttip = i['c1_name']
                item.setToolTip(ttip)
            self.ui.tableWidget.setItem(r, min_, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, i['whichmax'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            if i['whichmax'] is not None:
                ttip = i['c0_name']
                if i['whichmax'] == i['cid1']:
                    ttip = i['c1_name']
                item.setToolTip(ttip)
            self.ui.tableWidget.setItem(r, max_, item)
            if i['overlapindex'] is not None:
                item = QtWidgets.QTableWidgetItem(str(i['overlapindex'][0]))
                item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
                self.ui.tableWidget.setItem(r, overlap0, item)
                item = QtWidgets.QTableWidgetItem(str(i['overlapindex'][1]))
                item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
                self.ui.tableWidget.setItem(r, overlap1, item)
            if i['unionindex'] is not None:
                item = QtWidgets.QTableWidgetItem(str(i['unionindex'][0]))
                item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
                self.ui.tableWidget.setItem(r, union0, item)
                item = QtWidgets.QTableWidgetItem(str(i['unionindex'][1]))
                item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
                self.ui.tableWidget.setItem(r, union1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, i['distance'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, distance, item)
            item = QtWidgets.QTableWidgetItem(i['text_before'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, text_before, item)
            item = QtWidgets.QTableWidgetItem(i['text_overlap'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, text_overlap, item)
            item = QtWidgets.QTableWidgetItem(i['text_after'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, text_after, item)
            item = QtWidgets.QTableWidgetItem(str(i['owner']))
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, owner, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, i['ctid0'])
            item.setToolTip(i['ctid0_text'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, ctid0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, i['ctid1'])
            item.setToolTip(i['ctid1_text'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, ctid1, item)
            item = QtWidgets.QTableWidgetItem(i['coded_memo0'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, memo0, item)
            item = QtWidgets.QTableWidgetItem(i['coded_memo1'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, memo1, item)
        self.ui.tableWidget.resizeColumnsToContents()

    def export_exact_excel_file(self):
        """ Export exact match text codings for all codes as excel file.
        Output ordered by filename and code name ascending. """

        cur = self.app.conn.cursor()
        sql = "select fid, source.name, code_text.cid, code_name.name, seltext,pos0,pos1,code_text.owner from " \
              "code_text join code_name on code_name.cid=code_text.cid join source on source.id=code_text.fid " \
              "order by source.name, code_name.name"
        cur.execute(sql)
        res = cur.fetchall()
        coded_text0 = []
        keys = 'fid', 'filename', 'cid', 'codename', 'text', 'pos0', 'pos1', 'owner'
        for row in res:
            coded_text0.append(dict(zip(keys, row)))

        coded_text1 = deepcopy(coded_text0)
        result = []
        for i in coded_text0:
            tmp_result = []
            for j in coded_text1:
                if i != j and i['fid'] == j['fid'] and i['pos0'] == j['pos0'] and i['pos1'] == j['pos1']:
                    tmp_result.append(j)
            if tmp_result:
                result.append(i)
                # Remove matches from coded_text1 to avoid result duplications
                coded_text1.remove(i)
                for t in tmp_result:
                    result.append(t)
                    # Remove matches from coded_text1 to avoid result duplications
                    coded_text1.remove(t)
            if tmp_result:
                result.append({'fid': "", 'filename': "", 'cid': "", 'codename': "", 'text': "", 'pos0': "", 'pos1': "", 'owner': ""})
        if not result:
            msg = _("No exact matches found.")
            Message(self.app, _('No results'), msg, "information").exec()
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        # Column headings
        col_headings = ["Filename", "Codename", "pos0", "pos1", "Text", "Owner"]
        row = 1
        for col, code in enumerate(col_headings):
            ws.cell(column=col + 1, row=row, value=code)
        for row, data in enumerate(result):
            ws.cell(column=1, row=row + 2, value=data['filename'])
            ws.cell(column=2, row=row + 2, value=data['codename'])
            ws.cell(column=3, row=row + 2, value=data['pos0'])
            ws.cell(column=4, row=row + 2, value=data['pos1'])
            ws.cell(column=5, row=row + 2, value=data['text'])
            ws.cell(column=6, row=row + 2, value=data['owner'])
        filepath, ok = QtWidgets.QFileDialog.getSaveFileName(self,
                                                             _("Save Excel File"), self.app.settings['directory'],
                                                             "XLSX Files(*.xlsx)")
        # options=QtWidgets.QFileDialog.Option.DontUseNativeDialog)
        if filepath is None or not ok:
            return
        if filepath[-4:] != ".xlsx":
            filepath += ".xlsx"
        wb.save(filepath)
        msg = _("Report of exact matches for text codings by file and code") + "\n"
        msg += _("Each row contains filename, codename, pos0, pos1, text, owner.") + "\n"
        msg += _('Report exported to: ') + filepath
        Message(self.app, _('Report exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def export_excel_file(self):
        """ Export data as excel file plus csv stats file,
        The main file is called projectname_relations.xlsx.
        The summary file (if generated) is called projectname_relations_stats.csv
        """

        if not self.result_relations:
            return
        shortname = self.app.project_name.split(".qda")[0]
        filename = shortname + "_relations.xlsx"
        export_dir = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_dir.filepath
        if filepath is None:
            return
        col_names = ["Fid", _("Filename"), "Code0", "Code0 " + _("name"), "Code0_pos0", "Code0_pos1",
                     "Code1", "Code1 " + _("name"),
                     "Code1_pos0", "Code1_pos1", _("Relation"), _("Minimum"), _("Maximum"),
                     _("Overlap") + " 0", _("Overlap") + " 1", _("Union") + " 0",
                     _("Union") + " 1", _("Distance"), _("Text before"), _("Text overlap"), _("Text after"), _("Owner"),
                     "ctid0", "ctid1", "text0", "text1", _("Memo") + "0", _("Memo") + "1"]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Code Relations"
        wb.create_sheet("Statistics")
        ws2 = wb["Statistics"]
        row = 1
        for col, header in enumerate(col_names):
            ws.cell(column=col + 1, row=row, value=header)
        for row, data in enumerate(self.result_relations):
            ws.cell(column=1, row=row + 2, value=data['fid'])
            ws.cell(column=2, row=row + 2, value=data['file_name'])
            ws.cell(column=3, row=row + 2, value=data['cid0'])
            ws.cell(column=4, row=row + 2, value=data['c0_name'])
            ws.cell(column=5, row=row + 2, value=data['c0_pos0'])
            ws.cell(column=6, row=row + 2, value=data['c0_pos1'])
            ws.cell(column=7, row=row + 2, value=data['cid1'])
            ws.cell(column=8, row=row + 2, value=data['c1_name'])
            ws.cell(column=9, row=row + 2, value=data['c1_pos0'])
            ws.cell(column=10, row=row + 2, value=data['c1_pos1'])
            ws.cell(column=11, row=row + 2, value=data['relation'])
            ws.cell(column=12, row=row + 2, value=str(data['whichmin']).replace('None', ''))
            ws.cell(column=13, row=row + 2, value=str(data['whichmax']).replace('None', ''))
            if data['overlapindex']:
                ws.cell(column=14, row=row + 2, value=data['overlapindex'][0])
                ws.cell(column=15, row=row + 2, value=data['overlapindex'][1])
            else:
                ws.cell(column=14, row=row + 2, value='')
                ws.cell(column=15, row=row + 2, value='')
            if data['unionindex']:
                ws.cell(column=16, row=row + 2, value=data['unionindex'][0])
                ws.cell(column=17, row=row + 2, value=data['unionindex'][1])
            else:
                ws.cell(column=16, row=row + 2, value='')
                ws.cell(column=17, row=row + 2, value='')
            ws.cell(column=18, row=row + 2, value=str(data['distance']).replace('None', ''))
            ws.cell(column=19, row=row + 2, value=data['text_before'])
            ws.cell(column=20, row=row + 2, value=data['text_overlap'])
            ws.cell(column=21, row=row + 2, value=data['text_after'])
            ws.cell(column=22, row=row + 2, value=data['owner'])
            ws.cell(column=23, row=row + 2, value=data['ctid0'])
            ws.cell(column=24, row=row + 2, value=data['ctid1'])
            ws.cell(column=25, row=row + 2, value=data['ctid0_text'])
            ws.cell(column=26, row=row + 2, value=data['ctid1_text'])
            ws.cell(column=27, row=row + 2, value=data['coded_memo0'])
            ws.cell(column=28, row=row + 2, value=data['coded_memo1'])

        # Write statistical summary to statistics sheet
        stats_col_names = ["Code0", "Code0 " + _("name"), "Code1", "Code1 " + _("name"), "Count", _("Minimum"), "Q1",
                           "Median", "Q3", _("Maximum"), "Mean", "std dev"]
        row = 1
        for col, header in enumerate(stats_col_names):
            ws2.cell(column=col + 1, row=row, value=header)
        for row, r in enumerate(self.result_summary):
            ws2.cell(column=1, row=row + 2, value=r['cid0'])
            ws2.cell(column=2, row=row + 2, value=r['c0_name'])
            ws2.cell(column=3, row=row + 2, value=r['cid1'])
            ws2.cell(column=4, row=row + 2, value=r['c1_name'])
            ws2.cell(column=5, row=row + 2, value=str(r['count']))
            ws2.cell(column=6, row=row + 2, value=str(r['min']))
            ws2.cell(column=7, row=row + 2, value=str(r['quantiles'][0]))
            ws2.cell(column=8, row=row + 2, value=str(r['quantiles'][1]))
            ws2.cell(column=9, row=row + 2, value=str(r['quantiles'][2]))
            ws2.cell(column=10, row=row + 2, value=str(r['max']))
            ws2.cell(column=11, row=row + 2, value=str(r['mean']))
            ws2.cell(column=12, row=row + 2, value=str(r['stdev']))
        wb.save(filepath)
        msg = _("Code relations file exported to: ") + filepath
        Message(self.app, _('File export'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def closeEvent(self, event):
        """ Save splitter dimensions. """

        sizes = self.ui.splitter.sizes()
        self.app.settings['dialogcodecrossovers_splitter0'] = sizes[0]
        self.app.settings['dialogcodecrossovers_splitter1'] = sizes[1]

    # Statistics
    def summary_statistics(self):
        """ Show summary coding distance statistics.
         Called after the main results are produced. """

        self.result_summary = []
        rows = self.ui.tableWidget_statistics.rowCount()
        for r in range(0, rows):
            self.ui.tableWidget_statistics.removeRow(0)
        # Setup list of dictionaries with cid0 and cid1 as identifiers
        for i in self.result_relations:
            relation = {'cid0': i['cid0'], 'cid1': i['cid1']}
            if relation not in self.result_summary:
                self.result_summary.append(relation)
        # Fill each dictionary with data and statistics
        pandas_data = []
        for r in self.result_summary:
            data = []
            distances = []
            for res in self.result_relations:
                if r['cid0'] == res['cid0'] and r['cid1'] == res['cid1']:
                    r['c0_name'] = res['c0_name']
                    r['c1_name'] = res['c1_name']
                    data.append(res)
                    distances.append(res['distance'])
                    pandas_data.append([r['c0_name'] + ":" + r['c1_name'], res['distance']])
            r['data'] = data
            # Statistics descriptive summary
            r['count'] = len(data)
            r['max'] = max(distances)
            r['min'] = min(distances)
            r['mean'] = round(statistics.mean(distances), 5)
            try:
                r['stdev'] = round(statistics.stdev(distances), 5)
            except statistics.StatisticsError:
                r['stdev'] = ""
            try:
                r['quantiles'] = statistics.quantiles(distances, method='inclusive')
            except statistics.StatisticsError:
                r['quantiles'] = ["", "", ""]
        columns = [_("Code pair"), _("Distance (characters)")]
        self.dataframe = pd.DataFrame(data=pandas_data, columns=columns)
        self.fill_table_statistics()

    def create_boxplots(self):
        """ Create multiple boxplots. """

        if self.dataframe is None:
            return
        fig = px.box(self.dataframe, x="Code pair", y="Distance (characters)", title=_("Code relations"))
        fig.update_traces(quartilemethod="inclusive")  # or "inclusive", or "linear" by default
        fig.show()

    def fill_table_statistics(self):
        """ Fill statistics table with statistical summary of results.
         """

        col_stats_names = [_("Code") + " 0", _("Code") + " 1", _("Count"), "Min", "1st Q",
                           "Median", "3rd Q", "Max", "mean", "std dev"]
        self.ui.tableWidget_statistics.setColumnCount(len(col_stats_names))
        self.ui.tableWidget_statistics.setHorizontalHeaderLabels(col_stats_names)

        for row, rel in enumerate(self.result_summary):
            self.ui.tableWidget_statistics.insertRow(row)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['c0_name'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            item.setToolTip(f"cid:{rel['cid0']}")
            self.ui.tableWidget_statistics.setItem(row, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['c1_name'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            item.setToolTip(f"cid:{rel['cid1']}")
            self.ui.tableWidget_statistics.setItem(row, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['count'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget_statistics.setItem(row, 2, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['min'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget_statistics.setItem(row, 3, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['quantiles'][0])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget_statistics.setItem(row, 4, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['quantiles'][1])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget_statistics.setItem(row, 5, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['quantiles'][2])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget_statistics.setItem(row, 6, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['max'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget_statistics.setItem(row, 7, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['mean'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget_statistics.setItem(row, 8, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, rel['stdev'])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget_statistics.setItem(row, 9, item)

    def table_statistics_menu(self, position):
        """ Context menu to order rows. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        cell_value = ""
        try:
            row = self.ui.tableWidget_statistics.currentRow()
            col = self.ui.tableWidget_statistics.currentColumn()
            cell_value = self.ui.tableWidget_statistics.item(row, col).text()
        except AttributeError:
            # No table for table menu
            return
        action_sort_ascending = menu.addAction(_("Sort ascending"))
        action_sort_descending = menu.addAction(_("Sort descending"))
        action_filter_equals = menu.addAction(_("Filter equals: ") + cell_value)
        action_clear_filter = menu.addAction(_("Clear filter"))
        action = menu.exec(self.ui.tableWidget_statistics.mapToGlobal(position))
        if action == action_sort_ascending:
            self.ui.tableWidget_statistics.sortItems(col, QtCore.Qt.SortOrder.AscendingOrder)
        if action == action_sort_descending:
            self.ui.tableWidget_statistics.sortItems(col, QtCore.Qt.SortOrder.DescendingOrder)
        if action == action_clear_filter:
            for r in range(0, self.ui.tableWidget_statistics.rowCount()):
                self.ui.tableWidget_statistics.setRowHidden(r, False)
        if action == action_filter_equals:
            for r in range(0, self.ui.tableWidget_statistics.rowCount()):
                self.ui.tableWidget_statistics.setRowHidden(r, False)
                if self.ui.tableWidget_statistics.item(r, col).text() != cell_value:
                    self.ui.tableWidget_statistics.setRowHidden(r, True)

    def get_collapsed(self, item):
        """ On category collapse or expansion signal, find the collapsed parent category items.
        This will fill the self.app.collapsed_categories and is the expanded/collapsed tree is then replicated across
        other areas of the app. """

        if item.text(1)[:3] == "cid":
            return
        if not item.isExpanded() and item.text(1) not in self.app.collapsed_categories:
            self.app.collapsed_categories.append(item.text(1))
        if item.isExpanded() and item.text(1) in self.app.collapsed_categories:
            self.app.collapsed_categories.remove(item.text(1))

    def fill_tree(self):
        """ Fill tree widget, top level items are main categories and unlinked codes.
        """

        self.ui.treeWidget.clear()

        cats = copy(self.categories)
        codes = copy(self.codes)
        self.ui.treeWidget.clear()
        header = [_("Code Tree"), _("Id")]
        self.ui.treeWidget.setColumnCount(len(header))
        self.ui.treeWidget.setHeaderLabels(header)
        self.ui.treeWidget.header().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.ui.treeWidget.header().setStretchLastSection(False)
        # Add top level categories
        remove_list = []
        for c in cats:
            if c['supercatid'] is None:
                top_item = QtWidgets.QTreeWidgetItem([c['name'], 'catid:' + str(c['catid'])])
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                if f"catid:{c['catid']}" in self.app.collapsed_categories:
                    top_item.setExpanded(False)
                else:
                    top_item.setExpanded(True)
                remove_list.append(c)
        for item in remove_list:
            cats.remove(item)

        ''' Add child categories. Look at each unmatched category, iterate through tree to
        add as child then remove matched categories from the list. '''
        count = 0
        while len(cats) > 0 and count < 10000:
            remove_list = []
            for c in cats:
                it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
                item = it.value()
                while item:  # while there is an item in the list
                    # logger.debug("While: ", item.text(0), item.text(1), c['catid'], c['supercatid'])
                    if item.text(1) == 'catid:' + str(c['supercatid']):
                        child = QtWidgets.QTreeWidgetItem([c['name'], 'catid:' + str(c['catid'])])
                        child.setToolTip(0, c['name'])
                        item.addChild(child)
                        if f"catid:{c['catid']}" in self.app.collapsed_categories:
                            child.setExpanded(False)
                        else:
                            child.setExpanded(True)
                        remove_list.append(c)
                    it += 1
                    item = it.value()
            for item in remove_list:
                cats.remove(item)
            count += 1

        # Add unlinked codes as top level items
        remove_items = []
        for c in codes:
            if c['catid'] is None:
                top_item = QtWidgets.QTreeWidgetItem([c['name'], 'cid:' + str(c['cid'])])
                top_item.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                color = TextColor(c['color']).recommendation
                top_item.setForeground(0, QBrush(QtGui.QColor(color)))
                top_item.setFlags(
                    Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_items.append(c)
        for item in remove_items:
            codes.remove(item)

        # Add codes as children
        for c in codes:
            it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
            item = it.value()
            while item:
                if item.text(1) == 'catid:' + str(c['catid']):
                    child = QtWidgets.QTreeWidgetItem([c['name'], 'cid:' + str(c['cid'])])
                    child.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                    color = TextColor(c['color']).recommendation
                    child.setForeground(0, QBrush(QtGui.QColor(color)))
                    child.setFlags(
                        Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                    child.setToolTip(0, c['name'])
                    item.addChild(child)
                    c['catid'] = -1  # make unmatchable
                it += 1
                item = it.value()
        # self.ui.treeWidget.expandAll()
