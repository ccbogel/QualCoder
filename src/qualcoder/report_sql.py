# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com/
https://qualcoder-org.github.io/
"""

from PyQt6 import QtCore, QtGui
from PyQt6 import QtWidgets
from PyQt6.QtCore import Qt

import csv
from datetime import datetime
import logging
import openpyxl
import os
import qtawesome as qta
import sqlite3

from .GUI.ui_dialog_SQL import Ui_Dialog_sql
from .save_sql_query import DialogSaveSql
from .helpers import ExportDirectoryPathDialog, Message
from .highlighter import Highlighter

path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogSQL(QtWidgets.QDialog):
    """ Uses single inheritance, subclass QDialog and set up the user interface in
    the __init__() method.
    A gui to allow the user to enter sql queries and return results.
    Data outputs are as tab (or other) separated files.
    DEFAULT_SQL is listed at end of module for additional complex queries. """

    app = None
    schema = None
    parent_textEdit = None
    sql = ""
    stored_sqls = []  # a list of dictionaries of user created sql, as {index, sql}
    default_sqls = []  # a list of dictionaries of default sql, as {index, sql}
    file_data = []  # for file exports
    results = None  # SQL results
    queryTime = ""  # for label tooltip
    queryFilters = ""  # for label tooltip
    cell_value = ""
    row = -1
    col = -1

    def __init__(self, app_, parent_textedit):

        QtWidgets.QDialog.__init__(self)
        self.app = app_
        self.parent_textEdit = parent_textedit
        self.queryTime = ""
        self.queryFilters = ""

        # Set up the user interface from Designer.
        self.ui = Ui_Dialog_sql()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        doc_font = f'font: {self.app.settings["docfontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.tableWidget_results.setStyleSheet(doc_font)
        self.ui.treeWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        highlighter = Highlighter(self.ui.textEdit_sql)
        if self.app.settings['stylesheet'] in ("dark", "rainbow"):
            highlighter.create_rules(dark=True)
        self.ui.textEdit_sql.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.textEdit_sql.customContextMenuRequested.connect(self.sql_menu)
        self.ui.textEdit_sql.setTabChangesFocus(True)
        self.ui.textEdit_sql.setAcceptRichText(False)
        self.ui.tableWidget_results.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget_results.customContextMenuRequested.connect(self.table_menu)
        self.ui.tableWidget_results.setTabKeyNavigation(False)

        # Add tables and fields to treeWidget
        self.get_schema_update_tree_widget()
        self.ui.treeWidget.itemClicked.connect(self.get_item)
        self.ui.treeWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.treeWidget.customContextMenuRequested.connect(self.tree_menu)
        # qtawsome see: https://pictogrammers.com/library/mdi/
        self.ui.pushButton_runSQL.setIcon(qta.icon('mdi6.play', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_runSQL.clicked.connect(self.run_sql)
        self.ui.pushButton_csv.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_csv.clicked.connect(self.export_csv_file)
        self.ui.pushButton_excel.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_excel.clicked.connect(self.export_excel_file)

        self.ui.splitter.setSizes([20, 180])
        try:
            s0 = int(self.app.settings['dialogsql_splitter_h0'])
            s1 = int(self.app.settings['dialogsql_splitter_h1'])
            if s0 > 10 and s1 > 10:
                self.ui.splitter.setSizes([s0, s1])
        except KeyError:
            pass
        self.ui.splitter_2.setSizes([10, 290])
        try:
            s0 = int(self.app.settings['dialogsql_splitter_v0'])
            s1 = int(self.app.settings['dialogsql_splitter_v1'])
            if s0 > 10 and s1 > 10:
                self.ui.splitter_2.setSizes([s0, s1])
        except KeyError:
            pass
        self.ui.splitter.splitterMoved.connect(self.update_sizes)
        self.ui.splitter_2.splitterMoved.connect(self.update_sizes)

    def update_sizes(self):
        """ Called by splitter resized """

        sizes = self.ui.splitter.sizes()
        self.app.settings['dialogsql_splitter_h0'] = sizes[0]
        self.app.settings['dialogsql_splitter_h1'] = sizes[1]
        sizes = self.ui.splitter_2.sizes()
        self.app.settings['dialogsql_splitter_v0'] = sizes[0]
        self.app.settings['dialogsql_splitter_v1'] = sizes[1]

    def export_excel_file(self):
        """ Load result set and export to Excel file. """

        cur = self.app.conn.cursor()
        sql = self.ui.textEdit_sql.toPlainText()
        if "select" not in sql.lower():
            Message(self.app, _("No select query"), _("No data to export")).exec()
            return
        try:
            cur.execute(sql)
        except Exception as e:
            Message(self.app, _("SQL error"), str(e), "warning").exec()
            return

        results = cur.fetchall()
        header = []
        if cur.description is not None:
            header = list(map(lambda x: x[0], cur.description))  # Gets column names
        filename = "sql_report.xlsx"
        export_dir = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_dir.filepath
        if filepath is None:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        # Excel headers
        for col, col_name in enumerate(header):
            cell = ws.cell(row=1, column=col + 2)
            cell.value = col_name

        # Data
        for row, row_data in enumerate(results):
            for col, col_data in enumerate(row_data):
                cell = ws.cell(row=row + 2, column=col + 2)
                cell.value = col_data

        '''for c in range(0, col_count):
            for r in range(0, row_count):
                te = self.te[r][c]
                try:
                    data_text = te.toPlainText()
                except AttributeError:  # None type error
                    data_text = ""
                cell = ws.cell(row=r + 2, column=c + 2)
                cell.value = data_text'''

        wb.save(filepath)
        msg = _('Results exported: ') + filepath
        Message(self.app, _('Results exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def export_csv_file(self):
        """ Load result set and export results to a delimited .csv file
        using \r\n as line separators. """

        cur = self.app.conn.cursor()
        sql = self.ui.textEdit_sql.toPlainText()
        if "select" not in sql.lower():
            Message(self.app, _("No select query"), _("No data to export")).exec()
            return
        try:
            cur.execute(sql)
        except Exception as e:
            Message(self.app, _("SQL error"), str(e), "warning").exec()
            return
        results = cur.fetchall()
        header = []
        if cur.description is not None:
            header = list(map(lambda x: x[0], cur.description))  # Gets column names
        filename = "sql_report.csv"
        export_dir = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_dir.filepath
        if filepath is None:
            return
        quote_option = csv.QUOTE_MINIMAL
        if self.ui.checkBox_quote.isChecked():
            quote_option = csv.QUOTE_ALL
        delimiter_ = str(self.ui.comboBox_delimiter.currentText())
        if delimiter_ == "tab":
            delimiter_ = "\t"
        with open(filepath, 'wt', encoding='utf-8-sig') as export_file:
            csv_writer = csv.writer(export_file, delimiter=delimiter_, quoting=quote_option)
            csv_writer.writerow(header)
            for row in results:
                csv_writer.writerow(row)
        msg = _("SQL Results exported to: ") + filepath
        self.parent_textEdit.append(msg)
        self.parent_textEdit.append(_("Query:") + "\n" + sql)
        Message(self.app, _("CSV file export"), msg, "information").exec()

    def get_item(self):
        """ Get the selected table name or tablename.fieldname and add to the sql text
        at the current cursor position.
        Get a default query and replace sql text in text edit.
        Get a stored query and replace sql text in text edit """

        item_text = self.ui.treeWidget.currentItem().text(0)
        index = self.ui.treeWidget.currentIndex()
        # Check use stored sql to fill correct text for sql
        for s in self.stored_sqls:
            if index == s['index']:
                self.ui.textEdit_sql.clear()
                self.ui.textEdit_sql.setText(s['sql'])
                return
        for d in self.default_sqls:
            if index == d['index']:
                self.ui.textEdit_sql.clear()
                self.ui.textEdit_sql.setText(d['sql'])
                return
        if index.parent().row() != -1:  # There is a parent if not -1
            item_parent = self.ui.treeWidget.itemFromIndex(index.parent())
            item_parent_text = item_parent.text(0)
            item_text = f"{item_parent_text}.{item_text}"
        cursor = self.ui.textEdit_sql.textCursor()
        cursor.insertText(f" {item_text} ")

    def run_sql(self):
        """ Run the sql text and add the results to the results text edit. """

        # Clear tableWidget and file data
        num_rows = self.ui.tableWidget_results.rowCount()
        for row in range(0, num_rows):
            self.ui.tableWidget_results.removeRow(0)
        self.ui.tableWidget_results.setHorizontalHeaderLabels([""])
        self.file_data = []
        self.ui.label.setText(_("Running query. Please wait."))
        QtWidgets.QApplication.processEvents()  # stops gui freeze
        self.sql = self.ui.textEdit_sql.toPlainText()
        cur = self.app.conn.cursor()
        self.sql = str(self.sql)
        QtWidgets.QApplication.processEvents()  # Stops gui freeze
        try:
            time0 = datetime.now()
            cur.execute(self.sql)
            self.ui.label.setToolTip("")
            self.results = cur.fetchall()
            time1 = datetime.now()
            timediff = time1 - time0
            self.queryTime = f"Time:{timediff}"
            self.ui.label.setToolTip(self.queryTime)
            self.ui.label.setText(str(len(self.results)) + _(" rows"))
            # Extra messaging where rows will be zero
            if self.sql[0:12].upper() == "CREATE TABLE":
                self.ui.label.setText(_("Table created"))
                self.app.delete_backup = False
            if self.sql[0:12].upper() == "CREATE INDEX":
                self.ui.label.setText(_("Index created"))
                self.app.delete_backup = False
                self.app.conn.commit()
            if self.sql[0:6].upper() == "DELETE":
                self.ui.label.setText(str(cur.rowcount) + _(" rows deleted"))
                self.app.delete_backup = False
                self.app.conn.commit()
            if self.sql[0:6].upper() == "UPDATE":
                self.ui.label.setText(str(cur.rowcount) + _(" rows updated"))
                self.app.delete_backup = False
                self.app.conn.commit()
            col_names = []
            if cur.description is not None:
                col_names = list(map(lambda x: x[0], cur.description))  # gets column names
            self.ui.tableWidget_results.setColumnCount(len(col_names))
            self.ui.tableWidget_results.setHorizontalHeaderLabels(col_names)
            self.file_data.append(col_names)
            for row, row_results in enumerate(self.results):
                self.file_data.append(row_results)
                self.ui.tableWidget_results.insertRow(row)
                for col, value in enumerate(row_results):
                    item = QtWidgets.QTableWidgetItem()
                    item.setData(QtCore.Qt.ItemDataRole.DisplayRole, value)
                    item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled)
                    self.ui.tableWidget_results.setItem(row, col, item)
            self.ui.tableWidget_results.resizeColumnsToContents()
            # Keep column widths reasonable, 500 pixels max
            for i in range(self.ui.tableWidget_results.columnCount()):
                if self.ui.tableWidget_results.columnWidth(i) > 500:
                    self.ui.tableWidget_results.setColumnWidth(i, 500)
            self.ui.tableWidget_results.resizeRowsToContents()
            sql_string = str(self.sql).upper()
            if sql_string.find("CREATE ") == 0 or sql_string.find("DROP ") == 0 or sql_string.find("ALTER ") == 0:
                self.get_schema_update_tree_widget()
        except (sqlite3.OperationalError, sqlite3.IntegrityError) as e:
            Message(self.app, _("Error"), str(e), "warning").exec()
            self.ui.label.setText(_("SQL Error"))
            self.ui.label.setToolTip(str(e))
        self.results = None
        self.app.conn.commit()

    def get_schema_update_tree_widget(self):
        """ Get table schema from database, and update the tables_an_views tree widget.
        The schema needs to be updated when drop table or create queries are run. """

        self.stored_sqls = []
        self.schema = []
        table_dict = {}
        cur = self.app.conn.cursor()
        cur.execute("SELECT sql, type, name FROM sqlite_master WHERE type IN ('table', 'view') ")
        result = cur.fetchall()
        for row in result:
            table_name = row[2]
            field_results = cur.execute(f"PRAGMA table_info({table_name})")
            # each field is a tuple of cid, name, type (integer, text, ), notNull (1=notNull),
            # defaultValue(None usually), primaryKey(as integers 1 up, or 0)
            fields = [field for field in field_results]
            table_dict[table_name] = fields
        self.schema = table_dict

        # Fill tree widget with tables and views
        tables_and_views = [k for k in self.schema.keys()]
        tables_and_views.sort()
        self.ui.treeWidget.clear()
        for table_name in tables_and_views:
            top_item = QtWidgets.QTreeWidgetItem()
            top_item.setText(0, table_name)
            result = cur.execute(f"SELECT type FROM sqlite_master WHERE name='{table_name}' ")
            table_or_view = result.fetchone()[0]
            if table_or_view == "view":
                top_item.setBackground(0, QtGui.QBrush(Qt.GlobalColor.yellow, Qt.BrushStyle.Dense6Pattern))
            self.ui.treeWidget.addTopLevelItem(top_item)
            for field in self.schema[table_name]:
                field_item = QtWidgets.QTreeWidgetItem()
                if table_or_view == "view":
                    field_item.setBackground(0, QtGui.QBrush(Qt.GlobalColor.yellow, Qt.BrushStyle.Dense6Pattern))
                if field[5] > 0:
                    field_item.setForeground(0, QtGui.QBrush(Qt.GlobalColor.red))
                field_item.setText(0, field[1])
                top_item.addChild(field_item)

        # Add default sqls
        default_item = QtWidgets.QTreeWidgetItem()
        default_item.setText(0, _("Default Queries"))
        self.ui.treeWidget.addTopLevelItem(default_item)
        for query in DEFAULT_SQL:
            item = QtWidgets.QTreeWidgetItem()
            title = query.split('\n')[0]
            item.setText(0, title)
            default_item.addChild(item)
            self.default_sqls.append({'index': self.ui.treeWidget.indexFromItem(item), 'sql': query})

        # Add user stored queries
        sql = "select title, description, grouper, ssql from stored_sql order by grouper, title"
        cur.execute(sql)
        res = cur.fetchall()
        if not res:
            return
        ssql_item = QtWidgets.QTreeWidgetItem()
        ssql_item.setText(0, _("Saved Queries"))
        self.ui.treeWidget.addTopLevelItem(ssql_item)
        for r in res:
            item = QtWidgets.QTreeWidgetItem()
            item.setText(0, r[0])
            item.setToolTip(0, r[1])
            ssql_item.addChild(item)
            self.stored_sqls.append({'index': self.ui.treeWidget.indexFromItem(item), 'sql': r[3]})

    def tree_menu(self, position):
        """ Context menu for treewidget stored sql items. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        index = self.ui.treeWidget.currentIndex()
        delete_sql = None
        for i in self.stored_sqls:
            if i['index'] == index:
                delete_sql = menu.addAction(_("Delete stored sql"))
        action = menu.exec(self.ui.treeWidget.mapToGlobal(position))
        if action is not None and action == delete_sql:
            for i in range(len(self.stored_sqls)):
                if self.stored_sqls[i]['index'] == index:
                    title = self.ui.treeWidget.currentItem().text(0)
                    cur = self.app.conn.cursor()
                    cur.execute("delete from stored_sql where title=?", [title])
                    self.app.conn.commit()
                    del self.stored_sqls[i]
                    break
            self.get_schema_update_tree_widget()

    def sql_menu(self, position):
        """ Context menu to textedit_sql
         Cut Ctrl + X, Copy Ctrl + C, Paste Ctrl + V, Delete, Ctrl + A selectall. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_select_all = menu.addAction(_("Select all"))
        action_copy = menu.addAction(_("Copy"))
        action_paste = menu.addAction(_("Paste"))
        action_delete = menu.addAction(_("Delete"))
        action_select_all_from = menu.addAction("SELECT * FROM ")
        action_save_query = None
        if len(self.ui.textEdit_sql.toPlainText()) > 2:
            action_save_query = menu.addAction(_("Save query"))
        action = menu.exec(self.ui.textEdit_sql.mapToGlobal(position))
        cursor = self.ui.textEdit_sql.textCursor()
        if action is None:
            return
        if action == action_delete:
            text_ = cursor.selectedText()
            if text_ is None or text_ == "":
                return
            start = cursor.position()
            end = cursor.anchor()
            if start > end:
                tmp = end
                end = start
                start = tmp
            start_text = self.ui.textEdit_sql.toPlainText()[0:start]
            end_text = self.ui.textEdit_sql.toPlainText()[end:len(self.ui.textEdit_sql.toPlainText())]
            self.ui.textEdit_sql.setText(start_text + end_text)
        if action == action_paste:
            clipboard = QtWidgets.QApplication.clipboard()
            text_ = clipboard.text()
            cursor.insertText(text_)
        if action == action_copy:
            text_ = cursor.selectedText()
            text_ = str(text_)
            clipboard = QtWidgets.QApplication.clipboard()
            clipboard.setText(text_)
        if action == action_select_all:
            cursor.setPosition(0)
            cursor.setPosition(len(self.ui.textEdit_sql.toPlainText()),
                               QtGui.QTextCursor.MoveMode.KeepAnchor)
            self.ui.textEdit_sql.setTextCursor(cursor)
        if action == action_select_all_from:
            cursor.insertText("SELECT * FROM ")
        if action == action_save_query:
            self.save_query()

    def save_query(self):
        """ Save query in stored_sql table.
        The grouper is not really used, apart from ordering the queries. """

        ssql = self.ui.textEdit_sql.toPlainText()
        ui_save = DialogSaveSql(self.app)
        ui_save.ui.label.hide()
        ui_save.ui.lineEdit_group.hide()
        ui_save.exec()
        title = ui_save.name
        if title == "":
            msg = _("The query must have a name")
            Message(self.app, _("Cannot save"), msg).exec()
            return
        grouper = ui_save.grouper
        description = ui_save.description
        cur = self.app.conn.cursor()
        sql = "insert into stored_sql (title, description, grouper, ssql) values (?,?,?,?)"
        try:
            cur.execute(sql, [title, description, grouper, ssql])
            self.app.conn.commit()
        except Exception as e:
            Message(self.app, _("Cannot save"), str(e)).exec()
        self.get_schema_update_tree_widget()

    # Start of table results context menu section
    def table_menu(self, position):
        """ Context menu for table_results. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        try:
            self.row = self.ui.tableWidget_results.currentRow()
            self.col = self.ui.tableWidget_results.currentColumn()
            self.cell_value = str(self.ui.tableWidget_results.item(self.row, self.col).text())
        except AttributeError as e:
            logger.warning("No table for table menu: " + str(e))
            return

        action_show_all_rows = menu.addAction(_("Clear filter"))
        action_show_all_rows.triggered.connect(self.show_all_rows)
        action_filter_on_cell_value = menu.addAction(_("Filter equals: ") + str(self.cell_value))
        action_filter_on_cell_value.triggered.connect(self.filter_cell_value)
        action_filter_on_text_like = menu.addAction(_("Filter on text like"))
        action_filter_on_text_like.triggered.connect(self.filter_text_like)
        action_filter_on_text_starts_with = menu.addAction(_("Filter on text starts with"))
        action_filter_on_text_starts_with.triggered.connect(self.filter_text_starts_with)
        action_sort_ascending = menu.addAction(_("Sort ascending"))
        action_sort_ascending.triggered.connect(self.sort_ascending)
        action_sort_descending = menu.addAction(_("Sort descending"))
        action_sort_descending.triggered.connect(self.sort_descending)
        action = menu.exec(self.ui.tableWidget_results.mapToGlobal(position))

    def sort_ascending(self):
        """ Sort rows on selected column in ascending order. """

        self.ui.tableWidget_results.sortItems(self.col, QtCore.Qt.SortOrder.AscendingOrder)
        self.ui.label.setText(str(len(self.file_data) - 1) + _(" rows [") + self.file_data[0][self.col] + _(" asc]"))

    def sort_descending(self):
        """ Sort rows on selected column in descending order. """

        self.ui.tableWidget_results.sortItems(self.col, QtCore.Qt.SortOrder.DescendingOrder)
        self.ui.label.setText(str(len(self.file_data) - 1) + _(" rows [") + self.file_data[0][self.col] + _(" desc]"))

    def filter_text_like(self):
        """ Hide rows where cells in the column do not contain the text fragment. """

        text_, ok = QtWidgets.QInputDialog.getText(self, _("Text filter"), _("Text contains:"),
                                                   QtWidgets.QLineEdit.EchoMode.Normal, str(self.cell_value))
        if ok and text_ != '':
            for r in range(0, self.ui.tableWidget_results.rowCount()):
                if self.ui.tableWidget_results.item(r, self.col).text().find(text_) == -1:
                    self.ui.tableWidget_results.setRowHidden(r, True)
        self.ui.label.setText(str(len(self.file_data) - 1) + _(" rows [filtered]"))
        self.queryFilters += "\n" + self.ui.tableWidget_results.horizontalHeaderItem(self.col).text() + \
                             " like: " + text_
        self.ui.label.setToolTip(self.queryTime + self.queryFilters)

    def filter_text_starts_with(self):
        """ Hide rows where cells in the column do not contain the text start fragment. """

        text_, ok = QtWidgets.QInputDialog.getText(self, _("Text filter"), _("Text contains:"),
                                                   QtWidgets.QLineEdit.EchoMode.Normal, str(self.cell_value))
        if ok and text_ != '':
            for r in range(0, self.ui.tableWidget_results.rowCount()):
                if self.ui.tableWidget_results.item(r, self.col).text().startswith(text_) is False:
                    self.ui.tableWidget_results.setRowHidden(r, True)
        self.ui.label.setText(str(len(self.file_data) - 1) + _(" rows [filtered]"))
        self.ui.label.setToolTip(self.queryTime)
        self.queryFilters += "\n" + self.ui.tableWidget_results.horizontalHeaderItem(self.col).text() + \
                             _(" starts with: ") + text_
        self.ui.label.setToolTip(self.queryTime + self.queryFilters)

    def filter_cell_value(self):
        """ Hide rows that do not have the selected cell value. """

        for r in range(0, self.ui.tableWidget_results.rowCount()):
            if self.ui.tableWidget_results.item(r, self.col).text() != self.cell_value:
                self.ui.tableWidget_results.setRowHidden(r, True)
        self.ui.label.setText(str(len(self.file_data) - 1) + _(" rows [filtered]"))
        self.queryFilters += f"\n{self.ui.tableWidget_results.horizontalHeaderItem(self.col).text()}" + \
                             f" = {self.cell_value}"
        self.ui.label.setToolTip(self.queryTime + self.queryFilters)

    def show_all_rows(self):
        """ Remove all hidden rows. """

        for r in range(0, self.ui.tableWidget_results.rowCount()):
            self.ui.tableWidget_results.setRowHidden(r, False)
        self.ui.label.setText(str(len(self.file_data) - 1) + _(" rows"))
        self.queryFilters = ""
        self.ui.label.setToolTip(self.queryTime + self.queryFilters)


class NewTableWidget(QtWidgets.QTableWidget):
    """ This extends the table widget by adding a context menu and associated actions. """

    row = None
    col = None
    cell_value = None

    def __init__(self, parent=None):
        super(NewTableWidget, self).__init__(parent)

    def contextMenuEvent(self, event):

        menu = QtWidgets.QMenu(self)
        try:
            self.row = self.currentRow()
            self.col = self.currentColumn()
            self.cell_value = str(self.item(self.row, self.col).text())
        except AttributeError as e:
            logger.warning(f"No table for menu: {e}")
            return

        action_show_all_rows = menu.addAction(_("Clear filter"))
        action_show_all_rows.triggered.connect(self.show_all_rows)
        action_filter_on_cell_value = menu.addAction(_("Filter equals: ") + str(self.cell_value))
        action_filter_on_cell_value.triggered.connect(self.filter_cell_value)
        action_filter_on_text_like = menu.addAction(_("Filter on text like"))
        action_filter_on_text_like.triggered.connect(self.filter_text_like)
        action_filter_on_text_starts_with = menu.addAction(_("Filter on text starts with"))
        action_filter_on_text_starts_with.triggered.connect(self.filter_text_starts_with)
        action_sort_ascending = menu.addAction(_("Sort ascending"))
        action_sort_ascending.triggered.connect(self.sort_ascending)
        action_sort_descending = menu.addAction(_("Sort descending"))
        action_sort_descending.triggered.connect(self.sort_descending)
        menu.exec(event.globalPos())

    def sort_ascending(self):
        """ Sort rows on selected column in ascending order. """

        self.sortItems(self.col, QtCore.Qt.SortOrder.AscendingOrder)

    def sort_descending(self):
        """ Sort rows on selected column in descending order. """

        self.sortItems(self.col, QtCore.Qt.SortOrder.DescendingOrder)

    def filter_text_like(self):
        """ Hide rows where cells in the column do not contain the text fragment. """

        text_, ok = QtWidgets.QInputDialog.getText(self, _("Text filter"), _("Text contains:"),
                                                   QtWidgets.QLineEdit.EchoMode.Normal, str(self.cell_value))
        if ok and text_ != '':
            for r in range(0, self.rowCount()):
                if str(self.item(r, self.col).text()).find(text_) == -1:
                    self.setRowHidden(r, True)

    def filter_text_starts_with(self):
        """ Hide rows where cells in the column do not contain the text start fragment. """

        text_, ok = QtWidgets.QInputDialog.getText(self, _("Text filter"), _("Text contains:"),
                                                   QtWidgets.QLineEdit.EchoMode.Normal, str(self.cell_value))
        if ok and text_ != '':
            for r in range(0, self.rowCount()):
                if str(self.item(r, self.col).text()).startswith(text_) is False:
                    self.setRowHidden(r, True)

    def filter_cell_value(self):
        """ Hide rows that do not have the selected cell value. """

        for r in range(0, self.rowCount()):
            if str(self.item(r, self.col).text()) != self.cell_value:
                self.setRowHidden(r, True)

    def show_all_rows(self):
        """ Remove all hidden rows. """

        for r in range(0, self.rowCount()):
            self.setRowHidden(r, False)


class TableWidgetItem(QtWidgets.QTableWidgetItem):
    """ A sorting method that works. From:
        http://www.tagwith.com/question_868979_sort-string-column-in-pyqtqtablewidget-based-on-non-string-value
        With some modification for unicode and numerics """

    def __init__(self, value):
        super(TableWidgetItem, self).__init__(value)

    def __lt__(self, other):
        """ Not sure about the if isinstance statement. """

        if isinstance(other, TableWidgetItem):
            try:
                self_value = float(str(self.data(QtCore.Qt.ItemDataRole.EditRole)))
                other_value = float(str(other.data(QtCore.Qt.ItemDataRole.EditRole)))
                return self_value < other_value
            except ValueError:
                self_value = str(self.data(QtCore.Qt.ItemDataRole.EditRole))
                other_value = str(other.data(QtCore.Qt.ItemDataRole.EditRole))
                return self_value < other_value
        else:
            return QtWidgets.QTableWidgetItem.__lt__(self, other)


# Default queries
DEFAULT_SQL = ["-- CASE TEXT\nselect cases.name ,  substr(source.fulltext, case_text.pos0, case_text.pos1 -  case_text.pos0 ) as casetext \
from cases join case_text on  cases.caseid = case_text.caseid join source on source.id= case_text.fid \
where case_text.pos1 >0",
"-- SELECT CODED CASE TEXT BY CODE AND CASE ATTRIBUTE WHERE THE CASES ARE SECTIONS WITHIN THE SAME TEXT DOCUMENT \n\
SELECT cases.name as 'case name', attribute.name as 'Case attribute', attribute.value,\n\
source.name as 'text file name', code_name.name as 'code name', code_text.pos0, code_text.pos1, seltext \n\
from \n\
case_text join attribute on attribute.id=case_text.caseid join source on source.id = case_text.fid \
join code_text on code_text.fid=source.id join code_name on code_name.cid = code_text.cid \
join cases on cases.caseid=case_text.caseid \
where \n\
code_name.name = 'code1' -- change to code name of interest \n\
and attribute.name='District' -- change to case attribute name of interest \n\
and attribute.value='north'  -- change to case attribute value of interest \n\
and case_text.pos0 <= code_text.pos0 \n\
and case_text.pos1 >= code_text.pos1 ",
    '-- CATEGORY, CODES, FILE NAME, CODED TEXT\n\
select  code_cat.name as "category", code_name.name as "codename", source.name as "filename", code_text.seltext \n\
from code_name join  code_text on code_name.cid = code_text.cid\n\
join source on source.id=code_text.fid\n\
left join code_cat on code_name.catid=code_cat.catid\n\
order by code_cat.name, code_name.name asc\n\
-- UNCOMMENT "--" LINES BELOW TO GET DETAILS FOR A CODE AND OR A FILE\n\
-- AND codename="CODENAME" -- FILL CODENAME WITH CODE NAME\n\
-- AND source.name="FILENAME" -- FILL FILENAME WITH FILE NAME',
               '-- GET_CODING_TABLE\n-- Implementation of RQDA function\n\
select code_text.cid, code_text.fid, code_name.name as "codename", \
source.name as "filename", code_text.pos1 - code_text.pos0 as "CodingLength",\
code_text.pos0 as "index1", code_text.pos1 as "index2" \
from code_text join code_name on code_text.cid = code_name.cid \
join source on code_text.fid = source.id',

               '-- CODED TEXT WITH EACH CASE\nselect code_name.name as codename, cases.name as casename,\
 code_text.pos0, code_text.pos1, code_text.fid, seltext as "coded text", code_text.owner \
 from code_text join code_name on code_name.cid = code_text.cid \
join (case_text join cases on cases.caseid = case_text.caseid) on code_text.fid = case_text.fid \
where \n\
-- code_name.cid in ( code_ids ) -- provide your code ids \n\
-- and case_text.caseid in ( case_ids ) -- provide your case ids \n\
-- and \n\
(code_text.pos0 >= case_text.pos0 and code_text.pos1 <= case_text.pos1)',
               '-- CODED TEXT WITH EACH CASE, LIMITED BY CASE ATTRUBIUTES - EXAMPLE SQL\n\
               select distinct code_name.name as codename, cases.name as casename, code_text.pos0, code_text.pos1,\n\
 code_text.fid, seltext as "coded text", code_text.owner \n\
from code_text join code_name on code_name.cid = code_text.cid \n\
join (case_text join cases on cases.caseid = case_text.caseid) on code_text.fid = case_text.fid\n\
join attribute on  attribute.id =  case_text.caseid \n\
join attribute as attribute2 on  attribute2.id =  case_text.caseid \n\
 where \n\
 attribute.attr_type ="case" and attribute.name = "Gender" and attribute.value = "male" and -- first attribute\n\
 attribute2.attr_type ="case" and attribute2.name = "Age" and cast(attribute2.value as integer) < 30 and --second attribute \n\
-- code_name.name in ( "Aggression", "Increased workload") and -- uncomment and put code names in list here to filter by code names \n\
(code_text.pos0 >= case_text.pos0 and code_text.pos1 <= case_text.pos1)',

               '-- ALL OR SELECTED ANNOTATIONS\nselect annotation.anid as "Annotation ID" , annotation.memo as "Annotation text", \n\
annotation.fid as "File ID" , source.name as "File name", annotation.pos0 as "Start position", \n\
annotation.pos1 as "End position", annotation.owner as "Coder name", annotation.date as "Date" from annotation \n\
left join source on source.id = annotation.fid \n\
-- DISPLAY ANNOTATIONS FOR SELECTED FILE, UNCOMMENT THE FOLLOWING: \n\
-- AND source.name = "FILE NAME"',

               '-- ALL OR SELECTED CODINGS MEMOS\nselect code_text.memo as "Coding memo", code_text.cid as "Code ID", code_name.name as "Code name", \n\
code_text.fid as "File ID", source.name as "File name", code_text.owner as "Coder name", code_text.date as "Date", \n\
code_text.important as "Important(yes=1, no=0)" from source left join code_text on code_text.fid = source.id \n\
left join code_name on code_name.cid = code_text.cid where code_text.memo != "" \n\
-- TO DISPLAY CODING FOR SELECTED CODE OR FILE, UNCOMMENT THE FOLLOWING:\n\
-- AND code_name.name = "CODE NAME"  -- TO SELECT SPECIFIC CODE\n\
-- AND source.name = "FILE NAME" -- TO SELECT SPECIFIC FILE',
               '-- ALL OR SELECTED CATEGORY MEMOS\n\
select code_cat.memo as "Category memo", code_cat.catid as "Cat ID", code_cat.name as "Name", \n\
code_cat.owner as "Coder name", code_cat.date as "Date" from  code_cat \n\n\
-- TO FILTER UNCOMMENT THE FOLLOWING:\n\
-- MUST HAVE "where" TO START THIS SECTION\n\
-- where \n\
--- code_cat.name = "abilities"  -- TO SELECT SPECIFIC CATEGORY\n\
-- AND code_cat.memo  like "%filter words for memo text%" -- TO SELECT MEMOS CONTAINING SPECIFIED TEXT',

'-- FILES THAT ARE NOT CODED with code id 1\n\
select source.name from source where source.id not in\n\
(select code_text.fid from code_text where code_text.cid=1\n\
union select code_av.id from code_av where code_av.cid=1\n\
union select code_image.id from code_image where code_image.cid=1)',

'-- CODES NOT USED IN A FILE. Example using file id 1 presuming a text file.\n\
select code_name.name from code_name where code_name.cid not in\n\
-- Uncomment the appropriate line below for another file type if needed\n\
(select code_text.cid from code_text where code_text.fid=1)  -- comment out for another file type\n\
-- (select code_av.cid from code_av where code_av.id=1) -- uncomment for av files\n\
-- (select code_image.cid from code_image where code_image.id=1) -- uncomment for image files',

'-- FILTERED CONTAINER CODES\n\
-- This is useful when asking questions like:\n\
-- What does speaker 1 say? What do male respondents say?\n\
-- The containing codes can be quite long (speakerID, sex) but the content of interest can be quite short.\n\
-- All the codes applied are in a single column and appends to the bottom a list of all the codes used in the output.\n\
-- That list of codes then needs to be turned into column headers.\n\
-- Then in Excel you can do conditional values like this:\n\
-- =IF(ISNUMBER(SEARCH("," & E$1 & ",", "," & $D2 & ",")), 1, 0)\n\
-- Where E1 has the code of interest and column D has the comma delimited list of codes applied to that segment.\n\n\
WITH container_codes AS (\n\
-- EDIT THIS LIST ONLY: names of codes you want as CONTAINERS (AND logic)\n\
SELECT "CONTAINER_CODE_A" AS name -- change CONTAINER_CODE_A to your container code name\n\
UNION ALL\n\
SELECT "CONTAINER_CODE_B" -- add or remove lines as needed\n\
),\n\n\
segments AS (SELECT ct.ctid,ct.cid,ct.fid,ct.owner,ct.pos0,ct.pos1,ct.seltext FROM code_text ct),\n\
segments_with_names AS (SELECT s.*,cn.name AS code_name FROM segments s JOIN code_name cn ON cn.cid = s.cid),\n\n\
-- All coded segments that have a container code\n\
container_segments AS (SELECT swn.*FROM segments_with_names swn JOIN container_codes cc ON swn.code_name = cc.name),\n\
-- All containment relationships: inner segments inside container segments\n\
contained_raw AS (\n\
SELECT inner.ctid AS segment_ctid,inner.fid,inner.owner,inner.pos0,inner.pos1,inner.seltext AS segment_text,\n\
inner.code_name AS segment_code,outer.code_name AS container_code FROM segments_with_names AS inner\n\
JOIN container_segments AS outer\n\
ON inner.fid = outer.fid AND inner.owner = outer.owner AND inner.pos0 >= outer.pos0 AND inner.pos1 <= outer.pos1),\n\
\n\
-- Keep only segments that are inside ALL container codes (AND)\n\
segments_in_all_containers AS (\n\
SELECT segment_ctid, fid,owner,pos0,pos1,segment_text FROM contained_raw\n\
GROUP BY segment_ctid, fid, owner, pos0, pos1, segment_text\n\
HAVING COUNT(DISTINCT container_code) = (SELECT COUNT(*) FROM container_codes)\n\
),\n\n\
-- For those focal segments, attach ALL codes that occur within them\n\
segments_all_codes AS (\n\
SELECT s.segment_ctid,s.segment_text,sc.code_name FROM segments_in_all_containers s\n\
JOIN segments_with_names sc ON sc.fid = s.fid AND sc.owner = s.owner AND sc.pos0 >= s.pos0 AND sc.pos1 <= s.pos1\n\
)\n\n\
-- FINAL OUTPUT:\n\
-- (1) one row per segment_ctid, with comma-joined code list\n\
-- (2) one extra row ALL_SEGMENTS listing all codes seen\n\
SELECT segment_ctid,segment_text, GROUP_CONCAT(DISTINCT code_name) AS codes_applied FROM segments_all_codes\n\
GROUP BY segment_ctid, segment_text\n\n\
UNION ALL\n\n\
SELECT "ALL_SEGMENTS" AS segment_ctid, "(all segments)" AS segment_text,\n\
GROUP_CONCAT(DISTINCT code_name) AS codes_applied\n\
FROM segments_all_codes\n\
ORDER BY segment_ctid;',

'-- THREE OR MORE EXACTLY OVERLAPPING CODINGS IN A SELECTED FILE\n\
-- Everywhere use the same file id, otherwise results will be incorrect\n\
-- IDs for codes and files can be found by selecting Show IDs in Settings then go to Code Text, and Manage Files\n\n\
SELECT  code_name.name AS "codename", code_text.cid,  code_text.fid, source.name, code_text.pos0, code_text.pos1, \n\
code_text.seltext, code_text.memo AS "coding memo" \n\
FROM  code_name JOIN  code_text ON code_name.cid = code_text.cid JOIN source ON source.id = code_text.fid \n\
WHERE code_text.cid IN (1,2,3) -- INSERT IDs OF CODES OF INTEREST, (3 or more)\n\
AND  fid = 1 \n\
AND pos0 IN (SELECT  pos0 FROM code_text WHERE cid=1 AND fid=1) -- Put ID OF THE FIRST OBLIGATORY CODE \n\
AND pos1 IN (SELECT  pos1 FROM code_text WHERE cid=1 AND fid=1) \n\
AND pos0 IN (SELECT  pos0 FROM code_text WHERE cid=2 AND fid=1) -- PUT ID OF THE SECOND OBLIGATORY CODE \n\
AND pos1 IN (SELECT  pos1 FROM code_text WHERE cid=2 AND fid=1) \n\
AND pos0 IN (SELECT  pos0 FROM code_text WHERE cid=3 AND fid=1) -- PUT ID OF THE THIRD OBLIGATORY CODE \n\
AND pos1 IN (SELECT  pos1 FROM code_text WHERE cid=3 AND fid=1) \n\
-- Add more codes if needed by copying the above two "AND" statements and put cid=THE EXTRA CODE ID for each additional code\n\n\
-- Remove "--" from the statements below and add CODE IDS to exclude some overlapping codes\n\
-- e.g. Four different codes are used for some of these overlaps, \n\
-- and you want to show only the 3 selected above, providing the 4th code below is not overlapping\n\
-- AND NOT pos0 IN (SELECT pos0 FROM code_text WHERE cid=CODEID AND fid=1) \n\
-- AND NOT pos1 IN (SELECT pos1 FROM code_text WHERE cid=CODEID AND fid=1) \n\n\
-- Remove "--" before these statements, add CODE IDs of codes\n\
-- This is used to display the 3 overlapping  codings above, \n\
-- but only if one OR two more additional overlapping code are also applied to this same segment\n\
--AND (\n\
--(pos0 IN (SELECT pos0 FROM code_text WHERE cid=CODEID AND fid=1) AND pos1 IN (SELECT pos1 FROM code_text WHERE cid=CODEID AND fid=1))\n\
--OR (pos0 IN (SELECT pos0 FROM code_text WHERE cid=CODEID AND fid=1) AND pos1 IN (SELECT pos1 FROM code_text WHERE cid=CODEID AND fid=1))\n\
--)\n\n\
-- Below remove "--" to hide duplicate rows of overlapping codings\n\
-- GROUP BY pos0'
]
