# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com
"""

from PyQt6 import QtWidgets, QtCore
from PyQt6.QtCore import Qt

import csv
import datetime
import logging
from openpyxl import load_workbook
import os
import re
from shutil import copyfile
import sqlite3
# import sys
# import traceback

from .GUI.ui_dialog_import import Ui_Dialog_Import
from .helpers import Message

path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogImportSurvey(QtWidgets.QDialog):
    """ Import case and file attributes from a csv file. EXTEND LATER
    The first row must contain a header row of the attribute names.
    The first column must contain unique identifiers for each response (the cases)
    this then allows automatic assignment of attributes to each case
    Each column can be categorised as an attribute OR as qualitative.
    Text from each qualitative colums are treated as individual files and loaded into the
    source table.
    Some GUI elements cannot be changed to another language:
    Quote format: NONE, MINIMAL, ALL
    Field type: character, numeric qualitative
    """

    app = None
    fields = []
    fields_type = []
    delimiter = ""
    filepath = ""
    headerIndex = 0  # Table column index for header context menu actions
    data = []  # Obtained from csv file
    preexisting_fields = []  # attribute names already in database
    parent_textEdit = None
    success = False  # Ability to load file and has individual ids in first column

    def __init__(self, app, parent_text_edit):
        """ Need to comment out the connection accept signal line in ui_Dialog_Import.py.
         Otherwise, get a double-up of accept signals. """

        self.app = app
        self.parent_textEdit = parent_text_edit
        self.delimiter = ","
        self.fields = []
        self.filepath = ""
        self.success = True

        # Set up the user interface from Designer.
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_Import()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        self.ui.lineEdit_delimiter.setText(self.delimiter)
        self.ui.lineEdit_delimiter.textChanged.connect(self.options_changed)
        self.ui.comboBox_quote.currentIndexChanged.connect(self.options_changed)
        self.ui.tableWidget.setHorizontalHeaderLabels([""])
        self.ui.tableWidget.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget.horizontalHeader().customContextMenuRequested.connect(self.table_menu)
        self.ui.tableWidget.setTabKeyNavigation(False)

        cur = self.app.conn.cursor()
        cur.execute("select name from attribute_type where caseOrFile='case'")
        result = cur.fetchall()
        self.preexisting_fields = []
        for row in result:
            self.preexisting_fields.append({'name': row[0]})
        self.select_file()
        # print("FILE ", self.filepath)  # returns a List
        if self.filepath:
            self.prepare_fields()
            self.fill_table_widget()
        else:
            self.ui.groupBox.hide()
            self.ui.tableWidget.hide()
            self.ui.checkBox_collate.hide()
            self.ui.buttonBox.hide()
            self.ui.label_msg.setText(_("No survey selected."))
            self.ui.label_msg.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.close()

    def select_file(self):
        """ Select csv or Excel file """

        response = QtWidgets.QFileDialog.getOpenFileNames(None, _('Select survey file'),
                                                          self.app.settings['directory'], "(*.csv *.xlsx)",
                                                          options=QtWidgets.QFileDialog.Option.DontUseNativeDialog
                                                          )
        self.filepath = response[0]
        if not self.filepath:
            self.parent_textEdit.append(_("Survey not imported. Survey not a csv or xlsx file: "))
            self.success = False
            return
        self.filepath = self.filepath[0]  # A list of one name
        # Copy file into project folder
        name_split = os.path.split(self.filepath)[1]  # Tail
        filename = name_split[-1]
        destination = os.path.join(self.app.project_path, "documents", filename)
        copyfile(self.filepath, destination)

    def read_xlsx_file(self):
        """ Read the data from the xlsx file.
        Fill Class variables self.fields, self.data
        Called by: prepare_fields
        """

        self.data = []
        wb = load_workbook(filename=self.filepath)
        # To work with the first sheet (by name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        sheet = ws
        for value in sheet.iter_rows(values_only=True):
            # Some rows may be complete blank so ignore importation
            if (set(value)) != {None}:
                # Values are tuples, convert to list, and remove 'None' string
                row = [item if item else "" for item in value]
                self.data.append(row)
        # Get field names and replace blanks with a placeholder
        self.fields = []
        for i, field in enumerate(self.data[0]):
            if field != '':
                # Using str() method as f may be an Integer or Float
                self.fields.append(str(field))
            else:
                self.fields.append(f"Field_{i}")
        self.data = self.data[1:]
        # Widgets are not needed
        self.ui.lineEdit_delimiter.hide()
        self.ui.comboBox_quote.hide()
        self.ui.label_delimiter.hide()
        self.ui.label_quotefmt.hide()
        self.ui.label_information.hide()
        self.ui.groupBox.setMinimumSize(QtCore.QSize(0, 30))
        self.ui.groupBox.setMaximumSize(QtCore.QSize(16777215, 30))
        return True

    def read_csv_file(self):
        """ Read the data from the csv file.
         Fill Class variables self.fields, self.data
          Called by: prepare_fields, options_changed
        """

        self.data = []
        with open(self.filepath, 'r', newline='') as f:
            delimiter_ = self.ui.lineEdit_delimiter.text()
            if delimiter_ == '':
                msg = _("A column delimiter has not been set.")
                Message(self.app, _("Warning"), msg, "warning").exec()
                return False
            if delimiter_ in ('ta', 'tab'):
                delimiter_ = "\t"
            # The English text is in the GUI - do not translate with qt linguist
            quoting_ = csv.QUOTE_MINIMAL
            quote_type = self.ui.comboBox_quote.currentText()
            if quote_type == "NONE":
                quoting_ = csv.QUOTE_NONE
            if quote_type == "ALL":
                quoting_ = csv.QUOTE_ALL
            reader = csv.reader(f, delimiter=delimiter_, quoting=quoting_)
            try:
                for row in reader:
                    self.data.append(row)
            except csv.Error as err:
                logger.error(('file %s, line %d: %s' % (self.filepath, reader.line_num, err)))
                self.parent_textEdit.append(f"Row error: {reader.line_num}  {err}")
                return False
        # Get field names and replace blacks with a placeholder
        self.fields = []
        for i, field in enumerate(self.data[0]):
            if field != '':
                self.fields.append(str(field))
            else:
                self.fields.append(f"Field_{i}")
        self.data = self.data[1:]
        return True

    def prepare_fields(self):
        """ Check for a .csv or .xlsx extension.
        Determine number of fields. Load the data.
        Also called when import options changed. """

        self.fields = []
        self.fields_type = []
        self.data = []

        if self.filepath[-4:].lower() == ".csv":
            success = self.read_csv_file()
            if not success:
                self.parent_textEdit.append(_("Survey not imported.") + self.filepath)
                return False
        else:
            success = self.read_xlsx_file()
            if not success:
                self.parent_textEdit.append(_("Survey not imported.") + self.filepath)
                return False
        self.setWindowTitle(_("Importing from: ") + self.filepath)

        # Clean field names
        removes = r"!@#$%^&*()-+=[]{}\\|;:,.<>/?~`"
        for i in range(0, len(self.fields)):
            self.fields[i] = self.fields[i].replace('\t', '')
            self.fields[i] = self.fields[i].replace(r'\xa0', '')
            for r in removes:
                self.fields[i] = self.fields[i].replace(r, '')
            if self.fields[i] in self.preexisting_fields:
                self.fields[i] += "_DUPLICATED"
        # Default field type is character
        self.fields_type = ["character"] * len(self.fields)
        # Determine if field type is numeric
        for field in range(0, len(self.fields)):
            numeric = True
            for row in range(0, len(self.data)):
                try:
                    float(self.data[row][field])
                except (ValueError, IndexError):
                    # IndexError, Might be malformed csv, so presume numeric is False anyway
                    numeric = False
            if numeric:
                self.fields_type[field] = "numeric"

        # Estimate if field type is qualitative, based on at least 20 different character entries
        for field in range(1, len(self.fields)):
            if self.fields_type[field] == 'character':
                set_of_values = set()
                for row in range(0, len(self.data)):
                    value = ""
                    try:
                        value = self.data[row][field]
                    except IndexError as e:
                        msg = f"IndexError: [row] {row}   [field] {field}"
                        msg += f"\nlen(self.data) {len(self.data)}\n{e}"
                        logger.debug(msg)
                    set_of_values.add(value)
                if len(set_of_values) > 19:
                    try:
                        self.fields_type[field] = "qualitative"
                    except IndexError:
                        # Occurs if the delimiter or quoting incorrect. Or if the rows fields length does not match
                        pass
        # Check first column has unique identifiers
        ids = []
        for row in self.data:
            try:
                ids.append(row[0])
            except IndexError:
                # Occurs with csv import if wrong quote type selected
                ids.append("")
        ids_set = set(ids)
        if len(ids) > len(ids_set):
            fail_msg = _("There are duplicated identifiers in the first column.\nFile not imported")
            self.parent_textEdit.append(f"{self.filepath} {fail_msg}")
            return False
        msg = f"{_('Survey file:')} {self.filepath}\n"
        msg += f"{_('Fields:')} {len(self.fields)}. {_('Rows:')} {len(self.data)}"
        logger.info(msg)
        self.parent_textEdit.append(msg)
        return True

    def accept(self):
        """ Check the table details are valid and import the data into a new table or
        append to an existing table. """

        if not self.success:
            super(DialogImportSurvey, self).accept()
            return
        # Check for duplicate field names
        if len(self.fields) != len(set(self.fields)):
            msg = "There are duplicate attribute names."
            Message(self.app, _("Attribute name error"), msg, "warning").exec()
            logger.info(_("Survey Not Imported. Attribute duplicate name error: ") + msg)
            self.parent_textEdit.append(msg)
            self.fields = []
            return

        # Check for appropriate quote format
        row_length_error = ""
        for i, row in enumerate(self.data):
            if len(row) != len(self.fields):
                row_length_error += f"\nError row {i + 1} length does not match fields length"
        if row_length_error != "":
            msg = _("Number of fields does not match header\nPossible wrong quote format")
            msg += row_length_error
            logger.error(_("Survey not loaded: ") + msg)
            Message(self.app, _("Survey not loaded"), msg, "warning").exec()
            return
        self.insert_data()
        super(DialogImportSurvey, self).accept()

    def insert_data(self):
        """ Insert case, attributes, attribute values and qualitative text. """

        now_date = str(datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"))
        cur = self.app.conn.cursor()
        name_and_caseids = []
        for i, c in enumerate(self.data):
            try:
                self.ui.label_msg.setText(_("Inserting cases: " + str(i)))
                cur.execute("insert into cases (name,memo,owner,date) values(?,?,?,?)",
                            (c[0], "", self.app.settings['codername'], now_date))
                self.app.conn.commit()
                cur.execute("select last_insert_rowid()")
                name_and_caseids.append([c[0], cur.fetchone()[0]])
                QtWidgets.QApplication.processEvents()
            except sqlite3.IntegrityError as e:
                fail_msg = str(e) + _(
                    " - Duplicate case names, either in the file, or duplicates with existing cases in the project")
                logger.error(_("Survey not loaded: ") + fail_msg)
                Message(self.app, _('Survey not loaded'), fail_msg, "warning").exec()
                self.parent_textEdit.append(_("Survey not loaded: ") + fail_msg)
                return
        # Insert non-qualitative attribute types, except if they are already present
        sql = "select name from attribute_type where caseOrFile='case'"
        cur.execute(sql)
        result = cur.fetchall()
        existing_attr_names = [r[0] for r in result]
        sql = "insert into attribute_type (name,date,owner,memo, valueType, caseOrFile) values(?,?,?,?,?,?)"
        for col, name in enumerate(self.fields):
            if self.fields_type[col] != "qualitative" and col > 0:  # col==0 is the case identifier
                if name not in existing_attr_names:
                    logger.debug(f"{name} is not in case attribute_types. Adding.")
                    cur.execute(sql, (name, now_date, self.app.settings['codername'], "",
                                      self.fields_type[col], 'case'))
        self.app.conn.commit()

        # Look for pre-existing attributes that are not in the survey and insert blank value rows if present
        survey_field_names = []
        for col, field_name in enumerate(self.fields):
            if self.fields_type[col] != "qualitative" and col > 0:
                survey_field_names.append(field_name)
        for name in existing_attr_names:
            if name not in survey_field_names:
                for name_id in name_and_caseids:
                    sql = "insert into attribute (name, value, id, attr_type, date, owner) values (?,'',?,?,?,?)"
                    cur.execute(sql, (name, name_id[1], 'case', now_date, self.app.settings['codername']))
        self.app.conn.commit()

        # Insert non-qualitative values to each case using caseids
        sql = "insert into attribute (name, value, id, attr_type, date, owner) values (?,?,?,?,?,?)"
        for i, name_id in enumerate(name_and_caseids):
            self.ui.label_msg.setText(_("Inserting attributes to cases: ") + str(i))
            for val in self.data:
                if name_id[0] == val[0]:
                    for col in range(1, len(val)):
                        if self.fields_type[col] != "qualitative":
                            cur.execute(sql, (self.fields[col], val[col], name_id[1], 'case',
                                              now_date, self.app.settings['codername']))
            QtWidgets.QApplication.processEvents()
        self.app.conn.commit()

        # Create and insert qualitative codes from qualitative column names
        for field in range(1, len(self.fields)):  # column 0 is for identifiers
            # Create one text file combining each row, prefix [case identifier] to each row.
            if self.fields_type[field] == "qualitative" and self.fields[field] != "":
                try:
                    cur.execute("insert into code_name (name,memo,owner,date,catid,color) values(?,?,?,?,?,?)",
                                (self.fields[field], "", self.app.settings['codername'],
                                 datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"),
                                 None, "#B8B8B8"))
                    self.app.conn.commit()
                except Exception as e_:  # Codename might exist - sqlite.Integrityerror
                    print(e_)
                    logger.warning("Survey Insert code name from qual column " + str(e_))

        # Insert qualitative data into source table
        self.ui.label_msg.setText(_("Creating qualitative text file(s)"))
        source_sql = "insert into source(name,fulltext,memo,owner,date, mediapath) values(?,?,?,?,?, Null)"
        for field in range(1, len(self.fields)):  # column 0 is for identifiers
            # More robust method to get code cid, e.g. if appending to existing survey, and codes already present.
            cur.execute("select cid from code_name where name=?", [self.fields[field]])
            res_code_cid = cur.fetchone()  # Either [cid] or None

            case_text_list = []

            # Create one text file combining each row, prefix [case identifier] to each row.
            if self.fields_type[field] == "qualitative" and self.ui.checkBox_collate.isChecked():
                fulltext = ""
                for row in range(0, len(self.data)):
                    if self.data[row][field] != "":
                        fulltext += f"[{self.data[row][0]}] "
                        pos0 = len(fulltext) - 1
                        fulltext += f"{self.data[row][field]}\n\n"
                        pos1 = len(fulltext) - 2
                        case_text = [self.app.settings['codername'], now_date, "", pos0, pos1, name_and_caseids[row][1]]
                        case_text_list.append(case_text)
                # Add the current time to the file name to ensure uniqueness and to
                # Prevent sqlite Integrity Error. Do not use now_date which contains colons
                now = str(datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H-%M-%S"))
                qual_file_name = f"{self.fields[field]}_{now}"
                cur.execute(source_sql,
                            (f"{self.fields[field]}_{now}", fulltext, "", self.app.settings['codername'], now_date))
                self.app.conn.commit()
                cur.execute("select last_insert_rowid()")
                fid = cur.fetchone()[0]
                case_text_sql = "insert into case_text (owner, date, memo, pos0, pos1, caseid, fid) values(?,?,?,?,?,?,?)"
                for case_text in case_text_list:
                    case_text.append(fid)
                    cur.execute(case_text_sql, case_text)
                    # Insert code text for this qualitative column item
                    if res_code_cid:
                        try:
                            cur.execute("insert into code_text (cid,fid,seltext,pos0,pos1,owner,\
                                        memo,date, important) values(?,?,?,?,?,?,?,?,?)", (res_code_cid[0],
                                                                                           fid,
                                                                                           fulltext[case_text[3]:case_text[4]],
                                                                                           case_text[3],
                                                                                           case_text[4],
                                                                                           self.app.settings['codername'],
                                                                                           "",
                                                                                           now_date,
                                                                                           None))
                            self.app.conn.commit()
                        except Exception as e_:
                            print(e_)
                            logger.debug(e_)

                self.app.conn.commit()
                # Add doc to vectorstore
                if self.app.settings['ai_enable'] == 'True':
                    self.app.ai.sources_vectorstore.import_document(fid, qual_file_name, fulltext, update=True)

            # Create one text file per row, prefix [case identifier] to each row.
            if self.fields_type[field] == "qualitative" and not self.ui.checkBox_collate.isChecked():
                for row in range(0, len(self.data)):
                    qual_file_name = f"{self.data[row][0]}_{self.fields[field]}"
                    fulltext = f"{self.data[row][field]}"
                    cur.execute(source_sql,
                                (qual_file_name, fulltext, "", self.app.settings['codername'], now_date))
                    self.app.conn.commit()
                    cur.execute("select last_insert_rowid()")
                    fid = cur.fetchone()[0]
                    case_text_sql = "insert into case_text (owner, date, memo, pos0, pos1, caseid, fid) values(?,?,?,?,?,?,?)"
                    cur.execute(case_text_sql, [self.app.settings['codername'], now_date, "", 0, len(fulltext),
                                                name_and_caseids[row][1], fid])
                    self.app.conn.commit()
                    # Insert code text for this qualitative column
                    if res_code_cid:
                        try:
                            cur.execute("insert into code_text (cid,fid,seltext,pos0,pos1,owner,\
                                        memo,date, important) values(?,?,?,?,?,?,?,?,?)", (res_code_cid[0],
                                                                                           fid,
                                                                                           fulltext,
                                                                                           0,
                                                                                           len(fulltext),
                                                                                           self.app.settings['codername'],
                                                                                           "",
                                                                                           now_date,
                                                                                           None))
                            self.app.conn.commit()
                        except Exception as e_:
                            print(e_)
                            logger.debug(e_)

                    # Add doc to vectorstore
                    if self.app.settings['ai_enable'] == 'True':
                        self.app.ai.sources_vectorstore.import_document(fid, qual_file_name, fulltext, update=True)

        logger.info(_("Survey imported"))
        self.parent_textEdit.append(_("Survey imported."))
        Message(self.app, _("Survey imported"), _("Survey imported")).exec()
        self.app.delete_backup = False

    def options_changed(self):
        """ When import options are changed fill the table.
         Import options are: delimiter
         The delimiter can only be one character long """

        self.delimiter = str(self.ui.lineEdit_delimiter.text())
        if self.delimiter == "tb" or self.delimiter == "ta" or self.delimiter == "tab":
            self.delimiter = "\t"
        if len(self.delimiter) > 1 and self.delimiter != "\t":
            self.ui.lineEdit_delimiter.setText(self.delimiter[0:1])
            self.delimiter = self.delimiter[0:1]
        self.read_csv_file()
        self.fill_table_widget()

    def fill_table_widget(self):
        """ Fill table widget with data.
        Warn if an incorrect number of fields in the row. """

        num_fields_in_row_error = False
        num_rows = self.ui.tableWidget.rowCount()
        for row in range(0, num_rows):
            self.ui.tableWidget.removeRow(0)
        self.ui.tableWidget.setColumnCount(len(self.fields))
        for c, field in enumerate(self.fields):
            item = QtWidgets.QTableWidgetItem(f"{field}\n{self.fields_type[c]}\n")
            msg = "Right click to change column name or to change from character to qualitative"
            item.setToolTip(_(msg))
            self.ui.tableWidget.setHorizontalHeaderItem(c, item)
        self.ui.tableWidget.setRowCount(len(self.data))
        for row in range(0, len(self.data)):
            for col in range(0, len(self.fields)):
                value = "MISSING VALUE"
                try:
                    value = str(self.data[row][col])
                except IndexError:
                    num_fields_in_row_error = True
                item = QtWidgets.QTableWidgetItem(value)
                item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)  # Not editable
                self.ui.tableWidget.setItem(row, col, item)
        self.ui.tableWidget.resizeColumnsToContents()
        for col in range(0, self.ui.tableWidget.columnCount()):
            if self.ui.tableWidget.columnWidth(col) > 200:
                self.ui.tableWidget.setColumnWidth(col, 200)
        if num_fields_in_row_error and self.filepath[-4:].lower() == ".csv":
            msg = _("Number of fields in row error.") + "\n"
            msg += _("Use another quote format OR another delimiter")
            self.ui.label_msg.setText(msg)
            self.ui.comboBox_quote.setFocus(True)

    def table_menu(self, pos):
        """ Header context menu to change data types and set primary key(s) and change field names.
        The header index idea came from:
        param:
            pos: TableWidget position
        """

        self.headerIndex = self.ui.tableWidget.indexAt(pos)
        self.headerIndex = int(self.headerIndex.column())
        menu = QtWidgets.QMenu(self)
        action_change_fieldname = menu.addAction(_('Change fieldname'))
        action_change_fieldname.triggered.connect(self.change_fieldname)
        if self.fields_type[self.headerIndex] == "character" and self.headerIndex != 0:
            action_change_fieldname = menu.addAction(_('Change to Qualitative'))
            action_change_fieldname.triggered.connect(self.qualitative_field_type)
        if self.fields_type[self.headerIndex] in ('numeric', 'qualitative'):
            action_change_fieldname = menu.addAction(_('Change to Character'))
            action_change_fieldname.triggered.connect(self.character_field_type)
        menu.popup(self.ui.tableWidget.mapToGlobal(pos))

    # NOTE changes to field types are overwritten if quote type changed
    def qualitative_field_type(self):
        """ If the current field is listed as character, redefine it as qualitative.
        Qualitative data is stored in the source table in a generated text file. """

        self.fields_type[self.headerIndex] = 'qualitative'
        item_txt = f"{self.fields[self.headerIndex]}\n{self.fields_type[self.headerIndex]}\n"
        item = QtWidgets.QTableWidgetItem(item_txt)
        self.ui.tableWidget.setHorizontalHeaderItem(self.headerIndex, item)

    def character_field_type(self):
        """ If the current field is listed as numeric or qualitative, redefine it as character.
        """

        self.fields_type[self.headerIndex] = 'character'
        item_txt = f"{self.fields[self.headerIndex]}\n{self.fields_type[self.headerIndex]}\n"
        item = QtWidgets.QTableWidgetItem(item_txt)
        self.ui.tableWidget.setHorizontalHeaderItem(self.headerIndex, item)

    def change_fieldname(self):
        """ Change the fieldname. """

        fieldname, ok = QtWidgets.QInputDialog.getText(self, _("Change field name"), _("New name:"),
                                                       QtWidgets.QLineEdit.EchoMode.Normal,
                                                       self.fields[self.headerIndex])
        if not ok:
            return
        # Check for valid values
        if re.match(r"^[a-zA-Z_\s][a-zA-Z0-9_\s]*$", fieldname) is None or fieldname == "":
            msg = _("Name must contain only letters and numbers or '_' and must not start with a number")
            Message(self.app, _("Field name invalid"), msg, "warning").exec()
            return
        if fieldname in self.preexisting_fields or fieldname in self.fields:
            msg = fieldname + _(" Already in use")
            Message(self.app, _("Field name invalid."), msg, "warning").exec()
            return
        self.fields[self.headerIndex] = fieldname
        item_txt = f"{self.fields[self.headerIndex]}\n{self.fields_type[self.headerIndex]}"
        item = QtWidgets.QTableWidgetItem(item_txt)
        self.ui.tableWidget.setHorizontalHeaderItem(self.headerIndex, item)
