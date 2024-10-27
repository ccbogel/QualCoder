# -*- coding: utf-8 -*-

"""
Copyright (c) 2024 Colin Curtain

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com/
"""

from copy import deepcopy
import logging
import openpyxl
import os

from PyQt6 import QtGui, QtWidgets, QtCore
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QBrush

from .color_selector import TextColor
from .GUI.base64_helper import *
from .GUI.ui_report_codes_by_segments import Ui_DialogSegmentCodings
from .helpers import Message


path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogCodesBySegments(QtWidgets.QDialog):
    """ Get reports on coded text/images/audio/video using a range of variables:
        Files, Cases, Coders, text limiters, Attribute limiters.
        Export reports as plain text, ODT, html, xlsx or csv.

        Text context of a coded text portion is shown in the third splitter panel in a text edit.
        Case matrix is also shown in a qtablewidget in the third splitter pane.
        If a case matrix is displayed, the text-in-context method overrides it and replaces the matrix with the
        text in context.
    """

    app = None
    parent_textEdit = None
    code_names = []
    coders = [""]
    categories = []
    files = []
    cases = []
    results = []
    # Variables for search restrictions
    file_ids_string = ""
    case_ids_string = ""

    def __init__(self, app, parent_textedit):
        super(DialogCodesBySegments, self).__init__()
        self.app = app
        self.parent_textEdit = parent_textedit
        self.get_codes_categories_coders()
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_DialogSegmentCodings()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        font = f"font: {self.app.settings['fontsize']}pt "
        font += f'"{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        treefont = f'font: {self.app.settings["treefontsize"]}pt '
        treefont += f'"{self.app.settings["font"]}";'
        self.ui.treeWidget.setStyleSheet(treefont)
        self.ui.listWidget_files.setStyleSheet(treefont)
        self.ui.listWidget_files.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.ui.listWidget_cases.setStyleSheet(treefont)
        self.ui.listWidget_cases.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.ui.treeWidget.setSelectionMode(QtWidgets.QTreeWidget.SelectionMode.ExtendedSelection)
        self.ui.comboBox_coders.insertItems(0, self.coders)
        self.fill_tree()
        self.ui.pushButton_run_report.clicked.connect(self.search)
        pm = QtGui.QPixmap()
        pm.loadFromData(QtCore.QByteArray.fromBase64(play_icon), "png")
        self.ui.pushButton_run_report.setIcon(QtGui.QIcon(pm))
        pm = QtGui.QPixmap()
        pm.loadFromData(QtCore.QByteArray.fromBase64(doc_export_icon), "png")
        self.ui.pushButton_export_xlsx.setIcon(QtGui.QIcon(pm))
        self.ui.pushButton_export_xlsx.clicked.connect(self.export_xlsx_file)

        self.get_files_and_cases()
        self.ui.listWidget_files.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.listWidget_files.customContextMenuRequested.connect(self.listwidget_files_menu)
        self.ui.listWidget_cases.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.listWidget_cases.customContextMenuRequested.connect(self.listwidget_cases_menu)
        self.ui.treeWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.treeWidget.customContextMenuRequested.connect(self.treewidget_menu)
        self.ui.splitter.setSizes([200, 500])
        self.file_ids_string = ""
        self.case_ids_string = ""
        self.code_ids_string = ""
        self.code_columns = []
        self.results = []
        self.segment_rows = []
        self.horizontal_labels = []
        self.xlsx_results = []

    def get_files_and_cases(self):
        """ Get source files with additional details and fill files list widget.
        Get cases and fill case list widget
        Called from : init, manage_files.delete manage_files.delete_button_multiple_files
        """

        self.ui.listWidget_files.clear()
        self.files = self.app.get_text_filenames()
        item = QtWidgets.QListWidgetItem("")
        item.setToolTip(_("No file selection"))
        self.ui.listWidget_files.addItem(item)
        for f in self.files:
            item = QtWidgets.QListWidgetItem(f['name'])
            tt = ""
            if f['memo'] != "":
                tt = _("\nMEMO: ") + f['memo']
            item.setToolTip(tt)
            self.ui.listWidget_files.addItem(item)

        self.ui.listWidget_cases.clear()
        self.cases = self.app.get_casenames()
        item = QtWidgets.QListWidgetItem("")
        item.setToolTip(_("No case selection"))
        self.ui.listWidget_cases.addItem(item)
        for c in self.cases:
            tt = ""
            item = QtWidgets.QListWidgetItem(c['name'])
            if c['memo'] != "":
                tt = _("MEMO: ") + c['memo']
            item.setToolTip(tt)
            self.ui.listWidget_cases.addItem(item)

    def get_codes_categories_coders(self):
        """ Called from init, delete category. Load codes, categories, and coders. """

        self.code_names, self.categories = self.app.get_codes_categories()
        cur = self.app.conn.cursor()
        self.coders = []
        cur.execute("select distinct owner from code_text")
        result = cur.fetchall()
        self.coders = [""]
        for row in result:
            self.coders.append(row[0])

    def get_selected_files_and_cases(self):
        """ Fill file_ids and case_ids Strings used in the search.
        Clear attribute selection.
         Called by: search """

        selected_files = []
        self.file_ids_string = ""
        for item in self.ui.listWidget_files.selectedItems():
            selected_files.append(item.text())
            for f in self.files:
                if f['name'] == item.text():
                    self.file_ids_string += f",{f['id']}"
        if len(self.file_ids_string) > 0:
            self.file_ids_string = self.file_ids_string[1:]
        selected_cases = []
        self.case_ids_string = ""
        for item in self.ui.listWidget_cases.selectedItems():
            selected_cases.append(item.text())
            for c in self.cases:
                if c['name'] == item.text():
                    self.case_ids_string += f",{c['id']}"
        if len(self.case_ids_string) > 0:
            self.case_ids_string = self.case_ids_string[1:]

    def listwidget_files_menu(self, position):
        """ Context menu for file selection. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_all_files = menu.addAction(_("Select all files"))
        action_files_like = menu.addAction(_("Select files like"))
        action_files_none = menu.addAction(_("Select none"))
        action = menu.exec(self.ui.listWidget_files.mapToGlobal(position))
        if action == action_all_files:
            self.ui.listWidget_files.selectAll()
            self.ui.listWidget_files.item(0).setSelected(False)
        if action == action_files_none:
            for i in range(self.ui.listWidget_files.count()):
                self.ui.listWidget_files.item(i).setSelected(False)
        if action == action_files_like:
            # Input dialog narrow, so code below
            dialog = QtWidgets.QInputDialog(None)
            dialog.setStyleSheet("* {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
            dialog.setWindowTitle(_("Select some files"))
            dialog.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
            dialog.setInputMode(QtWidgets.QInputDialog.InputMode.TextInput)
            dialog.setLabelText(_("Show files containing text"))
            dialog.resize(200, 20)
            ok = dialog.exec()
            if not ok:
                return
            dlg_text = str(dialog.textValue())
            for i in range(self.ui.listWidget_files.count()):
                item_name = self.ui.listWidget_files.item(i).text()
                if dlg_text in item_name:
                    self.ui.listWidget_files.item(i).setSelected(True)
                else:
                    self.ui.listWidget_files.item(i).setSelected(False)

    def listwidget_cases_menu(self, position):
        """ Context menu for case selection. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_all_cases = menu.addAction(_("Select all cases"))
        action_cases_like = menu.addAction(_("Select cases like"))
        action_cases_none = menu.addAction(_("Select none"))
        action = menu.exec(self.ui.listWidget_cases.mapToGlobal(position))
        if action == action_all_cases:
            self.ui.listWidget_cases.selectAll()
            self.ui.listWidget_cases.item(0).setSelected(False)
        if action == action_cases_none:
            for i in range(self.ui.listWidget_cases.count()):
                self.ui.listWidget_cases.item(i).setSelected(False)
        if action == action_cases_like:
            # Input dialog narrow, so code below
            dialog = QtWidgets.QInputDialog(None)
            dialog.setStyleSheet("* {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
            dialog.setWindowTitle(_("Select some cases"))
            dialog.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
            dialog.setInputMode(QtWidgets.QInputDialog.InputMode.TextInput)
            dialog.setLabelText(_("Select cases containing text"))
            dialog.resize(200, 20)
            ok = dialog.exec()
            if not ok:
                return
            text_ = str(dialog.textValue())
            for i in range(self.ui.listWidget_cases.count()):
                item_name = self.ui.listWidget_cases.item(i).text()
                if text_ in item_name:
                    self.ui.listWidget_cases.item(i).setSelected(True)
                else:
                    self.ui.listWidget_cases.item(i).setSelected(False)

    def fill_tree(self):
        """ Fill tree widget, top level items are main categories and unlinked codes. """

        cats = deepcopy(self.categories)
        codes = deepcopy(self.code_names)
        self.ui.treeWidget.clear()
        self.ui.treeWidget.setColumnCount(4)
        self.ui.treeWidget.setHeaderLabels([_("Name"), "Id", _("Memo"), _("Count")])
        self.ui.treeWidget.header().setToolTip(_("Codes and categories"))
        if not self.app.settings['showids']:
            self.ui.treeWidget.setColumnHidden(1, True)
        else:
            self.ui.treeWidget.setColumnHidden(1, False)
        self.ui.treeWidget.header().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.ui.treeWidget.header().setStretchLastSection(False)
        # Add top level categories
        remove_list = []
        for c in cats:
            if c['supercatid'] is None:
                memo = ""
                if c['memo'] != "":
                    memo = _("Memo")
                top_item = QtWidgets.QTreeWidgetItem([c['name'], 'catid:' + str(c['catid']), memo])
                top_item.setToolTip(0, '')
                if len(c['name']) > 52:
                    top_item.setText(0, f"{c['name'][:25]}..{c['name'][-25:]}")
                    top_item.setToolTip(0, c['name'])
                top_item.setToolTip(2, c['memo'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_list.append(c)
        for item in remove_list:
            cats.remove(item)

        ''' Add child categories. Look at each unmatched category, iterate through tree
        to add as child then remove matched categories from the list. '''
        count = 0
        while len(cats) > 0 and count < 10000:
            remove_list = []
            for c in cats:
                it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
                item = it.value()
                count2 = 0
                while item and count2 < 10000:  # while there is an item in the list
                    if item.text(1) == 'catid:' + str(c['supercatid']):
                        memo = ""
                        if c['memo'] != "":
                            memo = "Memo"
                        child = QtWidgets.QTreeWidgetItem([c['name'], f"catid:{c['catid']}", memo])
                        child.setToolTip(0, '')
                        if len(c['name']) > 52:
                            child.setText(0, f"{c['name'][:25]}..{c['name'][-25:]}")
                            child.setToolTip(0, c['name'])
                        child.setToolTip(2, c['memo'])
                        item.addChild(child)
                        remove_list.append(c)
                    it += 1
                    item = it.value()
                    count2 += 1
            for item in remove_list:
                cats.remove(item)
            count += 1

        # Add unlinked codes as top level items
        remove_items = []
        for c in codes:
            if c['catid'] is None:
                memo = ""
                if c['memo'] != "":
                    memo = "Memo"
                top_item = QtWidgets.QTreeWidgetItem([c['name'], f"cid:{c['cid']}", memo])
                top_item.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                color = TextColor(c['color']).recommendation
                top_item.setForeground(0, QBrush(QtGui.QColor(color)))
                top_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                top_item.setToolTip(0, '')
                if len(c['name']) > 52:
                    top_item.setText(0, f"{c['name'][:25]}..{c['name'][-25:]}")
                    top_item.setToolTip(0, c['name'])
                top_item.setToolTip(2, c['memo'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_items.append(c)
        for item in remove_items:
            codes.remove(item)

        # Add codes as children
        for c in codes:
            it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
            item = it.value()
            count = 0
            while item and count < 10000:
                if item.text(1) == 'catid:' + str(c['catid']):
                    memo = ""
                    if c['memo'] != "":
                        memo = _("Memo")
                    child = QtWidgets.QTreeWidgetItem([c['name'], f"cid:{c['cid']}", memo])
                    child.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                    color = TextColor(c['color']).recommendation
                    child.setForeground(0, QBrush(QtGui.QColor(color)))
                    child.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                    child.setToolTip(0, '')
                    if len(c['name']) > 52:
                        child.setText(0, f"{c['name'][:25]}..{c['name'][-25:]}")
                        child.setToolTip(0, c['name'])
                    child.setToolTip(2, c['memo'])
                    item.addChild(child)
                    c['catid'] = -1  # make unmatchable
                it += 1
                item = it.value()
                count += 1
        self.ui.treeWidget.sortByColumn(0, QtCore.Qt.SortOrder.AscendingOrder)
        self.fill_code_counts_in_tree()
        self.ui.treeWidget.expandAll()

    def fill_code_counts_in_tree(self):
        """ Count instances of each code from all coders and all files. """

        cur = self.app.conn.cursor()
        sql = "select count(cid) from code_text where cid=? union "
        sql += "select count(cid) from code_av where cid=? union "
        sql += "select count(cid) from code_image where cid=?"
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        count = 0
        while item and count < 10000:
            if item.text(1)[0:4] == "cid:":
                cid = str(item.text(1)[4:])
                cur.execute(sql, [cid, cid, cid])  # , self.app.settings['codername']])
                result = cur.fetchall()
                total = 0
                for row in result:
                    total = total + row[0]
                if total > 0:
                    item.setText(3, str(total))
                else:
                    item.setText(3, "")
            it += 1
            item = it.value()
            count += 1

    def treewidget_menu(self, position):
        """ Menu to select all codes or other selection parameters. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_all = menu.addAction(_("Select all codes"))
        action_unselect = menu.addAction(_("Remove selections"))
        action_like = menu.addAction(_("Select codes like"))
        action = menu.exec(self.ui.treeWidget.mapToGlobal(position))
        if action == action_all:
            self.ui.treeWidget.selectAll()
        if action == action_unselect:
            selected = self.ui.treeWidget.selectedItems()
            for tree_item in selected:
                tree_item.setSelected(False)
        if action == action_like:
            # Need to unselect where mouse click occurred
            clicked_selected = self.ui.treeWidget.selectedItems()[0]
            clicked_selected.setSelected(False)
            # Input dialog narrow, so code below
            dialog = QtWidgets.QInputDialog(None)
            dialog.setStyleSheet("* {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
            dialog.setWindowTitle(_("Select some codes"))
            dialog.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
            dialog.setInputMode(QtWidgets.QInputDialog.InputMode.TextInput)
            dialog.setLabelText(_("Select codes containing text"))
            dialog.resize(200, 20)
            ok = dialog.exec()
            if not ok:
                return
            selection_text = str(dialog.textValue())
            tree_items = self.ui.treeWidget.findItems(selection_text, Qt.MatchFlag.MatchContains | Qt.MatchFlag.MatchRecursive, 0)
            for tree_item in tree_items:
                if 'cid' in tree_item.text(1):
                    tree_item.setSelected(True)

    def export_xlsx_file(self):
        """ Export report to xlsx file.
        """

        if not self.xlsx_results:
            return
        wb = openpyxl.Workbook()
        ws = wb.active

        row = 1
        for col, code in enumerate(self.horizontal_labels):
            ws.cell(column=col + 1, row=row, value=code)

        for row, xlsx_row in enumerate(self.xlsx_results):
            for col, data in enumerate(xlsx_row):
                ws.cell(column=1 + col, row=2 + row , value=data)

        filepath, ok = QtWidgets.QFileDialog.getSaveFileName(self,
                                                            _("Save Excel File"), self.app.settings['directory'],
                                                            "XLSX Files(*.xlsx)")
        # options=QtWidgets.QFileDialog.Option.DontUseNativeDialog)
        if filepath is None or not ok:
            return
        wb.save(filepath)
        msg = _('Report exported: ') + filepath
        Message(self.app, _('Report exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def recursive_set_selected(self, item):
        """ Set all children of this item to be selected if the item is selected.
        Recurse through any child categories.
        Called by: search
        """

        child_count = item.childCount()
        for i in range(child_count):
            if item.isSelected():
                item.child(i).setSelected(True)
            self.recursive_set_selected(item.child(i))

    def search(self):
        """ Make table data
        self.file_ids_string
        self.case_ids_string """

        self.get_selected_files_and_cases()
        self.results = []
        # Select all code items under selected categories
        self.recursive_set_selected(self.ui.treeWidget.invisibleRootItem())
        items = self.ui.treeWidget.selectedItems()
        self.code_ids_string = ""
        for i in items:
            if i.text(1)[0:3] == 'cid':
                self.code_ids_string += f",{i.text(1)[4:]}"
        self.code_ids_string = self.code_ids_string[1:]

        if self.case_ids_string == "":
            self.search_by_files()
        else:
            self.search_by_case()

        '''for r in self.results:
            print(r)'''
        # Collate results
        # Collate Codes - columns and segments - rows
        self.code_columns = []
        self.segment_rows = []
        for r in self.results:
            short_segment = {'caseid': r['caseid'], 'casename': r['casename'], 'fid': r['fid'], 'filename': r['filename'], 'pos0': r['pos0'], 'pos1': r['pos1'], 'text': r['text']}
            if short_segment not in self.segment_rows:
                self.segment_rows.append(short_segment)
            codename = r['codename']
            if codename not in self.code_columns:
                self.code_columns.append(codename)

        for row in self.segment_rows:
            row['codes'] = [0] * len(self.code_columns)

        for row in self.segment_rows:
            for res in self.results:
                if res['fid'] == row['fid'] and res['pos0'] == row['pos0'] and res['pos1'] == row['pos1']:
                    for i, codename in enumerate(self.code_columns):
                        if codename == res['codename']:
                            row['codes'][i] = 1

        self.fill_table()

    def search_by_files(self):

        coder = self.ui.comboBox_coders.currentText()
        search_text = self.ui.lineEdit.text()
        parameters = []
        cur = self.app.conn.cursor()
        # Coded text
        sql = "select code_name.name, color, source.name, pos0, pos1, seltext, "
        sql += "code_text.owner, fid, ifnull(code_text.memo,''), ifnull(code_name.memo,''), " \
               "ifnull(source.memo,''), ctid, code_name.cid, -1 as caseid, '' as casename "
        sql += " from code_text join code_name "
        sql += "on code_name.cid = code_text.cid join source on fid = source.id "
        sql += f"where code_name.cid in ({self.code_ids_string}) "
        sql += f"and source.id in ({self.file_ids_string}) "
        if coder != "":
            sql += " and code_text.owner=? "
            parameters.append(coder)
        if search_text != "":
            sql += " and seltext like ? "
            parameters.append(f"%{search_text}%")
        sql += " order by seltext"  # code_name.name, source.name, pos0"
        if not parameters:
            cur.execute(sql)
        else:
            cur.execute(sql, parameters)
        result = cur.fetchall()
        keys = 'codename', 'color', 'filename', 'pos0', 'pos1', 'text', 'coder', 'fid', 'coded_memo', \
            'codename_memo', 'source_memo', 'ctid', 'cid', 'caseid', 'casename'
        for row in result:
            tmp = dict(zip(keys, row))
            self.results.append(tmp)

    def search_by_case(self):

        coder = self.ui.comboBox_coders.currentText()
        search_text = self.ui.lineEdit.text()
        cur = self.app.conn.cursor()
        parameters = []

        # Coded text
        sql = "select code_name.name, color, cases.name, "
        sql += "code_text.pos0, code_text.pos1, seltext, code_text.owner, code_text.fid, "
        sql += "ifnull(cases.memo,''), ifnull(code_text.memo,''), ifnull(code_name.memo,''), "
        sql += "ifnull(source.memo,''), ctid, code_name.cid, "
        sql += "case_text.caseid, source.name as filename "
        sql += "from code_text join code_name on code_name.cid = code_text.cid "
        sql += "join (case_text join cases on cases.caseid = case_text.caseid) on "
        sql += "code_text.fid = case_text.fid "
        sql += "join source on source.id=code_text.fid "
        sql += f"where code_name.cid in ({self.code_ids_string}) "
        sql += f"and case_text.caseid in ({self.case_ids_string}) "
        if self.file_ids_string != "":
            sql += f" and code_text.fid in ({self.file_ids_string})"
        sql += "and (code_text.pos0 >= case_text.pos0 and code_text.pos1 <= case_text.pos1)"
        if coder != "":
            sql += " and code_text.owner=? "
            parameters.append(coder)
        if search_text != "":
            sql += " and seltext like ? "
            parameters.append(f"%{search_text}%")
        sql += " order by seltext"  # code_name.name, cases.name"
        if not parameters:
            cur.execute(sql)
        else:
            cur.execute(sql, parameters)
        results = cur.fetchall()
        keys = 'codename', 'color', 'casename', 'pos0', 'pos1', 'text', 'coder', 'fid', \
            'cases_memo', 'coded_memo', 'codename_memo', 'source_memo', 'ctid', 'cid', 'caseid', 'filename'
        for row in results:
            tmp = dict(zip(keys, row))
            self.results.append(tmp)

    def fill_table(self):
        """ Fill table and prepare xlsx result set. """

        num_rows = self.ui.tableWidget.rowCount()
        for row in range(0, num_rows):
            self.ui.tableWidget.removeRow(0)
        self.horizontal_labels = ["caseid", "case name", "fileid", "file name", "pos0", "pos1", "text"] + self.code_columns
        self.ui.tableWidget.setColumnCount(len(self.horizontal_labels))
        self.ui.tableWidget.setHorizontalHeaderLabels(self.horizontal_labels)
        font = f"font: {self.app.settings['fontsize']}pt "
        font += f'"{self.app.settings["font"]}";'
        self.ui.tableWidget.horizontalHeader().setStyleSheet("QHeaderView {font-size: 8pt; }")
        self.ui.tableWidget.setRowCount(len(self.segment_rows))

        self.xlsx_results = []

        for row, segment in enumerate(self.segment_rows):
            xlsx_row = []
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, segment['caseid'])
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled)
            item.setFont(QtGui.QFont(self.app.settings['font'], 8))
            self.ui.tableWidget.setItem(row, 0, item)
            xlsx_row.append(segment['caseid'])

            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, segment['casename'])
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled)
            item.setFont(QtGui.QFont(self.app.settings['font'], 8))
            self.ui.tableWidget.setItem(row, 1, item)
            xlsx_row.append(segment['casename'])

            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, segment['fid'])
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled)
            item.setFont(QtGui.QFont(self.app.settings['font'], 8))
            self.ui.tableWidget.setItem(row, 2, item)
            xlsx_row.append(segment['fid'])

            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, segment['filename'])
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled)
            item.setFont(QtGui.QFont(self.app.settings['font'], 8))
            self.ui.tableWidget.setItem(row, 3, item)
            xlsx_row.append(segment['filename'])

            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, segment['pos0'])
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled)
            item.setFont(QtGui.QFont(self.app.settings['font'], 8))
            self.ui.tableWidget.setItem(row, 4, item)
            xlsx_row.append(segment['pos0'])

            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, segment['pos1'])
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled)
            item.setFont(QtGui.QFont(self.app.settings['font'], 8))
            self.ui.tableWidget.setItem(row, 5, item)
            xlsx_row.append(segment['pos1'])

            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, segment['text'])
            item.setFont(QtGui.QFont(self.app.settings['font'], 6))
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled)
            self.ui.tableWidget.setItem(row, 6, item)
            xlsx_row.append(segment['text'])

            for col_pos, coding in enumerate(segment['codes']):
                item = QtWidgets.QTableWidgetItem()
                item.setData(QtCore.Qt.ItemDataRole.DisplayRole, coding)
                item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable | QtCore.Qt.ItemFlag.ItemIsEnabled)
                self.ui.tableWidget.setItem(row, 7 + col_pos, item)
                xlsx_row.append(coding)
            self.xlsx_results.append(xlsx_row)

        self.ui.tableWidget.resizeColumnsToContents()
        self.ui.tableWidget.setColumnWidth(6, 250)
        self.ui.tableWidget.resizeRowsToContents()


