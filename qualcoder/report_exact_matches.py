# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com/
"""

from copy import copy
import logging
import openpyxl
import os
import qtawesome as qta  # see: https://pictogrammers.com/library/mdi/

from PyQt6 import QtGui, QtWidgets, QtCore
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QBrush

from .color_selector import TextColor
from .GUI.ui_report_matching_segments import Ui_DialogMatchingTextSegments
from .helpers import DialogCodeInText, Message
from .report_attributes import DialogSelectAttributeParameters

path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogReportExactTextMatches(QtWidgets.QDialog):
    """ Based on code relation.
    Show exact match code overlaps.
    This is for text coding only. """

    app = None
    parent_textEdit = None
    coder_names = []
    categories = []
    codes = []
    coders = []
    files = []
    results_display = []
    excluded_codes = []
    excluded_icon = None

    def __init__(self, app, parent_textedit):

        self.results_display = []
        self.app = app
        self.parent_textEdit = parent_textedit

        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_DialogMatchingTextSegments()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        self.ui.pushButton_run.setIcon(qta.icon('mdi6.play', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_export.pressed.connect(self.export_excel_file)
        self.ui.pushButton_file_filter.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.3}]))
        self.excluded_icon = qta.icon('mdi6.window-close')
        self.get_data()

        try:
            s0 = int(self.app.settings['dialogcodecrossovers_splitter0'])
            s1 = int(self.app.settings['dialogcodecrossovers_splitter1'])
            if s0 > 10 and s1 > 10:
                self.ui.splitter.setSizes([s0, s1])
        except KeyError:
            pass
        info_font = f'font: 10pt "{self.app.settings["font"]}";'
        self.ui.label_instructions.setStyleSheet(info_font)
        self.ui.checkBox_any_matches.setStyleSheet(info_font)
        msg = _("Only matches where ALL codes are applied to the same text will be shown.")
        self.ui.label_instructions.setToolTip(msg)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        font = f'font: {self.app.settings["treefontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.treeWidget.setStyleSheet(font)
        self.ui.treeWidget.setSelectionMode(QtWidgets.QTreeWidget.SelectionMode.ExtendedSelection)
        self.fill_tree()
        self.ui.treeWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.treeWidget.customContextMenuRequested.connect(self.tree_menu)
        self.get_files_fill_list_widget()
        self.ui.listWidget_files.setSelectionMode(QtWidgets.QListWidget.SelectionMode.ExtendedSelection)
        self.ui.pushButton_file_filter.pressed.connect(self.select_files_by_parameters)
        self.ui.pushButton_run.pressed.connect(self.get_exact_text_matches)
        self.ui.tableWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget.customContextMenuRequested.connect(self.table_menu)

    def get_data(self):
        """ Called from init. gets code_names, categories and owner names.
        """

        self.coder_names = self.app.get_coder_names_in_project()
        self.codes, self.categories = self.app.get_codes_categories()
        sql = "select owner from  code_image union select owner from code_text union select owner from code_av"
        cur = self.app.conn.cursor()
        cur.execute(sql)
        result = cur.fetchall()
        self.coders = []
        for row in result:
            self.coders.append(row[0])
        self.ui.comboBox_coders.insertItems(0, self.coders)

    def get_files_fill_list_widget(self):
        """ Get source files with additional details and fill list widget.
        Add file type to dictionary for each file.
        """

        self.ui.listWidget_files.clear()
        self.files = self.app.get_filenames()
        # Fill additional details about each file in the memo
        cur = self.app.conn.cursor()
        sql = "select length(fulltext), mediapath from source where id=?"
        sql_text_codings = "select count(cid) from code_text where fid=?"
        sql_av_codings = "select count(cid) from code_av where id=?"
        sql_image_codings = "select count(cid) from code_image where id=?"
        for f in self.files:
            cur.execute(sql, [f['id'], ])
            res = cur.fetchone()
            if res is None:  # safety catch
                res = [0]
            tt = ""
            f['mediapath'] = res[1]
            if res[1] is None or res[1][0:5] == "docs:" or res[1][0:5] == "/docs":
                tt += _("Text file\n")
                tt += _("Characters: ") + str(res[0])
                f['type'] = 'text'
            if res[1] is not None and (res[1][0:7] == "images:" or res[1][0:7] == "/images"):
                tt += _("Image")
                f['type'] = 'image'
            if res[1] is not None and (res[1][0:6] == "audio:" or res[1][0:6] == "/audio"):
                tt += _("Audio")
                f['type'] = 'audio'
            if res[1] is not None and (res[1][0:6] == "video:" or res[1][0:6] == "/video"):
                tt += _("Video")
                f['type'] = 'video'
            cur.execute(sql_text_codings, [f['id']])
            txt_res = cur.fetchone()
            tt += _("\nCodings: ")
            if txt_res[0] > 0:
                tt += str(txt_res[0])
            item = QtWidgets.QListWidgetItem(f['name'])
            if f['memo'] != "":
                tt += _("\nMemo: ") + f['memo']
            item.setToolTip(tt)
            self.ui.listWidget_files.addItem(item)

    def select_files_by_parameters(self):
        """ Select files for report, by using file attributes. """

        Message(self.app, _('Work to do'), "Work in progress", "warning").exec()

    def get_exact_text_matches(self):
        """ Use selected, coder, selected files and two or more codes.
        All selected codes mus tbe present for the match.
        Create two sets of results:
            One row per coded segment per code; one row collating all codes at one coded segment.
        """

        selected_coder = self.ui.comboBox_coders.currentText()
        file_ids = []
        for file_item in self.ui.listWidget_files.selectedItems():
            for f in self.files:
                print(file_item.text(), f['name'])
                if f['name'] == file_item.text():
                    file_ids.append(f['id'])
        if not file_ids:
            msg = _("No file has been selected.")
            Message(self.app, _('No file'), msg, "warning").exec()
            return
        selected_codes = []
        items = self.ui.treeWidget.selectedItems()
        excluded_cids = []
        excluded_cids_string = ""
        for excluded in self.excluded_codes:
            excluded_cids.append(excluded[0])
            excluded_cids_string += f",{excluded[0]}"
        if len(excluded_cids_string) > 0:
            excluded_cids_string = excluded_cids_string[1:]
        for i in items:
            if i.text(1)[0:3] == 'cid':
                cid = int(i.text(1)[4:])
                if cid not in excluded_cids:
                    selected_codes.append(str(cid))
        if len(selected_codes) == 0:
            msg = _("No codes have been selected.")
            Message(self.app, _('No codes'), msg, "warning").exec()
            return
        if len(selected_codes) == 1:
            msg = _("Only one code has been selected.")
            Message(self.app, _('One code selected'), msg, "warning").exec()
            return
        selected_codes_string = ",".join(selected_codes)
        includes_text = self.ui.lineEdit_include.text()
        any_selected_codes = self.ui.checkBox_any_matches.isChecked()
        # if False - ALL selected codes must match
        self.results_display = []
        final_matches_list = []
        one_line_results_list = []
        cur = self.app.conn.cursor()
        for fid in file_ids:
            values = [selected_coder, fid]
            sql = "select code_text.cid, code_name.name, pos0,pos1, substr(source.fulltext,pos0, 1+pos1-pos0), "
            sql += " ifnull(code_text.memo,''), source.id, source.name, code_text.owner "
            sql += " from code_text join code_name on code_name.cid=code_text.cid "
            sql += " join source on source.id=code_text.fid "
            sql += f" where code_text.cid in ({selected_codes_string}) "
            if includes_text != "":
                sql += " and instr(substr(source.fulltext,pos0, 1+pos1-pos0), ?) > 0 "
                values = [includes_text, selected_coder, fid]
            sql += " and code_text.owner=? and code_text.fid=? "
            sql += " order by code_name.name, pos0"
            cur.execute(sql, values)
            coded_results = cur.fetchall()

            sql = f"select code_text.cid, pos0,pos1 from code_text where "
            sql += f" code_text.cid in ({excluded_cids_string}) "
            sql += " and code_text.owner=? and code_text.fid=?"
            cur.execute(sql, [selected_coder, fid])
            excludes_result = cur.fetchall()

            for c in coded_results:
                matching_codes_list = []
                # Get all coded matching text segment data
                for c2 in coded_results:
                    if c[2] == c2[2] and c[3] == c2[3] and c[0] != c2[0]:
                        matching_codes_list.append(c2)
                if matching_codes_list:
                    matching_codes_list.append(c)
                # Remove from result if the matching data is in the excludes list
                for excludes in excludes_result:
                    if c[2] == excludes[2] and c[3] == excludes[3]:
                        matching_codes_list = []
                # Sort lists by cid. Helps to remove duplicated differing order matches.
                matching_codes_list.sort()
                # checkbox NOT checked. So all exact matching codes must be present at coded segment.
                if not any_selected_codes and len(matching_codes_list) != len(selected_codes):
                    matching_codes_list = []
                # Add resulting list to final results. Check if there is a need to include specified text.
                if matching_codes_list and matching_codes_list not in final_matches_list:
                    final_matches_list.append(matching_codes_list)
                    # Create one line result
                    one_line_results = list(matching_codes_list[0])
                    one_line_results[0] = str(one_line_results[0])  # cid
                    #one_line_results[5] = ""  # coded memo
                    for row in range(1, len(matching_codes_list)):
                        one_line_results[0] += f", {str(matching_codes_list[row][0])}"  # cid
                        one_line_results[1] += f"|{matching_codes_list[row][1]}"  # codename
                        one_line_results[5] += f"|{matching_codes_list[row][5]}"  # coded segment memo
                    if one_line_results[5] == "||":
                        one_line_results[5] = ""
                    one_line_results_list.append(one_line_results)

        # Each rows displayed
        for match_list in final_matches_list:
            for match_item in match_list:
                self.results_display.append(match_item)
        if len(final_matches_list) == 0:
            msg = _("No exact matches found.") + "\n"
            if not any_selected_codes:
                msg += _("ALL selected codes need to be exactly overlapping.")
            Message(self.app, _('No results'), msg, "warning").exec()
            rows = self.ui.tableWidget.rowCount()
            for r in range(0, rows):
                self.ui.tableWidget.removeRow(0)
            return

        # Replace self.results_display if one line resulst is chosen
        if self.ui.checkBox_one_line.isChecked():
            self.results_display = one_line_results_list
        self.fill_table()

    #TODO
    '''def search_text(self):
        """ Search for text in the results. """

        search_text = self.ui.lineEdit_search_results.text()
        if search_text == "":
            return
        row_count = self.ui.tableWidget.rowCount()
        col_count = self.ui.tableWidget.columnCount()
        if row_count == 0:
            return
        current_row = 0
        current_col = 0
        try:
            current_row = self.ui.tableWidget.currentRow()
            current_col = self.ui.tableWidget.currentColumn()
        except AttributeError:
            # No table for table menu
            return
        cell_counter_pos = current_row * col_count + current_col
        found_row = -1
        found_col = -1
        for row in range(0, row_count):
            for col in range(0, col_count):
                try:
                    cell = self.ui.tableWidget.item(row, col).text()
                    if cell_counter_pos < row * col_count + col and found_row == -1 and search_text in cell:
                        found_row = row
                        found_col = col
                        break
                except AttributeError:
                    pass
                if found_row != -1:
                    break
        if found_row == -1:
            return
        self.ui.tableWidget.setCurrentCell(found_row, found_col)'''

    def table_menu(self, position):
        """ Context menu to show row text in original context, row ordering. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        cell_value = ""
        try:
            row = self.ui.tableWidget.currentRow()
            col = self.ui.tableWidget.currentColumn()
            cell_value = self.ui.tableWidget.item(row, col).text()
        except AttributeError:
            # No table for table menu
            return
        action_show_context = menu.addAction(_("View in context"))
        action_sort_ascending = menu.addAction(_("Sort ascending"))
        action_sort_descending = menu.addAction(_("Sort descending"))
        action_filter_equals = menu.addAction(_("Filter equals: ") + cell_value)
        action_filter_greater = menu.addAction(_("Filter greater or equals: ") + cell_value)
        action_filter_lower = menu.addAction(_("Filter lower or equals: ") + cell_value)
        action_clear_filter = menu.addAction(_("Clear filter"))
        action = menu.exec(self.ui.tableWidget.mapToGlobal(position))
        if action == action_show_context:
            self.show_context()
        if action == action_sort_ascending:
            self.ui.tableWidget.sortItems(col, QtCore.Qt.SortOrder.AscendingOrder)
        if action == action_sort_descending:
            self.ui.tableWidget.sortItems(col, QtCore.Qt.SortOrder.DescendingOrder)
        if action == action_clear_filter:
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
        if action == action_filter_equals:
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
                if self.ui.tableWidget.item(r, col).text() != cell_value:
                    self.ui.tableWidget.setRowHidden(r, True)
        if action == action_filter_greater:
            val_type = "str"
            if col in (0, 2, 3):
                val_type = "int"
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
                if val_type == "str" and self.ui.tableWidget.item(r, col).text() < cell_value:
                    self.ui.tableWidget.setRowHidden(r, True)
                if val_type == "int" and int(self.ui.tableWidget.item(r, col).text()) < int(cell_value):
                    self.ui.tableWidget.setRowHidden(r, True)
        if action == action_filter_lower:
            val_type = "str"
            if col in (0, 2, 3):
                val_type = "int"
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
                if val_type == "str" and self.ui.tableWidget.item(r, col).text() > cell_value:
                    self.ui.tableWidget.setRowHidden(r, True)
                if val_type == "int" and int(self.ui.tableWidget.item(r, col).text()) > int(cell_value):
                    self.ui.tableWidget.setRowHidden(r, True)

    def show_context(self):
        """ Show context of coding in dialog.
        Called by table_menu.
        """

        row = self.ui.tableWidget.currentRow()
        code_color = "#888888"
        if self.ui.tableWidget.item(row, 0).text().isdigit():
            for c in self.codes:
                if c['cid'] == int(self.ui.tableWidget.item(row, 0).text()):
                    code_color = c['color']
        # data: dictionary: codename, color, file_or_casename, pos0, pos1, text, coder, fid, file_or_case,
        # textedit_start, textedit_end
        data = {'codename': self.ui.tableWidget.item(row, 1).text(), 'color': code_color, 'text': '',
                'pos0':  int(self.ui.tableWidget.item(row, 2).text()),
                'pos1': int(self.ui.tableWidget.item(row, 3).text()),
                'fid': int(self.ui.tableWidget.item(row, 6).text()),
                'file_or_casename': self.ui.tableWidget.item(row, 7).text(),
                'coder': self.ui.tableWidget.item(row, 8).text(), 'file_or_case': 'File'}
        ui = DialogCodeInText(self.app, data)
        ui.exec()
        return

    def fill_table(self):
        """ A table of matching coded text segments. """

        cid = 0
        code_name = 1
        pos0 = 2
        pos1 = 3
        sel_text = 4
        coded_memo = 5
        fid = 6
        file_name = 7
        coder = 8

        col_names = ["cid", _("code name"), "pos0", "pos1", _("text"), _("Coded memo"), "fid", _("File name"), _("Coder")]
        self.ui.tableWidget.setColumnCount(len(col_names))
        self.ui.tableWidget.setHorizontalHeaderLabels(col_names)
        rows = self.ui.tableWidget.rowCount()
        for r in range(0, rows):
            self.ui.tableWidget.removeRow(0)
        for r, match_item in enumerate(self.results_display):
            self.ui.tableWidget.insertRow(r)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, match_item[cid])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, cid, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, match_item[code_name])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, code_name, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, match_item[pos0])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, pos0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, match_item[pos1])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, pos1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, match_item[sel_text])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, sel_text, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, match_item[coded_memo])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, coded_memo, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, match_item[fid])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, fid, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, match_item[file_name])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, file_name, item)
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.ItemDataRole.DisplayRole, match_item[coder])
            item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(r, coder, item)
        self.ui.tableWidget.hideColumn(fid)
        self.ui.tableWidget.hideColumn(coder)
        self.ui.tableWidget.resizeColumnsToContents()
        self.ui.tableWidget.setColumnWidth(sel_text, 1000)
        self.ui.tableWidget.resizeRowsToContents()

    def export_excel_file(self):
        """ Export exact match text codings for all codes as excel file.
        Output ordered by filename and code name ascending. """

        if len(self.results_display) == 0:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        # Column headings
        col_headings = ["cid", "Code name", "pos0", "pos1", "Text", "Coded memo", "File name"]
        row = 1
        for col, code in enumerate(col_headings):
            ws.cell(column=col + 1, row=row, value=code)
        for row, data in enumerate(self.results_display):
            ws.cell(column=1, row=row + 2, value=data[0])
            ws.cell(column=2, row=row + 2, value=data[1])
            ws.cell(column=3, row=row + 2, value=data[2])
            ws.cell(column=4, row=row + 2, value=data[3])
            ws.cell(column=5, row=row + 2, value=data[4])
            ws.cell(column=6, row=row + 2, value=data[5])
            ws.cell(column=7, row=row + 2, value=data[7])
        filepath, ok = QtWidgets.QFileDialog.getSaveFileName(self,
                                                             _("Save Excel File"), self.app.settings['directory'],
                                                             "XLSX Files(*.xlsx)")
        # options=QtWidgets.QFileDialog.Option.DontUseNativeDialog)
        if filepath is None or not ok:
            return
        if filepath[-4:] != ".xlsx":
            filepath += ".xlsx"
        wb.save(filepath)
        msg = _("Report of exact matches for text codings for file") + "\n"
        msg += _('Report exported to: ') + filepath
        Message(self.app, _('Report exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def fill_tree(self):
        """ Fill tree widget, top level items are main categories and unlinked codes.
        """

        self.ui.treeWidget.clear()

        cats = copy(self.categories)
        codes = copy(self.codes)
        self.ui.treeWidget.clear()
        header = [_("Code Tree"), _("Id")]
        self.ui.treeWidget.setColumnCount(len(header))
        self.ui.treeWidget.setHeaderLabels(header)
        self.ui.treeWidget.header().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.ui.treeWidget.header().setStretchLastSection(False)
        # add top level categories
        remove_list = []
        for c in cats:
            if c['supercatid'] is None:
                top_item = QtWidgets.QTreeWidgetItem([c['name'], 'catid:' + str(c['catid'])])
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_list.append(c)
        for item in remove_list:
            cats.remove(item)

        ''' Add child categories. Look at each unmatched category, iterate through tree to
        add as child then remove matched categories from the list. '''
        count = 0
        while len(cats) > 0 and count < 10000:
            remove_list = []
            for c in cats:
                it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
                item = it.value()
                while item:  # while there is an item in the list
                    # logger.debug("While: ", item.text(0), item.text(1), c['catid'], c['supercatid'])
                    if item.text(1) == 'catid:' + str(c['supercatid']):
                        child = QtWidgets.QTreeWidgetItem([c['name'], 'catid:' + str(c['catid'])])
                        child.setToolTip(0, c['name'])
                        item.addChild(child)
                        # logger.debug("Adding: " + c['name'])
                        remove_list.append(c)
                    it += 1
                    item = it.value()
            for item in remove_list:
                cats.remove(item)
            count += 1

        # Add unlinked codes as top level items
        remove_items = []
        for c in codes:
            if c['catid'] is None:
                top_item = QtWidgets.QTreeWidgetItem([c['name'], f"cid:{c['cid']}"])
                top_item.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                color = TextColor(c['color']).recommendation
                top_item.setForeground(0, QBrush(QtGui.QColor(color)))
                top_item.setFlags(
                    Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_items.append(c)
        for item in remove_items:
            codes.remove(item)

        # Add codes as children
        for c in codes:
            it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
            item = it.value()
            while item:
                if item.text(1) == f'catid:{c["catid"]}':
                    child = QtWidgets.QTreeWidgetItem([c['name'], f'cid:{c["cid"]}'])
                    child.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                    color = TextColor(c['color']).recommendation
                    child.setForeground(0, QBrush(QtGui.QColor(color)))
                    child.setFlags(
                        Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                    child.setToolTip(0, c['name'])
                    item.addChild(child)
                    c['catid'] = -1  # Make unmatchable
                it += 1
                item = it.value()
        self.ui.treeWidget.expandAll()

    def tree_menu(self, position):
        """ Context menu for treewidget code/category items.
        Add, rename, memo, move or delete code or category. Change code color.
        Assign selected text to current hovered code. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        selected = self.ui.treeWidget.currentItem()
        action_clear_selected = menu.addAction(_("Clear all"))
        action_select_all = menu.addAction(_("Select all"))
        action_exclude_code = None
        if selected is not None and selected.text(1)[0:3] == 'cid':
            action_exclude_code = menu.addAction(_("Exclude code"))

        action = menu.exec(self.ui.treeWidget.mapToGlobal(position))
        if action == action_clear_selected:
            selected = self.ui.treeWidget.selectedItems()
            for tree_item in selected:
                tree_item.setSelected(False)
            for item in self.excluded_codes:
                item[1].setIcon(0, QtGui.QIcon())
            self.excluded_codes = []
            return
        if action == action_select_all:
            self.ui.treeWidget.selectAll()
        if action == action_exclude_code:
            if selected.text(1)[0:3] != 'cid':
                return
            cid = int(selected.text(1)[4:])
            self.excluded_codes.append([cid, selected])
            selected.setIcon(0, self.excluded_icon)
