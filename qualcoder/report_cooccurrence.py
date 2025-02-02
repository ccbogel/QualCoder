# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
"""

import logging
import openpyxl
import os
import qtawesome as qta  # see: https://pictogrammers.com/library/mdi/

from PyQt6 import QtCore, QtWidgets, QtGui

from .GUI.ui_dialog_cooccurrence import Ui_Dialog_Coocurrence
from .helpers import ExportDirectoryPathDialog, Message


path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogReportCooccurrence(QtWidgets.QDialog):
    """ Provide a co-occurrence report.
    """

    app = None
    parent_tetEdit = None
    files = []

    def __init__(self, app, parent_text_edit):
        self.app = app
        self.parent_textEdit = parent_text_edit
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_Coocurrence()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        self.ui.pushButton_export.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export.pressed.connect(self.export_to_excel)
        self.ui.checkBox_hide_blanks.stateChanged.connect(self.show_or_hide_empty_rows_and_cols)
        tablefont = f'font: 7pt "{self.app.settings["font"]}";'
        self.ui.tableWidget.setStyleSheet(tablefont)  # should be smaller
        self.ui.tableWidget.cellClicked.connect(self.cell_selected)
        #self.ui.tableWidget.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        #self.ui.tableWidget.customContextMenuRequested.connect(self.table_menu)
        self.ui.splitter.setSizes([500, 0])

        self.codes = []
        self.categories = []
        self.result_relations = []
        self.max_count = 0
        self.data_counts = []
        self.data_colors = []
        self.data_details = []

        self.process_data()

    def process_data(self):
        """ Calculate the relations for selected codes for ALL coders (or only THIS coder - TODO).
        For text codings only. """

        code_names_str = ""
        code_ids_str = ""

        self.codes, self.categories = self.app.get_codes_categories()
        code_names_list = []
        for c in self.codes:
            code_names_list.append(c['name'])
            code_names_str += f"{c['name']}|"
            code_ids_str += f",{c['cid']}"
        code_ids_str = code_ids_str[1:]
        self.result_relations = []
        self.calculate_relations(code_ids_str)

        '''# Tmp for testing
        for r in self.result_relations:
            print(r)'''

        # Create data matrices zeroed, codes are ordered alphabetically by name
        self.data_counts = []
        self.data_colors = []
        self.data_details = []
        for row in self.codes:
            self.data_counts.append([0] * len(self.codes))
            self.data_colors.append([""] * len(self.codes))
            self.data_details.append(["."] * len(self.codes))

        '''print("Data details")
        for r in self.data_details:
            print(r)'''

        self.max_count = 0
        for r in self.result_relations:
            row_pos = code_names_list.index(r['c0_name'])
            col_pos = code_names_list.index(r['c1_name'])
            self.data_counts[row_pos][col_pos] += 1
            if self.data_counts[row_pos][col_pos] > self.max_count:
                self.max_count = self.data_counts[row_pos][col_pos]
            print(r['cid0'], r['c0_name'], r['ctid0'], r['cid1'], r['c1_name'], r['ctid1'], r['fid'], r['file_name'],
                  r['owners'], r['c0_pos0'], r['c0_pos1'], r['c1_pos0'], r['c1_pos1'])
            res_list = [r['cid0'], r['c0_name'], r['ctid0'], r['cid1'], r['c1_name'], r['ctid1'], r['fid'], r['file_name'],
                  r['owners'], r['c0_pos0'], r['c0_pos1'], r['c1_pos0'], r['c1_pos1'],
                        r['text_before'], r['text_overlap'], r['text_after']]
            if self.data_details[row_pos][col_pos] == ".":
                self.data_details[row_pos][col_pos] = [res_list]
            else:
                self.data_details[row_pos][col_pos].append(res_list)

        # Color heat map for spread across 5 colours
        colors = ["#F8E0E0", "#F6CECE", "#F5A9A9", "#F78181", "#FA5858"]  # light to dark red
        for row, row_data in enumerate(self.data_counts):
            for col, item_data in enumerate(row_data):
                if self.data_counts[row][col] > 0:
                    color_range_index = int(self.data_counts[row][col] / self.max_count * 5) - 1
                    if color_range_index < 0:
                        color_range_index = 0
                    self.data_colors[row][col] = colors[color_range_index]

        '''print("Summary counts table")
        for d in self.data_counts:
            print(d)
        print("Summary data details table")
        for d in self.data_details:
            print(d)'''

        self.fill_table()

    def export_to_excel(self):
        """ Export to Excel file. """

        filename = "Code_cooccurrence.xlsx"
        export_dir = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_dir.filepath
        if filepath is None:
            return


        # Excel headers
        header = []
        for code_ in self.codes:
            name_split_50 = [code_['name'][y - 50:y] for y in range(50, len(code_['name']) + 50, 50)]
            # header_labels.append(code_['name'])  # OLD, need line separators
            header.append("\n".join(name_split_50))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Counts"
        wb.create_sheet("Details")
        ws2 = wb["Details"]

        for col, col_name in enumerate(header):
            h_cell = ws.cell(row=1, column=col + 2)
            h_cell.value = col_name
            h_cell2 = ws2.cell(row=1, column=col + 2)
            h_cell2.value = col_name
            v_cell = ws.cell(row=col + 2, column=1)
            v_cell.value = col_name
            v_cell2 = ws2.cell(row=col + 2, column=1)
            v_cell2.value = col_name
        # Co-occurrence counts
        for row, row_data in enumerate(self.data_counts):
            for col, col_data in enumerate(row_data):
                cell = ws.cell(row=row + 2, column=col + 2)
                cell.value = col_data
        # Details
        # TODO
        v_cell = ws2.cell(row=2, column=2)
        v_cell.value = "TEST"

        wb.save(filepath)
        msg = _('Co-occurrence exported: ') + filepath
        Message(self.app, _('Co-occurrence exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def show_or_hide_empty_rows_and_cols(self):
        """ Unchecked - show all rows and columns.
        Checked - hide rows and columns with no code co-occurrences. """

        if self.ui.checkBox_hide_blanks.isChecked():
            for row, row_data in enumerate(self.data_counts):
                print(row_data)
                if sum(row_data) == 0:
                    self.ui.tableWidget.hideRow(row)

            for col in range(len(self.data_counts)):
                col_sum = 0
                for row, row_data in enumerate(self.data_counts):
                    col_sum += row_data[col]
                if col_sum == 0:
                    self.ui.tableWidget.hideColumn(col)

        if not self.ui.checkBox_hide_blanks.isChecked():
            for row, row_data in enumerate(self.data_counts):
                self.ui.tableWidget.showRow(row)
                self.ui.tableWidget.showColumn(row)


    def cell_selected(self):
        """ When the table widget memo cell is selected display the memo.
        Update memo text, or delete memo by clearing text.
        If a new memo, also show in table widget by displaying MEMO in the memo column. """

        row = self.ui.tableWidget.currentRow()
        col = self.ui.tableWidget.currentColumn()
        text = self.ui.tableWidget.item(row, col).text()
        if text == "":
            return
        self.ui.textEdit.clear()
        data_list = self.data_details[row][col]
        '''
        0 - 5 [r['cid0'], r['c0_name'], r['ctid0'], r['cid1'], r['c1_name'], r['ctid1'], 
        6 - 8 r['fid'], r['file_name'], r['owners'], 
        9 - 12 r['c0_pos0'], r['c0_pos1'], r['c1_pos0'], r['c1_pos1'],
        13 - 15 r['text_before'], r['text_overlap'], r['text_after']]
        '''
        for data in data_list:
            print(data)
            msg = f"Codes: {data[1]} ({data[9]} - {data[10]})| {data[4]} ({data[11]} - {data[12]})\n"
            msg += f"Coders: {data[8]}. (ctid0: {data[2]} | ctid1: {data[5]})\n"
            msg += f"File (fid {data[6]}): {data[7]}\n\n"
            self.ui.textEdit.append(msg)

            start_pos = len(self.ui.textEdit.toPlainText())
            msg = f"{data[13]}{data[14]}{data[15]}\n"
            self.ui.textEdit.append(msg)
            start_pos += len(data[13])
            end_pos = start_pos + len(data[14])
            cursor = self.ui.textEdit.textCursor()
            fmt = QtGui.QTextCharFormat()
            fmt.setUnderlineStyle(QtGui.QTextCharFormat.UnderlineStyle.SingleUnderline)
            if self.app.settings['stylesheet'] == 'dark':
                fmt.setUnderlineColor(QtGui.QColor("#FF0000"))
            else:
                fmt.setUnderlineColor(QtGui.QColor("#FF00000"))
            cursor.setPosition(start_pos, QtGui.QTextCursor.MoveMode.MoveAnchor)
            cursor.setPosition(end_pos, QtGui.QTextCursor.MoveMode.KeepAnchor)
            cursor.mergeCharFormat(fmt)

            msg = "========\n"
            self.ui.textEdit.append(msg)

        print("SP", self.ui.splitter.sizes())
        self.ui.splitter.setSizes([300, 200])



    '''def table_menu(self, position):
        """ Context menu for displaying table cell coding details.
        """

        row = self.ui.tableWidget.currentRow()
        col = self.ui.tableWidget.currentColumn()
        text = self.ui.tableWidget.item(row, col).text()
        print(row, col, text)

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_view = menu.addAction(_("View"))'''



    def fill_table(self):
        """ Fill table using code names alphabetically (case insensitive), using self.data """

        rows = self.ui.tableWidget.rowCount()
        for r in range(0, rows):
            self.ui.tableWidget.removeRow(0)

        header_labels = []
        for code_ in self.codes:
            name_split_50 = [code_['name'][y - 50:y] for y in range(50, len(code_['name']) + 50, 50)]
            # header_labels.append(code_['name'])  # OLD, needed line separators
            header_labels.append("\n".join(name_split_50))
        self.ui.tableWidget.setColumnCount(len(header_labels))
        self.ui.tableWidget.setHorizontalHeaderLabels(header_labels)
        self.ui.tableWidget.setRowCount(len(header_labels))
        self.ui.tableWidget.setVerticalHeaderLabels(header_labels)
        for row, row_data in enumerate(self.data_counts):
            for col, cell_data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem()
                if self.data_colors[row][col] != "":
                    item.setBackground(QtGui.QBrush(QtGui.QColor(self.data_colors[row][col])))
                    item.setForeground(QtGui.QBrush(QtGui.QColor("#000000")))
                if cell_data > 0:
                    item.setData(QtCore.Qt.ItemDataRole.DisplayRole, cell_data)
                item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
                self.ui.tableWidget.setItem(row, col, item)
        # self.ui.tableWidget.resizeColumnsToContents()  # Doesnt look great
        self.ui.tableWidget.resizeRowsToContents()



    def calculate_relations(self, code_ids_str):
        """ Calculate the relations for selected codes for all coders.
        For codings in code_text only.

        id1, id2, overlapindex, unionindex, distance, whichmin, whichmax, fid
        relation is 1 character: Inclusion, Overlap, Exact, (Proximity - not used)
        owners is a combination of: owner of ctid0 pipe owner of ctid1
        """

        selected_relations = ['E', 'I', 'O']
        file_ids_names = self.app.get_text_filenames()

        # Get codings for each selected text file
        cur = self.app.conn.cursor()
        for fid in file_ids_names:
            sql = "select fid, code_text.cid, pos0, pos1, name, ctid,seltext, ifnull(code_text.memo,''), code_text.owner " \
                  "from code_text " \
                  "join code_name on code_name.cid=code_text.cid where fid=? " \
                  "and code_text.cid in (" + code_ids_str + ") order by code_text.cid"
            cur.execute(sql, [fid['id']])
            result = cur.fetchall()
            coded = [row for row in result if row[0] == fid['id']]

            # Look at each code again other codes, when done remove from list of codes
            cid = 1
            pos0 = 2
            pos1 = 3
            name = 4
            ctid = 5
            seltext = 6
            coded_memo = 7
            owner = 8
            while len(coded) > 0:
                c0 = coded.pop()
                for c1 in coded:
                    if c0[cid] != c1[cid]:
                        relation = self.relation(c0, c1)
                        # Add extra details for output
                        relation['c0_name'] = c0[name]
                        relation['c1_name'] = c1[name]
                        relation['fid'] = fid['id']
                        relation['file_name'] = fid['name']
                        relation['c0_pos0'] = c0[pos0]
                        relation['c0_pos1'] = c0[pos1]
                        relation['c1_pos0'] = c1[pos0]
                        relation['c1_pos1'] = c1[pos1]
                        relation['owners'] = f"{c0[owner]}|{c1[owner]}"
                        relation['ctid0'] = c0[ctid]
                        relation['ctid0_text'] = c0[seltext]
                        relation['ctid1'] = c1[ctid]
                        relation['ctid1_text'] = c1[seltext]
                        relation['coded_memo0'] = c0[coded_memo]
                        relation['coded_memo1'] = c1[coded_memo]
                        # Append relation based on comboBox selection
                        if relation['relation'] in selected_relations:
                            self.result_relations.append(relation)

    def relation(self, c0, c1):
        """ Relation function as in RQDA

        whichmin is the code with the lowest pos0, or None if equal
        whichmax is the code with the highest pos1 or None if equal
        operlapindex is the combined lowest to the highest positions. Only used for E, O, P
        unionindex is the lowest and highest positions of the union of overlap. Only used for E, O

        Called by:
            calculate_relations_for_coder_and_selected_codes

        Returns:
        id1, id2, overlapindex, unionindex, distance, whichmin, min, whichmax, max, fid
        relation is 1 character: Inclusion, Overlap, Exact, Proximity
        actual text as before, overlap, after
        """

        # fid = 0
        cid = 1
        pos0 = 2
        pos1 = 3
        result = {"cid0": c0[cid], "cid1": c1[cid], "relation": "", "whichmin": None, "min": 0,
                  "whichmax": None, "max": 0, "overlapindex": None, "unionindex": None, "distance": None,
                  "text_before": "", "text_overlap": "", "text_after": ""}

        cur = self.app.conn.cursor()

        # Which min
        if c0[pos0] < c1[pos0]:
            result['whichmin'] = c0[cid]
            result['min'] = c0[pos0]
        if c1[pos0] < c0[pos0]:
            result['whichmin'] = c1[cid]
            result['min'] = c1[pos0]

        # Which max
        if c0[pos1] > c1[pos1]:
            result['whichmax'] = c0[cid]
            result['max'] = c0[pos1]
        if c1[pos1] > c0[pos1]:
            result['whichmax'] = c1[cid]
            result['max'] = c1[pos1]

        # Check for Exact
        if c0[pos0] == c1[pos0] and c0[pos1] == c1[pos1]:
            result['relation'] = "E"
            result['overlapindex'] = [c0[pos0], c0[pos1]]
            result['unionindex'] = [c0[pos0], c0[pos1]]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c0[pos1] - c0[pos0], c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
                result['text_before'] = ""
                result['text_after'] = ""
                result['distance'] = 0
            return result

        # Check for Proximity
        if c0[pos1] < c1[pos0]:
            result['relation'] = "P"
            result['distance'] = c1[pos0] - c0[pos1]
            result['text_overlap'] = ""
            result['text_before'] = ""
            result['text_after'] = ""
            return result
        if c0[pos0] > c1[pos1]:
            result['relation'] = "P"
            result['distance'] = c0[pos0] - c1[pos1]
            result['text_overlap'] = ""
            result['text_before'] = ""
            result['text_after'] = ""
            return result

        # Check for Inclusion
        # Exact has been resolved above
        # c0 inside c1
        if c0[pos0] >= c1[pos0] and c0[pos1] <= c1[pos1]:
            result['relation'] = "I"
            result['overlapindex'] = [c0[pos0], c0[pos1]]
            result['unionindex'] = [c0[pos0], c0[pos1]]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c0[pos1] - c0[pos0], c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, c0[pos0] - c1[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos1] + 1, c1[pos1] - c0[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

        # c1 inside c0
        if c1[pos0] >= c0[pos0] and c1[pos1] <= c0[pos1]:
            result['relation'] = "I"
            result['overlapindex'] = [c1[pos0], c1[pos1]]
            result['unionindex'] = [c1[pos0], c1[pos1]]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, c1[pos1] - c1[pos0], c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c1[pos0] - c0[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos1] + 1, c0[pos1] - c1[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

        # Check for Overlap
        # Should be all that is remaining
        # c0 overlaps on the right side, left side is not overlapping
        if c0[pos0] < c1[pos0] and c0[pos1] < c1[pos1]:
            '''print("c0 overlaps on the right side, left side is not overlapping")
            print("c0", c0)
            print("C1", c1)'''
            result['relation'] = "O"
            # Reorder lowest to highest
            result['overlapindex'] = sorted([c0[pos0], c1[pos1]])
            result['unionindex'] = sorted([c0[pos1], c1[pos0]])
            overlap_length = result['unionindex'][1] - result['unionindex'][0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, overlap_length, c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, c1[pos0] - c0[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos1] + 1, c1[pos1] - c0[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result

        # c1 overlaps on the right side, left side is not overlapping
        if c1[pos0] < c0[pos0] and c1[pos1] < c0[pos1]:
            result['relation'] = "O"
            result['overlapindex'] = sorted([c1[pos0], c0[pos1]])
            result['unionindex'] = sorted([c1[pos1], c0[pos0]])
            overlap_length = result['unionindex'][1] - result['unionindex'][0]
            '''print("TODO c1 overlaps on the right, the left side is not overlapping")
            print("C0", c0)
            print("C1", c1)'''
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c0[pos0] + 1, overlap_length, c0[0]])
            txt = cur.fetchone()
            if txt is not None:
                result['text_overlap'] = txt[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos0] + 1, c0[pos0] - c1[pos0], c0[0]])
            txt_before = cur.fetchone()
            if txt_before is not None:
                result['text_before'] = txt_before[0]
            cur.execute("select substr(fulltext,?,?) from source where source.id=?",
                        [c1[pos1] + 1, c0[pos1] - c1[pos1], c0[0]])
            txt_after = cur.fetchone()
            if txt_after is not None:
                result['text_after'] = txt_after[0]
            result['distance'] = 0
            return result