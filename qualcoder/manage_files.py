# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com/
"""

import csv
import datetime
import sqlite3
import ebooklib
from ebooklib import epub
import openpyxl
import PIL
from PIL import Image
import qtawesome as qta
from typing import Iterable, Any
from shutil import copyfile, move
from urllib.parse import urlparse
import webbrowser
import zipfile

from PyQt6 import QtCore, QtGui, QtWidgets
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams, LTTextLine
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage

from .GUI.ui_dialog_manage_files import Ui_Dialog_manage_files
from .add_attribute import DialogAddAttribute
from .add_item_name import DialogAddItemName
from .code_text import DialogCodeText  # for isinstance()
from .confirm_delete import DialogConfirmDelete
from .docx import opendocx, getdocumenttext
from .edit_textfile import DialogEditTextFile
from .helpers import ExportDirectoryPathDialog, Message, msecs_to_hours_mins_secs
from .html_parser import *
from .memo import DialogMemo
from .report_codes import DialogReportCodes  # for isInstance()
from .ris import Ris
from .select_items import DialogSelectItems
from .view_av import DialogViewAV, DialogCodeAV  # for isinstance update files
from .view_image import DialogViewImage, DialogCodeImage  # for isinstance update files
from .code_pdf import DialogCodePdf  # for isinstance update files

# If VLC not installed, it will not crash
vlc = None
try:
    import vlc
except Exception as e:
    print(e)

path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogManageFiles(QtWidgets.QDialog):
    """ View, import, export, rename and delete text files.
    Files are normally imported into the qda project folder.
    Option to link to external files.
    """

    NAME_COLUMN = 0
    MEMO_COLUMN = 1
    DATE_COLUMN = 2
    ID_COLUMN = 3
    CASE_COLUMN = 4
    ATTRIBUTE_START_COLUMN = 5

    def __init__(self, app, parent_text_edit, tab_coding, tab_reports):

        self.app = app
        self.parent_text_edit = parent_text_edit
        self.tab_coding = tab_coding  # Tab widget coding for updates
        self.tab_reports = tab_reports  # Tab widget reports for updates
        self.rows_hidden = False
        self.source = []  # Dictionaries of source files
        self.header_labels = []
        self.default_import_directory = os.path.expanduser("~")
        self.attribute_names = []  # list of dictionary name:value for AddAtribute dialog
        self.attribute_labels_ordered = []  # helps with filling table data
        self.files_renamed = []  # list of dictionaries of old and new names and fid
        self.pdf_page_text = ""  # Used when loading pdf text
        self.clipboard_text = ""  # Used to copy text into another cell
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_manage_files()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        self.default_import_directory = self.app.settings['directory']
        self.attribute_labels_ordered = []
        self.av_dialog_open = None  # Used for opened AV dialog
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        self.ui.pushButton_create.setIcon(qta.icon('mdi6.pencil-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_create.clicked.connect(self.create_text_file)
        self.ui.pushButton_view.setIcon(qta.icon('mdi6.magnify', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_view.clicked.connect(self.view)
        self.ui.pushButton_delete.setIcon(qta.icon('mdi6.delete-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_delete.clicked.connect(self.delete_button_multiple_files)
        self.ui.pushButton_import.setIcon(qta.icon('mdi6.file-document-plus-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_import.clicked.connect(self.import_files)
        self.ui.pushButton_link.setIcon(qta.icon('mdi6.link-variant', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_link.clicked.connect(self.link_files)
        self.ui.pushButton_import_from_linked.setIcon(qta.icon('mdi6.link-variant-minus', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_import_from_linked.clicked.connect(self.button_import_linked_file)
        self.ui.pushButton_export_to_linked.setIcon(qta.icon('mdi6.link-variant-plus', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export_to_linked.clicked.connect(self.button_export_file_as_linked_file)
        self.ui.pushButton_export.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export.clicked.connect(self.export)
        self.ui.pushButton_add_attribute.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_add_attribute.clicked.connect(self.add_attribute)
        self.ui.pushButton_export_attributes.setIcon(qta.icon('mdi6.file-export-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export_attributes.clicked.connect(self.export_attributes)
        self.ui.pushButton_undo.setIcon(qta.icon('mdi6.undo', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_undo.clicked.connect(self.undo_file_rename)
        self.ui.pushButton_bulk_rename.setIcon(qta.icon('mdi6.file-multiple-outline', options=[{'scale_factor': 1.2}]))
        self.ui.pushButton_bulk_rename.clicked.connect(self.bulk_rename_database_entry)
        self.ui.pushButton_help.setIcon(qta.icon('mdi6.help'))
        self.ui.pushButton_help.pressed.connect(self.help)
        self.ui.tableWidget.setTabKeyNavigation(False)
        self.ui.tableWidget.itemChanged.connect(self.cell_modified)
        self.ui.tableWidget.cellClicked.connect(self.cell_selected)
        self.ui.tableWidget.cellDoubleClicked.connect(self.cell_double_clicked)
        self.ui.tableWidget.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget.customContextMenuRequested.connect(self.table_menu)
        self.ui.tableWidget.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        self.ui.tableWidget.installEventFilter(self)
        self.ui.tableWidget.horizontalHeader().setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget.horizontalHeader().customContextMenuRequested.connect(self.table_header_menu)
        self.ui.tableWidget.horizontalHeader().setToolTip(_("Right click header row to hide columns"))
        self.load_file_data()

    @staticmethod
    def help():
        """ Open help for transcribe section in browser. """

        url = "https://github.com/ccbogel/QualCoder/wiki/3.2.-Files"
        webbrowser.open(url)

    def keyPressEvent(self, event):
        """ Used to activate buttons.
        Ctrl 0 to 9
        """
        key = event.key()
        mods = QtWidgets.QApplication.keyboardModifiers()
        # Ctrl 0 to 4
        if mods & QtCore.Qt.KeyboardModifier.ControlModifier:
            if key == QtCore.Qt.Key.Key_1:
                self.view()
                return
            if key == QtCore.Qt.Key.Key_2:
                self.import_files()
                return
            if key == QtCore.Qt.Key.Key_3:
                self.link_files()
                return
            if key == QtCore.Qt.Key.Key_4:
                self.create_text_file()
                return
            if key == QtCore.Qt.Key.Key_5:
                self.button_import_linked_file()
                return
            if key == QtCore.Qt.Key.Key_6:
                self.button_export_file_as_linked_file()
                return
            if key == QtCore.Qt.Key.Key_7:
                self.add_attribute()
                return
            if key == QtCore.Qt.Key.Key_8:
                self.export_attributes()
                return
            if key == QtCore.Qt.Key.Key_9:
                self.export()
                return
            if key == QtCore.Qt.Key.Key_0:
                self.help()
                return
            if key == QtCore.Qt.Key.Key_C:
                x = self.ui.tableWidget.currentRow()
                y = self.ui.tableWidget.currentColumn()
                self.clipboard_text = self.ui.tableWidget.item(x, y).text()
                return
            if key == QtCore.Qt.Key.Key_V:
                x = self.ui.tableWidget.currentRow()
                y = self.ui.tableWidget.currentColumn()
                self.ui.tableWidget.item(x, y).setText(self.clipboard_text)
                return

    def eventFilter(self, object_, event):
        """ Using this event filter to
        Ctrl + A to show all rows
        Ctrl + Z Undo the last  deletion.
        """

        if type(event) == QtGui.QKeyEvent:
            key = event.key()
            mod = event.modifiers()
            if key == QtCore.Qt.Key.Key_A and mod == QtCore.Qt.KeyboardModifier.ControlModifier:
                for r in range(0, self.ui.tableWidget.rowCount()):
                    self.ui.tableWidget.setRowHidden(r, False)
                self.rows_hidden = False
                return True
        return False

    def table_header_menu(self, position):
        """ Used to show and hide columns """

        index_at = self.ui.tableWidget.indexAt(position)
        header_index = int(index_at.column())
        menu = QtWidgets.QMenu(self)
        action_show_all_columns = menu.addAction(_("Show all columns"))
        action_hide_column = None
        if header_index > 0:
            action_hide_column = menu.addAction(_("Hide column"))
        action_hide_columns_starting = menu.addAction(_("Hide columns starting with"))
        action_show_columns_starting = menu.addAction(_("Show columns starting with"))
        action = menu.exec(self.ui.tableWidget.mapToGlobal(position))
        if action == action_show_all_columns:
            for c in range(0, self.ui.tableWidget.columnCount()):
                self.ui.tableWidget.setColumnHidden(c, False)
            if not self.app.settings['showids']:
                self.ui.tableWidget.setColumnHidden(self.ID_COLUMN, True)
            return
        if action == action_hide_column:
            self.ui.tableWidget.setColumnHidden(header_index, True)
            return
        if action == action_hide_columns_starting:
            msg = _("Hide columns starting with:")
            hide_filter, ok = QtWidgets.QInputDialog.getText(self, _("Hide Columns"), msg,
                                                             QtWidgets.QLineEdit.EchoMode.Normal)
            for c in range(1, self.ui.tableWidget.columnCount()):
                h_text = self.ui.tableWidget.horizontalHeaderItem(c).text()
                if len(h_text) >= len(hide_filter) and hide_filter == h_text[:len(hide_filter)]:
                    self.ui.tableWidget.setColumnHidden(c, True)
        if action == action_show_columns_starting:
            msg = _("Show columns starting with:")
            show_filter, ok = QtWidgets.QInputDialog.getText(self, _("Show Columns"), msg,
                                                             QtWidgets.QLineEdit.EchoMode.Normal)
            for c in range(4, self.ui.tableWidget.columnCount()):
                h_text = self.ui.tableWidget.horizontalHeaderItem(c).text()
                if len(h_text) >= len(show_filter) and show_filter == h_text[:len(show_filter)]:
                    self.ui.tableWidget.setColumnHidden(c, False)
                else:
                    self.ui.tableWidget.setColumnHidden(c, True)

    def table_menu(self, position):
        """ Context menu for displaying table rows in differing order,
        hiding table rows, assigning case to file, file rename, export import from linked. """

        row = self.ui.tableWidget.currentRow()
        col = self.ui.tableWidget.currentColumn()
        item_text = self.ui.tableWidget.item(row, col).text()
        # Use these next few lines to use for moving a linked file into or an internal file out of the project folder
        mediapath = None
        risid = None
        try:
            id_ = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())
        except AttributeError:
            # Occurs if a table cell is not clicked, but click occurs elsewhere in container
            return
        for s in self.source:
            if s['id'] == id_:
                mediapath = s['mediapath']
                risid = s['risid']
        # Action cannot be None otherwise may default to one of the actions below depending on column clicked
        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_view = menu.addAction(_("View"))
        action_view_original_text = None
        if mediapath is not None and len(mediapath) > 6 and (mediapath[:6] == '/docs/' or mediapath[:5] == 'docs:'):
            action_view_original_text = menu.addAction(_("view original text file"))
        action_filename_asc = None
        action_filename_desc = None
        action_type = None
        if col == self.NAME_COLUMN:
            action_filename_asc = menu.addAction(_("Order ascending"))
            action_filename_desc = menu.addAction(_("Order descending"))
            action_type = menu.addAction(_("File type order"))
        action_date_asc = None
        if col == self.DATE_COLUMN:
            action_date_asc = menu.addAction(_("Order ascending"))
        action_casename_asc = None
        action_assign_case = None
        if col == self.CASE_COLUMN:
            action_casename_asc = menu.addAction(_("Order ascending"))
            action_assign_case = menu.addAction(_("Assign case to file"))
        action_show_values_like = None
        action_hide_values_like = None
        if col != self.MEMO_COLUMN:
            action_show_values_like = menu.addAction(_("Show values like"))
            action_hide_values_like = menu.addAction(_("Hide values like"))
        action_equals_value = menu.addAction(_("Show this value"))
        action_order_by_value_asc = None
        action_order_by_value_desc = None
        action_date_picker = None
        action_ref_apa = None
        action_ref_vancouver = None
        if col > self.CASE_COLUMN:
            action_order_by_value_asc = menu.addAction(_("Order ascending"))
            action_order_by_value_desc = menu.addAction(_("Order descending"))
            if "date" in self.header_labels[col].lower():
                # Check that a character date can be entered
                cur = self.app.conn.cursor()
                cur.execute("select valuetype from attribute_type where caseOrFile='file' and name=?",
                            [self.header_labels[col], ])
                result = cur.fetchone()
                if result is not None and result[0] == "character":
                    action_date_picker = menu.addAction(_("Enter date"))
            if self.header_labels[col] in ("Ref_Authors", "Ref_Title", "Ref_Journal", "Ref_Type", "Ref_Year"):
                action_ref_apa = menu.addAction(_("Copy reference to clipboard. APA"))
                action_ref_vancouver = menu.addAction(_("Copy reference to clipboard. Vancouver"))
        action_rename = None
        action_export = None
        action_delete = None
        action_export_to_linked = None
        action_import_linked = None
        if col == self.NAME_COLUMN:
            action_rename = menu.addAction(_("Rename database entry"))
            action_export = menu.addAction(_("Export"))
            action_delete = menu.addAction(_("Delete"))
            if mediapath is None or mediapath == "" or (mediapath is not None and mediapath[0] == "/"):
                action_export_to_linked = menu.addAction(_("Move file to externally linked file"))
            if mediapath is not None and mediapath != "" and mediapath[0] != "/":
                action_import_linked = menu.addAction(_("Import linked file"))
        action_show_all = None
        if self.rows_hidden:
            action_show_all = menu.addAction(_("Show all rows Ctrl A"))
        action_url = None
        url_test = urlparse(item_text)
        if all([url_test.scheme, url_test.netloc]):
            action_url = menu.addAction(_("Open URL"))
        action = menu.exec(self.ui.tableWidget.mapToGlobal(position))
        if action is None:
            return
        if action == action_view:
            self.view()
            return
        if action == action_view_original_text:
            self.view_original_text_file(mediapath)
            return
        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
            return
        if action == action_import_linked:
            self.import_linked_file(id_, mediapath)
        if action == action_export_to_linked:
            self.export_file_as_linked_file(id_, mediapath)
        if action == action_export:
            self.export()
        if action == action_delete:
            self.delete()
        if action == action_rename:
            self.rename_database_entry()
        if action == action_assign_case:
            self.assign_case_to_file()
        if action == action_filename_asc:
            self.load_file_data()
        if action == action_filename_desc:
            self.load_file_data("filename desc")
        if action == action_date_asc:
            self.load_file_data("date")
        if action == action_type:
            self.load_file_data("filetype")
        if action == action_casename_asc:
            self.load_file_data("casename")
        if action == action_order_by_value_asc:
            self.load_file_data("attribute asc:" + self.header_labels[col])
        if action == action_order_by_value_desc:
            self.load_file_data("attribute desc:" + self.header_labels[col])
        if action == action_equals_value:
            # Hide rows that do not match this value
            item_to_compare = self.ui.tableWidget.item(row, col)
            compare_text = item_to_compare.text()
            for r in range(0, self.ui.tableWidget.rowCount()):
                item = self.ui.tableWidget.item(r, col)
                text_ = item.text()
                if compare_text != text_:
                    self.ui.tableWidget.setRowHidden(r, True)
            self.rows_hidden = True
            rows_showing = 0
            for r in range(self.ui.tableWidget.rowCount()):
                if not self.ui.tableWidget.isRowHidden(r):
                    rows_showing += 1
            self.ui.label_fcount.setText(f"Files: {rows_showing} / {self.ui.tableWidget.rowCount()}")
            return
        if action == action_show_values_like:
            text_value, ok = QtWidgets.QInputDialog.getText(self, _("Text filter"), _("Show values like:"),
                                                            QtWidgets.QLineEdit.EchoMode.Normal)
            self.rows_hidden = True
            if ok and text_value != '':
                for r in range(0, self.ui.tableWidget.rowCount()):
                    if self.ui.tableWidget.item(r, col).text().find(text_value) == -1:
                        self.ui.tableWidget.setRowHidden(r, True)
            rows_showing = 0
            for r in range(self.ui.tableWidget.rowCount()):
                if not self.ui.tableWidget.isRowHidden(r):
                    rows_showing += 1
            self.ui.label_fcount.setText(f"Files: {rows_showing} / {self.ui.tableWidget.rowCount()}")
            return
        if action == action_hide_values_like:
            text_value, ok = QtWidgets.QInputDialog.getText(self, _("Text filter"), _("Hide values like:"),
                                                            QtWidgets.QLineEdit.EchoMode.Normal)
            self.rows_hidden = True
            if ok and text_value != '':
                for r in range(0, self.ui.tableWidget.rowCount()):
                    if self.ui.tableWidget.item(r, col).text().find(text_value) != -1:
                        self.ui.tableWidget.setRowHidden(r, True)
            rows_showing = 0
            for r in range(self.ui.tableWidget.rowCount()):
                if not self.ui.tableWidget.isRowHidden(r):
                    rows_showing += 1
            self.ui.label_fcount.setText(f"Files: {rows_showing} / {self.ui.tableWidget.rowCount()}")
            return
        if action == action_show_all:
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
            self.rows_hidden = False
            self.ui.label_fcount.setText(f"Files: {self.ui.tableWidget.rowCount()}")
            return
        if action == action_url:
            webbrowser.open(item_text)
            return
        if action == action_date_picker:
            ui_memo = DialogMemo(self.app, "Date selector", "", "hide")
            ui_memo.ui.textEdit.hide()
            calendar = QtWidgets.QCalendarWidget()
            ui_memo.ui.gridLayout.addWidget(calendar, 0, 0, 1, 1)
            ok = ui_memo.exec()
            if ok:
                selected_date = calendar.selectedDate().toString("yyyy-MM-dd")
                self.ui.tableWidget.setItem(row, col, QtWidgets.QTableWidgetItem(selected_date))
            return
        if action == action_ref_apa:
            ris_obj = Ris(self.app)
            ris_obj.get_references(selected_ris=risid)
            apa = ris_obj.refs
            if not apa:
                return
            cb = QtWidgets.QApplication.clipboard()
            cb.setText(apa[0]['apa'].replace("\n", " "))
        if action == action_ref_vancouver:
            ris_obj = Ris(self.app)
            ris_obj.get_references(selected_ris=risid)
            vancouver = ris_obj.refs
            if not vancouver:
                return
            cb = QtWidgets.QApplication.clipboard()
            cb.setText(vancouver[0]['vancouver'].replace("\n", " "))

    def view_original_text_file(self, mediapath):
        """ View original text file.
         param:
         mediapath: String '/docs/' for internal 'docs:/' for external """

        if mediapath[:6] == "/docs/":
            media_path = self.app.project_path + "/documents/" + mediapath[6:]
            webbrowser.open(media_path)
            return
        if mediapath[:5] == "docs:":
            media_path = mediapath[5:]
            webbrowser.open(media_path)
            return
        logger.error("Cannot open text file in browser " + mediapath)
        print(f"manage_files.view_original_text_file. Cannot open text file in browser {mediapath}")

    def assign_case_to_file(self):
        """ Assign one or more cases to file. """

        row = self.ui.tableWidget.currentRow()
        fid = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())
        casenames = self.app.get_casenames()
        ui = DialogSelectItems(self.app, casenames, _("Delete files"), "multi")
        ok = ui.exec()
        if not ok:
            return
        selection = ui.get_selected()
        if not selection:
            return
        cur = self.app.conn.cursor()
        cur.execute("select fulltext from source where id=?", [fid])
        res = cur.fetchone()
        len_text = 0
        if res is not None and res[0] is not None:
            len_text = len(res[0])
        for case_ in selection:
            # Check if already linked file to case
            cur.execute("select * from case_text where caseid = ? and fid=? and pos0=? and pos1=?",
                        (case_['id'], fid, 0, len_text))
            result = cur.fetchall()
            if len(result) == 0:
                sql = "insert into case_text (caseid, fid, pos0, pos1, owner, date, memo) values(?,?,?,?,?,?,?)"
                cur.execute(sql, (case_['id'], fid, 0, len_text, self.app.settings['codername'],
                                  datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"), ""))
                self.app.conn.commit()
        # Visual feedback
        cases_text = self.get_cases_by_filename(self.ui.tableWidget.item(row, self.NAME_COLUMN).text())
        self.ui.tableWidget.item(row, self.CASE_COLUMN).setText(cases_text)

    def rename_database_entry(self):
        """ Rename the database entry of the file. """

        row = self.ui.tableWidget.currentRow()
        if row == -1:
            return
        existing_name = self.ui.tableWidget.item(row, self.NAME_COLUMN).text()
        filenames = []
        for s in self.source:
            filenames.append({'name': s['name']})
        ui = DialogAddItemName(self.app, filenames, _("Rename database entry"), existing_name)
        ui.ui.lineEdit.setText(existing_name)
        ui.exec()
        new_name = ui.get_new_name()
        if new_name is None:
            return
        cur = self.app.conn.cursor()
        cur.execute("update source set name=? where name=?", [new_name, existing_name])
        self.app.conn.commit()
        self.parent_text_edit.append(_("Renamed database file entry: ") + f"{existing_name} -> {new_name}")
        entry = {'old_name': existing_name, 'name': new_name,
                 'fid': int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())}
        self.files_renamed.append(entry)
        self.ui.pushButton_undo.setEnabled(True)
        self.load_file_data()
        self.app.delete_backup = False
        self.update_files_in_dialogs()
        # update doc in vectorstore
        id = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())
        if self.app.settings['ai_enable'] == 'True':
            docs = self.app.get_file_texts(file_ids=[id])
            self.app.ai.sources_vectorstore.import_document(docs[0]['id'], docs[0]['name'], docs[0]['fulltext'], True)

    def undo_file_rename(self):
        """ Undo file name rename. """

        if len(self.files_renamed) == 0:
            self.ui.pushButton_undo.setEnabled(False)
            # Could occur when file deleted
            return
        ui = DialogSelectItems(self.app, self.files_renamed, _("Undo file rename"), "single")
        ok = ui.exec()
        if not ok:
            return
        selection = ui.get_selected()
        if not selection:
            return
        filenames = self.app.get_filenames()
        for f in filenames:
            if f['name'] == selection['old_name']:
                Message(self.app, _("Cannot undo"), _("Another file has this name"), "warning").exec()
                self.files_renamed = [x for x in self.files_renamed if not (selection['fid'] == x.get('fid'))]
                if len(self.files_renamed) == 0:
                    self.ui.pushButton_undo.setEnabled(False)
                return
        cur = self.app.conn.cursor()
        cur.execute("update source set name=? where name=?", [selection['old_name'], selection['name']])
        self.app.conn.commit()
        self.parent_text_edit.append(_("Reversed renamed database file entry: ") +
                                     f"{selection['name']} -> {selection['old_name']}")
        self.load_file_data()
        if self.app.settings['ai_enable'] == 'True':
            self.app.ai.sources_vectorstore.update_vectorstore()
        self.files_renamed = [x for x in self.files_renamed if not (selection['fid'] == x.get('fid'))]
        if len(self.files_renamed) == 0:
            self.ui.pushButton_undo.setEnabled(False)

    def bulk_rename_database_entry(self):
        """ Bulk Rename source name database entries of the selected files. """

        rows = self.ui.tableWidget.rowCount()
        selected_rows = []
        for row in range(0, rows):
            if not self.ui.tableWidget.isRowHidden(row):
                selected_rows.append([int(self.ui.tableWidget.item(row, self.ID_COLUMN).text()),
                                      self.ui.tableWidget.item(row, self.NAME_COLUMN).text()])
        if not selected_rows:
            return
        # Sort selected rows by their id (order of entry) to ensure sequential renaming
        selected_rows.sort()
        # Display the rename dialog and ask for a base name
        additem = DialogAddItemName(self.app, [], _("Bulk Rename of database file name entries"),
                                    "Give a prefix for the names for all the displayed rows.\n"
                                    "e.g. prefix_001, prefix_002 ...")
        additem.ui.lineEdit.setText("prefix")
        ok = additem.exec()
        if not ok:
            return
        prefix_name = additem.get_new_name()
        if not prefix_name:
            return
        # Perform renaming for all visible rows
        err_msg = ""
        msg = ""
        cur = self.app.conn.cursor()
        for index, row in enumerate(selected_rows):
            fid = row[0]
            existing_name = row[1]
            new_name = f"{prefix_name}_{str(index + 1).zfill(3)}"  # Zero-padded to 3 digits
            # Update the database with the new name
            msg = ""
            try:
                cur.execute("update source set name=? where name=?", [new_name, existing_name])
                self.app.conn.commit()
                msg += f'{_("Renamed database file entry:")} {existing_name} -> {new_name}\n'
            except sqlite3.IntegrityError:
                err_msg += f'_("Bulk Rename. Not renamed in use:") {existing_name}\n'

            # Logging and tracking the renamed entry
            entry = {'old_name': existing_name, 'name': new_name, 'fid': fid}
            self.files_renamed.append(entry)
        self.parent_text_edit.append(msg + err_msg)
        # Updating vectorstore
        if self.app.settings['ai_enable'] == 'True':
            self.app.ai.sources_vectorstore.update_vectorstore()

        self.ui.pushButton_undo.setEnabled(True)
        self.load_file_data()
        self.app.delete_backup = False
        self.update_files_in_dialogs()

    def button_export_file_as_linked_file(self):
        """ User presses button to export current row's file.
         Only to work with an exportable file. """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        row = self.ui.tableWidget.currentRow()
        if row == -1:
            return
        mediapath = None
        id_ = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())
        for s in self.source:
            if s['id'] == id_:
                mediapath = s['mediapath']
        if id_ is None or mediapath is None:
            return
        if mediapath is None or (mediapath is not None and mediapath[0] == "/"):
            self.export_file_as_linked_file(id_, mediapath)

    def export_file_as_linked_file(self, id_, mediapath):
        """ Move an internal project file into an external location as a linked file.
        #TODO Do not export text files as linked files. e.g. internally created in database, or
        docx, txt, md, odt files.

        params:
            id_ : the file id, Integer
            mediapath: stored path to media, will be None for text files, or String
        """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        options = QtWidgets.QFileDialog.Option.DontResolveSymlinks | QtWidgets.QFileDialog.Option.ShowDirsOnly
        directory = QtWidgets.QFileDialog.getExistingDirectory(None,
                                                               _("Select directory to save file"),
                                                               self.app.last_export_directory, options)
        if directory == "":
            return
        if directory != self.app.last_export_directory:
            self.app.last_export_directory = directory
        destination = self.app.last_export_directory
        file_directory = ""
        if mediapath is not None and mediapath[:6] == "/docs/":
            mediapath = "/documents/" + mediapath[6:]
        if mediapath is not None:
            file_directory = mediapath.split('/')[1]  # as [0] will be blank
            destination = directory + "/" + mediapath.split('/')[-1]
        if mediapath is None:
            # Some older text files, and QC internally created text Db entries have None as mediapath
            cur = self.app.conn.cursor()
            cur.execute("select name from source where id=?", [id_, ])
            name = cur.fetchone()[0]
            file_directory = "documents"
            mediapath = "/documents/" + name
            destination = os.path.join(directory, name)
        msg = f'{_("Export to")} {destination}\n'
        try:
            move(self.app.project_path + mediapath, destination)
        except Exception as err:
            logger.warning(str(err))
            Message(self.app, _("Cannot export"), _("Cannot export as linked file\n") + str(err), "warning").exec()
            return
        new_mediapath = ""
        if file_directory == "documents":
            new_mediapath = "docs:" + destination
        if file_directory == "images":
            new_mediapath = "images:" + destination
        if file_directory == "audio":
            new_mediapath = "audio:" + destination
        if file_directory == "video":
            new_mediapath = "video:" + destination
        cur = self.app.conn.cursor()
        cur.execute("update source set mediapath=? where id=?", [new_mediapath, id_])
        self.parent_text_edit.append(msg)
        self.app.conn.commit()
        self.update_files_in_dialogs()
        self.load_file_data()
        self.app.delete_backup = False

    def button_import_linked_file(self):
        """ User presses button to import a linked file into the project folder.
        Only to work with an importable file. """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        row = self.ui.tableWidget.currentRow()
        if row == -1:
            return
        mediapath = None
        id_ = int(self.ui.tableWidget.item(row, self.ID_COLUMN).text())
        for s in self.source:
            if s['id'] == id_:
                mediapath = s['mediapath']
        if id_ is None or mediapath is None:
            return
        if mediapath is not None and mediapath[0] != "/":
            self.import_linked_file(id_, mediapath)

    def import_linked_file(self, id_, mediapath):
        """ Import a linked file into the project folder, and change mediapath details. """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        name_split1 = mediapath.split(":")[1]
        filename = name_split1.split('/')[-1]
        if mediapath[0:6] == "audio:":
            copyfile(mediapath[6:], self.app.project_path + "/audio/" + filename)
            mediapath = '/audio/' + filename
        if mediapath[0:6] == "video:":
            copyfile(mediapath[6:], self.app.project_path + "/video/" + filename)
            mediapath = '/video/' + filename
        if mediapath[0:7] == "images:":
            copyfile(mediapath[7:], self.app.project_path + "/images/" + filename)
            mediapath = '/images/' + filename
        # This must be the last if statement as mediapath can be None
        if mediapath[0:5] == "docs:":
            copyfile(mediapath[5:], f"{self.app.project_path}/documents/{filename}")
            mediapath = None
        cur = self.app.conn.cursor()
        cur.execute("update source set mediapath=? where id=?", [mediapath, id_])
        self.app.conn.commit()
        self.update_files_in_dialogs()
        self.load_file_data()
        self.app.delete_backup = False

    def check_attribute_placeholders(self):
        """ Files can be added after attributes are in the project.
         Need to add placeholder attribute values for these, if missing.
         Also,if a file is deleted, check and remove any isolated attribute values. """

        cur = self.app.conn.cursor()
        sql = "select id from source "
        cur.execute(sql)
        sources = cur.fetchall()
        sql = 'select name from attribute_type where caseOrFile ="file"'
        cur.execute(sql)
        attr_types = cur.fetchall()
        insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
        for source in sources:
            for attribute in attr_types:
                sql = "select value from attribute where id=? and name=?"
                cur.execute(sql, (source[0], attribute[0]))
                res = cur.fetchone()
                if res is None:
                    placeholders = [attribute[0], source[0], datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    self.app.settings['codername']]
                    cur.execute(insert_sql, placeholders)
        self.app.conn.commit()

        # Check and delete attribute values where file has been deleted
        attribute_to_del_sql = "SELECT distinct attribute.id FROM  attribute where \
        attribute.id not in (select source.id from source) order by attribute.id asc"
        cur.execute(attribute_to_del_sql)
        res = cur.fetchall()
        for r in res:
            cur.execute("delete from attribute where attr_type='file' and id=?", [r[0], ])
            self.app.conn.commit()

    def export_attributes(self):
        """ Export attributes from table to an Excel file. """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        shortname = self.app.project_name.split(".qda")[0]
        filename = f"{shortname}_file_attributes.xlsx"
        exp_dlg = ExportDirectoryPathDialog(self.app, filename)
        filepath = exp_dlg.filepath
        if filepath is None:
            return
        cols = self.ui.tableWidget.columnCount()
        rows = self.ui.tableWidget.rowCount()
        header = [self.ui.tableWidget.horizontalHeaderItem(i).text() for i in range(0, cols)]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "File Attributes"
        for col, col_name in enumerate(header):
            h_cell = ws.cell(row=1, column=col + 1)
            h_cell.value = col_name
        for row in range(rows):
            for col in range(cols):
                cell = ws.cell(row=row + 2, column=col + 1)
                data = ""
                try:
                    data = self.ui.tableWidget.item(row, col).text()
                except AttributeError:
                    pass
                cell.value = data
        wb.save(filepath)
        msg = _("File attributes exported to: ") + filepath
        Message(self.app, _('File Export'), msg).exec()
        self.parent_text_edit.append(msg)

    def load_file_data(self, order_by=""):
        """ Documents images and audio contain the filetype suffix.
        No suffix implies the 'file' was imported from a survey question or created internally.
        This also fills out the table header labels with file attribute names.
        Db versions < 5: Files with the '.transcribed' suffix mean they are associated with audio and
        video files.
        Db version 5+: av_text_id links the text file to the audio/video
        Obtain some file metadata to use in table tooltip.
        Fills table after data is loaded.
        param:
            order_by: string ""= name asc, "filename desc" = name desc,
            "date" = date, "filetype" = mediapath,
                "casename" = by alphabetic casename
                "attribute:attribute name" selected attribute - ascending
                "attribute desc: attribute name ttribute - descending
        """

        # check a placeholder attribute is present for the file, add if missing
        self.check_attribute_placeholders()
        self.source = []
        cur = self.app.conn.cursor()
        placeholders = None
        # Default alphabetic order
        sql = "select name, id, fulltext, mediapath, ifnull(memo,''), owner, date, av_text_id, risid from source " \
              "order by upper(name)"
        if order_by == "filename desc":
            sql += " desc"
        if order_by == "date":
            sql = "select name, id, fulltext, mediapath, ifnull(memo,''), owner, date, av_text_id, risid from source " \
                  "order by date, upper(name)"
        if order_by == "filetype":
            sql = "select name, id, fulltext, mediapath, ifnull(memo,''), owner, date, av_text_id, risid from source " \
                  "order by mediapath"
        if order_by == "casename":
            sql = "select distinct source.name, source.id, source.fulltext, source.mediapath, ifnull(source.memo,''), "
            sql += "source.owner, source.date, av_text_id "
            sql += "from source left join case_text on source.id=case_text.fid "
            sql += "left join cases on cases.caseid=case_text.caseid "
            sql += "order by cases.name, source.name "

        if order_by[:14] == "attribute asc:":
            attribute_name = order_by[14:]
            # Two types of ordering character or numeric
            cur.execute("select valuetype from attribute_type where name=?", [attribute_name])
            attr_type = cur.fetchone()[0]
            sql = "select source.name, source.id, fulltext, mediapath, ifnull(source.memo,''), source.owner, "
            sql += "source.date, av_text_id, risid from source join attribute on attribute.id = source.id "
            sql += " where attribute.attr_type = 'file' and attribute.name=? "
            if attr_type == "character":
                sql += "order by lower(attribute.value) asc "
            else:
                sql += "order by cast(attribute.value as numeric) asc"
            placeholders = [attribute_name]

        if order_by[:15] == "attribute desc:":
            attribute_name = order_by[15:]
            # two types of ordering character or numeric
            cur.execute("select valuetype from attribute_type where name=?", [attribute_name])
            attr_type = cur.fetchone()[0]
            sql = "select source.name, source.id, fulltext, mediapath, ifnull(source.memo,''), source.owner, "
            sql += "source.date, av_text_id, risid from source join attribute on attribute.id = source.id "
            sql += " where attribute.attr_type = 'file' and attribute.name=? "
            if attr_type == "character":
                sql += "order by lower(attribute.value) desc "
            else:
                sql += "order by cast(attribute.value as numeric) desc"
            placeholders = [attribute_name]

        if placeholders is not None:
            cur.execute(sql, placeholders)
        else:
            cur.execute(sql)
        result = cur.fetchall()
        for row in result:
            icon, metadata = self.get_icon_and_metadata(row[1])
            self.source.append({'name': row[0], 'id': row[1], 'fulltext': row[2],
                                'mediapath': row[3], 'memo': row[4], 'owner': row[5], 'date': row[6],
                                'av_text_id': row[7], 'risid': row[8], 'metadata': metadata, 'icon': icon,
                                'case': self.get_cases_by_filename(row[0]),
                                'attributes': []})

        self.header_labels = [_("Name"), _("Memo"), _("Date"), _("Id"), _("Case")]
        # Attributes
        sql = "select name from attribute_type where caseOrFile='file' order by upper(name)"
        cur.execute(sql)
        attribute_names_res = cur.fetchall()
        self.attribute_names = []  # For AddAttribute dialog
        self.attribute_labels_ordered = []  # Help filling table more quickly
        for att_name in attribute_names_res:
            print(att_name[0])
            self.header_labels.append(att_name[0])
            self.attribute_labels_ordered.append(att_name[0])
            self.attribute_names.append({'name': att_name[0]})  # For AddAttribute dialog
        # Add list of attribute values to files, order matches header columns
        sql = "select ifnull(value, '') from attribute where attr_type='file' and attribute.name=? and id=?"
        for s in self.source:
            for att_name in self.attribute_labels_ordered:
                cur.execute(sql, [att_name, s['id']])
                res = cur.fetchone()
                if res:
                    tmp = res[0]
                    # For nicer display
                    if att_name == "Ref_authors":
                        tmp = tmp.replace(";", "\n")
                    s['attributes'].append(tmp)
        # Get reference for file, Vancouver and APA style
        # TODO

        self.fill_table()

    def get_icon_and_metadata(self, id_):
        """ Get metadata used in table tooltip.
        Called by: create_text_file, load_file_data
        param:
            id_  : integer source.id
        """

        cur = self.app.conn.cursor()
        cur.execute("select name, fulltext, mediapath from source where id=?", [id_])
        res = cur.fetchone()
        metadata = res[0] + "\n"
        icon = QtGui.QIcon(qta.icon('mdi6.text-box-outline'))
        # Check if text file is a transcription and add details
        cur.execute("select name from source where av_text_id=?", [id_])
        transcript_res = cur.fetchone()
        if transcript_res is not None:
            metadata += _("Transcript for: ") + f"{transcript_res[0]}\n"
            metadata += _("Characters: ") + str(len(res[1]))
            icon = QtGui.QIcon(qta.icon('mdi6.text'))
        if res[1] is not None and len(res[1]) > 0 and res[2] is None:
            metadata += _("Characters: ") + str(len(res[1]))
            return icon, metadata
        if res[2] is None:
            logger.debug("empty media path error")
            return icon, metadata
        if res[1] is not None and len(res[1]) > 5 and res[2][:6] == "/docs/":
            metadata += _("Characters: ") + str(len(res[1]))
            return icon, metadata
        if res[1] is not None and len(res[1]) > 5 and res[2][:5] == "docs:":
            metadata += _("Characters: ") + str(len(res[1]))
            icon = QtGui.QIcon(qta.icon('mdi6.text-box-check-outline'))
            return icon, metadata

        abs_path = ""
        if 'audio:' == res[2][0:6]:
            abs_path = res[2][6:]
        elif 'video:' == res[2][0:6]:
            abs_path = res[2][6:]
        elif 'images:' == res[2][0:7]:
            abs_path = res[2][7:]
        else:
            abs_path = self.app.project_path + res[2]

        if res[2][:8] == "/images/":
            icon = QtGui.QIcon(qta.icon('mdi6.image-outline'))
            try:
                image = Image.open(abs_path)
                w, h = image.size
            except (FileNotFoundError, PIL.UnidentifiedImageError):
                metadata += _("Cannot locate media. ") + abs_path
                return icon, metadata
            metadata += f"W: {w} x H: {h}"
        if res[2][:7] == "images:":
            icon = QtGui.QIcon(qta.icon('mdi6.image-check-outline'))
            try:
                image = Image.open(abs_path)
                w, h = image.size
            except (FileNotFoundError, PIL.UnidentifiedImageError, AttributeError):
                metadata += _("Cannot locate media. ") + abs_path
                return icon, metadata
            metadata += f"W: {w} x H: {h}"
        if res[2][:7] == "/video/":
            icon = QtGui.QIcon(qta.icon('mdi6.video-outline'))
        if res[2][:6] == "video:":
            icon = QtGui.QIcon(qta.icon('mdi6.video-check-outline'))
        if res[2][:7] == "/audio/":
            icon = QtGui.QIcon(qta.icon('mdi6.play'))
        if res[2][:6] == "audio:":
            icon = QtGui.QIcon(qta.icon('mdi6.play-protected-content'))
        if res[2][:6] in ("/audio", "audio:", "/video", "video:"):
            if not os.path.exists(abs_path):
                metadata += _("Cannot locate media. ") + abs_path
                return icon, metadata
            if vlc:
                try:
                    try:
                        instance = vlc.Instance()
                    except NameError as name_err:
                        # NameError: no function 'libvlc_new'
                        logger.error(f"vlc.Instance: {name_err}")
                        return icon, f"Cannot use vlc. {name_err}"
                    media = instance.media_new(abs_path)
                    media.parse()
                    msecs = media.get_duration()
                    duration_txt = msecs_to_hours_mins_secs(msecs)
                    metadata += _("Duration: ") + duration_txt
                    return icon, metadata
                except AttributeError as err:
                    logger.warning(str(err))
                    metadata += _("Cannot locate media. ") + abs_path + "\n" + str(err)
                    return icon, metadata
            else:
                metadata += _("Cannot get media duration.\nVLC not installed.")
                return icon, metadata
        bytes_ = 0
        try:
            bytes_ = os.path.getsize(abs_path)
        except OSError as e_:
            print(e_)
        metadata += f"\nBytes: {bytes_}"
        if 1024 < bytes_ < 1024 * 1024:
            metadata += f"  {int(bytes_ / 1024)}KB"
        if bytes_ > 1024 * 1024:
            metadata += f"  {int(bytes_ / 1024 / 1024)}MB"
        # Get case names linked to the file
        txt = self.get_cases_by_filename(res[0])
        if txt != "":
            metadata += f'\n{_("Case linked:")}\n{txt}'
        return icon, metadata

    def get_cases_by_filename(self, name):
        """ Called by get_icon_and_metadata, get_file_data
        param: name String of filename """

        cur = self.app.conn.cursor()
        # Case_text is the table, but this also links av and images
        sql = "select distinct cases.name from cases join case_text on case_text.caseid=cases.caseid "
        sql += "join source on source.id=case_text.fid where source.name=? "
        text_ = ""
        cur.execute(sql, [name])
        res = cur.fetchall()
        if res:
            for r in res:
                text_ += f"{r[0]};"
            text_ = text_[:-1]
        return text_

    def add_attribute(self):
        """ When add button pressed, opens the AddAtribute dialog to get new attribute text.
        Then get the attribute type through a dialog.
        AddAttribute dialog checks for duplicate attribute name.
        New attribute is added to the model and database.
        Reserved attribute words - used for imported references:
        Ref_Type (Type of Reference) – character variable
        Ref_Author (authors list) – character
        Ref_Title – character
        Ref_Year (of publication) – numeric
        Ref_Journal - character
        """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        ui = DialogAddAttribute(self.app)
        ok = ui.exec()
        if not ok:
            return
        name = ui.new_name
        value_type = ui.value_type
        if name == "":
            return
        self.attribute_names.append({'name': name})
        # update attribute_type list and database
        now_date = str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        try:
            cur = self.app.conn.cursor()
            cur.execute("insert into attribute_type (name,date,owner,memo,caseOrFile, valuetype) values(?,?,?,?,?,?)",
                        (name, now_date, self.app.settings['codername'], "", 'file', value_type))
            sql = "select id from source"
            cur.execute(sql)
            ids = cur.fetchall()
            for id_ in ids:
                sql = "insert into attribute (name, value, id, attr_type, date, owner) values (?,?,?,?,?,?)"
                cur.execute(sql, (name, "", id_[0], 'file', now_date, self.app.settings['codername']))
            self.app.conn.commit()
            self.app.delete_backup = False
        except Exception as e_:
            print(e_)
            logger.debug(str(e_))
            self.app.conn.rollback()  # Revert all changes
            raise
        self.load_file_data()
        self.fill_table()
        self.parent_text_edit.append(f'{_("Attribute added to files:")} {name}, {_("type")}: {value_type}')

    def cell_double_clicked(self):
        """ View file """

        y = self.ui.tableWidget.currentColumn()
        if y == self.NAME_COLUMN:
            self.view()

    def cell_selected(self):
        """ When the table widget memo cell is selected display the memo.
        Update memo text, or delete memo by clearing text.
        If a new memo, also show in table widget by displaying MEMO in the memo column. """

        x = self.ui.tableWidget.currentRow()
        y = self.ui.tableWidget.currentColumn()
        self.ui.label_file.setText(f"{_('File')}: {self.source[x]['name']}")
        if y == self.MEMO_COLUMN:
            name = self.source[x]['name'].lower()
            cur = self.app.conn.cursor()
            # Need to dynamically get the memo text in case it has been changed in a coding dialog
            cur.execute('select memo from source where id=?', [self.source[x]['id']])
            self.source[x]['memo'] = cur.fetchone()[0]
            if name[-5:] == ".jpeg" or name[-4:] in ('.jpg', '.png', '.gif'):
                ui = DialogMemo(self.app, _("Memo for file ") + self.source[x]['name'],
                                self.source[x]['memo'])
                ui.exec()
                self.source[x]['memo'] = ui.memo
                cur.execute('update source set memo=? where id=?', (ui.memo, self.source[x]['id']))
                self.app.conn.commit()
            else:
                ui = DialogMemo(self.app, _("Memo for file ") + self.source[x]['name'],
                                self.source[x]['memo'])
                ui.exec()
                self.source[x]['memo'] = ui.memo
                cur = self.app.conn.cursor()
                cur.execute('update source set memo=? where id=?', (ui.memo, self.source[x]['id']))
                self.app.conn.commit()
            if self.source[x]['memo'] == "":
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem())
            else:
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem("Memo"))

    def cell_modified(self):
        """ Attribute values can be changed.
        """

        x = self.ui.tableWidget.currentRow()
        y = self.ui.tableWidget.currentColumn()

        # Update attribute value
        if y > self.CASE_COLUMN:
            value = str(self.ui.tableWidget.item(x, y).text()).strip()
            attribute_name = self.header_labels[y]
            cur = self.app.conn.cursor()
            # Check numeric for numeric attributes, clear "" if it cannot be cast
            cur.execute("select valuetype from attribute_type where caseOrFile='file' and name=?", (attribute_name,))
            result = cur.fetchone()
            if result is None:
                return
            if result[0] == "numeric":
                try:
                    float(value)
                except ValueError:
                    self.ui.tableWidget.item(x, y).setText("")
                    value = ""
                    msg = _("This attribute is numeric")
                    Message(self.app, _("Warning"), msg, "warning").exec()
            cur.execute("update attribute set value=? where id=? and name=? and attr_type='file'",
                        (value, self.source[x]['id'], attribute_name))
            self.app.conn.commit()

            # Update self.source[attributes]
            # Add list of attribute values to files, order matches header columns
            sql = "select ifnull(value, '') from attribute where attr_type='file' and attribute.name=? and id=?"
            self.source[x]['attributes'] = []
            for att_name in self.attribute_labels_ordered:
                cur.execute(sql, [att_name, self.source[x]['id']])
                res = cur.fetchone()
                if res:
                    tmp = res[0]
                    # For nicer display
                    if att_name == "Ref_authors":
                        tmp = tmp.replace(";", "\n")
                    self.source[x]['attributes'].append(tmp)

            self.app.delete_backup = False
            self.ui.tableWidget.resizeColumnsToContents()

    def view(self):
        """ View and edit text file contents.
        Alternatively view an image, audio or video media. """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        x = self.ui.tableWidget.currentRow()
        self.ui.tableWidget.selectRow(x)
        if self.source[x]['mediapath'] is not None and 'docs:' != self.source[x]['mediapath'][0:5]:
            if len(self.source[x]['mediapath']) > 6 and self.source[x]['mediapath'][:7] in ("/images", "images:"):
                self.view_image(x)
                return
            if len(self.source[x]['mediapath']) > 5 and self.source[x]['mediapath'][:6] in ("/video", "video:"):
                self.view_av(x)
                return
            if len(self.source[x]['mediapath']) > 5 and self.source[x]['mediapath'][:6] in ("/audio", "audio:"):
                self.view_av(x)
                return
        ui = DialogEditTextFile(self.app, self.source[x]['id'])
        ui.exec()
        # Get fulltext if changed (for metadata)
        cur = self.app.conn.cursor()
        cur.execute("select fulltext from source where id=?", [self.source[x]['id']])
        res = cur.fetchone()
        fulltext = ""
        if res is not None:
            fulltext = res[0]
        self.source[x]['fulltext'] = fulltext

    def view_av(self, x):
        """ View an audio or video file. Edit the memo. Edit the transcript file.
        Added try block in case VLC bindings do not work.
        Uses a non-modal dialog.

        param:
            x  :  row number Integer
        """

        if not vlc:
            msg = _("VLC not installed cannot play audio or video.")
            Message(self.app, _('View AV error'), msg, "warning").exec()
            return
        # Check media exists
        abs_path = ""
        if self.source[x]['mediapath'][0:6] in ('/audio', '/video'):
            abs_path = self.app.project_path + self.source[x]['mediapath']
        if self.source[x]['mediapath'][0:6] in ('audio:', 'video:'):
            abs_path = self.source[x]['mediapath'][6:]
        if not os.path.exists(abs_path):
            self.parent_text_edit.append(_("Bad link or non-existent file ") + abs_path)
            return
        try:
            ui = DialogViewAV(self.app, self.source[x])
            # ui.exec()  # this dialog does not display well on Windows 10 so trying .show()
            # The vlc window becomes unmovable and not resizable
            self.av_dialog_open = ui
            ui.show()
            if self.source[x]['memo'] == "":
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem())
            else:
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem(_("Memo")))
        except Exception as err:
            logger.warning(str(err))
            Message(self.app, _('view AV error'), str(err), "warning").exec()
            self.av_dialog_open = None
            return

    def view_image(self, x):
        """ View an image file and edit the image memo.

        param:
            x  :  row number Integer
        """

        # Check image exists
        abs_path = ""
        if self.source[x]['mediapath'][:7] == "images:":
            abs_path = self.source[x]['mediapath'][7:]
        else:
            abs_path = self.app.project_path + self.source[x]['mediapath']
        if not os.path.exists(abs_path):
            self.parent_text_edit.append(_("Bad link or non-existent file ") + abs_path)
            return
        ui = DialogViewImage(self.app, self.source[x])
        ui.exec()
        memo = ui.ui.textEdit.toPlainText()
        if self.source[x]['memo'] != memo:
            self.source[x]['memo'] = memo
            cur = self.app.conn.cursor()
            cur.execute('update source set memo=? where id=?', (self.source[x]['memo'],
                                                                self.source[x]['id']))
            self.app.conn.commit()
        if self.source[x]['memo'] == "":
            self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem())
        else:
            self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem(_("Memo")))

    def create_text_file(self):
        """ Create a new text file by entering text into the dialog.
        Implements the QtDesigner memo dialog. """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        ui = DialogAddItemName(self.app, self.source, _('New File'), _('Enter file name'))
        ui.exec()
        name = ui.get_new_name()
        if name is None:
            return

        # Create entry details to add to self.source and to database
        item = {'name': name, 'id': -1, 'fulltext': '', 'memo': "",
                'owner': self.app.settings['codername'], 'date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'mediapath': None, 'icon': None, 'metadata': '', 'case': ""}
        # Update database
        cur = self.app.conn.cursor()
        cur.execute("insert into source(name,fulltext,mediapath,memo,owner,date) values(?,?,?,?,?,?)",
                    (
                        item['name'], item['fulltext'], item['mediapath'], item['memo'], item['owner'],
                        item['date']))
        self.app.conn.commit()
        cur.execute("select last_insert_rowid()")
        id_ = cur.fetchone()[0]
        item['id'] = id_
        ui = DialogEditTextFile(self.app, id_)
        ui.exec()
        icon, metadata = self.get_icon_and_metadata(id_)
        item['icon'] = icon
        item['metadata'] = metadata
        item['attributes'] = []
        item['risid'] = None
        # Add file attribute placeholders
        att_sql = 'select name from attribute_type where caseOrFile ="file"'
        cur.execute(att_sql)
        attr_types = cur.fetchall()
        insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
        for a in attr_types:
            placeholders = [a[0], id_, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            self.app.settings['codername']]
            cur.execute(insert_sql, placeholders)
            self.app.conn.commit()
            item['attributes'].append('')
        self.update_files_in_dialogs()
        self.parent_text_edit.append(_("File created: ") + item['name'])
        self.source.append(item)
        self.fill_table()
        self.app.delete_backup = False

    def link_files(self):
        """ Trigger to link to file location. """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        self.import_files(True)

    def import_files(self, link=False):
        """ Import files and store into relevant directories (documents, images, audio, video).
        Convert documents to plain text and store this in data.qda
        Can import from plain text files, also import from html, odt, docx and md.
        md is text Markdown format.
        Note importing from html, odt, docx all formatting is lost.
        Imports images as jpg, jpeg, png which are stored in an images directory.
        Imports audio as mp3, wav, m4a which are stored in an audio directory.
        Imports video as mp4, mov, ogg, wmv which are stored in a video directory.

        param:
            link:   False - files are imported into project folder,
                    True- files are linked and not imported
        """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        response = QtWidgets.QFileDialog.getOpenFileNames(None, _('Open file'),
                                                          self.default_import_directory,
                                                          options=QtWidgets.QFileDialog.Option.DontUseNativeDialog
                                                          )
        imports = response[0]
        if not imports:
            return
        known_file_type = False
        name_split = imports[0].split("/")
        temp_filename = name_split[-1]
        self.default_import_directory = imports[0][0:-len(temp_filename)]
        pdf_msg = ""
        for import_path in imports:
            link_path = ""
            if link:
                link_path = import_path
            # Check file size, any files over 2Gb are linked and not imported internally
            fileinfo = os.stat(import_path)
            if fileinfo.st_size >= 2147483647:
                link_path = import_path
            # Need process events, if many large files are imported, leaves the FileDialog open and covering the screen.
            QtWidgets.QApplication.processEvents()
            filename = import_path.split("/")[-1]
            destination = self.app.project_path
            if import_path.split('.')[-1].lower() in ('docx', 'odt', 'txt', 'htm', 'html', 'epub', 'md'):
                destination += f"/documents/{filename}"
                if link_path == "":
                    copyfile(import_path, destination)
                    self.load_file_text(import_path)
                else:
                    self.load_file_text(import_path, f"docs:{link_path}")
                known_file_type = True
            if import_path.split('.')[-1].lower() == 'pdf':
                destination += f"/documents/{filename}"
                if link_path == "":
                    copyfile(import_path, destination)
                    self.load_file_text(import_path)
                else:
                    self.load_file_text(import_path, f"docs:{link_path}")
                known_file_type = True
            # Media files
            if import_path.split('.')[-1].lower() in ('jpg', 'jpeg', 'png'):
                if link_path == "":
                    destination += f"/images/{filename}"
                    copyfile(import_path, destination)
                    self.load_media_reference(f"/images/{filename}")
                else:
                    self.load_media_reference(f"images:{link_path}")
                known_file_type = True
            if import_path.split('.')[-1].lower() in ('wav', 'mp3', 'm4a'):
                if link_path == "":
                    destination += f"/audio/{filename}"
                    copyfile(import_path, destination)
                    self.load_media_reference(f"/audio/{filename}")
                else:
                    self.load_media_reference(f"audio:{link_path}")
                known_file_type = True
            if import_path.split('.')[-1].lower() in ('mkv', 'mov', 'mp4', 'ogg', 'wmv'):
                if link_path == "":
                    destination += f"/video/{filename}"
                    copyfile(import_path, destination)
                    self.load_media_reference(f"/video/{filename}")
                else:
                    self.load_media_reference(f"video:{link_path}")
                known_file_type = True
            if not known_file_type:
                Message(self.app, _('Unknown file type'),
                        _("Trying to import as text") + f":\n{import_path}", "warning")
                destination += "/documents/" + filename
                if link_path == "":
                    try:
                        self.load_file_text(import_path)
                    except Exception as err:
                        print(err)
                        logger.warning(str(err))
                    try:
                        copyfile(import_path, destination)
                    except OSError as err:
                        logger.warning(str(err))
                        Message(self.app, _('Unknown file type'), _("Cannot import file") + f":\n{import_path}",
                                "warning")
                else:
                    try:
                        self.load_file_text(import_path, f"docs:{link_path}")
                    except Exception as err:
                        logger.warning(str(err))
                        Message(self.app, _('Unknown file type'), _("Cannot import file") + f":\n{import_path}",
                                "warning")
        if pdf_msg != "":
            self.parent_text_edit.append(pdf_msg)
        self.load_file_data()
        self.fill_table()
        self.app.delete_backup = False
        self.update_files_in_dialogs()

    def update_files_in_dialogs(self):
        """ Update files list in any opened dialogs:
         DialogReportCodes, DialogCodeText, DialogCodeAV, DialogCodeImage """

        contents = self.tab_coding.layout()
        if contents is not None:
            for i in reversed(range(contents.count())):
                c = contents.itemAt(i).widget()
                if isinstance(c, DialogCodeImage):
                    c.get_files()
                if isinstance(c, DialogCodeAV):
                    c.get_files()
                if isinstance(c, DialogCodeText):
                    c.get_files()
                if isinstance(c, DialogCodePdf):
                    c.get_files()
        contents = self.tab_reports.layout()
        if contents is not None:
            # Examine widgets in layout
            for i in reversed(range(contents.count())):
                c = contents.itemAt(i).widget()
                if isinstance(c, DialogReportCodes):
                    c.get_files_and_cases()

    def load_media_reference(self, mediapath):
        """ Load media reference information for all file types.

        param:
            mediapath: QualCoder project folder path OR external link path to file
                       External link path contains prefix 'docs:', 'images:, 'audio:', 'video:'
        """

        # check for duplicated filename and update model, widget and database
        name_split = mediapath.split("/")
        filename = name_split[-1]
        if any(d['name'] == filename for d in self.source):
            QtWidgets.QMessageBox.warning(self, _('Duplicate file'), _("Duplicate filename.\nFile not imported"))
            return
        entry = {'name': filename, 'id': -1, 'fulltext': None, 'memo': "", 'mediapath': mediapath,
                 'owner': self.app.settings['codername'], 'date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                 'av_text_id': None}
        cur = self.app.conn.cursor()
        cur.execute("insert into source(name,memo,owner,date, mediapath, fulltext) values(?,?,?,?,?,?)",
                    (
                        entry['name'], entry['memo'], entry['owner'], entry['date'], entry['mediapath'],
                        entry['fulltext']))
        self.app.conn.commit()
        cur.execute("select last_insert_rowid()")
        id_ = cur.fetchone()[0]
        entry['id'] = id_
        msg = entry['name']
        if ':' in mediapath:
            msg += _(" linked")
        else:
            msg += _(" imported.")
        self.parent_text_edit.append(msg)
        self.source.append(entry)

        # Create an empty transcription file for audio and video
        if mediapath[:6] in ("/audio", "audio:", "/video", "video:"):
            entry = {'name': filename + ".txt", 'id': -1, 'fulltext': "", 'mediapath': None, 'memo': "",
                     'owner': self.app.settings['codername'],
                     'date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     'av_text_id': None}
            cur = self.app.conn.cursor()
            cur.execute("insert into source(name,fulltext,mediapath,memo,owner,date) values(?,?,?,?,?,?)",
                        (entry['name'], entry['fulltext'], entry['mediapath'], entry['memo'], entry['owner'],
                         entry['date']))
            self.app.conn.commit()
            cur.execute("select last_insert_rowid()")
            tr_id = cur.fetchone()[0]
            entry['id'] = tr_id
            # Update av file entry with av_text_id link to this text file
            cur.execute("update source set av_text_id=? where id=?", [tr_id, id_])
            self.app.conn.commit()
            
            # add doc to vectorstore
            if self.app.settings['ai_enable'] == 'True':
                self.app.ai.sources_vectorstore.import_document(entry['id'], entry['name'], entry['fulltext'], update=True)

            # Add file attribute placeholders
            att_sql = 'select name from attribute_type where caseOrFile ="file"'
            cur.execute(att_sql)
            attr_types = cur.fetchall()
            insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
            for a in attr_types:
                placeholders = [a[0], tr_id, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                self.app.settings['codername']]
                cur.execute(insert_sql, placeholders)
                self.app.conn.commit()

            self.parent_text_edit.append(entry['name'] + _(" created."))
            self.source.append(entry)

    def load_file_text(self, import_file, link_path=""):
        """ Import from file types of odt, docx pdf, epub, txt, html, htm.
        Implement character detection for txt imports.
        Loading pdf text. I have removed additional line breaks. See commented sections below.
        Removing these allows the pdf to be coded in Code_text and Code_pdf without positional shifting problems.

        param:
            import_file: filepath of file to be imported, String
            link_path:  filepath of file to be linked, String
        """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        text_ = ""
        # Import from odt
        if import_file[-4:].lower() == ".odt":
            text_ = self.convert_odt_to_text(import_file)
            text_ = text_.replace("\n", "\n\n")  # add line to paragraph spacing for visual format
        # Import from docx
        if import_file[-5:].lower() == ".docx":
            document = opendocx(import_file)
            list_ = getdocumenttext(document)
            text_ = "\n\n".join(list_)  # add line to paragraph spacing for visual format
        # Import from epub
        if import_file[-5:].lower() == ".epub":
            book = epub.read_epub(import_file)
            for d in book.get_items_of_type(ebooklib.ITEM_DOCUMENT):
                try:
                    bytes_ = d.get_body_content()
                    string = bytes_.decode('utf-8')
                    text_ += html_to_text(string) + "\n\n"  # add line to paragraph spacing for visual format
                except TypeError as err:
                    logger.debug("ebooklib get_body_content error " + str(err))
        # Import PDF
        if import_file[-4:].lower() == '.pdf':
            pdf_file = open(import_file, 'rb')
            resource_manager = PDFResourceManager()
            laparams = LAParams()
            # laparams.char_margin = 1.0
            # laparams.word_margin = 1.0
            device = PDFPageAggregator(resource_manager, laparams=laparams)
            interpreter = PDFPageInterpreter(resource_manager, device)
            pages_generator = PDFPage.get_pages(pdf_file)  # Generator PDFpage objects
            text_ = ""
            for i, page in enumerate(pages_generator):
                self.pdf_page_text = ""
                interpreter.process_page(page)
                layout = device.get_result()
                for lobj in layout:
                    self.get_item_and_hierarchy(page, lobj)
                text_ += self.pdf_page_text

        # Import from html
        if import_file[-5:].lower() == ".html" or import_file[-4:].lower() == ".htm":
            import_errors = 0
            with open(import_file, "r", encoding="utf-8", errors="surrogateescape") as sourcefile:
                html_text = ""
                while 1:
                    line = sourcefile.readline()
                    if not line:
                        break
                    html_text += line
                text_ = html_to_text(html_text)
                if import_errors > 0:
                    Message(self.app, _("Warning"), str(import_errors) + _(" lines not imported"), "warning").exec()
        # Try importing as a plain text file.
        # TODO https://stackoverflow.com/questions/436220/how-to-determine-the-encoding-of-text
        # ==> suggestion: use the new lib "charset_normalizer"  
        # coding = chardet.detect(file.content).get('encoding')
        # text = file.content[:10000].decode(coding)
        if text_ == "":
            import_errors = 0
            try:
                # can get UnicodeDecode Error on Windows so using error handler
                with open(import_file, "r", encoding="utf-8", errors="backslashreplace") as sourcefile:
                    while 1:
                        line = sourcefile.readline()
                        if not line:
                            break
                        try:
                            text_ += line
                        except Exception as err:
                            logger.warning("Importing plain text file, line ignored: " + str(err))
                            import_errors += 1
                    if text_[0:6] == "\ufeff":  # associated with notepad files
                        text_ = text_[6:]
            except Exception as err:
                logger.warning(str(err))
                Message(self.app, _("Warning"), _("Cannot import") + f"{import_file}\n{err}",
                        "warning").exec()
                return
            if import_errors > 0:
                Message(self.app, _("Warning"), str(import_errors) + _(" lines not imported"), "warning").exec()
                logger.warning(f"{import_file}: {import_errors} " + _("lines not imported"))
        # Import of text file did not work
        if text_ == "":
            Message(self.app, _("Warning"),
                    _("Cannot import ") + str(import_file) + "\nPlease check if the file is empty.", "warning").exec()
            return
        # Final checks: check for duplicated filename and update model, widget and database
        name_split = import_file.split("/")
        filename = name_split[-1]
        if any(d['name'] == filename for d in self.source):
            QtWidgets.QMessageBox.warning(self, _('Duplicate file'),
                                          _("Duplicate filename.\nFile not imported"))
            return

        # Internal storage
        mediapath = "/docs/" + filename
        if link_path != "":
            mediapath = link_path
        entry = {'name': filename, 'id': -1, 'fulltext': text_, 'mediapath': mediapath, 'memo': "",
                 'owner': self.app.settings['codername'], 'date': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        cur = self.app.conn.cursor()
        cur.execute("insert into source(name,fulltext,mediapath,memo,owner,date) values(?,?,?,?,?,?)",
                    (
                        entry['name'], entry['fulltext'], entry['mediapath'], entry['memo'], entry['owner'],
                        entry['date']))
        self.app.conn.commit()
        cur.execute("select last_insert_rowid()")
        id_ = cur.fetchone()[0]
        entry['id'] = id_

        # Add file attribute placeholders
        att_sql = 'select name from attribute_type where caseOrFile ="file"'
        cur.execute(att_sql)
        attr_types = cur.fetchall()
        insert_sql = "insert into attribute (name, attr_type, value, id, date, owner) values(?,'file','',?,?,?)"
        for a in attr_types:
            placeholders = [a[0], id_, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            self.app.settings['codername']]
            cur.execute(insert_sql, placeholders)
            self.app.conn.commit()
            
        # add doc to vectorstore
        if self.app.settings['ai_enable'] == 'True':
            self.app.ai.sources_vectorstore.import_document(entry['id'], entry['name'], entry['fulltext'], update=True)
            
        msg = entry['name']
        if link_path == "":
            msg += _(" imported")
        else:
            msg += _(" linked")
        self.parent_text_edit.append(msg)
        self.source.append(entry)

    # Pdf loading method
    def get_item_and_hierarchy(self, page, lobj: Any):
        """ Get text item details add to page_dict, with descendants.
        Use LTextLine as this object can be parsed in Code_pdf for font size and colour.
        """

        if isinstance(lobj, LTTextLine):  # Do not use LTTextBox
            obj_text = lobj.get_text()
            # Fix Pdfminer recognising invalid unicode characters.
            obj_text = obj_text.replace(u"\uE002", "Th")
            obj_text = obj_text.replace(u"\uFB01", "fi")
            self.pdf_page_text += obj_text
        if isinstance(lobj, Iterable):
            for obj in lobj:
                self.get_item_and_hierarchy(page, obj)

    def convert_odt_to_text(self, import_file):
        """ Convert odt to very rough equivalent with headings, list items and tables for
        html display in qTextEdits. """

        odt_file = zipfile.ZipFile(import_file)
        data = str(odt_file.read('content.xml'))  # bytes class to string
        # https://stackoverflow.com/questions/18488734/python3-unescaping-non-ascii-characters
        data = str(bytes([ord(char) for char in data.encode("utf_8").decode("unicode_escape")]), "utf_8")
        data_start = data.find("</text:sequence-decls>")
        data_end = data.find("</office:text>")
        if data_start == -1 or data_end == -1:
            logger.warning("ODT IMPORT ERROR")
            return ""
        data = data[data_start + 22: data_end]
        data = data.replace('</text:index-title-template>', '')
        data = data.replace('</text:index-entry-span>', '')
        data = data.replace('</text:table-of-content-entry-template>', '')
        data = data.replace('</text:index-title>', '')
        data = data.replace('</text:index-body>', '')
        data = data.replace('</text:table-of-contents>', '')
        data = data.replace('</text:table-of-content-source>', '')
        data = data.replace('<text:h', '\n<text:h')
        data = data.replace('</text:h>', '\n\n')
        data = data.replace('</text:list-item>', '\n')
        data = data.replace('</text:span>', '')
        data = data.replace('</text:p>', '\n')
        data = data.replace('</text:a>', ' ')
        data = data.replace('</text:list>', '')
        data = data.replace('</text:sequence>', '')
        data = data.replace('<text:list-item>', '')
        data = data.replace('<table:table table:name=', '\n=== TABLE ===\n<table:table table:name=')
        data = data.replace('</table:table>', '=== END TABLE ===\n')
        data = data.replace('</table:table-cell>', '\n')
        data = data.replace('</table:table-row>', '')
        data = data.replace('<draw:image', '\n=== IMG ===<draw:image')
        data = data.replace('</draw:frame>', '\n')
        text_ = ""
        tagged = False
        for i in range(0, len(data)):
            if data[i: i + 6] == "<text:" or data[i: i + 7] == "<table:" or data[i: i + 6] == "<draw:":
                tagged = True
            if not tagged:
                text_ += data[i]
            if data[i] == ">":
                tagged = False
        text_ = text_.replace("&apos;", "'")
        text_ = text_.replace("&quot;", '"')
        text_ = text_.replace("&gt;", '>')
        text_ = text_.replace("&lt;", '<')
        text_ = text_.replace("&amp;", '&')
        return text_

    def export(self):
        """ Export selected file to selected directory.
        If an imported file was from a docx, odt, pdf, html, epub then export the original file
        If the file was created within QualCoder (so only in the database), export as plain text.
        Can only export ONE file at time, due to tableWidget single selection mode
        Can only export file that was imported into the project folder.
        Need to check for this condition.
        """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        index_list = self.ui.tableWidget.selectionModel().selectedIndexes()
        rows = [i.row() for i in index_list]
        rows = list(set(rows))  # duplicate rows due to multiple columns
        if len(rows) == 0:
            return
        # Currently single selection mode in tableWidget, 1 row only, so rows[0]
        if self.source[rows[0]]['mediapath'] is not None and ':' in self.source[rows[0]]['mediapath'] \
                and (self.source[rows[0]]['fulltext'] is None or self.source[rows[0]]['fulltext'] == ""):
            msg = _("This is an external linked file") + "\n"
            msg += self.source[rows[0]]['mediapath'].split(':')[1]
            Message(self.app, _('Cannot export'), msg, "warning").exec()
            return
        # Currently can only export ONE file at time, due to tableWidget single selection mode
        row = rows[0]
        # Warn of export of text representation of linked files (e.g. odt, docx, txt, md, pdf)
        text_rep = False
        if self.source[row]['mediapath'] is not None and (':' in self.source[row]['mediapath']) \
                and self.source[row]['fulltext'] != "":
            msg = _("This is a linked file. Will export text representation.") + "\n"
            msg += self.source[row]['mediapath'].split(':')[1]
            Message(self.app, _("Can export text"), msg, "warning").exec()
            text_rep = True

        filename = self.source[row]['name']
        if self.source[row]['mediapath'] is None or self.source[row]['mediapath'][0:5] == 'docs:':
            filename = filename + ".txt"
        exp_dialog = ExportDirectoryPathDialog(self.app, filename)
        destination = exp_dialog.filepath
        if destination is None:
            return
        msg = _("Export to ") + f"{destination}\n"

        # Export audio, video, picture files
        if self.source[row]['mediapath'] is not None and self.source[row]['mediapath'][
                                                         0:6] != "/docs/" and text_rep is False:
            file_path = self.app.project_path + self.source[row]['mediapath']
            try:
                copyfile(file_path, destination)
                msg += f"{destination}\n"
                Message(self.app, _("Files exported"), msg).exec()
                self.parent_text_edit.append(filename + _(" exported to ") + msg)
            except FileNotFoundError:
                Message(self.app, _("Error"), _("File not found")).exec()
            return

        # Export pdf, docx, odt, epub, html files if located in documents directory, and text representation
        document_stored = os.path.exists(self.app.project_path + "/documents/" + self.source[row]['name'])
        if document_stored and (
                self.source[row]['mediapath'] is None or self.source[row]['mediapath'][0:6] == "/docs/"):
            try:
                copyfile(self.app.project_path + "/documents/" + self.source[row]['name'], destination)
                filedata = self.source[row]['fulltext']
                with open(f"{destination}.txt", 'w', encoding='utf-8-sig') as file_:
                    file_.write(filedata)
                msg += f"{destination}\n"
                Message(self.app, _("Files exported"), msg).exec()
                self.parent_text_edit.append(filename + _(" exported to ") + msg)
            except FileNotFoundError as err:
                logger.warning(str(err))
                print(err)
            return
        # Export transcribed files, user created text files, text representations of linked files
        if (self.source[row]['mediapath'] is None or self.source[row]['mediapath'][
                                                     0:5] == 'docs:') and not document_stored:
            filedata = self.source[row]['fulltext']
            with open(destination, 'w', encoding='utf-8-sig') as file_:
                file_.write(filedata)
            msg += f"{destination}\n"
        Message(self.app, _("Files exported"), msg).exec()
        self.parent_text_edit.append(filename + _(" exported to ") + msg)

    def delete_button_multiple_files(self):
        """ Delete files from database and update model and widget.
        Also, delete files from sub-directories, if not externally linked.

        Called by: delete button.
        """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        ui = DialogSelectItems(self.app, self.source, _("Delete files"), "multi")
        ok = ui.exec()
        if not ok:
            return
        selection = ui.get_selected()
        if not selection:
            return
        names = ""
        for selected in selection:
            names = f"{names}{selected['name']}\n"
        ui = DialogConfirmDelete(self.app, names)
        ok = ui.exec()
        if not ok:
            return

        msg = ""
        cur = self.app.conn.cursor()
        for s in selection:
            msg += _("Deleted file: ") + s['name'] + "\n"
            self.files_renamed = [x for x in self.files_renamed if not (s['id'] == x.get('fid'))]
            # Delete text source
            if s['mediapath'] is None or s['mediapath'][0:5] == 'docs:' or s['mediapath'][0:6] == '/docs/':
                try:
                    if s['mediapath'] is None:
                        # Legacy for older < 3.4 QualCoder projects
                        os.remove(self.app.project_path + "/documents/" + s['name'])
                    if s['mediapath'][0:6] == '/docs/':
                        os.remove(self.app.project_path + "/documents/" + s['name'][6:])
                except OSError as err:
                    logger.warning(_("Deleting file error: ") + str(err))
                # Delete stored coded sections and source details
                cur.execute("delete from source where id = ?", [s['id']])
                cur.execute("delete from code_text where fid = ?", [s['id']])
                cur.execute("delete from annotation where fid = ?", [s['id']])
                cur.execute("delete from case_text where fid = ?", [s['id']])
                cur.execute("delete from attribute where attr_type ='file' and id=?", [s['id']])
                self.app.conn.commit()
                # Delete from vectorstore
                self.app.ai.sources_vectorstore.delete_document(s['id'])    
            
            # Delete image, audio or video source
            if s['mediapath'] is not None and s['mediapath'][0:5] != 'docs:' and s['mediapath'][0:6] != '/docs/':
                # Get linked transcript file id
                cur.execute("select av_text_id from source where id=?", [s['id']])
                res = cur.fetchone()
                av_text_id = res[0]
                # Remove avid links in code_text
                sql = "select avid from code_av where id=?"
                cur.execute(sql, [s['id']])
                avids = cur.fetchall()
                sql = "update code_text set avid=null where avid=?"
                for avid in avids:
                    cur.execute(sql, [avid[0]])
                self.app.conn.commit()
                # Remove project folder file, if internally stored
                if ':' not in s['mediapath']:
                    filepath = self.app.project_path + s['mediapath']
                    try:
                        os.remove(filepath)
                    except OSError as err:
                        logger.warning(_("Deleting file error: ") + str(err))
                # Delete stored coded sections and source details
                cur.execute("delete from source where id = ?", [s['id']])
                cur.execute("delete from code_image where id = ?", [s['id']])
                cur.execute("delete from code_av where id = ?", [s['id']])
                cur.execute("delete from attribute where attr_type='file' and id=?", [s['id']])
                # Just in case, added this line
                cur.execute("delete from case_text where fid = ?", [s['id']])
                self.app.conn.commit()

                # Delete linked transcription text file
                if av_text_id is not None:
                    cur.execute("delete from source where id = ?", [res[0]])
                    cur.execute("delete from code_text where fid = ?", [res[0]])
                    cur.execute("delete from annotation where fid = ?", [res[0]])
                    cur.execute("delete from case_text where fid = ?", [res[0]])
                    cur.execute("delete from attribute where attr_type ='file' and id=?", [res[0]])
                    self.app.conn.commit()
                    # Delete from vectorstore
                    self.app.ai.sources_vectorstore.delete_document(res[0])    

        self.update_files_in_dialogs()
        self.check_attribute_placeholders()
        self.parent_text_edit.append(msg)
        self.load_file_data()
        self.fill_table()
        self.app.delete_backup = False

    def delete(self):
        """ Delete one file from database and update model and widget.
        Deletes only one file due to table single selection mode
        Also, delete the file from subdirectories, if not externally linked.
        Called by: right-click table context menu.
        """

        if self.av_dialog_open is not None:
            self.av_dialog_open.mediaplayer.stop()
            self.av_dialog_open = None
        index_list = self.ui.tableWidget.selectionModel().selectedIndexes()
        rows = [i.row() for i in index_list]
        rows = list(set(rows))  # duplicate rows due to multiple columns
        if len(rows) == 0:
            return
        names = ""
        names = f"{names}{self.source[rows[0]]['name']}\n"
        ui = DialogConfirmDelete(self.app, names)
        ok = ui.exec()
        if not ok:
            return

        cur = self.app.conn.cursor()
        row = rows[0]
        file_id = self.source[row]['id']
        # Delete text source
        if self.source[row]['mediapath'] is None or self.source[row]['mediapath'][0:5] == 'docs:' or \
                self.source[row]['mediapath'][0:6] == '/docs/':
            try:
                if self.source[row]['mediapath']:
                    # Legacy for older QualCoder Projects < 3.3
                    os.remove(self.app.project_path + "/documents/" + self.source[row]['name'])
                if self.source[row]['mediapath'] is not None and self.source[row]['mediapath'][0:6] == '/docs/':
                    os.remove(self.app.project_path + "/documents/" + self.source[row]['mediapath'][6:])
            except OSError as err:
                logger.warning(_("Deleting file error: ") + str(err))
            # Delete stored coded sections and source details
            cur.execute("delete from source where id = ?", [file_id])
            cur.execute("delete from code_text where fid = ?", [file_id])
            cur.execute("delete from annotation where fid = ?", [file_id])
            cur.execute("delete from case_text where fid = ?", [file_id])
            cur.execute("delete from attribute where attr_type ='file' and id=?", [file_id])
            self.app.conn.commit()
            # Delete from vectorstore
            self.app.ai.sources_vectorstore.delete_document(file_id)    

        # Delete image, audio or video source
        # (why not simply use 'else' instead of this complicated second if-clause?)
        if self.source[row]['mediapath'] is not None and self.source[row]['mediapath'][0:5] != 'docs:' and \
                self.source[row]['mediapath'][0:6] != '/docs/':
            # Get linked transcript file id
            cur.execute("select av_text_id from source where id=?", [file_id])
            res = cur.fetchone()
            av_text_id = res[0]
            # Remove avid links in code_text
            sql = "select avid from code_av where id=?"
            cur.execute(sql, [file_id])
            avids = cur.fetchall()
            sql = "update code_text set avid=null where avid=?"
            for avid in avids:
                cur.execute(sql, [avid[0]])
            self.app.conn.commit()
            # Remove folder file, if internally stored
            if ':' not in self.source[row]['mediapath']:
                filepath = self.app.project_path + self.source[row]['mediapath']
                try:
                    os.remove(filepath)
                except OSError as err:
                    logger.warning(_("Deleting file error: ") + str(err))
            # Delete stored coded sections and source details
            cur.execute("delete from source where id = ?", [file_id])
            cur.execute("delete from code_image where id = ?", [file_id])
            cur.execute("delete from code_av where id = ?", [file_id])
            cur.execute("delete from attribute where attr_type='file' and id=?", [file_id])
            self.app.conn.commit()
            # Delete from vectorstore (this should not be necessary since it's not a text file, but just to be sure...)
            self.app.ai.sources_vectorstore.delete_document(file_id)    

            # Delete transcription text file
            if av_text_id is not None:
                cur.execute("delete from source where id = ?", [res[0]])
                cur.execute("delete from code_text where fid = ?", [res[0]])
                cur.execute("delete from annotation where fid = ?", [res[0]])
                cur.execute("delete from case_text where fid = ?", [res[0]])
                cur.execute("delete from attribute where attr_type ='file' and id=?", [res[0]])
                self.app.conn.commit()
                # Delete from vectorstore
                self.app.ai.sources_vectorstore.delete_document(res[0])

        self.files_renamed = [x for x in self.files_renamed if not (file_id == x.get('fid'))]
        self.update_files_in_dialogs()
        self.check_attribute_placeholders()
        self.parent_text_edit.append(_("Deleted file: ") + self.source[row]['name'])
        self.load_file_data()
        self.app.delete_backup = False

    def get_tooltip_values(self, attribute_name):
        """ Get values to display in tooltips for the value list column.
        param: attribute_name : String """

        tt = ""
        cur = self.app.conn.cursor()
        sql_val_type = 'select valuetype from attribute_type where caseOrFile="file" and name=?'
        cur.execute(sql_val_type, [attribute_name])
        res_val_type = cur.fetchone()
        value_type = "character"
        if res_val_type is not None:
            value_type = res_val_type[0]
        if value_type == "numeric":
            sql = 'select min(cast(value as real)), max(cast(value as real)) from attribute where name=? and ' \
                  'attr_type="file"'
            cur.execute(sql, [attribute_name])
            res = cur.fetchone()
            tt = _("Minimum: ") + f"{res[0]}\n"
            tt += _("Maximum: ") + str(res[1])
        if value_type == "character":
            sql = 'select distinct value from attribute where name=? and attr_type="file" and length(value)>0 limit 10'
            cur.execute(sql, [attribute_name])
            res = cur.fetchall()
            for r in res:
                tt += f"\n{r[0]}"
            if len(tt) > 1:
                tt = tt[1:]
        return tt

    def fill_table(self):
        """ Fill the table widget with file details. """

        self.ui.tableWidget.blockSignals(True)
        self.ui.tableWidget.setColumnCount(len(self.header_labels))
        self.ui.tableWidget.setHorizontalHeaderLabels(self.header_labels)
        self.ui.tableWidget.horizontalHeader().setStretchLastSection(False)
        rows = self.ui.tableWidget.rowCount()
        for r in range(0, rows):
            self.ui.tableWidget.removeRow(0)
        for row, data in enumerate(self.source):
            self.ui.tableWidget.insertRow(row)
            icon = data['icon']
            name_item = QtWidgets.QTableWidgetItem(data['name'])
            name_item.setIcon(icon)
            # Having un-editable file names helps with assigning icons
            name_item.setFlags(name_item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            # Externally linked - add link details to tooltip
            name_tt = data['metadata']
            if data['mediapath'] is not None and ':' in data['mediapath']:
                name_tt += _("\nExternally linked file:\n")
                name_tt += data['mediapath']
            name_item.setToolTip(name_tt)
            self.ui.tableWidget.setItem(row, self.NAME_COLUMN, name_item)
            date_item = QtWidgets.QTableWidgetItem(data['date'])
            date_item.setFlags(date_item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, self.DATE_COLUMN, date_item)
            memo_string = ""
            if data['memo'] != "":
                memo_string = _("Memo")
            memo_item = QtWidgets.QTableWidgetItem(memo_string)
            if data['memo'] != "":
                memo_item.setToolTip(data['memo'])
            memo_item.setFlags(date_item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, self.MEMO_COLUMN, memo_item)
            fid = data['id']
            if fid is None:
                fid = ""
            iditem = QtWidgets.QTableWidgetItem(str(fid))
            iditem.setFlags(iditem.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, self.ID_COLUMN, iditem)
            case_item = QtWidgets.QTableWidgetItem(data['case'])
            case_item.setFlags(case_item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, self.CASE_COLUMN, case_item)
            # Add the attribute values
            # TODO consider using role type for numerics
            for offset, attribute in enumerate(data['attributes']):
                item = QtWidgets.QTableWidgetItem(attribute)
                self.ui.tableWidget.setItem(row, self.ATTRIBUTE_START_COLUMN + offset, item)
                if self.attribute_labels_ordered[offset] in (
                    "Ref_Authors", "Ref_Title", "Ref_Type", "Ref_Year", "Ref_Journal"):
                    item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
        # Resize columns and rows
        self.ui.tableWidget.hideColumn(self.ID_COLUMN)
        if self.app.settings['showids']:
            self.ui.tableWidget.showColumn(self.ID_COLUMN)
        self.ui.tableWidget.resizeColumnsToContents()
        for i in range(self.ui.tableWidget.columnCount()):
            if self.ui.tableWidget.columnWidth(i) > 500:
                self.ui.tableWidget.setColumnWidth(i, 500)
        self.ui.tableWidget.resizeRowsToContents()
        # self.ui.tableWidget.verticalHeader().setVisible(False)
        # Add statistics tooltips to table headers for attributes
        for i, attribute_name in enumerate(self.attribute_labels_ordered):
            tt = self.get_tooltip_values(attribute_name)
            self.ui.tableWidget.horizontalHeaderItem(self.ATTRIBUTE_START_COLUMN + i).setToolTip(
                _("Right click header row to hide columns") + "\n" + tt)

        self.ui.label_fcount.setText(_("Files: ") + str(len(self.source)))
        self.ui.tableWidget.blockSignals(False)
