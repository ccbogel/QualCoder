# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
"""

import logging
from math import isclose
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import os
import qtawesome as qta  # see: https://pictogrammers.com/library/mdi/

from PyQt6 import QtCore, QtWidgets, QtGui

from .GUI.ui_comparison_table import Ui_Dialog_Comparisons
from .helpers import ExportDirectoryPathDialog, Message, DialogCodeInText, DialogCodeInImage, DialogCodeInAV, msecs_to_hours_mins_secs
from .select_items import DialogSelectItems


path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogReportComparisonTable(QtWidgets.QDialog):
    """ Provide a co-occurrence table with codes rows and files columns. """

    def __init__(self, app, parent_text_edit):
        self.app = app
        self.parent_textEdit = parent_text_edit
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_Comparisons()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)

        self.ui.pushButton_export.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export.pressed.connect(self.export_to_excel)
        self.ui.pushButton_select_files.setIcon(qta.icon('mdi6.file-multiple', options=[{'scale_factor': 1.2}]))
        self.ui.pushButton_select_files.pressed.connect(self.select_files)
        self.ui.pushButton_select_cases.setIcon(qta.icon('mdi6.briefcase-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_select_cases.pressed.connect(self.select_cases)
        self.ui.pushButton_select_attributes.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_select_attributes.pressed.connect(self.select_attribute)
        self.ui.label_arrow.setPixmap(qta.icon('mdi6.arrow-right').pixmap(24, 24))
        self.ui.pushButton_select_codes.setIcon(qta.icon('mdi6.text', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_select_codes.pressed.connect(self.select_codes)
        self.ui.pushButton_select_categories.setIcon(qta.icon('mdi6.file-tree', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_select_categories.pressed.connect(self.select_categories)
        self.ui.checkBox_hide_blanks.stateChanged.connect(self.show_or_hide_empty_rows_and_cols)
        self.ui.listWidget.itemPressed.connect(self.show_list_item)
        #self.ui.listWidget.setSelectionMode()
        tablefont = f'font: 10pt "{self.app.settings["font"]}";'
        self.ui.tableWidget.setStyleSheet(tablefont)  # should be smaller
        self.ui.tableWidget.cellClicked.connect(self.cell_selected)
        #self.ui.tableWidget.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        #self.ui.tableWidget.customContextMenuRequested.connect(self.table_menu)
        self.ui.splitter.setSizes([500, 0])

        self.codes, self.categories = self.app.get_codes_categories()
        self.files = self.app.get_text_filenames()
        # Get attributes
        sql = "select name, valuetype, caseOrFile,0,0 from attribute_type where caseOrFile!='journal'"
        cur = self.app.conn.cursor()
        cur.execute(sql)
        self.attributes = []
        keys = 'true_name', 'valuetype', 'caseOrFile', 'min', 'max'
        for row in cur.fetchall():
            self.attributes.append(dict(zip(keys, row)))
        for attribute in self.attributes:
            attribute['name'] = f"{attribute['true_name']} [{attribute['caseOrFile']}]"
            if attribute['valuetype'] == 'numeric':
                sql = "select cast(value as real) from attribute where name=? and attr_type=? and value is not null order by cast(value as real) asc"
                cur.execute(sql, [attribute['true_name'], attribute['caseOrFile']])
                res = cur.fetchall()
                range = [r[0] for r in res]
                if range:
                    attribute['min'] = range[0]
                    attribute['max'] = range[-1]

        self.data = []
        self.max_count = 0
        self.data_counts = []
        self.data_colors = []
        self.data_list_widget = []   # Used to transfer data from list widget item to DialogCodeIn...

    def clear_table_and_data(self):
        self.data = []
        self.max_count = 0
        self.data_counts = []
        self.data_colors = []
        self.data_list_widget = []   # Used to transfer data from list widget item to DialogCodeIn...
        self.ui.tableWidget.setRowCount(0)
        self.ui.tableWidget.setColumnCount(0)
        self.ui.listWidget.clear()

    def select_attribute(self):
        """ Select an attribute.
        {'name': 'Age', 'valuetype': 'numeric', 'memo': '', 'caseOrFile': 'case'}
        {'name': 'gender', 'valuetype': 'character', 'memo': '', 'caseOrFile': 'case'}
        """

        ui = DialogSelectItems(self.app, self.attributes, _("Select Attribute"), "single")
        ok = ui.exec()
        if not ok:
            return
        attribute = ui.get_selected()
        split_value = None
        if attribute['valuetype'] == 'numeric':
            title = f"{attribute['name']}[{attribute['caseOrFile']}]"
            msg = f"Enter split number (Min: {attribute['min']} - Max: {attribute['max']}):"
            split_value, ok = QtWidgets.QInputDialog.getDouble(self, title, msg)
            if not ok or not split_value:
                self.clear_table_and_data()
                return

        self.files = []
        cur = self.app.conn.cursor()
        if attribute['caseOrFile'] == 'case' and attribute['valuetype'] == 'character':
            sql = "select fid, source.name, cases.name, value from attribute join cases on cases.caseid=attribute.id " \
                  "join case_text on cases.caseid=case_text.caseid " \
                  "join source on source.id=case_text.fid " \
                  "where attr_type='case' and attribute.name=? " \
                  "order by value, cases.name, source.name asc"
            cur.execute(sql, [attribute['true_name']])
            res = cur.fetchall()
            for r in res:
                self.files.append({'id': r[0], 'name': f"{attribute['true_name']}: {r[3]}\nCase: {r[2]}\n{r[1]}", 'memo': ""})

        if attribute['caseOrFile'] == 'case' and attribute['valuetype'] == 'numeric':
            sql = "select fid, source.name, cases.name, cast(value as real) from attribute " \
                  "join cases on cases.caseid=attribute.id " \
                  "join case_text on cases.caseid=case_text.caseid " \
                  "join source on source.id=case_text.fid " \
                  "where attr_type='case' and attribute.name=? " \
                  "order by cast(value as real), cases.name, source.name asc"
            cur.execute(sql, [attribute['true_name']])
            res = cur.fetchall()
            for r in res:
                attr_split_msg = attribute['true_name'] + " "
                if r[3] is None or r[3] < split_value:
                    attr_split_msg += f"< {split_value}"
                elif isclose(r[3], split_value):
                    attr_split_msg += f"= {split_value}"
                else:
                    attr_split_msg += f"> {split_value}"
                self.files.append({'id': r[0], 'name': f"{attr_split_msg}\nCase: {r[2]}\n{r[1]}", 'memo': ""})

        if attribute['caseOrFile'] == 'file' and attribute['valuetype'] == 'character':
            sql = "select source.id, source.name, value from attribute " \
                  "join source on source.id=attribute.id " \
                  "where attr_type='file' and attribute.name=? " \
                  "order by value, source.name asc"
            cur.execute(sql, [attribute['true_name']])
            res = cur.fetchall()
            for r in res:
                self.files.append({'id': r[0], 'name': f"{attribute['true_name']}: {r[2]}\n{r[1]}", 'memo': ""})

        if attribute['caseOrFile'] == 'file' and attribute['valuetype'] == 'numeric':
            sql = "select source.id, source.name, cast(value as real) from attribute " \
                  "join source on source.id=attribute.id " \
                  "where attr_type='file' and attribute.name=? " \
                  "order by cast(value as real), source.name asc"
            cur.execute(sql, [attribute['true_name']])
            res = cur.fetchall()
            for r in res:
                attr_split_msg = attribute['true_name'] + " "
                if r[2] is None or r[2] < split_value:
                    attr_split_msg += f"< {split_value}"
                elif isclose(r[2], split_value):
                    attr_split_msg += f"= {split_value}"
                else:
                    attr_split_msg += f"> {split_value}"
                self.files.append({'id': r[0], 'name': f"{attr_split_msg}\n{r[1]}", 'memo': ""})

        if not self.files:
            self.clear_table_and_data()
            return
        self.process_files_data()

    def select_files(self):
        """ Select files, stored in self.file_ids_names, then load data and fill table. """

        selection_list = [{'id': -1, 'name': ''}]
        for file_name in self.app.get_filenames():
            selection_list.append(file_name)
        ui = DialogSelectItems(self.app, selection_list, _("Select files"), "multi")
        ok = ui.exec()
        if not ok:
            return
        selected = ui.get_selected()
        if not selected or selected[0]['name'] == '':
            self.files = self.app.get_filenames()
            Message(self.app, _("Files selected"), _("All files selected")).exec()
        else:
            self.files = selected
            msg = ""
            for item in self.files:
                msg += f"{item['name']}\n"
            Message(self.app, _("Files selected"), msg).exec()
        self.process_files_data()

    def process_files_data(self):
        """ Calculate the relations for selected codes for ALL coders (or only THIS coder - TODO).
        For text? codings only.
        Rows as codes, columns as files.

        Data items:
        For DialogCodeInText
                dictionary {codename, color, file_or_casename, pos0, pos1, text, coder, fid, file_or_case,
                            textedit_start, textedit_end}
        For DialogCodeInImage
                dictionary {codename, color, file_or_casename, x1, y1, width, height, coder,
                 mediapath, fid, memo, file_or_case}
        """

        self.ui.checkBox_hide_blanks.setChecked(False)
        self.ui.splitter.setSizes([500, 0])

        # Create data matrices zeroed, codes are ordered alphabetically by name
        self.data_counts = []
        self.max_count = 0
        self.data_colors = []
        self.data = []
        for row in self.codes:
            self.data_counts.append([0] * len(self.files))
            self.data_colors.append([""] * len(self.files))
            self.data.append(["."] * len(self.files))

        cur = self.app.conn.cursor()
        for row, code_ in enumerate(self.codes):
            for col, file_ in enumerate(self.files):
                # Text results
                sql = "select source.id,source.name, code_text.cid, code_name.name, code_name.color, pos0, pos1," \
                      "ctid,seltext, ifnull(code_text.memo,''), " \
                      "code_text.owner, 'file', 'text' from code_text " \
                      "join code_name on code_name.cid=code_text.cid " \
                      "join source on code_text.fid=source.id " \
                      "where code_text.fid=? and code_text.cid=? order by code_text.ctid"
                cur.execute(sql, [file_['id'], code_['cid']])
                results_text = cur.fetchall()
                keys_text = 'fid', 'file_or_casename', 'cid', 'codename', 'color', 'pos0', 'pos1', 'ctid', 'text', \
                    'memo', 'owner', 'file_or_case', 'result_type'
                text_data = []
                for res in results_text:
                    text_data.append(dict(zip(keys_text, res)))

                # Image results
                sql = "select source.id,source.name, code_image.cid, code_name.name, code_name.color, x1,y1," \
                      "width,height, ifnull(code_image.memo,''), " \
                      "code_image.owner, mediapath, 'file', 'image' from code_image " \
                      "join code_name on code_name.cid=code_image.cid " \
                      "join source on code_image.id=source.id " \
                      "where code_image.id=? and code_image.cid=? order by code_image.imid"
                cur.execute(sql, [file_['id'], code_['cid']])
                results_image = cur.fetchall()
                keys_image = 'fid', 'file_or_casename', 'cid', 'codename', 'color', 'x1', 'y1', 'width', 'height', \
                    'memo', 'owner', 'mediapath', 'file_or_case', 'result_type'
                image_data = []
                for res in results_image:
                    image_data.append(dict(zip(keys_image, res)))

                # Audio /video results
                sql = "select source.id,source.name, code_av.cid, code_name.name, code_name.color, pos0,pos1, " \
                      "ifnull(code_av.memo,''), " \
                      "code_av.owner, mediapath, 'file', 'av' from code_av " \
                      "join code_name on code_name.cid=code_av.cid " \
                      "join source on code_av.id=source.id " \
                      "where code_av.id=? and code_av.cid=? order by code_av.avid"
                cur.execute(sql, [file_['id'], code_['cid']])
                results_av = cur.fetchall()
                keys_av = 'fid', 'file_or_casename', 'cid', 'codename', 'color', 'pos0', 'pos1', \
                    'memo', 'owner', 'mediapath', 'file_or_case', 'result_type'
                av_data = []
                for res in results_av:
                    av_data.append(dict(zip(keys_av, res)))

                result_length = len(results_text) + len(results_image) + len(results_av)
                if result_length > self.max_count:
                    self.max_count = result_length
                self.data_counts[row][col] = result_length

                self.data[row][col] = text_data + image_data + av_data

        # Color heat map for spread across 5 colours
        colors = ["#F8E0E0", "#F6CECE", "#F5A9A9", "#F78181", "#FA5858"]  # light to dark red
        for row, row_data in enumerate(self.data_counts):
            for col, item_data in enumerate(row_data):
                if self.data_counts[row][col] > 0:
                    color_range_index = int(self.data_counts[row][col] / self.max_count * 5) - 1
                    if color_range_index < 0:
                        color_range_index = 0
                    self.data_colors[row][col] = colors[color_range_index]
        self.fill_table(self.files)

    def select_cases(self):
        """ Select cases to display relevant files. """

        cases = self.app.get_casenames()
        selection_list = [{'id': -1, 'name': ''}]
        for case in cases:
            selection_list.append(case)
        ui = DialogSelectItems(self.app, selection_list, _("Select cases"), "multi")
        ok = ui.exec()
        if not ok:
            return
        selected = ui.get_selected()
        if selected and selected[0]['name'] != '':
            cases = []
            for item in selected:
                cases.append(item)
        self.files = []
        cur = self.app.conn.cursor()
        sql = "select case_text.fid, source.name, source.memo from case_text " \
              "join source on source.id=case_text.fid where case_text.caseid=?"
        for case in cases:
            cur.execute(sql, [case['id']])
            res = cur.fetchall()
            for r in res:
                self.files.append({'id': r[0], 'name': f"{case['name']}\n{r[1]}", 'memo': r[2]})
        msg = f"Selection\nCases: {len(cases)}. Files: {len(self.files)}"
        Message(self.app, _("Selection"), msg).exec()
        if not self.files:
            self.clear_table_and_data()
            return
        self.process_files_data()

    def select_codes(self):
        """ Select codes. """

        selection_list = [{'id': -1, 'name': ''}]
        codes, categories = self.app.get_codes_categories()
        for code_ in codes:
            selection_list.append(code_)
        ui = DialogSelectItems(self.app, selection_list, _("Select codes"), "multi")
        ok = ui.exec()
        if not ok:
            return
        selected = ui.get_selected()
        if not selected or selected[0]['name'] == '':
            #selected = deepcopy(self.codes)
            self.codes = codes
            Message(self.app, _("Codes selected"), _("All codes selected")).exec()
        else:
            msg = ""
            self.codes = []
            for selected_code in selected:
                self.codes.append(selected_code)
                msg += f"{selected_code['name']}\n"
            Message(self.app, _("Codes selected"), msg).exec()
        self.clear_table_and_data()

    def select_categories(self):
        """ Select all codes in selected categories. """

        selection_list = [{'id': -1, 'name': ''}]
        codes, categories = self.app.get_codes_categories()
        for category in categories:
            selection_list.append(category)
        ui = DialogSelectItems(self.app, selection_list, _("Select one category"), "single")
        ok = ui.exec()
        if not ok:
            return
        category = ui.get_selected()
        self.codes = self.get_children_of_category(category)
        for code_ in self.codes:
            code_['name'] = f"{category['name']}:\n{code_['name']}"
        self.clear_table_and_data()

    def get_children_of_category(self, node):
        """ Get child categories and codes of this category node.
        Only keep the category or code name. Used to reposition TextGraphicsItems on moving a category.

        Args:
             node : Dictionary of category

        Returns:
             child_codes : List of Dictionaries
        """

        child_names = []
        codes, categories = self.app.get_codes_categories()

        """ Create a list of this category (node) and all its category children.
        Maximum depth of 200. """
        selected_categories = [node]
        i = 0  # Ensure an exit from loop
        new_model_changed = True
        while categories != [] and new_model_changed and i < 200:
            new_model_changed = False
            append_list = []
            for n in selected_categories:
                for m in categories:
                    if m['supercatid'] == n['catid']:
                        append_list.append(m)
                        child_names.append(m['name'])
            for n in append_list:
                selected_categories.append(n)
                categories.remove(n)
                new_model_changed = True
            i += 1
        categories = selected_categories
        # Remove codes that are not associated with these categories
        selected_codes = []
        for cat in categories:
            for code in codes:
                if code['catid'] == cat['catid']:
                    selected_codes.append(code)
        return selected_codes

    def export_to_excel(self):
        """ Export to Excel file. """

        filename = "Code_comparisons.xlsx"
        export_dir = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_dir.filepath
        if filepath is None:
            return

        # Excel row headers
        row_header = []
        for code_ in self.codes:
            name_split_50 = [code_['name'][y - 50:y] for y in range(50, len(code_['name']) + 50, 50)]
            row_header.append("\n".join(name_split_50))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Counts"
        wb.create_sheet("Details")
        ws2 = wb["Details"]
        for row, row_name in enumerate(row_header):
            v_cell = ws.cell(row=row + 2, column=1)
            v_cell.value = row_name
            v_cell2 = ws2.cell(row=row + 2, column=1)
            v_cell2.value = row_name

        # Excel column headers
        for col, file_ in enumerate(self.files):
            h_cell = ws.cell(row=1, column=col + 2)
            h_cell.value = file_['name']
            h_cell2 = ws2.cell(row=1, column=col + 2)
            h_cell2.value = file_['name']
            ws.column_dimensions[get_column_letter(col + 1)].width = 20
            ws2.column_dimensions[get_column_letter(col + 1)].width = 20

        # Co-occurrence counts
        for row, row_data in enumerate(self.data_counts):
            for col, col_data in enumerate(row_data):
                cell = ws.cell(row=row + 2, column=col + 2)
                cell.value = col_data
                if self.data_colors[row][col] != "":
                    cell.fill = PatternFill(start_color=self.data_colors[row][col][1:], end_color=self.data_colors[row][col][1:], fill_type="solid")
                # Details list
                if self.data[row][col] == ".":
                    continue
                details = ""
                for data in self.data[row][col]:
                    details += f"Code:{data['codename']} File/Case: {data['file_or_casename']}\n"
                    if data['result_type'] == "text":
                        details += f"Pos0: {data['pos0']} - Pos1: {data['pos1']}\n"
                        details += f"{data['text']}"
                    if data['result_type'] == "av":
                        details += f"Pos0: {data['pos0']} - Pos1: {data['pos1']} msecs\n"
                        details += f"Memo: {data['memo']}"
                    if data['result_type'] == "image":
                        details += f"X, Y: {data['x1']}, {data['y1']}. Width: {data['width']}. Height: {data['height']}\n"
                        details += f"Memo: {data['memo']}"
                    details += f"\n========\n"
                d_cell = ws2.cell(row=row + 2, column=col + 2)
                d_cell.value = details
                d_cell.alignment = Alignment(wrap_text=True, vertical='top')

        wb.save(filepath)
        msg = _('Co-occurrence exported: ') + filepath
        Message(self.app, _('Co-occurrence exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def fill_table(self, column_header):
        """ Fill table using code names alphabetically (case insensitive) as rows
        header columns can be files, or ... MORE ?
        using self.data

        Args:
            column_header: List of dictionary items containing 'name' for the table columns header
        """

        rows = self.ui.tableWidget.rowCount()
        for r in range(0, rows):
            self.ui.tableWidget.removeRow(0)
        column_header_labels = []
        for item in column_header:
            column_header_labels.append(item['name'])
        self.ui.tableWidget.setColumnCount(len(column_header_labels))
        self.ui.tableWidget.setHorizontalHeaderLabels(column_header_labels)
        row_header = []
        for code_ in self.codes:
            # row_header.append(code_['name'])  # original
            name_split_50 = [code_['name'][y - 50:y] for y in range(50, len(code_['name']) + 50, 50)]
            row_header.append("\n".join(name_split_50))
        self.ui.tableWidget.setRowCount(len(row_header))
        self.ui.tableWidget.setVerticalHeaderLabels(row_header)
        for row, row_data in enumerate(self.data_counts):
            for col, cell_data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem()
                if self.data_colors[row][col] != "":
                    item.setBackground(QtGui.QBrush(QtGui.QColor(self.data_colors[row][col])))
                    item.setForeground(QtGui.QBrush(QtGui.QColor("#000000")))
                if cell_data > 0:
                    item.setData(QtCore.Qt.ItemDataRole.DisplayRole, cell_data)
                item.setFlags(item.flags() ^ QtCore.Qt.ItemFlag.ItemIsEditable)
                self.ui.tableWidget.setItem(row, col, item)
        self.ui.tableWidget.resizeColumnsToContents()  # Doesnt look great
        self.ui.tableWidget.resizeRowsToContents()

    def show_or_hide_empty_rows_and_cols(self):
        """ Unchecked - show all rows and columns.
        Checked - hide rows and columns with no code co-occurrences. """

        if not self.data_counts:
            return
        if self.ui.checkBox_hide_blanks.isChecked():
            for row, row_data in enumerate(self.data_counts):
                if sum(row_data) == 0:
                    self.ui.tableWidget.hideRow(row)

            for col in range(len(self.data_counts[0])):
                col_sum = 0
                for row, row_data in enumerate(self.data_counts):
                    col_sum += row_data[col]
                if col_sum == 0:
                    self.ui.tableWidget.hideColumn(col)
        if not self.ui.checkBox_hide_blanks.isChecked():
            for row, row_data in enumerate(self.data_counts):
                self.ui.tableWidget.showRow(row)
                self.ui.tableWidget.showColumn(row)

    def cell_selected(self):
        """ When the table widget memo cell is selected display the memo.
        Update memo text, or delete memo by clearing text.
        If a new memo, also show in table widget by displaying MEMO in the memo column.

        Text data keys = 'fid', 'file_or_casename', 'cid', 'codename', 'color', 'pos0', 'pos1', 'ctid', 'text',
                    'memo', 'owner', 'file_or_case', 'result_type'

        """

        row = self.ui.tableWidget.currentRow()
        col = self.ui.tableWidget.currentColumn()
        text = self.ui.tableWidget.item(row, col).text()
        if text == "":
            self.ui.listWidget.clear()
            self.data_list_widget = []
            return
        self.data_list_widget = self.data[row][col]
        self.ui.listWidget.clear()
        for row, data in enumerate(self.data_list_widget):
            display = f"{data['file_or_casename']} | {data['codename']}\n"
            if data['result_type'] == "text":
                display += f"{data['pos0']} - {data['pos1']}. Coder: {data['owner']}\n"
                display += f"{data['text']}"
            if data['result_type'] == "image":
                display += f"Image X: {data['x1']}, Y: {data['y1']}. Width: {data['width']}, Height: {data['height']}. "
                display += f"Coder: {data['owner']}\nMemo: {data['memo']}"
            if data['result_type'] == "av":
                display += f"A/V: {msecs_to_hours_mins_secs(data['pos0'])} - {msecs_to_hours_mins_secs(data['pos1'])}. Coder: {data['owner']}\n"
                display += f"{data['memo']}"
            list_item = QtWidgets.QListWidgetItem()
            list_item.setText(display)
            self.ui.listWidget.insertItem(row, display)
        self.ui.splitter.setSizes([300, 200])

    def show_list_item(self):
        """  """

        row = self.ui.listWidget.currentIndex().row()
        data = self.data_list_widget[row]
        if data['result_type'] == "text":
            ui = DialogCodeInText(self.app, data)
            ui.exec()
        if data['result_type'] == "image":
            ui = DialogCodeInImage(self.app, data)
            ui.exec()
        if data['result_type'] == "av":
            ui = DialogCodeInAV(self.app, data)
            ui.exec()


    ''' TODO for future expansion maybe.
    def table_menu(self, position):
        """ Context menu for displaying table cell coding details.
        """

        row = self.ui.tableWidget.currentRow()
        col = self.ui.tableWidget.currentColumn()
        text = self.ui.tableWidget.item(row, col).text()
        print(row, col, text)

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_view = menu.addAction(_("View"))'''

