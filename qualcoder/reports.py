# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com/
"""

from copy import copy
import datetime
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill
import os
import qtawesome as qta

from PyQt6 import QtGui, QtWidgets, QtCore
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QBrush

from .color_selector import TextColor
from .GUI.ui_dialog_report_comparisons import Ui_Dialog_reportComparisons
from .GUI.ui_dialog_report_code_frequencies import Ui_Dialog_reportCodeFrequencies
from .helpers import Message, ExportDirectoryPathDialog
from .information import DialogInformation
from .select_items import DialogSelectItems

path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogReportCodeFrequencies(QtWidgets.QDialog):
    """ Show code and category frequencies, overall and for each coder.
    This is for text, image and av coding. """

    app = None
    parent_textEdit = None
    coders = []
    categories = []
    codes = []
    coded = []  # to refactor name
    file_ids = []

    def __init__(self, app, parent_textedit):

        self.app = app
        self.parent_textEdit = parent_textedit
        self.get_data()
        self.calculate_code_frequencies()
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_reportCodeFrequencies()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        self.ui.pushButton_exporttext.pressed.connect(self.export_text_file)
        self.ui.pushButton_exporttext.setIcon(qta.icon('mdi6.export-variant', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_export_excel.pressed.connect(self.export_excel_file)
        self.ui.pushButton_export_excel.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.3}]))
        self.ui.pushButton_select_files.pressed.connect(self.select_files)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        font = f'font: {self.app.settings["treefontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.treeWidget.setStyleSheet(font)
        self.ui.treeWidget.setSelectionMode(QtWidgets.QTreeWidget.SelectionMode.ExtendedSelection)
        self.fill_tree()
        self.ui.radioButton.clicked.connect(self.sort_by_alphabet)
        self.ui.radioButton_2.clicked.connect(self.sort_by_totals)

    def select_files(self):
        """ Report code frequencies for all files or selected files. """

        filenames = self.app.get_filenames()
        if len(filenames) == 0:
            return
        ui = DialogSelectItems(self.app, filenames, _("Select files to view"), "many")
        ok = ui.exec()
        tooltip = _("Files selected: ")
        self.file_ids = []
        if ok:
            selected_files = ui.get_selected()  # List of dictionaries
            files_text = ""
            for row in selected_files:
                self.file_ids.append(row['id'])
                files_text += f"\n{row['name']}"
            files_text = files_text[2:]
            tooltip += files_text
            if len(self.file_ids) > 0:
                self.ui.pushButton_select_files.setToolTip(tooltip)
        self.get_data()
        self.calculate_code_frequencies()
        self.fill_tree()

    def get_data(self):
        """ Called from init.
        Calls calculate_code_frequency - for each code.
        Adds a list item that is ready to be used by the treeWidget to display multiple
        columns with the coder frequencies.
        Not using  app.get_codes_categories method as this adds extra columns for each end user
        """

        cur = self.app.conn.cursor()
        self.categories = []
        cur.execute("select name, catid, owner, date, ifnull(memo,''), supercatid from code_cat order by lower(name)")
        result = cur.fetchall()
        for row in result:
            self.categories.append({'name': row[0], 'catid': row[1], 'owner': row[2],
                                    'date': row[3], 'memo': row[4], 'supercatid': row[5],
                                    'display_list': [row[0], 'catid:' + str(row[1])]})
        self.codes = []
        cur.execute("select name, ifnull(memo,''), owner, date, cid, catid, color from code_name order by lower(name)")
        result = cur.fetchall()
        for row in result:
            self.codes.append({'name': row[0], 'memo': row[1], 'owner': row[2], 'date': row[3],
                               'cid': row[4], 'catid': row[5], 'color': row[6],
                               'display_list': [row[0], 'cid:' + str(row[4])]})
        self.coders = []
        cur.execute("select distinct owner from code_text union select distinct owner from code_image union "
                    "select distinct owner from code_av")
        result = cur.fetchall()
        self.coders = []
        for row in result:
            self.coders.append(row[0])
        self.coded = []
        if True:
            cur.execute("select cid, owner, fid from code_text")
            result = cur.fetchall()
            for row in result:
                if row[2] in self.file_ids or self.file_ids == []:
                    self.coded.append(row)
            cur.execute("select cid, owner, id from code_image")
            result = cur.fetchall()
            for row in result:
                if row[2] in self.file_ids or self.file_ids == []:
                    self.coded.append(row)
            cur.execute("select cid, owner, id from code_av")
            result = cur.fetchall()
            for row in result:
                if row[2] in self.file_ids or self.file_ids == []:
                    self.coded.append(row)

    def calculate_code_frequencies(self):
        """ Calculate the frequency of each code for all coders and the total.
        Add a list item to each code that can be used to display in treeWidget.
        For codings in code_image, code_text.
        """

        for c in self.codes:
            total = 0
            for cn in self.coders:
                count = 0
                for cit in self.coded:
                    if cit[1] == cn and cit[0] == c['cid']:
                        count += 1
                        total += 1
                c['display_list'].append(count)
            c['display_list'].append(total)

        # Add the number of codes directly under each category to the category
        for cat in self.categories:
            # magic 3 = cat name, cat id and total columns
            cat_list = [0] * (len(self.coders) + 3)
            for c in self.codes:
                if c['catid'] == cat['catid']:
                    for i in range(2, len(c['display_list'])):
                        cat_list[i] += c['display_list'][i]
            cat_list = cat_list[2:]
            for count in cat_list:
                cat['display_list'].append(count)

        # find leaf categories, add to above categories, and gradually remove leaves
        # until only top categories are left
        sub_cats = copy(self.categories)
        counter = 0
        while len(sub_cats) > 0 or counter < 10000:
            leaf_list = []
            branch_list = []
            for c in sub_cats:
                for c2 in sub_cats:
                    if c['catid'] == c2['supercatid']:
                        branch_list.append(c)
            for cat in sub_cats:
                if cat not in branch_list:
                    leaf_list.append(cat)
            # Add totals for each coder and overall total to higher category
            for leaf_cat in leaf_list:
                for cat in self.categories:
                    if cat['catid'] == leaf_cat['supercatid']:
                        for i in range(2, len(cat['display_list'])):
                            cat['display_list'][i] += leaf_cat['display_list'][i]
                sub_cats.remove(leaf_cat)
            counter += 1

        header = ["Code Tree", "Id"]
        for coder in self.coders:
            header.append(coder)
        header.append("Total")

    def sort_by_totals(self, ):
        """ Sort by totals descending. """

        self.get_data()
        self.calculate_code_frequencies()
        self.categories = sorted(self.categories, key=lambda i: (i['display_list'][-1]), reverse=True)
        self.codes = sorted(self.codes, key=lambda i: (i['display_list'][-1]), reverse=True)
        self.fill_tree()

    def sort_by_alphabet(self, ):
        """ Sort alphabtically ascending. """

        self.get_data()
        self.calculate_code_frequencies()
        self.categories = sorted(self.categories, key=lambda i: (i['display_list'][0]))
        self.codes = sorted(self.codes, key=lambda i: (i['display_list'][0]))
        self.fill_tree()

    def depthgauge(self, item):
        """ Get depth for treewidget item. """

        depth = 0
        while item.parent() is not None:
            item = item.parent()
            depth += 1
        return depth

    def export_text_file(self):
        """ Export coding frequencies to text file. """

        filename = "Code_frequencies.txt"
        e = ExportDirectoryPathDialog(self.app, filename)
        filepath = e.filepath
        if filepath is None:
            return
        text_ = _("Code frequencies") + "\n"
        text_ += f"{self.app.project_name}\n"
        text_ += _("Date: ") + datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S") + "\n"
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        item_total_position = 1 + len(self.coders)
        while item:
            self.depthgauge(item)
            cat = False
            if item.text(1).split(':')[0] == "catid":
                cat = True
            prefix = ""
            for i in range(0, self.depthgauge(item)):
                prefix += "--"
            if cat:
                text_ += f"\n{prefix}" + _("Category: ") + item.text(0)
                text_ += f", Frequency: {item.text(item_total_position)}"
            else:
                text_ += f"\n{prefix}" + _("Code: ") + item.text(0)
                text_ += _(", Frequency: ") + item.text(item_total_position)
            it += 1
            item = it.value()
        with open(filepath, 'w', encoding='utf-8-sig') as file_:
            file_.write(text_)
        msg = _("Coding frequencies text file exported to: ") + filepath
        Message(self.app, _('Text file Export'), msg).exec()
        self.parent_textEdit.append(msg)

    def export_excel_file(self):
        """ Export data as excel. """

        header = [_("Code Tree"), "Id"]
        for coder in self.coders:
            header.append(coder)
        header.append("Total")
        wb = openpyxl.Workbook()
        ws = wb.active
        # Column headings
        for col, code in enumerate(header):
            ws.cell(column=col + 1, row=1, value=code)
            ws.cell(column=col + 1, row=1).font = Font(b=True)
        # Data
        data = []
        code_colors = []
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        while item:
            row = []
            for i in range(0, len(header)):
                row.append(item.text(i))
            if row[1][:3] == "cid":
                cid = int(row[1][4:])
                for code_ in self.codes:
                    if cid == code_['cid']:
                        row[0] = code_['name']  # Full not abbreviated name
                        code_colors.append(code_['color'][1:])
            else:
                code_colors.append(None)
            it += 1
            item = it.value()
            data.append(row)

        for row, data_row in enumerate(data):
            for col, datum in enumerate(data_row):
                ws.cell(column=col + 1, row=row + 2, value=datum)
            if code_colors[row]:
                pf = PatternFill(start_color=code_colors[row], end_color=code_colors[row], fill_type="solid")
                ws.cell(column=1, row=row + 2).fill = pf
        filename = "Code_frequencies.xlsx"
        export_dir = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_dir.filepath
        if filepath is None:
            return
        wb.save(filepath)
        msg = _("Coding frequencies exported to: ") + filepath
        Message(self.app, _('File export'), msg).exec()
        self.parent_textEdit.append(msg)

    def fill_tree(self):
        """ Fill tree widget, top level items are main categories and unlinked codes. """

        self.ui.treeWidget.clear()
        cats = copy(self.categories)
        codes = copy(self.codes)
        self.ui.treeWidget.clear()
        header = [_("Code Tree"), "Id"]
        for coder in self.coders:
            header.append(coder)
        header.append("Total")
        self.ui.treeWidget.setColumnCount(len(header))
        self.ui.treeWidget.setHeaderLabels(header)
        if not self.app.settings['showids']:
            self.ui.treeWidget.setColumnHidden(1, True)
        else:
            self.ui.treeWidget.setColumnHidden(1, False)
        self.ui.treeWidget.header().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.ui.treeWidget.header().setStretchLastSection(False)
        # add top level categories
        remove_list = []
        for c in cats:
            if c['supercatid'] is None:
                display_list = []
                for i in c['display_list']:
                    display_list.append(str(i))
                if len(display_list[0]) > 62:  # Keep category name short
                    display_list[0] = display_list[0][:30] + '..' + display_list[0][-30:]
                top_item = QtWidgets.QTreeWidgetItem(display_list)
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_list.append(c)
        for item in remove_list:
            cats.remove(item)
        ''' Add child categories. Look at each unmatched category, iterate through tree to
        add as child then remove matched categories from the list. '''
        count = 0
        while not (len(cats) < 1 or count > 10000):
            remove_list = []
            for c in cats:
                it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
                item = it.value()
                while item:  # while there is an item in the list
                    if item.text(1) == f'catid:{c["supercatid"]}':
                        display_list = []
                        for i in c['display_list']:
                            display_list.append(str(i))
                        if len(display_list[0]) > 62:  # Keep category name short
                            display_list[0] = display_list[0][:30] + '..' + display_list[0][-30:]
                        child = QtWidgets.QTreeWidgetItem(display_list)
                        child.setToolTip(0, c['name'])
                        item.addChild(child)
                        remove_list.append(c)
                    it += 1
                    item = it.value()
            for item in remove_list:
                cats.remove(item)
            count += 1

        # Add unlinked codes as top level items
        remove_items = []
        for c in codes:
            if c['catid'] is None:
                display_list = []
                for i in c['display_list']:
                    display_list.append(str(i))
                if len(display_list[0]) > 62:  # Keep code name short
                    display_list[0] = f"{display_list[0][:30]}..{display_list[0][-30:]}"
                top_item = QtWidgets.QTreeWidgetItem(display_list)
                top_item.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                color = TextColor(c['color']).recommendation
                top_item.setForeground(0, QBrush(QtGui.QColor(color)))
                top_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_items.append(c)
        for item in remove_items:
            codes.remove(item)

        # Add codes as children
        for c in codes:
            it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
            item = it.value()
            while item:
                if item.text(1) == f'catid:{c["catid"]}':
                    display_list = []
                    for i in c['display_list']:
                        display_list.append(str(i))
                    if len(display_list[0]) > 62:  # Keep code name short
                        display_list[0] = f"{display_list[0][:30]}..{display_list[0][-30:]}"
                    child = QtWidgets.QTreeWidgetItem(display_list)
                    child.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                    color = TextColor(c['color']).recommendation
                    child.setForeground(0, QBrush(QtGui.QColor(color)))
                    child.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                    child.setToolTip(0, c['name'])
                    item.addChild(child)
                    c['catid'] = -1  # Make unmatchable
                it += 1
                item = it.value()
        self.ui.treeWidget.sortByColumn(0, QtCore.Qt.SortOrder.AscendingOrder)
        self.ui.treeWidget.expandAll()


class DialogReportCoderComparisons(QtWidgets.QDialog):
    """ Compare coded text sequences between coders using Cohen's Kappa. """

    def __init__(self, app, parent_textedit):

        self.app = app
        self.parent_textEdit = parent_textedit
        self.text_data = ""
        self.excel_data = []
        self.coders = []
        self.selected_coders = []
        self.file_summaries = []
        self.codes = []
        self.categories = []
        self.get_data()
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_reportComparisons()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        self.ui.pushButton_run.setEnabled(False)
        self.ui.pushButton_run.pressed.connect(self.calculate_statistics)
        self.ui.pushButton_run.setIcon(qta.icon('mdi6.play', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_clear.pressed.connect(self.clear_selection)
        self.ui.pushButton_clear.setIcon(qta.icon('mdi6.refresh', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export.pressed.connect(self.export_excel)  # export_text_file)
        self.ui.pushButton_export.setToolTip(_("Export Excel"))
        self.ui.pushButton_export.setIcon(qta.icon('mdi6.export', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_help1.setIcon(qta.icon('mdi6.help'))
        self.ui.pushButton_help1.pressed.connect(self.information)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        font = f'font: {self.app.settings["treefontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.treeWidget.setStyleSheet(font)
        self.ui.treeWidget.setSelectionMode(QtWidgets.QTreeWidget.SelectionMode.ExtendedSelection)
        self.ui.comboBox_coders.insertItems(0, self.coders)
        self.ui.comboBox_coders.currentTextChanged.connect(self.coder_selected)
        if len(self.coders) == 3:  # includes empty slot
            self.ui.comboBox_coders.setCurrentIndex(1)
            self.ui.comboBox_coders.setCurrentIndex(2)
        self.fill_tree()

    def get_data(self):
        """ Called from init. gets coders, codes, categories, file_summaries.
        Images and A/V files are not used. """

        self.codes, self.categories = self.app.get_codes_categories()
        cur = self.app.conn.cursor()
        cur.execute("select id, length(fulltext) from source where fulltext is not null")
        self.file_summaries = cur.fetchall()
        sql = "select owner from  code_image union select owner from code_text union select owner from code_av"
        cur.execute(sql)
        result = cur.fetchall()
        self.coders = [""]
        for row in result:
            self.coders.append(row[0])

    def coder_selected(self):
        """ Select coders for comparison - only two coders can be selected. """

        coder = self.ui.comboBox_coders.currentText()
        if coder == "":
            return
        if len(self.selected_coders) == 0:
            self.selected_coders.append(coder)
            self.ui.label_selections.setText(coder)
        if len(self.selected_coders) == 1 and self.selected_coders[0] != coder:
            self.selected_coders.append(coder)
            coder1 = self.ui.label_selections.text()
            self.ui.label_selections.setText(f"{coder1} , {coder}")
        if len(self.selected_coders) == 2:
            self.ui.pushButton_run.setEnabled(True)

    def clear_selection(self):
        """ Clear the coder selection and tree widget statistics.
        text(1) Catid/cid, text(2) Agree%, text(3) A and B %, text(4) Not A Not B %
        text (5) Disagree %, text(6) AgreeCodedOnly%, text(7) Kappa
        """

        self.selected_coders = []
        self.ui.pushButton_run.setEnabled(False)
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        while item:  # while there is an item in the list
            if item.text(1)[0:4] == 'cid:':
                item.setText(2, "")
                item.setText(3, "")
                item.setText(4, "")
                item.setText(5, "")
                item.setText(6, "")
                item.setText(7, "")
            it += 1
            item = it.value()
        self.ui.label_selections.setText(_("No coders selected"))

    def export_excel(self):
        """ Export to Excel. """

        filename = "Coder_comparison.xlsx"
        export_path = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_path.filepath
        if filepath is None:
            return

        wb = openpyxl.Workbook()
        ws = wb.active

        #file_.write(f"{self.app.project_name}\n")
        #file_.write(_("Date: ") + datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S") + "\n")
        #file_.write(self.comparisons)

        ws.cell(column=1, row=1, value=f"Coder Comparison: {self.selected_coders[0]}, {self.selected_coders[1]}")

        headings = ["Code tree", "Agree %", "A and B %", "Not A Not B %", "Disagree %", "Agree coded only %", "Kappa"]
        for col, heading in enumerate(headings):
            ws.cell(column=col + 1, row=2, value=heading)
            ws.cell(column=col + 1, row=2).font = Font(b=True)
        for row in range(len(self.excel_data)):
            for col in range(7):
                ws.cell(column=col + 1, row=3 + row, value=self.excel_data[row][col])
            if self.excel_data[row][7] != "Category":
                pf = PatternFill(start_color=self.excel_data[row][7], end_color=self.excel_data[row][7], fill_type="solid")
                ws.cell(column=1, row=3 + row).fill = pf

        wb.save(filepath)
        msg = _('Coder comparisons report exported: ') + filepath
        Message(self.app, _('Report exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def export_text_file(self):
        """ OLD Export coding comparison statistics to text file. """

        filename = "Coder_comparison.txt"
        export_path = ExportDirectoryPathDialog(self.app, filename)
        filepath = export_path.filepath
        if filepath is None:
            return
        with open(filepath, 'w', encoding="'utf-8-sig'") as file_:
            file_.write(f"{self.app.project_name}\n")
            file_.write(_("Date: ") + datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S") + "\n")
            file_.write(self.text_data)
        msg = _("Coder comparison text file exported to: ") + filepath
        Message(self.app, _('Text file export'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def calculate_statistics(self):
        """ Iterate through tree widget, for all cids
        For each code calculate the two-coder comparison statistics. """

        self.text_data = "====" + _("CODER COMPARISON") + "====\n" + _("Selected coders: ")
        self.text_data += self.selected_coders[0] + ", " + self.selected_coders[1] + "\n"
        self.excel_data = []
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        row = 0
        while item:
            if item.text(1)[0:4] == 'cid:':
                excel_row = []
                agreement = self.calculate_agreement_for_code(int(item.text(1)[4:]))
                item.setText(2, f"{agreement['agreement']}%")
                item.setText(3, f"{agreement['dual_percent']}%")
                item.setText(4, f"{agreement['uncoded_percent']}%")
                item.setText(5, f"{agreement['disagreement']}%")
                item.setText(6, f"{agreement['agree_coded_only']}%")
                item.setText(7, f"{agreement['kappa']}")
                self.text_data += f"\n{item.text(0)} ({item.text(1)})\n"
                excel_row.append(item.text(0))
                self.text_data += _("agreement: ") + f"{agreement['agreement']}%"
                excel_row.append(f"{agreement['agreement']}%")
                self.text_data += _(", dual coded: ") + f"{agreement['dual_percent']}%"
                excel_row.append(f"{agreement['dual_percent']}%")
                self.text_data += _(", uncoded: ") + f"{agreement['uncoded_percent']}%"
                excel_row.append(f"{agreement['uncoded_percent']}%")
                self.text_data += _(", disagreement: ") + f"{agreement['disagreement']}%"
                excel_row.append(f"{agreement['disagreement']}%")
                self.text_data += _(", agree coded only: ") + f"{agreement['agree_coded_only']}%"
                excel_row.append(f"{agreement['agree_coded_only']}%")
                self.text_data += f", Kappa: {agreement['kappa']}"
                excel_row.append(agreement['kappa'])
                # Fix codename, add color
                cid = int(item.text(1)[4:])
                for code_ in self.codes:
                    if cid == code_['cid']:
                        excel_row.append(code_['color'][1:])
                self.excel_data.append(excel_row)
            else:  # Category
                self.excel_data.append([item.text(0), "", "", "", "", "", "", "Category"])
            it += 1
            item = it.value()
            row += 1

    def calculate_agreement_for_code(self, cid):
        """ Calculate the two-coder statistics for this cid
        Percentage agreement.
        Get the start and end positions in all files (source table) for this cid.

        self.file_summaries item [0] = id, [1] = full text length
        Look at each file separately to ge the commonly coded text.
        Each character that is coded by coder 1 or coder 2 is incremented, resulting in a list of 0, 1, 2
        where 0 is no codings at all, 1 is coded by only one coder and 2 is coded by both coders.
        'Disagree%':'','A not B':'','B not A':'', coded only:'' ,'K':''

        param:
            cid : integer source file id

        """

        # coded0 and coded1 are the total characters coded by coder 0 and coder 1
        total = {'dual_coded': 0, 'single_coded': 0, 'uncoded': 0, 'characters': 0, 'coded0': 0, 'coded1': 0,
                 'agree_coded_only': 0.0, 'agreement': 0.0, 'disagreement': 0.0, 'uncoded_percent': 0.0}
        # Loop through each source file
        cur = self.app.conn.cursor()
        sql = "select pos0,pos1,fid from code_text where fid=? and cid=? and owner=?"
        for f in self.file_summaries:
            cur.execute(sql, [f[0], cid, self.selected_coders[0]])
            result0 = cur.fetchall()
            cur.execute(sql, [f[0], cid, self.selected_coders[1]])
            result1 = cur.fetchall()
            # Determine the same characters coded by both coders, by adding 1 to each coded character
            char_list = [0] * f[1]
            for coded in result0:
                for char in range(coded[0], coded[1]):
                    try:
                        char_list[char] += 1
                        total['coded0'] += 1
                    except IndexError as e_:
                        msg = "DialogReportCoderComparisons.calculate_agreement_for_code "
                        msg += f"{e_} fid:{f[0]} len_text:{f[1]} pos1:{coded[1]}"
                        msg += f" cid:{cid} coder:{self.selected_coders[0]}"
                        print(msg)
                        logger.error(msg)
                        self.parent_textEdit.append(msg)
            for coded in result1:
                for char in range(coded[0], coded[1]):
                    try:
                        char_list[char] += 1
                        total['coded1'] += 1
                    except IndexError as e_:
                        msg = "DialogReportCoderComparisons.calculate_agreement_for_code "
                        msg += f"{e_} fid:{f[0]} len_text:{f[1]} pos1:{coded[1]}"
                        msg += " cid:" + str(cid) + " coder:" + self.selected_coders[0]
                        print(msg)
                        logger.error(msg)
                        self.parent_textEdit.append(msg)
            uncoded = 0
            single_coded = 0
            dual_coded = 0
            for char in char_list:
                if char == 0:
                    uncoded += 1
                if char == 1:
                    single_coded += 1
                if char == 2:
                    dual_coded += 1
            total['dual_coded'] += dual_coded
            total['single_coded'] += single_coded
            total['uncoded'] += uncoded
            total['characters'] += f[1]

        if total['characters'] != 0:
            total['agreement'] = round(100 * (total['dual_coded'] + total['uncoded']) / total['characters'], 2)
            total['dual_percent'] = round(100 * total['dual_coded'] / total['characters'], 2)
            total['uncoded_percent'] = round(100 * total['uncoded'] / total['characters'], 2)
            total['disagreement'] = round(100 - total['agreement'], 2)
            try:
                total['agree_coded_only'] = round(100 * total['dual_coded'] / (total['dual_coded'] + total['single_coded']), 2)
            except ZeroDivisionError:
                total['agree_coded_only'] = "zero div"
        else:
            total['agreement'] = "zero div"
            total['dual_percent'] = "zero div"
            total['uncoded_percent'] = "zero div"
            total['disagreement'] = "zero div"
            total['agree_coded_only'] = "zero div"
        # Cohen's Kappa
        '''
        https://en.wikipedia.org/wiki/Cohen%27s_kappa

        k = Po - Pe     Po is proportionate agreement (both coders coded this text / all coded text))
            -------     Pe is probability of random agreement
            1  - Pe

            Pe = Pyes + Pno
            Pyes = proportion Yes by A multiplied by proportion Yes by B
                 = total['coded0']/total_coded * total['coded1]/total_coded

            Pno = proportion No by A multiplied by proportion No by B
                = (total_coded - total['coded0']) / total_coded * (total_coded - total['coded1]) / total_coded

        IMMEDIATE BELOW IS INCORRECT - RESULTS IN THE TOTAL AGREEMENT SCORE
        Po = total['agreement'] / 100
        Pyes = total['coded0'] / total['characters'] * total['coded1'] / total['characters']
        Pno = (total['characters'] - total['coded0']) / total['characters'] * (total['characters'] - total['coded1']) / 
            total['characters']

        BELOW IS BETTER - ONLY LOOKS AT PROPORTIONS OF CODED CHARACTERS
        NEED TO CONFIRM THIS IS THE CORRECT APPROACH
        '''
        total['kappa'] = "zerodiv"
        try:
            unique_codings = total['coded0'] + total['coded1'] - total['dual_coded']
            Po = total['dual_coded'] / unique_codings
            Pyes = total['coded0'] / unique_codings * total['coded1'] / unique_codings
            Pno = (unique_codings - total['coded0']) / unique_codings * (
                        unique_codings - total['coded1']) / unique_codings
            Pe = Pyes * Pno
            kappa = round((Po - Pe) / (1 - Pe), 4)
            total['kappa'] = kappa
        except ZeroDivisionError:
            pass
        return total

    def fill_tree(self):
        """ Fill tree widget, top level items are main categories and unlinked codes. """

        cats = copy(self.categories)
        codes = copy(self.codes)
        self.ui.treeWidget.clear()
        self.ui.treeWidget.setColumnCount(7)
        self.ui.treeWidget.setHeaderLabels(
            [_("Code Tree"), "Id", "Agree %", "A and B %", "Not A Not B %", "Disagree %", "Agree coded only %", "Kappa"])
        self.ui.treeWidget.hideColumn(1)
        if self.app.settings['showids']:
            self.ui.treeWidget.showColumn(1)
        self.ui.treeWidget.header().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.ui.treeWidget.header().setStretchLastSection(False)
        # Add top level categories
        remove_list = []
        for c in cats:
            if c['supercatid'] is None:
                top_item = QtWidgets.QTreeWidgetItem([c['name'], f'catid:{c["catid"]}'])
                if len(c['name']) > 62:
                    top_item.setText(0, f"{c['name'][:30]}..{c['name'][-30:]}")
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_list.append(c)
        for item in remove_list:
            cats.remove(item)
        ''' Add child categories. Look at each unmatched category, iterate through tree to
        add as child then remove matched categories from the list. '''
        count = 0
        while len(cats) > 0 or count < 10000:
            remove_list = []
            for c in cats:
                it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
                item = it.value()
                while item:  # while there is an item in the list
                    if item.text(1) == f'catid:{c["supercatid"]}':
                        child = QtWidgets.QTreeWidgetItem([c['name'], f'catid:{c["catid"]}'])
                        if len(c['name']) > 62:
                            child.setText(0, f"{c['name'][:30]}..{c['name'][-30:]}")
                        child.setToolTip(0, c['name'])
                        item.addChild(child)
                        remove_list.append(c)
                    it += 1
                    item = it.value()
            for item in remove_list:
                cats.remove(item)
            count += 1
        # Add unlinked codes as top level items
        remove_items = []
        for c in codes:
            if c['catid'] is None:
                top_item = QtWidgets.QTreeWidgetItem([c['name'], f'cid:{c["cid"]}'])
                if len(c['name']) > 62:
                    top_item.setText(0, f"{c['name'][:30]}..{c['name'][-30:]}")
                top_item.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                color = TextColor(c['color']).recommendation
                top_item.setForeground(0, QBrush(QtGui.QColor(color)))
                top_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                top_item.setToolTip(0, c['name'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_items.append(c)
        for item in remove_items:
            codes.remove(item)

        # Add codes as children
        for c in codes:
            it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
            item = it.value()
            while item:
                if item.text(1) == f'catid:{c["catid"]}':
                    child = QtWidgets.QTreeWidgetItem([c['name'], f'cid:{c["cid"]}'])
                    if len(c['name']) > 62:
                        child.setText(0, f"{c['name'][:30]}..{c['name'][-30:]}")
                    child.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                    color = TextColor(c['color']).recommendation
                    child.setForeground(0, QBrush(QtGui.QColor(color)))
                    child.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                    child.setToolTip(0, c['name'])
                    item.addChild(child)
                    c['catid'] = -1  # Make unmatchable
                it += 1
                item = it.value()
        self.ui.treeWidget.sortByColumn(0, QtCore.Qt.SortOrder.AscendingOrder)
        self.ui.treeWidget.expandAll()

    def information(self):
        """ Provide statistical help information. """

        ui = DialogInformation(self.app, "Statistics information", "")
        ui.setHtml(info)
        ui.exec()


info = "<b>Agree %</b>" \
       "<p>Calculated across all text files as the (total dual coded plus the total uncoded) / total characters</p>" \
       "<b>A and B %</b><p>Calculated as the total dual coded characters / total characters</p>" \
       "<b>Not A Not B %</b><p>The characters not coded by either coder / total characters</p>" \
       "<b>Disagree %</b><p>Is 100% minus the total agreement percent.</p>" \
       "<b>Agree coded only %</b><p>Is the dual coded characters divided by the dual coded and single coded characters" \
       "</p>" \
       "<b>Kappa</b><p>Used to measure inter-rater reliability. " \
       "Calculations are based on this site https://en.wikipedia.org/wiki/Cohen%27s_kappa</p>"
