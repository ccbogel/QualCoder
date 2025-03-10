# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
"""

import csv
import datetime
import logging
import openpyxl
from openpyxl import load_workbook
import os
import qtawesome as qta
from urllib.parse import urlparse
import webbrowser

from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtCore import Qt

from .add_attribute import DialogAddAttribute
from .add_item_name import DialogAddItemName
from .case_file_manager import DialogCaseFileManager
from .confirm_delete import DialogConfirmDelete
from .GUI.ui_dialog_cases import Ui_Dialog_cases
from .helpers import Message, ExportDirectoryPathDialog

from .memo import DialogMemo
from .view_av import DialogViewAV
from .view_image import DialogViewImage

path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogCases(QtWidgets.QDialog):
    """ Create, edit and delete cases.
    Assign entire text files or portions of files to cases.
    Assign attributes to cases. """

    NAME_COLUMN = 0  # Also primary key
    MEMO_COLUMN = 1
    ID_COLUMN = 2
    FILES_COLUMN = 3
    ATTRIBUTE_START_COLUMN = 4
    cases = []

    def __init__(self, app, parent_text_edit):

        self.app = app
        self.parent_text_edit = parent_text_edit
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_cases()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        doc_font = f'font: {self.app.settings["docfontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.textBrowser.setStyleSheet(doc_font)
        self.ui.pushButton_add.setIcon(qta.icon('mdi6.pencil-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_add.clicked.connect(self.add_case)
        self.ui.pushButton_delete.setIcon(qta.icon('mdi6.delete-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_delete.clicked.connect(self.delete_case)
        self.ui.pushButton_file_manager.setIcon(qta.icon('mdi6.text-box-multiple-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_file_manager.pressed.connect(self.case_file_manager)
        self.ui.tableWidget.itemChanged.connect(self.cell_modified)
        self.ui.tableWidget.cellClicked.connect(self.cell_selected)
        self.ui.tableWidget.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget.customContextMenuRequested.connect(self.table_menu)
        self.ui.pushButton_add_attribute.setIcon(qta.icon('mdi6.variable', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_add_attribute.clicked.connect(self.add_attribute)
        self.ui.pushButton_import_cases.setIcon(qta.icon('mdi6.file-import-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_import_cases.clicked.connect(self.import_cases_and_attributes)
        self.ui.pushButton_export_attributes.setIcon(qta.icon('mdi6.file-export-outline', options=[{'scale_factor': 1.4}]))
        self.ui.pushButton_export_attributes.clicked.connect(self.export_attributes)
        self.ui.pushButton_help.setIcon(qta.icon('mdi6.help'))
        self.ui.pushButton_help.pressed.connect(self.help)
        self.ui.textBrowser.setText("")
        self.ui.textBrowser.setAutoFillBackground(True)
        self.ui.textBrowser.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.textBrowser.customContextMenuRequested.connect(self.link_clicked)
        self.ui.textBrowser.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.textBrowser.customContextMenuRequested.connect(self.text_edit_menu)
        self.insert_nonexisting_attribute_placeholders()
        self.ui.tableWidget.itemSelectionChanged.connect(self.count_selected_items)
        self.ui.tableWidget.horizontalHeader().setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.tableWidget.horizontalHeader().customContextMenuRequested.connect(self.table_header_menu)
        self.ui.tableWidget.installEventFilter(self)
        self.ui.tableWidget.setTabKeyNavigation(False)

        self.header_labels = []
        self.attribute_labels_ordered = []
        self.source = []
        self.cases = []
        self.case_text = []
        self.display_text_links = []  # Clickable links for A/V images as dictionaries of pos0, pos1, file id
        self.attributes = []
        self.selected_case = None
        self.selected_file = None
        self.clipboard_text = ""  # Used to copy text into another cell

        self.load_cases_data()
        self.fill_table()
        # Initial resize of table columns
        self.ui.tableWidget.resizeColumnsToContents()
        self.ui.splitter.setSizes([1, 0])
        self.eventFilterTT = ToolTipEventFilter()
        self.ui.textBrowser.installEventFilter(self.eventFilterTT)

    def keyPressEvent(self, event):
        """ Used to activate buttons.
        Ctrl 0 to 6 for some functions and Ctrl C and Ctrl V for copy paste
        """
        key = event.key()
        mods = QtWidgets.QApplication.keyboardModifiers()
        # Ctrl 0 to 6
        if mods & QtCore.Qt.KeyboardModifier.ControlModifier:
            if key == QtCore.Qt.Key.Key_1:
                self.add_case()
                return
            if key == QtCore.Qt.Key.Key_2:
                self.import_cases_and_attributes()
                return
            if key == QtCore.Qt.Key.Key_3:
                self.case_file_manager()
                return
            if key == QtCore.Qt.Key.Key_4:
                self.add_attribute()
                return
            if key == QtCore.Qt.Key.Key_5:
                self.export_attributes()
                return
            if key == QtCore.Qt.Key.Key_6:
                self.delete_case()
                return
            if key == QtCore.Qt.Key.Key_0:
                self.help()
                return
            if key == QtCore.Qt.Key.Key_C:
                x = self.ui.tableWidget.currentRow()
                y = self.ui.tableWidget.currentColumn()
                self.clipboard_text = self.ui.tableWidget.item(x, y).text()
                if self.clipboard_text is None:
                    self.clipboard_text = ""
                return
            if key == QtCore.Qt.Key.Key_V:
                x = self.ui.tableWidget.currentRow()
                y = self.ui.tableWidget.currentColumn()
                self.ui.tableWidget.item(x, y).setText(self.clipboard_text)
                return

    def eventFilter(self, object_, event):
        """ Using this event filter to
        Ctrl + A to show all rows
        """

        if type(event) == QtGui.QKeyEvent:
            key = event.key()
            mod = event.modifiers()
            if key == QtCore.Qt.Key.Key_A and mod == QtCore.Qt.KeyboardModifier.ControlModifier:
                for r in range(0, self.ui.tableWidget.rowCount()):
                    self.ui.tableWidget.setRowHidden(r, False)
                return True
        return False

    def insert_nonexisting_attribute_placeholders(self):
        """ Check attribute placeholder is present in attribute table.
        An error in earlier qualcoder versions did not fill these placeholders.
        Fix if not present.
        Cases are a list of dictionaries.
        Attributes are a list of tuples(name,value,id)
        """

        now_date = str(datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"))
        cur = self.app.conn.cursor()
        cur.execute("select name from attribute_type where caseOrFile='case'")
        attribute_names = cur.fetchall()
        for c in self.cases:
            for att_name in attribute_names:
                cur.execute("select value from attribute where id=? and name=? and attr_type='case'",
                            [c['caseid'], att_name[0]])
                res = cur.fetchone()
                if res is None:
                    cur.execute("insert into attribute (value,id,name,attr_type, date,owner) values(?,?,?,'case',?,?)",
                                ("", c['caseid'], att_name[0], now_date, self.app.settings['codername']))
                    self.app.conn.commit()

    @staticmethod
    def help():
        """ Open help for transcribe section in browser. """

        url = "https://github.com/ccbogel/QualCoder/wiki/3.3.-Cases"
        webbrowser.open(url)

    # Revise
    def count_selected_items(self):
        """ Clear the text edit if multiple rows are selected.
         return:
            item_count """

        indexes = self.ui.tableWidget.selectedIndexes()
        ix = [i.row() for i in indexes]
        i = len(set(ix))
        if i > 1:
            self.ui.textBrowser.clear()
            self.ui.splitter.setSizes([100, 0])
        return i

    def export_attributes(self):
        """ Export attributes from table as an Excel file. """

        shortname = self.app.project_name.split(".qda")[0]
        filename = shortname + "_case_attributes.xlsx"
        e = ExportDirectoryPathDialog(self.app, filename)
        filepath = e.filepath
        if filepath is None:
            return
        cols = self.ui.tableWidget.columnCount()
        rows = self.ui.tableWidget.rowCount()
        header = [self.ui.tableWidget.horizontalHeaderItem(i).text() for i in range(0, cols)]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Case Attributes"
        for col, col_name in enumerate(header):
            h_cell = ws.cell(row=1, column=col + 1)
            h_cell.value = col_name
        for row in range(rows):
            for col in range(cols):
                cell = ws.cell(row=row + 2, column=col + 1)
                data = ""
                try:
                    data = self.ui.tableWidget.item(row, col).text()
                except AttributeError:
                    pass
                cell.value = data
        wb.save(filepath)
        '''with open(filepath, mode='w') as f: # OLD csv save
            writer = csv.writer(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header)
            for r in range(0, rows):
                data = []
                for c in range(0, cols):
                    # Table cell may be a None type
                    cell = ""
                    try:
                        cell = self.ui.tableWidget.item(r, c).text()
                    except AttributeError:
                        pass
                    data.append(cell)
                writer.writerow(data)'''
        msg = _("Case attributes file exported to: ") + filepath
        Message(self.app, _('File export'), msg).exec()
        self.parent_text_edit.append(msg)

    def load_cases_data(self, order_by="asc"):
        """ Load case (to maximum) and attribute details from database. Display in tableWidget.
        Cases are a list of dictionaries.
        Attributes are a list of tuples(name,value,id)
        """

        self.source = []
        self.cases = []
        self.case_text = []
        result = []

        cur = self.app.conn.cursor()
        cur.execute("select name, id, fulltext, mediapath, ifnull(memo,''), owner, date, av_text_id from source")
        file_result = cur.fetchall()
        for row in file_result:
            self.source.append({'name': row[0], 'id': row[1], 'fulltext': row[2],
                                'mediapath': row[3], 'memo': row[4], 'owner': row[5], 'date': row[6],
                                'av_text_id': row[7]})
        if order_by == "asc":
            # Odd error with null caseid in the past
            sql = "select name, ifnull(memo,''), owner, date, ifnull(caseid,'') from cases "
            sql += "order by name asc"
            cur.execute(sql)
            result = cur.fetchall()
        if order_by == "desc":
            # Odd error with null in the past
            sql = "select name, ifnull(memo,''), owner, date, ifnull(caseid,'') from cases "
            sql += "order by name desc"
            cur.execute(sql)
            result = cur.fetchall()
        if order_by[:14] == "attribute asc:":
            attribute_name = order_by[14:]
            cur.execute("select valuetype from attribute_type where name=?", [attribute_name])
            attr_type = cur.fetchone()[0]
            # Odd error with null caseid in the past
            sql = "select cases.name, ifnull(memo,''), cases.owner, cases.date, ifnull(caseid,'') from cases "
            sql += "join attribute on attribute.id = cases.caseid "
            sql += "where attribute.attr_type = 'case' and attribute.name=? "
            if attr_type == "character":
                sql += "order by lower(attribute.value) asc "
            else:
                sql += "order by cast(attribute.value as numeric) asc"
            cur.execute(sql, [attribute_name])
            result = cur.fetchall()
        if order_by[:15] == "attribute desc:":
            attribute_name = order_by[15:]
            cur.execute("select valuetype from attribute_type where name=?", [attribute_name])
            attr_type = cur.fetchone()[0]
            # Odd error with null caseid in the past
            sql = "select cases.name, ifnull(memo,''), cases.owner, cases.date, ifnull(caseid,'') from cases "
            sql += "join attribute on attribute.id = cases.caseid "
            sql += "where attribute.attr_type = 'case' and attribute.name=? "
            if attr_type == "character":
                sql += "order by lower(attribute.value) desc "
            else:
                sql += "order by cast(attribute.value as numeric) desc"
            cur.execute(sql, [attribute_name])
            result = cur.fetchall()

        for row in result:
            sql = "select distinct case_text.fid, source.name from case_text join source on case_text.fid=source.id "
            sql += "where caseid=? order by source.name asc"
            cur.execute(sql, [row[4], ])
            files_res = cur.fetchall()
            self.cases.append({'name': row[0], 'memo': row[1], 'owner': row[2], 'date': row[3],
                               'caseid': row[4], 'files': files_res, 'attributes': []})
        cur.execute("select name from attribute_type where caseOrFile='case' order by upper(name)")
        attribute_names_res = cur.fetchall()
        self.header_labels = ["Name", "Memo", "Id", "Files"]
        self.attribute_labels_ordered = []
        for att_name in attribute_names_res:
            self.header_labels.append(att_name[0])
            self.attribute_labels_ordered.append(att_name[0])
        # Add list if attribute values to cases, order matches header columns
        sql = "select ifnull(value, '') from attribute where attr_type='case' and attribute.name=? and id=?"
        for a in self.attribute_labels_ordered:
            for i, c in enumerate(self.cases):
                cur.execute(sql, [a, c['caseid']])
                res = cur.fetchone()
                if res:
                    c['attributes'].append(res[0])
        self.fill_table()

    def update_label(self):
        """ Update label when loading data, adding or deleting cases. """

        cur = self.app.conn.cursor()
        cur.execute("select count(caseid) from cases")
        total_cases = cur.fetchone()[0]
        msg = _("Cases: ") + f"{total_cases} "
        self.ui.label_cases.setText(msg)

    def add_attribute(self):
        """ When add button pressed, opens the addItem dialog to get new attribute text.
        Then get the attribute type through a dialog.
        AddItem dialog checks for duplicate attribute name.
        New attribute is added to the model and database. """

        add_ui = DialogAddAttribute(self.app)
        ok = add_ui.exec()
        if not ok or add_ui.new_name == "":
            return
        name = add_ui.new_name
        value_type = add_ui.value_type

        # Update attribute_type list and database
        now_date = str(datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"))
        sql = "insert into attribute_type (name,date,owner,memo,caseOrFile, valuetype) values(?,?,?,?,?,?)"
        cur = self.app.conn.cursor()
        cur.execute(sql, (name, now_date, self.app.settings['codername'], "", 'case', value_type))
        self.app.conn.commit()
        sql = "select caseid from cases"
        cur.execute(sql)
        case_ids = cur.fetchall()
        for id_ in case_ids:
            sql = "insert into attribute (name, value, id, attr_type, date, owner) values (?,?,?,?,?,?)"
            cur.execute(sql, (name, "", id_[0], 'case', now_date, self.app.settings['codername']))
        self.app.conn.commit()
        self.load_cases_data()
        self.fill_table()
        self.parent_text_edit.append(_("Attribute added to cases: ") + f"{name}, {_('type:')} {value_type}")
        self.app.delete_backup = False

    def import_cases_and_attributes(self):
        """ Get user chosen file as xlxs or csv for importation """

        if self.cases:
            logger.warning(_("Cases have already been created."))
        filename, ok = QtWidgets.QFileDialog.getOpenFileName(None,
                                                             _('Select cases file'),
                                                             self.app.settings['directory'],
                                                             "(*.csv *.CSV *.xlsx *.XLSX)",
                                                             options=QtWidgets.QFileDialog.Option.DontUseNativeDialog
                                                             )
        if filename == "":
            return
        if filename[-4:].lower() == ".csv":
            self.import_csv(filename)
        if filename[-5:].lower() == ".xlsx":
            self.import_xlsx(filename)

    def import_xlsx(self, filepath):
        """ Import from a xlsx file with the cases and any attributes.
        The file must have a header row which details the attribute names.
        The first column must have the case ids.
        The attribute types are calculated from the data.
        """

        data = []
        wb = load_workbook(filename=filepath)
        sheet = wb.active
        for value in sheet.iter_rows(values_only=True):
            # Some rows may be blank so ignore importation
            if (set(value)) != {None}:
                # Values are tuples, convert to list, and remove 'None' string
                row = [item if item else "" for item in value]
                data.append(row)
        now_date = str(datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"))
        # Get field names and replace blanks with a placeholder
        fields = []
        for i, f in enumerate(data[0]):
            if f != '':
                fields.append(f)
            else:
                fields.append("Field_" + str(i))
        data = data[1:]
        # Insert cases
        cur = self.app.conn.cursor()
        for v in data:
            item = {'name': v[0], 'memo': "", 'owner': self.app.settings['codername'],
                    'date': now_date}
            try:
                sql = "insert into cases (name,memo,owner,date) values(?,?,?,?)"
                cur.execute(sql, (item['name'], item['memo'], item['owner'], item['date']))
                self.app.conn.commit()
                cur.execute("select last_insert_rowid()")
                item['caseid'] = cur.fetchone()[0]
                self.cases.append(item)
            except Exception as e:
                logger.error("item:" + str(item) + ", " + str(e))
        # Determine attribute type
        attribute_value_type = ["character"] * len(fields)
        for col, att_name in enumerate(fields):
            numeric = True
            for val in data:
                try:
                    float(val[col])
                except ValueError:
                    numeric = False
            if numeric:
                attribute_value_type[col] = "numeric"
        # Insert attribute types
        for col, att_name in enumerate(fields):
            if col > 0:
                try:
                    sql = "insert into attribute_type (name,date,owner,memo, valueType, caseOrFile) values(?,?,?,?,?,?)"
                    cur.execute(sql, (att_name, now_date, self.app.settings['codername'], "",
                                      attribute_value_type[col], 'case'))
                    self.app.conn.commit()
                except Exception as e:
                    logger.error(_("attribute:") + f"{att_name}, {e}")
        # Insert attributes
        sql = "select name, caseid from cases"
        cur.execute(sql)
        name_and_ids = cur.fetchall()
        for n_i in name_and_ids:
            for v in data:
                if n_i[0] == v[0]:
                    for col in range(1, len(v)):
                        sql = "insert into attribute (name, value, id, attr_type, date, owner) values (?,?,?,?,?,?)"
                        cur.execute(sql, (fields[col], v[col], n_i[1], 'case',
                                          now_date, self.app.settings['codername']))
        self.app.conn.commit()
        self.load_cases_data()
        self.fill_table()
        msg = _("Cases and attributes imported from: ") + filepath
        self.app.delete_backup = False
        self.parent_text_edit.append(msg)
        logger.info(msg)

    def import_csv(self, filepath):
        """ Import from a csv file with the cases and any attributes.
        The csv file must have a header row which details the attribute names.
        The csv file must be comma delimited. The first column must have the case ids.
        The attribute types are calculated from the data.
        """

        values = []
        with open(filepath, 'r', newline='') as f:
            reader = csv.reader(f, delimiter=',', quoting=csv.QUOTE_MINIMAL)
            try:
                for row in reader:
                    values.append(row)
            except csv.Error as e:
                logger.warning(('file %s, line %d: %s' % (filepath, reader.line_num, e)))
        if len(values) <= 1:
            logger.info(_("Cannot import from csv, only one row in file"))
            return
        now_date = str(datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"))
        fields = values[0]
        data = values[1:]
        # Insert cases
        cur = self.app.conn.cursor()
        for v in data:
            item = {'name': v[0], 'memo': "", 'owner': self.app.settings['codername'],
                    'date': now_date}
            try:
                sql = "insert into cases (name,memo,owner,date) values(?,?,?,?)"
                cur.execute(sql, (item['name'], item['memo'], item['owner'], item['date']))
                self.app.conn.commit()
                cur.execute("select last_insert_rowid()")
                item['caseid'] = cur.fetchone()[0]
                self.cases.append(item)
            except Exception as e:
                logger.error(f"item: {item}, {e}")
        # Determine attribute type
        attribute_value_type = ["character"] * len(fields)
        for col, att_name in enumerate(fields):
            numeric = True
            for val in data:
                try:
                    float(val[col])
                except ValueError:
                    numeric = False
            if numeric:
                attribute_value_type[col] = "numeric"
        # Insert attribute types
        for col, att_name in enumerate(fields):
            if col > 0:
                try:
                    sql = "insert into attribute_type (name,date,owner,memo, \
                    valueType, caseOrFile) values(?,?,?,?,?,?)"
                    cur.execute(sql, (att_name, now_date, self.app.settings['codername'], "",
                                      attribute_value_type[col], 'case'))
                    self.app.conn.commit()
                except Exception as e:
                    logger.error(_("attribute:") + f"{att_name}, {e}")
        # Insert attributes
        sql = "select name, caseid from cases"
        cur.execute(sql)
        name_and_ids = cur.fetchall()
        for n_i in name_and_ids:
            for v in data:
                if n_i[0] == v[0]:
                    for col in range(1, len(v)):
                        sql = "insert into attribute (name, value, id, attr_type, date, owner) values (?,?,?,?,?,?)"
                        cur.execute(sql, (fields[col], v[col], n_i[1], 'case',
                                          now_date, self.app.settings['codername']))
        self.app.conn.commit()
        self.load_cases_data()
        self.fill_table()
        msg = _("Cases and attributes imported from: ") + filepath
        self.app.delete_backup = False
        self.parent_text_edit.append(msg)
        logger.info(msg)

    def add_case(self):
        """ When add case button pressed, open addItem dialog to get the case name.
        AddItem dialog checks for duplicate case name.
        New case is added to the model and database.
        Attribute placeholders are assigned to the database for this new case. """

        ui = DialogAddItemName(self.app, self.cases, _("Case"), _("Enter case name"))
        ui.exec()
        case_name = ui.get_new_name()
        if case_name is None:
            return
        # update case list and database
        item = {'name': case_name, 'memo': "", 'owner': self.app.settings['codername'],
                'date': datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"), 'files': [],
                'attributes': []}
        cur = self.app.conn.cursor()
        sql = "insert into cases (name,memo,owner,date) values(?,?,?,?)"
        cur.execute(sql, (item['name'], item['memo'], item['owner'], item['date']))
        self.app.conn.commit()
        cur.execute("select last_insert_rowid()")
        item['caseid'] = cur.fetchone()[0]
        # Add placeholder attribute values
        cur.execute("select name, valuetype from attribute_type where caseOrFile='case'")
        atts = cur.fetchall()
        for att in atts:
            cur.execute("insert into attribute(name,attr_type,value,id,date,owner) \
                values (?,?,?,?,?,?)",
                        (att[0], "case", "", item['caseid'], item['date'], item['owner']))
            item['attributes'].append('')
        self.app.conn.commit()
        self.cases.append(item)
        self.fill_table()
        self.parent_text_edit.append(_("Case added: ") + item['name'])
        self.app.delete_backup = False

    def delete_case(self):
        """ When delete button pressed, case is deleted from model and database. """

        table_rows_to_delete = []  # for table widget ids
        case_names_to_delete = ""  # for confirmDelete Dialog
        ids_to_delete = []  # for ids for cases and db

        for itemWidget in self.ui.tableWidget.selectedItems():
            table_rows_to_delete.append(int(itemWidget.row()))
            ids_to_delete.append(int(self.ui.tableWidget.item(itemWidget.row(),
                                                              self.ID_COLUMN).text()))
            case_names_to_delete = case_names_to_delete + "\n" + str(self.ui.tableWidget.item(itemWidget.row(),
                                                                                              self.NAME_COLUMN).text())
        table_rows_to_delete.sort(reverse=True)
        if len(case_names_to_delete) == 0:
            return
        ui = DialogConfirmDelete(self.app, case_names_to_delete)
        ok = ui.exec()
        if not ok:
            return
        for id_ in ids_to_delete:
            for c in self.cases:
                if c['caseid'] == id_:
                    cur = self.app.conn.cursor()
                    cur.execute("delete from cases where caseid=?", [id_])
                    cur.execute("delete from case_text where caseid=?", [id_])
                    cur.execute("delete from attribute where id=? and attr_type='case'", [id_])
                    self.app.conn.commit()
                    self.parent_text_edit.append(f"Case deleted: {c['name']}")
        self.load_cases_data()
        self.app.delete_backup = False
        self.fill_table()

    def cell_modified(self):
        """ If the case name has been changed in the table widget update the database.
         Cells that can be changed directly are the case name, and attributes. """

        x = self.ui.tableWidget.currentRow()
        y = self.ui.tableWidget.currentColumn()
        if y == self.NAME_COLUMN:  # update case name
            new_text = str(self.ui.tableWidget.item(x, y).text()).strip()
            # check that no other case name has this text and this is not empty
            update = True
            if new_text == "":
                update = False
            for c in self.cases:
                if c['name'] == new_text:
                    update = False
            if update:
                cur = self.app.conn.cursor()
                cur.execute("update cases set name=? where caseid=?", (new_text, self.cases[x]['caseid']))
                self.app.conn.commit()
                self.cases[x]['name'] = new_text
            else:  # put the original text in the cell
                self.ui.tableWidget.item(x, y).setText(self.cases[x]['name'])
        if y > 2:  # update attribute value
            now_date = str(datetime.datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S"))
            value = str(self.ui.tableWidget.item(x, y).text()).strip()
            attribute_name = self.header_labels[y]
            cur = self.app.conn.cursor()
            # Check numeric for numeric attributes, clear "" if cannot be cast
            cur.execute("select valuetype from attribute_type where caseOrFile='case' and name=?", [attribute_name])
            result = cur.fetchone()
            if result is None:
                return
            if result[0] == "numeric":
                try:
                    float(value)
                except ValueError:
                    self.ui.tableWidget.item(x, y).setText("")
                    value = ""
                    msg = _("This attribute is numeric")
                    Message(self.app, _("Warning"), msg, "warning").exec()
            # Check attribute row is present before updating
            cur.execute("select value from attribute where id=? and name=? and attr_type='case'",
                        [self.cases[x]['caseid'], attribute_name])
            res = cur.fetchone()
            if res is None:
                cur.execute("insert into attribute (value,id,name,attr_type, date,owner) values(?,?,?,'case',?,?)",
                            (value, self.cases[x]['caseid'], attribute_name, now_date, self.app.settings['codername']))
                self.app.conn.commit()
            cur.execute("update attribute set value=?, date=?, owner=? where id=? and name=? and attr_type='case'",
                        (value, now_date, self.app.settings['codername'], self.cases[x]['caseid'], attribute_name))
            self.app.conn.commit()

        # Update self.cases[attributes]
        # Add list of attribute values to files, order matches header columns
        sql = "select ifnull(value, '') from attribute where attr_type='case' and attribute.name=? and id=?"
        self.cases[x]['attributes'] = []
        cur = self.app.conn.cursor()
        for att_name in self.attribute_labels_ordered:
            cur.execute(sql, [att_name, self.cases[x]['caseid']])
            res = cur.fetchone()
            if res:
                self.cases[x]['attributes'].append(res[0])

        self.app.delete_backup = False

    def cell_selected(self):
        """ Indicate memo is present, update memo text, or delete memo by clearing text.
        Open case_files_manager if files column is selected.
        """

        self.ui.textBrowser.clear()
        x = self.ui.tableWidget.currentRow()
        y = self.ui.tableWidget.currentColumn()
        if x == -1:
            self.selected_case = None
            self.case_text = []
            return
        self.selected_case = self.cases[x]
        if self.count_selected_items() > 1:
            return

        # logger.debug("Selected case: " + str(self.selected_case['id']) +" "+self.selected_case['name'])'''
        # get case_text for this file
        if self.selected_file is not None:
            # logger.debug("File Selected: " + str(self.selected_file['id'])+"  "+self.selected_file['file'])
            self.case_text = []
            cur = self.app.conn.cursor()
            cur.execute("select caseid, fid, pos0, pos1, owner, date, memo from case_text where fid = ? and caseid = ?",
                        [self.selected_file['id'], self.selected_case['caseid']])
            result = cur.fetchall()
            for row in result:
                self.case_text.append({'caseid': row[0], 'fid': row[1], 'pos0': row[2],
                                       'pos1': row[3], 'owner': row[4], 'date': row[5], 'memo': row[6]})

        if y == self.MEMO_COLUMN:
            ui = DialogMemo(self.app, _("Memo for case ") + self.cases[x]['name'],
                            self.cases[x]['memo'])
            ui.exec()
            self.cases[x]['memo'] = ui.memo
            cur = self.app.conn.cursor()
            cur.execute('update cases set memo=? where caseid=?', (self.cases[x]['memo'], self.cases[x]['caseid']))
            self.app.conn.commit()
            if self.cases[x]['memo'] == "" or self.cases[x]['memo'] is None:
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem())
            else:
                self.ui.tableWidget.setItem(x, self.MEMO_COLUMN, QtWidgets.QTableWidgetItem(_("Memo")))
            self.app.delete_backup = False
        if y == self.FILES_COLUMN:
            self.case_file_manager()

    def case_file_manager(self):
        """ Link files to cases.
         Called by click in files column in table or by button. """

        x = self.ui.tableWidget.currentRow()
        if x == -1:
            return
        ui = DialogCaseFileManager(self.app, self.parent_text_edit, self.cases[x])
        ui.exec()
        # Reload files count
        cur = self.app.conn.cursor()
        sql = "select distinct case_text.fid, source.name from case_text join source on case_text.fid=source.id where "
        sql += "caseid=? order by source.name asc"
        cur.execute(sql, [self.cases[x]['caseid'], ])
        files = cur.fetchall()
        self.cases[x]['files'] = files
        self.fill_table()

    def table_header_menu(self, position):
        """ Used to show and hide columns """

        index_at = self.ui.tableWidget.indexAt(position)
        header_index = int(index_at.column())
        menu = QtWidgets.QMenu(self)
        action_show_all_columns = menu.addAction(_("Show all columns"))
        action_hide_column = None
        if header_index > 0:
            action_hide_column = menu.addAction(_("Hide column"))
        action_hide_columns_starting = menu.addAction(_("Hide columns starting with"))
        action_show_columns_starting = menu.addAction(_("Show columns starting with"))
        action = menu.exec(self.ui.tableWidget.mapToGlobal(position))
        if action == action_show_all_columns:
            for c in range(0, self.ui.tableWidget.columnCount()):
                self.ui.tableWidget.setColumnHidden(c, False)
            if not self.app.settings['showids']:
                self.ui.tableWidget.setColumnHidden(self.ID_COLUMN, True)
            return
        if action == action_hide_column:
            self.ui.tableWidget.setColumnHidden(header_index, True)
            return
        if action == action_hide_columns_starting:
            msg = _("Hide columns starting with:")
            hide_col, ok = QtWidgets.QInputDialog.getText(self, _("Hide Columns"), msg,
                                                            QtWidgets.QLineEdit.EchoMode.Normal)
            for c in range(1, self.ui.tableWidget.columnCount()):
                h_text = self.ui.tableWidget.horizontalHeaderItem(c).text()
                if len(h_text) >= len(hide_col) and hide_col == h_text[:len(hide_col)]:
                    self.ui.tableWidget.setColumnHidden(c, True)
            return
        if action == action_show_columns_starting:
            msg = _("Show columns starting with:")
            show_col, ok = QtWidgets.QInputDialog.getText(self, _("Show Columns"), msg,
                                                            QtWidgets.QLineEdit.EchoMode.Normal)
            for c in range(3, self.ui.tableWidget.columnCount()):
                h_text = self.ui.tableWidget.horizontalHeaderItem(c).text()
                if len(h_text) >= len(show_col) and show_col == h_text[:len(show_col)]:
                    self.ui.tableWidget.setColumnHidden(c, False)
                else:
                    self.ui.tableWidget.setColumnHidden(c, True)

    def table_menu(self, position):
        """ Context menu for displaying table rows in differing order
         and hiding table rows. """

        row = self.ui.tableWidget.currentRow()
        col = self.ui.tableWidget.currentColumn()
        item_text = ""
        try:
            item_text = self.ui.tableWidget.item(row, col).text()
        except AttributeError:  # NoneType error
            pass

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_asc = None
        action_desc = None
        action_view_case = None
        if col == 0:
            action_view_case = menu.addAction(_("View case"))
            action_asc = menu.addAction(_("Order ascending"))
            action_desc = menu.addAction(_("Order descending"))
        action_show_values_like = menu.addAction(_("Show values like"))
        action_equals_value = menu.addAction(_("Show this value"))
        action_order_by_value_asc = None
        action_order_by_value_desc = None
        if col >= self.ATTRIBUTE_START_COLUMN:
            action_order_by_value_asc = menu.addAction(_("Order ascending"))
            action_order_by_value_desc = menu.addAction(_("Order descending"))
        action_show_all = menu.addAction(_("Show all rows Ctrl A"))
        action_url = None
        url_test = urlparse(item_text)
        if all([url_test.scheme, url_test.netloc]):
            action_url = menu.addAction(_("Open URL"))
        action = menu.exec(self.ui.tableWidget.mapToGlobal(position))
        if action is None:
            return
        if action == action_view_case:
            self.ui.splitter.setSizes([1, 1])
            self.view_case_files()
            return
        if action == action_asc:
            self.load_cases_data("asc")
            self.fill_table()
        if action == action_desc:
            self.load_cases_data("desc")
            self.fill_table()
        if action == action_order_by_value_asc:
            self.load_cases_data("attribute asc:" + self.header_labels[col])
        if action == action_order_by_value_desc:
            self.load_cases_data("attribute desc:" + self.header_labels[col])
        if action == action_equals_value:
            # Hide rows that do not match this value
            item_to_compare = self.ui.tableWidget.item(row, col)
            if item_to_compare is None:
                item_to_compare = ""
            compare_text = item_to_compare.text()
            for r in range(0, self.ui.tableWidget.rowCount()):
                item = self.ui.tableWidget.item(r, col)
                text_ = item.text()
                if compare_text != text_:
                    self.ui.tableWidget.setRowHidden(r, True)
            return
        if action == action_show_values_like:
            text_value, ok = QtWidgets.QInputDialog.getText(self, _("Text filter"), _("Show values like:"),
                                                       QtWidgets.QLineEdit.EchoMode.Normal)
            if ok and text_value != '':
                for r in range(0, self.ui.tableWidget.rowCount()):
                    if self.ui.tableWidget.item(r, col).text().find(text_value) == -1:
                        self.ui.tableWidget.setRowHidden(r, True)
            return
        if action == action_show_all:
            for r in range(0, self.ui.tableWidget.rowCount()):
                self.ui.tableWidget.setRowHidden(r, False)
        if action == action_url:
            webbrowser.open(item_text)

    def fill_table(self):
        """ Fill the table widget with case details. """

        self.update_label()
        self.ui.tableWidget.blockSignals(True)
        self.ui.tableWidget.setColumnCount(len(self.header_labels))
        rows = self.ui.tableWidget.rowCount()
        for c in range(0, rows):
            self.ui.tableWidget.removeRow(0)
        self.ui.tableWidget.setHorizontalHeaderLabels(self.header_labels)
        self.ui.tableWidget.setRowCount(len(self.cases))
        for row in range(0, len(self.cases)):
            c = self.cases[row]
            self.ui.tableWidget.setItem(row, self.NAME_COLUMN,
                                        QtWidgets.QTableWidgetItem(c['name']))
            item = QtWidgets.QTableWidgetItem("")
            if c['memo'] != "":
                item = QtWidgets.QTableWidgetItem(_("Memo"))
            item.setToolTip(_("Click to edit memo"))
            self.ui.tableWidget.setItem(row, self.MEMO_COLUMN, item)
            item = QtWidgets.QTableWidgetItem(str(c['caseid']))
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled)
            self.ui.tableWidget.setItem(row, self.ID_COLUMN, item)
            # Number of files assigned to case
            item = QtWidgets.QTableWidgetItem(str(len(c['files'])))
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled)
            item.setToolTip(_("Click to manage files for this case"))
            self.ui.tableWidget.setItem(row, self.FILES_COLUMN, item)
            # Add attribute values to their columns
            for offset, attribute in enumerate(c['attributes']):
                item = QtWidgets.QTableWidgetItem(attribute)
                self.ui.tableWidget.setItem(row, self.ATTRIBUTE_START_COLUMN + offset, item)
        self.ui.tableWidget.verticalHeader().setVisible(False)
        self.ui.tableWidget.resizeRowsToContents()
        self.ui.tableWidget.hideColumn(self.ID_COLUMN)
        if self.app.settings['showids']:
            self.ui.tableWidget.showColumn(self.ID_COLUMN)
        # Add statistics tooltips to table headers for attributes
        for i, attribute_name in enumerate(self.attribute_labels_ordered):
            tt = self.get_tooltip_values(attribute_name)
            self.ui.tableWidget.horizontalHeaderItem(self.ATTRIBUTE_START_COLUMN + i).setToolTip(_("Right click header row to hide columns") + f"\n{tt}")
        self.ui.tableWidget.blockSignals(False)

    def get_tooltip_values(self, attribute_name):
        """ Get values to display in tooltips for the value list column.
        param: attribute_name : String """

        tt = ""
        cur = self.app.conn.cursor()
        sql_val_type = 'select valuetype from attribute_type where caseOrFile="case" and name=?'
        cur.execute(sql_val_type, [attribute_name])
        res_val_type = cur.fetchone()
        value_type = "character"
        if res_val_type is not None:
            value_type = res_val_type[0]
        if value_type == "numeric":
            sql = 'select min(cast(value as real)), max(cast(value as real)) from attribute where name=? and ' \
                  'attr_type="case"'
            cur.execute(sql, [attribute_name])
            res = cur.fetchone()
            tt = f"{_('Minimum:')} {res[0]}\n{_('Maximum:')} {res[1]}"
        if value_type == "character":
            sql = 'select distinct value from attribute where name=? and attr_type="case" and length(value)>0 limit 10'
            cur.execute(sql, [attribute_name])
            res = cur.fetchall()
            for r in res:
                tt += f"\n{r[0]}"
            if len(tt) > 1:
                tt = tt[1:]
        return tt

    def view_case_files(self):
        """ View all the text associated with this case.
        Add links to open image and A/V files. """

        row = self.ui.tableWidget.currentRow()
        if row == -1:
            return
        if self.selected_case is None:
            return
        self.ui.textBrowser.clear()
        self.ui.textBrowser.setPlainText("")
        self.selected_file = None
        self.ui.label_filename.setText(_("Viewing text of case: ") + str(self.cases[row]['name']))
        display_text = []
        self.display_text_links = []
        cur = self.app.conn.cursor()
        cur.execute(
            "select caseid, fid, pos0, pos1, owner, date, memo from case_text where caseid = ? order by fid, pos0",
            [self.selected_case['caseid'], ])
        result = cur.fetchall()
        for row in result:
            case_text = ""
            filename = ""
            mediapath = ""
            av_text_id = None
            for src in self.source:
                if src['id'] == row[1] and src['fulltext'] is not None:
                    case_text = src['fulltext'][int(row[2]):int(row[3])]
                    filename = src['name']
                if src['id'] == row[1] and src['fulltext'] is None:
                    filename = src['name']
                    mediapath = src['mediapath']
                    av_text_id = src['av_text_id']
            display_text.append({'caseid': row[0], 'fid': row[1], 'pos0': row[2],
                                 'pos1': row[3], 'owner': row[4], 'date': row[5], 'memo': row[6],
                                 'text': case_text, 'name': filename, 'mediapath': mediapath,
                                 'av_text_id': av_text_id})

        format_reg = QtGui.QTextCharFormat()
        brush = QtGui.QBrush(QtGui.QColor(QtCore.Qt.GlobalColor.black))
        if self.app.settings['stylesheet'] in ('dark', 'rainbow'):
            brush = QtGui.QBrush(QtGui.QColor("#eeeeee"))
        format_reg.setForeground(brush)

        format_bold = QtGui.QTextCharFormat()
        format_bold.setFontWeight(QtGui.QFont.Weight.Bold)
        brush_bold = QtGui.QBrush(QtGui.QColor(QtCore.Qt.GlobalColor.black))
        if self.app.settings['stylesheet'] in ('dark', 'rainbow'):
            brush_bold = QtGui.QBrush(QtGui.QColor("#eeeeee"))
        format_bold.setForeground(brush_bold)

        format_blue = QtGui.QTextCharFormat()
        format_blue.setFontWeight(QtGui.QFont.Weight.Bold)
        # This blue colour good on dark and light background
        format_blue.setForeground(QtGui.QBrush(QtGui.QColor("#9090e3")))

        cursor = self.ui.textBrowser.textCursor()
        for c in display_text:
            if c['mediapath'] is None or c['mediapath'] == '' or c['mediapath'][:5] == "docs:":  # text source
                header = f"\n{_('File:')} {c['name']}, {_('Characters:')} {c['pos1']} - {c['pos0']}"
                pos0 = len(self.ui.textBrowser.toPlainText())
                self.ui.textBrowser.append(header)
                cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
                pos1 = len(self.ui.textBrowser.toPlainText())
                cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
                cursor.setCharFormat(format_bold)
                pos0 = len(self.ui.textBrowser.toPlainText())
                self.ui.textBrowser.append(c['text'])
                cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
                pos1 = len(self.ui.textBrowser.toPlainText())
                cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
                cursor.setCharFormat(format_reg)

            if c['mediapath'][:7] in ("/images", "images:"):
                header = f"\n{_('Image:')} {c['name']}"
                pos0 = len(self.ui.textBrowser.toPlainText())
                self.ui.textBrowser.append(header)
                cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
                pos1 = len(self.ui.textBrowser.toPlainText())
                cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
                cursor.setCharFormat(format_blue)
                data = {'pos0': pos0, 'pos1': pos1, 'id': c['fid'], 'mediapath': c['mediapath'],
                        'owner': c['owner'], 'date': c['date'], 'memo': c['memo'], 'name': c['name']}
                self.display_text_links.append(data)

            if c['mediapath'][:6] in ("/audio", "audio:", "/video", "video:"):
                header = f"\n{_('AV media:')} {c['name']}"
                pos0 = len(self.ui.textBrowser.toPlainText())
                self.ui.textBrowser.append(header)
                cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
                pos1 = len(self.ui.textBrowser.toPlainText())
                cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
                cursor.setCharFormat(format_blue)
                data = {'pos0': pos0, 'pos1': pos1, 'id': c['fid'], 'mediapath': c['mediapath'],
                        'owner': c['owner'], 'date': c['date'], 'memo': c['memo'], 'name': c['name'],
                        'av_text_id': c['av_text_id']}
                self.display_text_links.append(data)
        self.eventFilterTT.set_positions(self.display_text_links)  # uses pos0, pos1

    def link_clicked(self, position):
        """ View image or audio/video media in dialog.
        For A/V, added try block in case VLC bindings do not work.
        Also check existence of media, as particularly, linked files may have bad links. """

        cursor = self.ui.textBrowser.cursorForPosition(position)
        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_link = None
        for item in self.display_text_links:
            if item['pos0'] <= cursor.position() <= item['pos1']:
                action_link = menu.addAction(_("Open"))
        action = menu.exec(self.ui.textBrowser.mapToGlobal(position))
        if action is None:
            return

        for item in self.display_text_links:
            if item['pos0'] <= cursor.position() <= item['pos1']:
                ui = None
                if item['mediapath'][:6] == "video:":
                    abs_path = item['mediapath'].split(':')[1]
                    if not os.path.exists(abs_path):
                        return
                    ui = DialogViewAV(self.app, item)
                if item['mediapath'][:6] == "/video":
                    abs_path = self.app.project_path + item['mediapath']
                    if not os.path.exists(abs_path):
                        return
                    ui = DialogViewAV(self.app, item)
                if item['mediapath'][:6] == "audio:":
                    abs_path = item['mediapath'].split(':')[1]
                    if not os.path.exists(abs_path):
                        return
                    ui = DialogViewAV(self.app, item)
                if item['mediapath'][0:6] == "/audio":
                    abs_path = self.app.project_path + item['mediapath']
                    if not os.path.exists(abs_path):
                        return
                    ui = DialogViewAV(self.app, item)
                if item['mediapath'][0:7] == "images:":
                    abs_path = item['mediapath'].split(':')[1]
                    if not os.path.exists(abs_path):
                        return
                    ui = DialogViewImage(self.app, item)
                if item['mediapath'][0:7] == "/images":
                    abs_path = self.app.project_path + item['mediapath']
                    if not os.path.exists(abs_path):
                        return
                    ui = DialogViewImage(self.app, item)
                ui.exec()

    def text_edit_menu(self, position):
        """ Context menu for text Edit. Select all, Copy. """

        menu = QtWidgets.QMenu()
        action_select_all = menu.addAction(_("Select all"))
        action_copy = menu.addAction(_("Copy"))
        action = menu.exec(self.ui.textBrowser.mapToGlobal(position))
        if action == action_select_all:
            self.ui.textBrowser.selectAll()
        if action == action_copy:
            selected_text = self.ui.textBrowser.textCursor().selectedText()
            cb = QtWidgets.QApplication.clipboard()
            cb.setText(selected_text)


class ToolTipEventFilter(QtCore.QObject):
    """ Used to add a dynamic tooltip for the textBrowser.
    The tool top text is presented according to its position in the text.
    """

    media_data = None

    def set_positions(self, media_data):
        """ Code_text contains the positions for the tooltip to be displayed.

        param:
            media_data: List of dictionaries of the text contains: pos0, pos1
        """

        self.media_data = media_data

    def eventFilter(self, receiver, event):
        # QtGui.QToolTip.showText(QtGui.QCursor.pos(), tip)
        # Added check for media_data, it may be None
        if event.type() == QtCore.QEvent.Type.ToolTip and self.media_data:
            help_event = event
            cursor = receiver.cursorForPosition(help_event.pos())
            pos = cursor.position()
            receiver.setToolTip("")
            for item in self.media_data:
                if item['pos0'] <= pos <= item['pos1']:
                    receiver.setToolTip(_("Right click to view"))
        # Call Base Class Method to Continue Normal Event Processing
        return super(ToolTipEventFilter, self).eventFilter(receiver, event)
