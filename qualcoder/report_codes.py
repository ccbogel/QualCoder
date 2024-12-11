# -*- coding: utf-8 -*-

"""
This file is part of QualCoder.

QualCoder is free software: you can redistribute it and/or modify it under the
terms of the GNU Lesser General Public License as published by the Free Software
Foundation, either version 3 of the License, or (at your option) any later version.

QualCoder is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
See the GNU General Public License for more details.

You should have received a copy of the GNU Lesser General Public License along with QualCoder.
If not, see <https://www.gnu.org/licenses/>.

Author: Colin Curtain (ccbogel)
https://github.com/ccbogel/QualCoder
https://qualcoder.wordpress.com/
"""
import sqlite3
from copy import deepcopy
import csv
import logging
import openpyxl
import os
from PIL import Image
import qtawesome as qta  # See https://pictogrammers.com/library/mdi/
import re
from shutil import copyfile

from PyQt6 import QtGui, QtWidgets, QtCore
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QBrush

from .color_selector import TextColor
from .confirm_delete import DialogConfirmDelete
from .GUI.ui_dialog_report_codings import Ui_Dialog_reportCodings
from .helpers import Message, msecs_to_hours_mins_secs, DialogCodeInImage, DialogCodeInAV, DialogCodeInText, \
    ExportDirectoryPathDialog
from .report_attributes import DialogSelectAttributeParameters
from .select_items import DialogSelectItems

# If VLC not installed, it will not crash
vlc = None
try:
    import vlc
except Exception as e:
    print(e)

path = os.path.abspath(os.path.dirname(__file__))
logger = logging.getLogger(__name__)


class DialogReportCodes(QtWidgets.QDialog):
    """ Get reports on coded text/images/audio/video using a range of variables:
        Files, Cases, Coders, text limiters, Attribute limiters.
        Export reports as plain text, ODT, html, xlsx or csv.

        Text context of a coded text portion is shown in the third splitter panel in a text edit.
        Case matrix is also shown in a qtablewidget in the third splitter pane.
        If a case matrix is displayed, the text-in-context method overrides it and replaces the matrix with the
        text in context.
    """

    app = None
    parent_textEdit = None
    code_names = []
    coders = [""]
    categories = []
    files = []
    cases = []
    results = []
    # html results need media links {imagename, QImage, char_pos, avname, av0, av1, avtext}
    html_links = []
    te = []  # Matrix (table) [row][col] of textEditWidget results
    # Variables for search restrictions
    file_ids_string = ""
    case_ids_string = ""
    attributes = []
    attribute_file_ids = []
    attribute_case_ids = []
    attributes_msg = ""
    # Text positions in the main textEdit for right-click context menu to View original file
    text_links = []
    # Text positions in the matrix textEdits for right-click context menu to View original file
    # list of dictionaries of row, col, textEdit, list of links
    matrix_links = []

    def __init__(self, app, parent_textedit, tab_coding):
        super(DialogReportCodes, self).__init__()
        self.app = app
        self.parent_textEdit = parent_textedit
        self.tab_coding = tab_coding
        self.get_codes_categories_coders()
        QtWidgets.QDialog.__init__(self)
        self.ui = Ui_Dialog_reportCodings()
        self.ui.setupUi(self)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        font = f'font: {self.app.settings["fontsize"]}pt "{self.app.settings["font"]}";'
        self.setStyleSheet(font)
        tree_font = f'font: {self.app.settings["treefontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.treeWidget.setStyleSheet(tree_font)
        doc_font = f'font: {self.app.settings["docfontsize"]}pt "{self.app.settings["font"]}";'
        self.ui.textEdit.setStyleSheet(doc_font)
        self.ui.treeWidget.installEventFilter(self)  # For H key
        self.ui.listWidget_files.setStyleSheet(tree_font)
        self.ui.listWidget_files.installEventFilter(self)  # For H key
        self.ui.listWidget_files.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.ui.listWidget_cases.setStyleSheet(tree_font)
        self.ui.listWidget_cases.installEventFilter(self)  # For H key
        self.ui.listWidget_cases.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.ui.treeWidget.setSelectionMode(QtWidgets.QTreeWidget.SelectionMode.ExtendedSelection)
        self.ui.comboBox_coders.insertItems(0, self.coders)
        self.fill_tree()
        self.ui.pushButton_run_report.clicked.connect(self.search)
        self.ui.pushButton_run_report.setIcon(qta.icon('mdi6.play'))
        self.ui.label_exports.setPixmap(qta.icon('mdi6.export').pixmap(22, 22))
        self.ui.pushButton_attributeselect.setIcon(qta.icon('mdi6.line-scan'))
        self.ui.pushButton_search_next.setIcon(qta.icon('mdi6.arrow-right'))
        self.ui.pushButton_search_next.pressed.connect(self.search_results_next)
        options = ["", _("Top categories by case"), _("Top categories by file"), _("Categories by case"),
                   _("Categories by file"), _("Codes by case"), _("Codes by file")]
        self.ui.comboBox_matrix.addItems(options)
        self.ui.label_memos.setPixmap(qta.icon('mdi6.text-box-outline').pixmap(22, 22))
        options = [_("None"), _("Also code memos"), _("Also coded memos"), _("Also all memos"), _("Only memos"),
                   _("Only coded memos"), _("Annotations"), _("Codebook memos")]
        self.ui.comboBox_memos.addItems(options)
        cur = self.app.conn.cursor()
        sql = "select count(name) from attribute_type"
        cur.execute(sql)
        res = cur.fetchone()
        if res[0] == 0:
            self.ui.pushButton_attributeselect.setEnabled(False)
        self.ui.pushButton_attributeselect.clicked.connect(self.select_attributes)
        self.ui.comboBox_export.currentIndexChanged.connect(self.export_option_selected)
        self.ui.comboBox_export.setEnabled(False)
        self.ui.textEdit.installEventFilter(self)
        self.ui.textEdit.setReadOnly(True)
        self.ui.textEdit.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.textEdit.customContextMenuRequested.connect(self.text_edit_menu)
        self.ui.splitter.setSizes([100, 200, 0])
        try:
            s0 = int(self.app.settings['dialogreportcodes_splitter0'])
            s1 = int(self.app.settings['dialogreportcodes_splitter1'])
            if s0 > 10 and s1 > 10:
                self.ui.splitter.setSizes([s0, s1, 0])
            v0 = int(self.app.settings['dialogreportcodes_splitter_v0'])
            if v0 < 10:
                v0 = 10
            v1 = int(self.app.settings['dialogreportcodes_splitter_v1'])
            if v1 < 10:
                v1 = 10
            v2 = int(self.app.settings['dialogreportcodes_splitter_v2'])
            if v2 < 10:
                v2 = 10
            self.ui.splitter_vert.setSizes([v0, v1, v2])
        except KeyError:
            pass
        self.ui.splitter.splitterMoved.connect(self.splitter_sizes)
        self.ui.splitter_vert.splitterMoved.connect(self.splitter_sizes)
        self.get_files_and_cases()
        self.ui.listWidget_files.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.listWidget_files.customContextMenuRequested.connect(self.listwidget_files_menu)
        self.ui.listWidget_cases.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.listWidget_cases.customContextMenuRequested.connect(self.listwidget_cases_menu)
        self.ui.treeWidget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.ui.treeWidget.customContextMenuRequested.connect(self.treewidget_menu)
        self.eventFilterTT = ToolTipEventFilter()
        self.ui.textEdit.installEventFilter(self.eventFilterTT)

    def splitter_sizes(self):
        """ Detect size changes in splitter and store in app.settings variable. """

        sizes = self.ui.splitter.sizes()
        self.app.settings['dialogreportcodes_splitter0'] = sizes[0]
        self.app.settings['dialogreportcodes_splitter1'] = sizes[1]
        sizes_vert = self.ui.splitter_vert.sizes()
        self.app.settings['dialogreportcodes_splitter_v0'] = max(sizes_vert[0], 10)
        self.app.settings['dialogreportcodes_splitter_v1'] = max(sizes_vert[1], 10)
        self.app.settings['dialogreportcodes_splitter_v2'] = max(sizes_vert[2], 10)

    def get_files_and_cases(self):
        """ Get source files with additional details and fill files list widget.
        Get cases and fill case list widget
        Called from : init, manage_files.delete manage_files.delete_button_multiple_files
        """

        self.ui.listWidget_files.clear()
        self.files = self.app.get_filenames()
        # Fill additional details about each file in the memo
        cur = self.app.conn.cursor()
        sql = "select length(fulltext), mediapath from source where id=?"
        sql_text_codings = "select count(cid) from code_text where fid=?"
        sql_av_codings = "select count(cid) from code_av where id=?"
        sql_image_codings = "select count(cid) from code_image where id=?"
        item = QtWidgets.QListWidgetItem("")
        item.setToolTip(_("No file selection"))
        self.ui.listWidget_files.addItem(item)
        for f in self.files:
            cur.execute(sql, [f['id'], ])
            res = cur.fetchone()
            if res is None:  # safety catch
                res = [0]
            tt = ""
            if res[1] is None or res[1][0:5] == "docs:":
                tt += _("Text file\n")
                tt += _("Characters: ") + str(res[0])
            if res[1] is not None and (res[1][0:7] == "images:" or res[1][0:7] == "/images"):
                tt += _("Image")
            if res[1] is not None and (res[1][0:6] == "audio:" or res[1][0:6] == "/audio"):
                tt += _("Audio")
            if res[1] is not None and (res[1][0:6] == "video:" or res[1][0:6] == "/video"):
                tt += _("Video")
            cur.execute(sql_text_codings, [f['id']])
            txt_res = cur.fetchone()
            cur.execute(sql_av_codings, [f['id']])
            av_res = cur.fetchone()
            cur.execute(sql_image_codings, [f['id']])
            img_res = cur.fetchone()
            tt += _("\nCodings: ")
            if txt_res[0] > 0:
                tt += str(txt_res[0])
            if av_res[0] > 0:
                tt += str(av_res[0])
            if img_res[0] > 0:
                tt += str(img_res[0])
            item = QtWidgets.QListWidgetItem(f['name'])
            if f['memo'] != "":
                tt += _("\nMEMO: ") + f['memo']
            item.setToolTip(tt)
            self.ui.listWidget_files.addItem(item)

        self.ui.listWidget_cases.clear()
        self.cases = self.app.get_casenames()
        item = QtWidgets.QListWidgetItem("")
        item.setToolTip(_("No case selection"))
        self.ui.listWidget_cases.addItem(item)
        for c in self.cases:
            tt = ""
            item = QtWidgets.QListWidgetItem(c['name'])
            if c['memo'] != "":
                tt = _("MEMO: ") + c['memo']
            item.setToolTip(tt)
            self.ui.listWidget_cases.addItem(item)

    def get_codes_categories_coders(self):
        """ Called from init, delete category. Load codes, categories, and coders. """

        self.code_names, self.categories = self.app.get_codes_categories()
        cur = self.app.conn.cursor()
        self.coders = []
        cur.execute("select distinct owner from code_text")
        result = cur.fetchall()
        self.coders = [""]
        for row in result:
            self.coders.append(row[0])

    def get_selected_files_and_cases(self):
        """ Fill file_ids and case_ids Strings used in the search.
         Called by: search """

        selected_files = []
        self.file_ids_string = ""
        for item in self.ui.listWidget_files.selectedItems():
            selected_files.append(item.text())
            for f in self.files:
                if f['name'] == item.text():
                    self.file_ids_string += f",{f['id']}"
        if len(self.file_ids_string) > 0:
            self.file_ids_string = self.file_ids_string[1:]
        selected_cases = []
        self.case_ids_string = ""
        for item in self.ui.listWidget_cases.selectedItems():
            selected_cases.append(item.text())
            for c in self.cases:
                if c['name'] == item.text():
                    self.case_ids_string += f",{c['id']}"
        if len(self.case_ids_string) > 0:
            self.case_ids_string = self.case_ids_string[1:]

    def listwidget_files_menu(self, position):
        """ Context menu for file selection. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_all_files = menu.addAction(_("Select all files"))
        action_files_like = menu.addAction(_("Select files like"))
        action_files_none = menu.addAction(_("Select none"))
        action = menu.exec(self.ui.listWidget_files.mapToGlobal(position))
        if action == action_all_files:
            self.ui.listWidget_files.selectAll()
            self.ui.listWidget_files.item(0).setSelected(False)
        if action == action_files_none:
            for i in range(self.ui.listWidget_files.count()):
                self.ui.listWidget_files.item(i).setSelected(False)
        if action == action_files_like:
            # Input dialog narrow, so code below
            dialog = QtWidgets.QInputDialog(None)
            dialog.setStyleSheet("* {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
            dialog.setWindowTitle(_("Select some files"))
            dialog.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
            dialog.setInputMode(QtWidgets.QInputDialog.InputMode.TextInput)
            dialog.setLabelText(_("Show files containing text"))
            dialog.resize(200, 20)
            ok = dialog.exec()
            if not ok:
                return
            dlg_text = str(dialog.textValue())
            for i in range(self.ui.listWidget_files.count()):
                item_name = self.ui.listWidget_files.item(i).text()
                if dlg_text in item_name:
                    self.ui.listWidget_files.item(i).setSelected(True)
                else:
                    self.ui.listWidget_files.item(i).setSelected(False)

    def listwidget_cases_menu(self, position):
        """ Context menu for case selection. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_all_cases = menu.addAction(_("Select all cases"))
        action_cases_like = menu.addAction(_("Select cases like"))
        action_cases_none = menu.addAction(_("Select none"))
        action = menu.exec(self.ui.listWidget_cases.mapToGlobal(position))
        if action == action_all_cases:
            self.ui.listWidget_cases.selectAll()
            self.ui.listWidget_cases.item(0).setSelected(False)
        if action == action_cases_none:
            for i in range(self.ui.listWidget_cases.count()):
                self.ui.listWidget_cases.item(i).setSelected(False)
        if action == action_cases_like:
            # Input dialog narrow, so code below
            dialog = QtWidgets.QInputDialog(None)
            dialog.setStyleSheet("* {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
            dialog.setWindowTitle(_("Select some cases"))
            dialog.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
            dialog.setInputMode(QtWidgets.QInputDialog.InputMode.TextInput)
            dialog.setLabelText(_("Select cases containing text"))
            dialog.resize(200, 20)
            ok = dialog.exec()
            if not ok:
                return
            text_ = str(dialog.textValue())
            for i in range(self.ui.listWidget_cases.count()):
                item_name = self.ui.listWidget_cases.item(i).text()
                if text_ in item_name:
                    self.ui.listWidget_cases.item(i).setSelected(True)
                else:
                    self.ui.listWidget_cases.item(i).setSelected(False)

    def fill_tree(self):
        """ Fill tree widget, top level items are main categories and unlinked codes. """

        cats = deepcopy(self.categories)
        codes = deepcopy(self.code_names)
        self.ui.treeWidget.clear()
        self.ui.treeWidget.setColumnCount(4)
        self.ui.treeWidget.setHeaderLabels([_("Name"), "Id", _("Memo"), _("Count")])
        self.ui.treeWidget.header().setToolTip(_("Codes and categories"))
        if not self.app.settings['showids']:
            self.ui.treeWidget.setColumnHidden(1, True)
        else:
            self.ui.treeWidget.setColumnHidden(1, False)
        self.ui.treeWidget.header().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.ui.treeWidget.header().setStretchLastSection(False)
        # Add top level categories
        remove_list = []
        for c in cats:
            if c['supercatid'] is None:
                memo = ""
                if c['memo'] != "":
                    memo = _("Memo")
                top_item = QtWidgets.QTreeWidgetItem([c['name'], 'catid:' + str(c['catid']), memo])
                top_item.setToolTip(0, '')
                if len(c['name']) > 52:
                    top_item.setText(0, f"{c['name'][:25]}..{c['name'][-25:]}")
                    top_item.setToolTip(0, c['name'])
                top_item.setToolTip(2, c['memo'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_list.append(c)
        for item in remove_list:
            cats.remove(item)

        ''' Add child categories. Look at each unmatched category, iterate through tree
        to add as child then remove matched categories from the list. '''
        count = 0
        while len(cats) > 0 and count < 10000:
            remove_list = []
            for c in cats:
                it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
                item = it.value()
                count2 = 0
                while item and count2 < 10000:  # while there is an item in the list
                    if item.text(1) == 'catid:' + str(c['supercatid']):
                        memo = ""
                        if c['memo'] != "":
                            memo = "Memo"
                        child = QtWidgets.QTreeWidgetItem([c['name'], f"catid:{c['catid']}", memo])
                        child.setToolTip(0, '')
                        if len(c['name']) > 52:
                            child.setText(0, f"{c['name'][:25]}..{c['name'][-25:]}")
                            child.setToolTip(0, c['name'])
                        child.setToolTip(2, c['memo'])
                        item.addChild(child)
                        remove_list.append(c)
                    it += 1
                    item = it.value()
                    count2 += 1
            for item in remove_list:
                cats.remove(item)
            count += 1

        # Add unlinked codes as top level items
        remove_items = []
        for c in codes:
            if c['catid'] is None:
                memo = ""
                if c['memo'] != "":
                    memo = "Memo"
                top_item = QtWidgets.QTreeWidgetItem([c['name'], f"cid:{c['cid']}", memo])
                top_item.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                color = TextColor(c['color']).recommendation
                top_item.setForeground(0, QBrush(QtGui.QColor(color)))
                top_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                top_item.setToolTip(0, '')
                if len(c['name']) > 52:
                    top_item.setText(0, f"{c['name'][:25]}..{c['name'][-25:]}")
                    top_item.setToolTip(0, c['name'])
                top_item.setToolTip(2, c['memo'])
                self.ui.treeWidget.addTopLevelItem(top_item)
                remove_items.append(c)
        for item in remove_items:
            codes.remove(item)

        # Add codes as children
        for c in codes:
            it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
            item = it.value()
            count = 0
            while item and count < 10000:
                if item.text(1) == 'catid:' + str(c['catid']):
                    memo = ""
                    if c['memo'] != "":
                        memo = _("Memo")
                    child = QtWidgets.QTreeWidgetItem([c['name'], f"cid:{c['cid']}", memo])
                    child.setBackground(0, QBrush(QtGui.QColor(c['color']), Qt.BrushStyle.SolidPattern))
                    color = TextColor(c['color']).recommendation
                    child.setForeground(0, QBrush(QtGui.QColor(color)))
                    child.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                    child.setToolTip(0, '')
                    if len(c['name']) > 52:
                        child.setText(0, f"{c['name'][:25]}..{c['name'][-25:]}")
                        child.setToolTip(0, c['name'])
                    child.setToolTip(2, c['memo'])
                    item.addChild(child)
                    c['catid'] = -1  # make unmatchable
                it += 1
                item = it.value()
                count += 1
        self.ui.treeWidget.sortByColumn(0, QtCore.Qt.SortOrder.AscendingOrder)
        self.fill_code_counts_in_tree()
        self.ui.treeWidget.expandAll()

    def fill_code_counts_in_tree(self):
        """ Count instances of each code from all coders and all files. """

        cur = self.app.conn.cursor()
        sql = "select count(cid) from code_text where cid=? union "
        sql += "select count(cid) from code_av where cid=? union "
        sql += "select count(cid) from code_image where cid=?"
        it = QtWidgets.QTreeWidgetItemIterator(self.ui.treeWidget)
        item = it.value()
        count = 0
        while item and count < 10000:
            if item.text(1)[0:4] == "cid:":
                cid = str(item.text(1)[4:])
                cur.execute(sql, [cid, cid, cid])  # , self.app.settings['codername']])
                result = cur.fetchall()
                total = 0
                for row in result:
                    total = total + row[0]
                if total > 0:
                    item.setText(3, str(total))
                else:
                    item.setText(3, "")
            it += 1
            item = it.value()
            count += 1

    def treewidget_menu(self, position):
        """ Menu to select all codes or other selection parameters. """

        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
        action_all = menu.addAction(_("Select all codes"))
        action_unselect = menu.addAction(_("Remove selections"))
        action_like = menu.addAction(_("Select codes like"))
        action = menu.exec(self.ui.treeWidget.mapToGlobal(position))
        if action == action_all:
            self.ui.treeWidget.selectAll()
        if action == action_unselect:
            selected = self.ui.treeWidget.selectedItems()
            for tree_item in selected:
                tree_item.setSelected(False)
        if action == action_like:
            # Need to unselect where mouse click occurred
            clicked_selected = self.ui.treeWidget.selectedItems()[0]
            clicked_selected.setSelected(False)
            # Input dialog narrow, so code below
            dialog = QtWidgets.QInputDialog(None)
            dialog.setStyleSheet("* {font-size:" + str(self.app.settings['fontsize']) + "pt} ")
            dialog.setWindowTitle(_("Select some codes"))
            dialog.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
            dialog.setInputMode(QtWidgets.QInputDialog.InputMode.TextInput)
            dialog.setLabelText(_("Select codes containing text"))
            dialog.resize(200, 20)
            ok = dialog.exec()
            if not ok:
                return
            selection_text = str(dialog.textValue())
            tree_items = self.ui.treeWidget.findItems(selection_text, Qt.MatchFlag.MatchContains | Qt.MatchFlag.MatchRecursive, 0)
            for tree_item in tree_items:
                if 'cid' in tree_item.text(1):
                    tree_item.setSelected(True)

    def export_option_selected(self):
        """ ComboBox export option selected. """

        text_ = self.ui.comboBox_export.currentText()
        if text_ == "":
            return
        if text_ == "html":
            self.export_html_file()
        if text_ == "odt":
            self.export_odt_file()
        if text_ == "txt":
            self.export_text_file()
        if text_ == "csv":
            self.export_csv_file()
        if text_ == "xlsx":
            self.export_xlsx_file()
        self.ui.comboBox_export.setCurrentIndex(0)
        if self.te:
            reply = QtWidgets.QMessageBox.question(self, _("Export Matrix"), _("Export matrix results"),
                                                   QtWidgets.QMessageBox.StandardButton.Yes,
                                                   QtWidgets.QMessageBox.StandardButton.No)
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                self.export_matrix()

    def export_matrix(self):
        """ Export matrix as xlsx spreadsheet. """

        row_count = self.ui.tableWidget.rowCount()
        col_count = self.ui.tableWidget.columnCount()
        if row_count == 0 or col_count == 0:
            return
        filename = "Report_matrix.xlsx"
        exp_dlg = ExportDirectoryPathDialog(self.app, filename)
        filepath = exp_dlg.filepath
        if filepath is None:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        # Column header
        for c in range(0, col_count):
            cell = ws.cell(row=1, column=c + 2)
            cell.value = self.ui.tableWidget.horizontalHeaderItem(c).text()
        # Row header
        for r in range(0, row_count):
            cell = ws.cell(row=r + 2, column=1)
            cell.value = self.ui.tableWidget.verticalHeaderItem(r).text()
        # Data
        for c in range(0, col_count):
            for r in range(0, row_count):
                te = self.te[r][c]
                try:
                    data_text = te.toPlainText()
                except AttributeError:  # None type error
                    data_text = ""
                cell = ws.cell(row=r + 2, column=c + 2)
                cell.value = data_text
        wb.save(filepath)
        msg = _('Matrix exported: ') + filepath
        Message(self.app, _('Matrix exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def export_text_file(self):
        """ Export report to a plain text file with .txt ending.
        QTextWriter supports plaintext, ODF and HTML.
        BUT QTextWriter does not support utf-8-sig
        """

        if len(self.ui.textEdit.document().toPlainText()) == 0:
            return
        filepath, ok = QtWidgets.QFileDialog.getSaveFileName(self,
                                                            _("Save Text File"), self.app.settings['directory'],
                                                            "Text Files(*.txt)")
        # options=QtWidgets.QFileDialog.Option.DontUseNativeDialog)
        if filepath is None or not ok:
            return
        if filepath[-4:] != ".txt":
            filepath += ".txt"
        ''' https://stackoverflow.com/questions/39422573/python-writing-weird-unicode-to-csv
        Using a byte order mark so that other software recognises UTF-8
        '''
        data = self.ui.textEdit.toPlainText()
        f = open(filepath, 'w', encoding='utf-8-sig')
        f.write(data)
        f.close()
        msg = _('Report exported: ') + filepath
        Message(self.app, _('Report exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def export_odt_file(self):
        """ Export report to open document format with .odt ending.
        QTextWriter supports plaintext, ODF and HTML .
        """

        if len(self.ui.textEdit.document().toPlainText()) == 0:
            return
        filepath, ok = QtWidgets.QFileDialog.getSaveFileName(self,
                                                            _("Save Open Document Text File"), self.app.settings['directory'],
                                                            "ODT Files(*.odt)")
        # options=QtWidgets.QFileDialog.Option.DontUseNativeDialog)
        if filepath is None or not ok:
            return
        if filepath[-4:] != ".odt":
            filepath += ".odt"

        if filepath is None:
            return
        tw = QtGui.QTextDocumentWriter()
        tw.setFileName(filepath)
        tw.setFormat(b'ODF')  # byte array needed for Windows 10
        tw.write(self.ui.textEdit.document())
        msg = _("Report exported: ") + filepath
        self.parent_textEdit.append(msg)
        Message(self.app, _('Report exported'), msg, "information").exec()

    def export_csv_file(self):
        """ Export report to csv file. Comma delimited and all cells quoted.
        Columns file/case, coder, coded text/img/av, id, codename, categories ... {file variables ... case variables}
        Draw data from self.results
        Checkbox for optionally exporting file and case variables
        """

        if not self.results:
            return

        # Column headings
        col_headings = ["File/case", "Coder", "Coded", "Id", "Codename", "Coded_Memo"]
        # Number of categories, for category column headings
        total_categories = 0
        for data in self.results:
            if len(self.categories_of_code(data['cid'])) > total_categories:
                total_categories = len(self.categories_of_code(data['cid']))
        if total_categories > 0:
            col_headings += ["Category"] * total_categories

        cur = self.app.conn.cursor()

        # Number of file and case variables, for variable column headings
        file_variables = []
        case_variables = []
        if self.ui.checkBox_variables.isChecked():
            cur.execute("select name from attribute_type where caseOrFile='file' order by name")
            result = cur.fetchall()
            for var_heading in result:
                col_headings.append("FileVar_" + var_heading[0])
                file_variables.append(var_heading[0])
            # Number of case variables, for variable column headings
            cur.execute("select name from attribute_type where caseOrFile='case' order by name")
            result = cur.fetchall()
            for var_heading in result:
                col_headings.append("CaseVar_" + var_heading[0])
                case_variables.append(var_heading[0])

        # Create data rows
        csv_data = []
        for row, data in enumerate(self.results):
            csv_data_row = []
            csv_data_row.append(data['file_or_casename'])  # col 0
            csv_data_row.append(data['coder'])  # col 1
            coding_id = ""
            if data['result_type'] == 'text':
                coding_id = f"ctid:{data['ctid']}"
                csv_data_row.append(data['text'])
            if data['result_type'] == 'image':
                coding_id = f"imid:{data['imid']}"
                csv_data_row.append("image")
            if data['result_type'] == 'av':
                coding_id = f"avid:{data['avid']}"
                csv_data_row.append("a/v")
            csv_data_row.append(coding_id)  # col 3
            csv_data_row.append(data['codename'])  # col 4
            csv_data_row.append(data['coded_memo'])  # col 5
            categories = self.categories_of_code(data['cid'])
            for i, category in enumerate(categories):
                csv_data_row.append(category)

            if self.ui.checkBox_variables.isChecked():
                # File variables
                for file_var_pos, file_var_name in enumerate(file_variables):
                    cur.execute("select value from attribute where attr_type='file' and name=? and id=?"
                                , [file_var_name, data['fid']])
                    value = ""
                    file_var_value = cur.fetchone()
                    if file_var_value:  # Could potentially be None
                        value = file_var_value[0]
                    csv_data_row.append(value)
                # Case variables
                for case_var_pos, case_var_name in enumerate(case_variables):
                    cur.execute("select value from attribute where attr_type='case' and name=? and id=?"
                                , [case_var_name, data['caseid']])
                    value = ""
                    case_var_value = cur.fetchone()
                    if case_var_value:  # Could potentially be None
                        value = case_var_value[0]
                    csv_data_row.append(value)
            csv_data.append(csv_data_row)

        filepath, ok = QtWidgets.QFileDialog.getSaveFileName(self,
                                                             _("Save CSV File"), self.app.settings['directory'],
                                                             "CSV Files(*.csv)")
        # options=QtWidgets.QFileDialog.Option.DontUseNativeDialog)
        if filepath is None or not ok:
            return
        if filepath[-4:] != ".csv":
            filepath += ".csv"
        with open(filepath, 'w', encoding='utf-8-sig', newline='') as csvfile:
            filewriter = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
            filewriter.writerow(col_headings)
            for row in csv_data:
                filewriter.writerow(row)

        msg = _("Each row contains filename, coder, coded, codename and categories.") + "\n"
        if self.ui.checkBox_variables.isChecked():
            msg += _("And file and case variables") + "\n"
        msg += _('Report exported: ') + filepath
        Message(self.app, _('Report exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def export_xlsx_file(self):
        """ Export report to xlsx file.
        Columns file/case, coder, coded text/img/av, id, codename, categories ... {file variables ... case variables}
        Draw data from self.results
        Checkbox for optionally exporting file and case variables
        """

        if not self.results:
            return
        wb = openpyxl.Workbook()
        ws = wb.active

        # Column headings
        col_headings = ["File/case", "Coder", "Coded", "Id", "Codename", "Coded_Memo"]
        # Number of categories, for category column headings
        total_categories = 0
        for data in self.results:
            if len(self.categories_of_code(data['cid'])) > total_categories:
                total_categories = len(self.categories_of_code(data['cid']))
        if total_categories > 0:
            col_headings += ["Category"] * total_categories

        cur = self.app.conn.cursor()

        # Number of file variables, for variable column headings
        file_variables = []
        case_variables = []
        if self.ui.checkBox_variables.isChecked():
            cur.execute("select name from attribute_type where caseOrFile='file' order by name")
            result = cur.fetchall()
            for var_heading in result:
                col_headings.append("FileVar_" + var_heading[0])
                file_variables.append(var_heading[0])
            # Number of case variables, for variable column headings
            cur.execute("select name from attribute_type where caseOrFile='case' order by name")
            result = cur.fetchall()
            for var_heading in result:
                col_headings.append("CaseVar_" + var_heading[0])
                case_variables.append(var_heading[0])

        row = 1
        for col, col_heading in enumerate(col_headings):
            ws.cell(column=col + 1, row=row, value=col_heading)

        # Fill Excel Worksheet
        for row, data in enumerate(self.results):
            ws.cell(column=1, row=row + 2, value=data['file_or_casename'])
            ws.cell(column=2, row=row + 2, value=data['coder'])
            coding_id = ""
            if data['result_type'] == 'text':
                coding_id = f"ctid:{data['ctid']}"
                ws.cell(column=3, row=row + 2, value=data['text'])
            if data['result_type'] == 'image':
                coding_id = f"imid:{data['imid']}"
                ws.cell(column=3, row=row + 2, value="image")
            if data['result_type'] == 'av':
                coding_id = f"avid:{data['avid']}"
                ws.cell(column=3, row=row + 2, value="a/v")
            ws.cell(column=4, row=row + 2, value=coding_id)
            ws.cell(column=5, row=row + 2, value=data['codename'])
            ws.cell(column=6, row=row + 2, value=data['coded_memo'])
            categories = self.categories_of_code(data['cid'])
            for i, category in enumerate(categories):
                ws.cell(column=7 + i, row=row + 2, value=category)

            if self.ui.checkBox_variables.isChecked():
                # File variables
                file_vars_start_column = 7 + total_categories
                for file_var_pos, file_var_name in enumerate(file_variables):
                    cur.execute("select value from attribute where attr_type='file' and name=? and id=?"
                                , [file_var_name, data['fid']])
                    value = ""
                    file_var_value = cur.fetchone()
                    if file_var_value:  # Could potentially be None
                        value = file_var_value[0]
                    ws.cell(column=file_vars_start_column + file_var_pos, row=row + 2, value=value)
                # Case variables
                case_vars_start_column = 7 + total_categories + len(file_variables)
                for case_var_pos, case_var_name in enumerate(case_variables):
                    cur.execute("select value from attribute where attr_type='case' and name=? and id=?"
                                , [case_var_name, data['caseid']])
                    value = ""
                    case_var_value = cur.fetchone()
                    if case_var_value:  # Could potentially be None
                        value = case_var_value[0]
                    ws.cell(column=case_vars_start_column + case_var_pos, row=row + 2, value=value)

        filepath, ok = QtWidgets.QFileDialog.getSaveFileName(self,
                                                            _("Save Excel File"), self.app.settings['directory'],
                                                            "XLSX Files(*.xlsx)")
        # options=QtWidgets.QFileDialog.Option.DontUseNativeDialog)
        if filepath is None or not ok:
            return
        wb.save(filepath)
        msg = _("Each row contains filename, coder, coded, codename and categories.") + "\n"
        msg += _('Report exported: ') + filepath
        Message(self.app, _('Report exported'), msg, "information").exec()
        self.parent_textEdit.append(msg)

    def categories_of_code(self, cid):
        """ Get parent categories of this code.

        param: cid : Integer of code id
        return: category_names : List
        """

        code_ = None
        for c in self.code_names:
            if c['cid'] == cid:
                code_ = c
        if not code_:
            return []
        if not code_['catid']:
            return []
        catid = code_['catid']
        category_names = []
        more = True
        counter = 0
        while more and counter < 1000:
            for category in self.categories:
                if catid == category['catid']:
                    category_names.append(category['name'])
                    catid = category['supercatid']
                    if not catid:
                        more = False
            counter += 1
        return category_names

    def export_html_file(self):
        """ Export report to a html file. Create folder of images and change refs to the
        folder.
        Uses self.html_links
        """

        if len(self.ui.textEdit.document().toPlainText()) == 0:
            return
        filepath, ok = QtWidgets.QFileDialog.getSaveFileName(self,
                                                            _("Save HTML File"), self.app.settings['directory'],
                                                            "HTML Files(*.html)")
        # options=QtWidgets.QFileDialog.Option.DontUseNativeDialog)
        if filepath is None or not ok:
            return
        if filepath[-5:] != ".html":
            filepath += ".html"

        tw = QtGui.QTextDocumentWriter()
        tw.setFileName(filepath)
        tw.setFormat(b'HTML')  # byte array needed for Windows 10
        # tw.setCodec(QTextCodec.codecForName('UTF-8'))  # for Windows 10
        tw.write(self.ui.textEdit.document())
        need_media_folders = False
        for item in self.html_links:
            if item['image'] is not None or item['avname'] is not None:
                need_media_folders = True
        html_folder_path = ""
        html_relative_path = ""
        if need_media_folders:
            # Create folder with sub-folders for images, audio and video
            html_folder_path = filepath[:-5]
            html_relative_path = html_folder_path.split("/")[-1]
            try:
                os.mkdir(html_folder_path)
                os.mkdir(html_folder_path + "/images")
                os.mkdir(html_folder_path + "/audio")
                os.mkdir(html_folder_path + "/video")
            except Exception as err:
                logger.warning(_("html folder creation error ") + str(err))
                Message(self.app, _("Folder creation"), f"{html_folder_path} {_('error ')} {err}", "critical").exec()
                return
        try:
            with open(filepath, 'r') as f:
                html = f.read()
        except Exception as err:
            logger.warning(_('html file reading error:') + str(err))
            return

        # Change html links to reference the html folder
        start_pos = 0
        for item in self.html_links:
            if item['imagename'] is not None:
                # What if linked?
                filename = item['imagename'].replace('/images/', '')
                img_path = f"{html_folder_path}/images/{filename}"
                img_relative_link = f"{html_relative_path}/images/{filename}"
                item['image'].save(img_path)
                html = html.replace(item['imagename'], img_relative_link)
            if item['avname'] is not None:
                # Add audio/video to html folder
                mediatype = "video"
                if item['avname'][0:6] in ("/audio", "audio:"):
                    mediatype = "audio"
                # Remove link prefix and note if link or not
                linked = False
                av_path = item['avname']
                if av_path[0:6] == "video:":
                    av_path = av_path[6:]
                    linked = True
                if av_path[0:6] == "audio:":
                    linked = True
                    av_path = av_path[6:]
                relative_link = ""
                filename = av_path.split('/')[-1]
                # Copy non-linked a/v file to html folder
                if not (linked and os.path.isfile(html_folder_path + av_path)):
                    copyfile(self.app.project_path + item['avname'], html_folder_path + av_path)
                    relative_link = f"{html_relative_path}/video/{filename}"
                # Copy Linked video file to html folder
                if mediatype == "video" and linked:
                    av_destination = f"{html_folder_path}/video/{filename}"
                    relative_link = f"{html_relative_path}/video/{filename}"
                    if not os.path.isfile(av_destination):
                        copyfile(av_path, av_destination)
                # Copy Linked audio file to html folder
                if mediatype == "audio" and linked:
                    audio_destination = f"{html_folder_path}/audio/{filename}"
                    relative_link = f"{html_relative_path}/audio/{filename}"
                    if not os.path.isfile(audio_destination):
                        copyfile(av_path, audio_destination)

                # Create html to display media time positions
                extension = item['avname'][item['avname'].rfind('.') + 1:]
                html_controls = f"</p>\n<{mediatype} controls>"
                html_controls += f'<source src="{relative_link}'
                html_controls += f'#t={item["av0"]},{item["av1"]}"'
                html_controls += f' type="{mediatype}/{extension}">'
                html_controls += f'</{mediatype}><p>\n'
                search_string = f"File: {filename},  Coder: "
                coded_pos = html.find(search_string, start_pos)
                next_p_pos = html.find("<p ", coded_pos)
                html_tmp = html[:next_p_pos] + html_controls
                start_pos = len(html_tmp)
                html_tmp += html[next_p_pos:]
                html = html_tmp

        with open(filepath, 'w', encoding='utf-8-sig') as f:
            f.write(html)
        msg = _("Report exported to: ") + filepath
        if need_media_folders:
            msg += f"\n{_('Media folder:')} {html_folder_path}"
        self.parent_textEdit.append(msg)
        Message(self.app, _('Report exported'), msg, "information").exec()

    def eventFilter(self, object_, event):
        """ Used to detect key events in the textedit.
        H Hide / Unhide top groupbox
        Ctrl F Search box focus
        """

        if type(event) == QtGui.QKeyEvent:
            key = event.key()
            mod = event.modifiers()
            # Hide unHide top groupbox
            if key == QtCore.Qt.Key.Key_H and (self.ui.textEdit.hasFocus() or self.ui.treeWidget.hasFocus() or
                                               self.ui.listWidget_files.hasFocus() or
                                               self.ui.listWidget_cases.hasFocus()):
                self.ui.groupBox.setHidden(not (self.ui.groupBox.isHidden()))
                return True
            # Ctrl + F jump to search box
            if key == QtCore.Qt.Key.Key_F and mod == QtCore.Qt.KeyboardModifier.ControlModifier:
                self.ui.lineEdit_search_results.setFocus()
                self.ui.groupBox.setHidden(False)
                return True
        return False

    def recursive_set_selected(self, item):
        """ Set all children of this item to be selected if the item is selected.
        Recurse through any child categories.
        Called by: search
        """

        child_count = item.childCount()
        for i in range(child_count):
            if item.isSelected():
                item.child(i).setSelected(True)
            self.recursive_set_selected(item.child(i))

    def search_codebook(self):
        """ Codebook display with memos. """

        self.te = []
        self.ui.tableWidget.setColumnCount(0)
        self.ui.tableWidget.setRowCount(0)
        self.ui.checkBox_matrix_transpose.setChecked(False)
        self.ui.comboBox_matrix.setCurrentIndex(0)
        self.ui.splitter.setSizes([200, 400, 0])
        self.ui.listWidget_cases.clearSelection()
        self.ui.listWidget_files.clearSelection()
        self.ui.textEdit.clear()
        self.ui.textEdit.append(_("Codebook with memos\n=================\n"))

        # Select all code items under selected categories
        self.recursive_set_selected(self.ui.treeWidget.invisibleRootItem())
        cur = self.app.conn.cursor()
        results = []
        for i in self.ui.treeWidget.selectedItems():
            if i.text(1)[0:5] == "catid":
                cur.execute("select name, ifnull(memo,'') from code_cat where catid=?", [i.text(1)[6:]])
                res = cur.fetchone()
                if res is not None:
                    results.append([_("Category: "), res[0], res[1]])
            if i.text(1)[0:3] == 'cid':
                cur.execute("select name, ifnull(memo,''), color from code_name where cid=?", [i.text(1)[4:]])
                res = cur.fetchone()
                if res is not None:
                    results.append([_("Code: "), res[0], res[1], res[2]])
        for r in results:
            if r[0] != "Code: ":
                self.ui.textEdit.append(f"{r[0]}{r[1]}\n{r[2]}\n")
            else:
                cursor = self.ui.textEdit.textCursor()
                fmt = QtGui.QTextCharFormat()
                pos0 = len(self.ui.textEdit.toPlainText())
                self.ui.textEdit.append(r[0] + r[1])
                cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
                pos1 = len(self.ui.textEdit.toPlainText())
                cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
                brush = QBrush(QtGui.QColor(r[3]))
                fmt.setBackground(brush)
                text_brush = QBrush(QtGui.QColor(TextColor(r[3]).recommendation))
                fmt.setForeground(text_brush)
                cursor.setCharFormat(fmt)
                self.ui.textEdit.append(f"{r[2]}\n")
        self.ui.comboBox_export.setEnabled(True)

    def search_annotations(self):
        """ Find and display annotations from selected text files. """

        self.te = []
        self.ui.tableWidget.setColumnCount(0)
        self.ui.tableWidget.setRowCount(0)
        self.ui.checkBox_matrix_transpose.setChecked(False)
        self.ui.comboBox_matrix.setCurrentIndex(0)
        self.ui.splitter.setSizes([200, 400, 0])

        # Get variables for search: search text, coders, codes, files,cases, attributes
        coder = self.ui.comboBox_coders.currentText()
        self.html_links = []  # For html file output with media
        search_text = self.ui.lineEdit.text()
        self.get_selected_files_and_cases()
        if self.file_ids_string == "":
            Message(self.app, _("Warning"), _("No files selected for annotations")).exec()
            return
        self.ui.treeWidget.clearSelection()
        self.ui.listWidget_cases.clearSelection()

        cur = self.app.conn.cursor()
        sql = "select anid, fid, source.name, pos0, pos1, annotation.memo, annotation.owner, annotation.date, "
        sql += "substr(fulltext, pos0 + 1, pos1 - pos0) as subtext "
        sql += "from annotation join source on source.id=annotation.fid "
        sql += f"where source.fulltext is not null and fid in ({self.file_ids_string}) "
        # Coder limiter
        values = []
        if coder != "":
            sql += " and annotation.owner=? "
            values.append(coder)
        if search_text != "":
            sql += " and instr(subtext, ?) is not null "
            values.append(search_text)
        sql += " order by source.name, anid asc"
        if not values:
            cur.execute(sql)
        else:
            cur.execute(sql, values)
        res = cur.fetchall()
        annotes = []
        keys = "anid", "fid", "filename", "pos0", "pos1", "annotation", "owner", "date", "text"
        for row in res:
            annotes.append(dict(zip(keys, row)))

        self.ui.textEdit.clear()
        # Display search parameters
        self.ui.textEdit.append(_("Annotation search parameters") + "\n==========")
        if coder == "":
            self.ui.textEdit.append(_("Coder: All coders"))
        else:
            self.ui.textEdit.append(_("Coder: ") + coder)
        if search_text != "":
            self.ui.textEdit.append(_("Search text: ") + search_text)
        self.ui.textEdit.append(_("Files:"))
        cur.execute(
            "select name from source where id in (" + self.file_ids_string + ") and source.fulltext is not null order by name")
        res = cur.fetchall()
        file_txt = ""
        for r in res:
            file_txt += f"{r[0]}, "
        self.ui.textEdit.append(file_txt)
        self.ui.textEdit.append("==========")
        for a in annotes:
            txt = f"\n{_('File')}: {a['filename']} anid: {a['anid']} "
            txt += f"{_('DATE:')} {a['date'][0:10]} {_('Coder:')} {a['owner']}, "
            txt += f"{_('Position')}: {a['pos0']} - {a['pos1']}\n"
            txt += f"{_('TEXT')}: {a['text']}\n"
            txt += f"{_('ANNOTATION')}: {a['annotation']}"
            self.ui.textEdit.append(txt)
        self.ui.comboBox_export.setEnabled(True)

    def select_attributes(self):
        """ Select files based on attribute selections.
        Attribute results are a dictionary of:
        first item is a Boolean AND or OR list item
        Followed by each attribute list item
        """

        # Clear ui
        self.attribute_file_ids = []
        self.attribute_case_ids = []
        self.ui.pushButton_attributeselect.setToolTip(_("Attributes"))
        self.ui.splitter.setSizes([300, 300, 0])
        # Remove any selected case or file ids
        self.file_ids_string = ""
        for i in range(self.ui.listWidget_files.count()):
            self.ui.listWidget_files.item(i).setSelected(False)
        self.case_ids_string = ""
        for i in range(self.ui.listWidget_cases.count()):
            self.ui.listWidget_cases.item(i).setSelected(False)

        attr_ui = DialogSelectAttributeParameters(self.app)
        attr_ui.fill_parameters(self.attributes)
        temp_attributes = deepcopy(self.attributes)
        self.attributes = []
        ok = attr_ui.exec()
        if not ok:
            self.attributes = temp_attributes
            self.ui.pushButton_attributeselect.setIcon(qta.icon('mdi6.line-scan'))
            self.ui.pushButton_attributeselect.setToolTip(_("Attributes"))
            if self.attributes:
                self.ui.pushButton_attributeselect.setIcon(qta.icon('mdi6.variable'))
            return
        # As List containing (1) list of attributes, within (2) [List of attributes, boolean type]
        self.attributes = attr_ui.parameters
        if len(self.attributes) == 1:  # The and /or boolean operator only
            self.ui.pushButton_attributeselect.setIcon(qta.icon('mdi6.line-scan'))
            self.ui.pushButton_attributeselect.setToolTip(_("Attributes"))
            return
        self.ui.pushButton_attributeselect.setIcon(qta.icon('mdi6.variable'))
        self.ui.pushButton_attributeselect.setToolTip(attr_ui.tooltip_msg)
        self.attributes_msg = attr_ui.tooltip_msg
        # Used ..?
        self.attribute_file_ids = attr_ui.result_file_ids
        self.attribute_case_ids = attr_ui.result_case_ids

    def search(self):
        """ Search for selected codings.
        There are four main search pathways.
        1:  file selection only.
        2: case selection combined with files selection. (No files selected presumes ALL files)
        3: attribute selection, which may include files or cases.
        4. codebook memo selection
        """

        memo_choice = self.ui.comboBox_memos.currentText()
        if memo_choice == _("Annotations"):
            self.search_annotations()
            return
        if memo_choice == _("Codebook memos"):
            self.search_codebook()
            return

        # Get variables for search: codes, files,cases, attribute file ids
        self.get_selected_files_and_cases()

        # Select all code items under selected categories
        self.recursive_set_selected(self.ui.treeWidget.invisibleRootItem())
        items = self.ui.treeWidget.selectedItems()
        if len(items) == 0:
            msg = _("No codes have been selected.")
            Message(self.app, _('No codes'), msg, "warning").exec()
            return
        if self.file_ids_string == "" and self.case_ids_string == "" and self.attributes == []:
            msg = _("No files, cases found.")
            Message(self.app, _('Nothing selected'), msg, "warning").exec()
            return

        prog_dialog = QtWidgets.QProgressDialog("Running", "", 1, 5, None)
        prog_dialog.setWindowTitle(_("Searching"))
        prog_dialog.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowCloseButtonHint)
        prog_dialog.setAutoClose(True)
        prog_dialog.setValue(1)
        prog_dialog.show()
        QtCore.QCoreApplication.processEvents()

        # Clear results output
        rows = self.ui.tableWidget.rowCount()
        for r in range(0, rows):
            self.ui.tableWidget.removeRow(0)
        self.ui.comboBox_export.setEnabled(True)
        self.ui.textEdit.clear()
        self.te = []
        self.html_links = []  # For html file output with media

        # Add search terms to textEdit
        if memo_choice == _("Only memos"):
            self.ui.textEdit.insertPlainText(_("Only memos shown. Coded data not shown.") + "\n")
        if memo_choice == _("Only coded memos"):
            self.ui.textEdit.insertPlainText(_("Coded memos shown if available. Coded data not shown.") + "\n")
        self.ui.textEdit.insertPlainText(_("Search parameters") + "\n==========\n")
        coder = self.ui.comboBox_coders.currentText()
        if coder == "":
            self.ui.textEdit.insertPlainText(f"{_('Coding by: All coders')}\n")
        else:
            self.ui.textEdit.insertPlainText(f"{_('Coding by: ')}{coder}\n")
        codes_string = f"{_('Codes: ')}\n"
        codes_count = 0
        for i in items:
            if i.text(1)[0:3] == 'cid':
                codes_count += 1
                codes_string += i.text(0) + ". "
        codes_string += f"{_('Codes: ')}{codes_count} / {len(self.code_names)}"
        self.ui.textEdit.insertPlainText(codes_string)

        cur = self.app.conn.cursor()
        parameters_display = ""
        if self.attributes:
            parameters_display += f"\n{_('Attributes:')}\n {self.attributes_msg}\n"
            if not (self.attribute_file_ids and self.attribute_case_ids):
                parameters_display += "No cases or files match attribute selection\n"
        if self.attribute_file_ids or self.attribute_case_ids:
            self.file_ids_string = ""
            self.case_ids_string = ""
            for fid in self.attribute_file_ids:
                self.file_ids_string += f",{fid}"
            self.file_ids_string = self.file_ids_string[1:]
            for caseid in self.attribute_case_ids:
                self.case_ids_string += f",{caseid}"
            self.case_ids_string = self.case_ids_string[1:]

            for i in range(self.ui.listWidget_files.count()):
                self.ui.listWidget_files.item(i).setSelected(False)
            for i in range(self.ui.listWidget_cases.count()):
                self.ui.listWidget_cases.item(i).setSelected(False)

        if self.file_ids_string != "":
            parameters_display += _("\nFiles:\n")
            cur.execute(f"select name from source where id in ({self.file_ids_string}) order by name")
            res = cur.fetchall()
            for r in res:
                parameters_display += f"{r[0]}, "
            parameters_display += f"{_(' Files: ')} {len(res)} / {len(self.files)}"
        if self.case_ids_string != "":
            parameters_display += _("\nCases:\n")
            cur.execute(f"select name from cases where caseid in ({self.case_ids_string}) order by name")
            res = cur.fetchall()
            for r in res:
                parameters_display += f"{r[0]}, "
        self.ui.textEdit.insertPlainText(f"{parameters_display}\n")
        if self.ui.lineEdit.text() != "":
            self.ui.textEdit.insertPlainText(f"\n{_('Search text: ')} {self.ui.lineEdit.text()}\n")
        self.ui.textEdit.insertPlainText("\n==========\n")

        # Get selected codes as comma separated String of cids
        code_ids_string = ""
        for i in items:
            if i.text(1)[0:3] == 'cid':
                code_ids_string += f",{i.text(1)[4:]}"
        code_ids_string = code_ids_string[1:]
        self.html_links = []
        self.results = []

        # FILES SEARCH, ALSO ATTRIBUTES FILE IDS SEARCH
        if self.file_ids_string != "" and self.case_ids_string == "":
            self.search_by_files(code_ids_string)
        # CASES AND FILES SEARCH
        # Default to all files if none are selected, otherwise limit to the selected files
        if self.case_ids_string != "":
            self.search_by_cases(code_ids_string)

        QtCore.QCoreApplication.processEvents()
        prog_dialog.setValue(2)
        # Trim results for option: Only coded memos
        if self.ui.comboBox_memos.currentText() in ("Only memos", "Only coded memos"):
            tmp = []
            for r in self.results:
                if r['coded_memo'] != "":
                    tmp.append(r)
            self.results = tmp
        # Organise results
        self.sort_search_results()
        self.fill_text_edit_with_search_results()
        # Clean up for next search. Except attributes list, keep attributes selection active.
        self.attribute_file_ids = []
        self.file_ids_string = ""
        self.case_ids_string = ""
        self.attributes_msg = ""
        self.ui.pushButton_attributeselect.setToolTip(_("Attributes"))
        del prog_dialog

    def search_by_files(self, code_ids):
        """ Search by files and if attributes file ids are selected.
        Called by search() if self.file_ids_string is not empty and self.case_ids_string is empty

        :param: code_ids : String comma separated ids
        """

        coder = self.ui.comboBox_coders.currentText()
        search_text = self.ui.lineEdit.text()
        important = self.ui.checkBox_important.isChecked()
        parameters = []
        cur = self.app.conn.cursor()
        # Coded text
        sql = "select code_name.name, color, source.name, pos0, pos1, seltext, "
        sql += "code_text.owner, fid, ifnull(code_text.memo,''), ifnull(code_name.memo,''), " \
               "ifnull(source.memo,''), ctid, code_name.cid "
        sql += " from code_text join code_name "
        sql += "on code_name.cid = code_text.cid join source on fid = source.id "
        sql += f"where code_name.cid in ({code_ids}) "
        sql += f"and source.id in ({self.file_ids_string}) "
        if coder != "":
            sql += " and code_text.owner=? "
            parameters.append(coder)
        if search_text != "":
            sql += " and seltext like ? "
            parameters.append("%" + str(search_text) + "%")
        if important:
            sql += " and code_text.important=1 "
        sql += " order by code_name.name, source.name, pos0"
        if not parameters:
            cur.execute(sql)
        else:
            cur.execute(sql, parameters)
        result = cur.fetchall()
        keys = 'codename', 'color', 'file_or_casename', 'pos0', 'pos1', 'text', 'coder', 'fid', 'coded_memo', \
            'codename_memo', 'source_memo', 'ctid', 'cid'
        for row in result:
            tmp = dict(zip(keys, row))
            tmp['result_type'] = 'text'
            tmp['file_or_case'] = "File"
            tmp['pretext'] = ""
            tmp['posttext'] = ""
            tmp['caseid'] = -1  # Need a placeholder if export vars is checked
            self.results.append(tmp)
        if self.ui.checkBox_text_context.isChecked():
            self.get_prettext_and_posttext()

        # Coded images
        parameters = []
        sql = "select code_name.name, color, source.name, x1, y1, width, height,"
        sql += "code_image.owner, source.mediapath, source.id, ifnull(code_image.memo,''), "
        sql += "code_name.memo, ifnull(source.memo,''), imid, code_name.cid "
        sql += " from code_image join code_name "
        sql += "on code_name.cid = code_image.cid join source on code_image.id = source.id "
        sql += f"where code_name.cid in ({code_ids}) "
        sql += f"and source.id in ({self.file_ids_string}) "
        if coder != "":
            sql += " and code_image.owner=? "
            parameters.append(coder)
        if search_text != "":
            sql += " and code_image.memo like ? "
            parameters.append("%" + str(search_text) + "%")
        if important:
            sql += " and code_image.important=1 "
        sql += " order by code_name.name, source.name, x1"
        if not parameters:
            cur.execute(sql)
        else:
            cur.execute(sql, parameters)
        result = cur.fetchall()
        keys = 'codename', 'color', 'file_or_casename', 'x1', 'y1', 'width', 'height', 'coder', 'mediapath', \
            'fid', 'coded_memo', 'codename_memo', 'source_memo', 'imid', 'cid'
        for row in result:
            tmp = dict(zip(keys, row))
            tmp['result_type'] = 'image'
            tmp['file_or_case'] = "File"
            tmp['caseid'] = -1  # Need a placeholder if export vars is checked
            self.results.append(tmp)

        # Coded audio and video, also looks for search_text in coded segment memo
        parameters = []
        sql = "select code_name.name, color, source.name, pos0, pos1, ifnull(code_av.memo,''), "
        sql += " code_av.owner, source.mediapath, source.id, ifnull(code_name.memo,''), ifnull(source.memo,''), " \
               "avid, code_name.cid"
        sql += " from code_av join code_name "
        sql += "on code_name.cid = code_av.cid join source on code_av.id = source.id "
        sql += f"where code_name.cid in ({code_ids}) "
        sql += f"and source.id in ({self.file_ids_string}) "
        if coder != "":
            sql += " and code_av.owner=? "
            parameters.append(coder)
        if search_text != "":
            sql += " and code_av.memo like ? "
            parameters.append("%" + str(search_text) + "%")
        if important:
            sql += " and code_av.important=1 "
        sql += " order by code_name.name, source.name, pos0"
        if not parameters:
            cur.execute(sql)
        else:
            cur.execute(sql, parameters)
        result = cur.fetchall()
        keys = 'codename', 'color', 'file_or_casename', 'pos0', 'pos1', 'coded_memo', 'coder', 'mediapath', 'fid', \
            'codename_memo', 'source_memo', 'avid', 'cid'
        for row in result:
            tmp = dict(zip(keys, row))
            tmp['result_type'] = 'av'
            tmp['file_or_case'] = "File"
            text_ = str(tmp['file_or_casename']) + " "
            if len(tmp['coded_memo']) > 0:
                text_ += "\nMEMO: " + tmp['coded_memo']
            text_ += " " + msecs_to_hours_mins_secs(tmp['pos0']) + " - " + msecs_to_hours_mins_secs(tmp['pos1'])
            tmp['text'] = text_
            tmp['caseid'] = -1  # Need a placeholder if export vars is checked
            self.html_links.append({'imagename': None, 'image': None,
                                    'avname': tmp['mediapath'], 'av0': str(int(tmp['pos0'] / 1000)),
                                    'av1': str(int(tmp['pos1'] / 1000)), 'avtext': text_})
            self.results.append(tmp)

    def search_by_cases(self, code_ids):
        """ Search by cases and if attributes file ids are selected.
        Called by search() if self.case_ids_string is not empty.
        Also uses self.file_ids_string to limit results

        :param: code_ids : String comma separated ids
        """

        coder = self.ui.comboBox_coders.currentText()
        search_text = self.ui.lineEdit.text()
        important = self.ui.checkBox_important.isChecked()
        cur = self.app.conn.cursor()
        parameters = []

        # Coded text
        sql = "select code_name.name, color, cases.name, cases.caseid, "
        sql += "code_text.pos0, code_text.pos1, seltext, code_text.owner, code_text.fid, "
        sql += "ifnull(cases.memo,''), ifnull(code_text.memo,''), ifnull(code_name.memo,''), "
        sql += "ifnull(source.memo,''), ctid, code_name.cid "
        sql += "from code_text join code_name on code_name.cid = code_text.cid "
        sql += "join (case_text join cases on cases.caseid = case_text.caseid) on "
        sql += "code_text.fid = case_text.fid "
        sql += "join source on source.id=code_text.fid "
        sql += f"where code_name.cid in ({code_ids}) "
        sql += f"and case_text.caseid in ({self.case_ids_string}) "
        if self.file_ids_string != "":
            sql += f" and code_text.fid in ({self.file_ids_string})"
        sql += "and (code_text.pos0 >= case_text.pos0 and code_text.pos1 <= case_text.pos1)"
        if coder != "":
            sql += " and code_text.owner=? "
            parameters.append(coder)
        if search_text != "":
            sql += " and seltext like ? "
            parameters.append("%" + str(search_text) + "%")
        if important:
            sql += " and code_text.important=1 "
        sql += " order by code_name.name, cases.name"
        if not parameters:
            cur.execute(sql)
        else:
            cur.execute(sql, parameters)
        results = cur.fetchall()
        keys = 'codename', 'color', 'file_or_casename', 'caseid', 'pos0', 'pos1', 'text', 'coder', 'fid', \
            'cases_memo', 'coded_memo', 'codename_memo', 'source_memo', 'ctid', 'cid'
        for row in results:
            tmp = dict(zip(keys, row))
            tmp['result_type'] = 'text'
            tmp['file_or_case'] = "Case"
            tmp['pretext'] = ""
            tmp['posttext'] = ""
            self.results.append(tmp)
        if self.ui.checkBox_text_context.isChecked():
            self.get_prettext_and_posttext()

        # Coded images
        parameters = []
        sql = "select code_name.name, color, cases.name, cases.caseid, "
        sql += "x1, y1, width, height, code_image.owner,source.mediapath, source.id, "
        sql += "ifnull(code_image.memo,''), ifnull(cases.memo,''), ifnull(code_name.memo,''), "
        sql += "ifnull(source.memo,''), imid, code_name.cid "
        sql += "from code_image join code_name on code_name.cid = code_image.cid "
        sql += "join (case_text join cases on cases.caseid = case_text.caseid) on "
        sql += "code_image.id = case_text.fid "
        sql += " join source on case_text.fid = source.id "
        sql += f"where code_name.cid in ({code_ids}) "
        sql += f"and case_text.caseid in ({self.case_ids_string}) "
        if self.file_ids_string != "":
            sql += f" and source.id in ({self.file_ids_string})"
        if coder != "":
            sql += " and code_image.owner=? "
            parameters.append(coder)
        if search_text != "":
            sql += " and code_image.memo like ? "
            parameters.append("%" + str(search_text) + "%")
        if important:
            sql += " and code_image.important=1 "
        sql += " order by code_name.name, cases.name"
        if not parameters:
            cur.execute(sql)
        else:
            cur.execute(sql, parameters)
        image_results = cur.fetchall()
        keys = ('codename', 'color', 'file_or_casename', 'caseid', 'x1', 'y1', 'width', 'height', 'coder',
                'mediapath', 'fid', 'coded_memo', 'case_memo', 'codename_memo', 'source_memo', 'imid', 'cid')
        for row in image_results:
            tmp = dict(zip(keys, row))
            tmp['result_type'] = 'image'
            tmp['file_or_case'] = "Case"
            self.results.append(tmp)

        # Coded audio and video
        parameters = []
        av_sql = "select distinct code_name.name, color, cases.name as case_name, cases.caseid, "
        av_sql += "code_av.pos0, code_av.pos1, code_av.owner,source.mediapath, source.id, "
        av_sql += "ifnull(code_av.memo,'') as coded_memo, ifnull(cases.memo,'') as case_memo, "
        av_sql += "ifnull(code_name.memo,''), ifnull(source.memo,''), avid, "
        av_sql += "code_name.cid "
        av_sql += "from code_av join code_name on code_name.cid = code_av.cid "
        av_sql += "join (case_text join cases on cases.caseid = case_text.caseid) on "
        av_sql += "code_av.id = case_text.fid "
        av_sql += " join source on case_text.fid = source.id "
        av_sql += f"where code_name.cid in ({code_ids}) "
        av_sql += f"and case_text.caseid in ({self.case_ids_string}) "
        if self.file_ids_string != "":
            av_sql += f" and source.id in ({self.file_ids_string})"
        if coder != "":
            av_sql += " and code_av.owner=? "
            parameters.append(coder)
        if search_text != "":
            av_sql += " and code_av.memo like ? "
            parameters.append("%" + str(search_text) + "%")
        if important:
            sql += " and code_av.important=1 "
        sql += " order by code_name.name, cases.name"
        if not parameters:
            cur.execute(av_sql)
        else:
            cur.execute(av_sql, parameters)
        av_results = cur.fetchall()
        keys = 'codename', 'color', 'file_or_casename', 'caseid', 'pos0', 'pos1', 'coder', 'mediapath', \
            'fid', 'coded_memo', 'case_memo', 'codename_memo', 'source_memo', 'avid', 'cid'
        for row in av_results:
            tmp = dict(zip(keys, row))
            tmp['result_type'] = 'av'
            tmp['file_or_case'] = "Case"
            tmp_text = f"{tmp['file_or_casename']} "
            if len(tmp['coded_memo']) > 0:
                tmp_text += f"\nMEMO: {tmp['coded_memo']}"
            tmp_text += f" {msecs_to_hours_mins_secs(tmp['pos0'])} - {msecs_to_hours_mins_secs(tmp['pos1'])}"
            tmp['text'] = tmp_text
            self.html_links.append({'imagename': None, 'image': None,
                                    'avname': tmp['mediapath'], 'av0': str(int(tmp['pos0'] / 1000)),
                                    'av1': str(int(tmp['pos1'] / 1000)), 'avtext': tmp_text})
            self.results.append(tmp)

    def sort_search_results(self):
        """ Sort results by alphabet or by code count, ascending or descending. """

        sort_by = self.ui.comboBox_sort.currentText()
        if sort_by == "A - z":
            self.results = sorted(self.results, key=lambda i_: i_['codename'])
            return
        if sort_by == "Z - a":
            self.results = sorted(self.results, key=lambda i_: i_['codename'], reverse=True)
            return

        # Order code names by frequency
        # Get unique code names
        tmp_names = [r['codename'] for r in self.results]
        codenames = list(set(tmp_names))
        # Create list dictionary of code name and code count
        name_and_count = []
        for codename in codenames:
            count = 0
            for r in self.results:
                if r['codename'] == codename:
                    count += 1
            name_and_count.append({'codename': codename, 'count': count})
        tmp_results = []
        if sort_by == "1 - 10":
            small_to_large = sorted(name_and_count, key=lambda d: d['count'])
            for s in small_to_large:
                for r in self.results:
                    if s['codename'] == r['codename']:
                        tmp_results.append(r)
            self.results = tmp_results
            return
        if sort_by == "10 - 1":
            large_to_small = sorted(name_and_count, key=lambda d: d['count'], reverse=True)
            for s in large_to_small:
                for r in self.results:
                    if s['codename'] == r['codename']:
                        tmp_results.append(r)
            self.results = tmp_results
            return

    def get_prettext_and_posttext(self):
        """ Get surrounding text 200 characters.
        When context checkbox is checked """

        cur = self.app.conn.cursor()
        for r in self.results:
            # Pre text
            pre_text_length = self.app.settings['report_text_context_characters']
            if r['pos0'] > pre_text_length - 1:
                pre_text_start = r['pos0'] - pre_text_length + 1  # sqlite strings start at 1 not 0
            else:
                pre_text_start = 1  # sqlite strings start at 1 not 0
                pre_text_length = r['pos0']  # sqlite strings start at 1 not 0, so this length is OK
            if pre_text_start < 1:
                pre_text_start = 1
            sql = "select substr(fulltext,?,?) from source where id=?"
            cur.execute(sql, [pre_text_start, pre_text_length, r['fid']])
            res_pre = cur.fetchone()
            if res_pre is not None:
                r['pretext'] = res_pre[0]
            # Post text
            post_text_start = r['pos1'] + 1  # sqlite strings start at 1 not 0
            post_text_length = self.app.settings['report_text_context_characters']
            sql = "select substr(fulltext,?,?) from source where id=?"
            cur.execute(sql, [post_text_start, post_text_length, r['fid']])
            res_post = cur.fetchone()
            if res_post is not None:
                r['posttext'] = res_post[0]

    def text_code_count_and_percent(self):
        """ First part of results, fill code counts and text percentages.
        Text percentages is total of coded text divided by total of text source characters. """

        # Get file text lengths for the text files from the files in the results
        file_ids = []
        code_names = []
        for r in self.results:
            if r['result_type'] == 'text':
                file_ids.append(r['fid'])
                code_names.append(r['codename'])
        file_ids = list(set(file_ids))
        code_names = list(set(code_names))
        code_names.sort()
        cur = self.app.conn.cursor()
        sql = "select id, length(fulltext), name from source where fulltext is not null and id=? order by name"
        file_lengths = []
        for id_ in file_ids:
            cur.execute(sql, [id_])
            res = cur.fetchone()
            res_dict = {"fid": res[0], "length": res[1], "filename": res[2]}
            file_lengths.append(res_dict)
        # Stats results dictionary preparation
        stats = []
        for c in code_names:
            for f in file_lengths:
                stats.append({'codename': c, 'fid': f['fid'], 'filetextlength': f['length'],
                              'filename': f['filename'], 'codecount': 0,
                              'codetextlength': 0, 'percent': 0})
        # Stats results calculated
        """
        {codename , color , file_or_casename , pos0 , pos1 , text , coder, fid, 
        coded_memo codename_memo, source_memo, result_type, file_or_case': 'File'}
        """
        for st in stats:
            for r in self.results:
                if st['codename'] == r['codename'] and st['fid'] == r['fid']:
                    st['codecount'] += 1
                    st['codetextlength'] += len(r['text'])
                    # 2 decimal places
                    st['percent'] = round((st['codetextlength'] / st['filetextlength']) * 100, 2)
        final_stats = []
        for st in stats:
            if st['codecount'] > 0:
                final_stats.append(st)
        msg = _("Text code statistics:")
        for st in final_stats:
            msg += f"\n{st['codename']} | {st['filename']} | {_('Count:')} {st['codecount']} | "
            msg += f"{_('Percent of file:')} {st['percent']}%"
        msg += "\n========"
        if len(final_stats) == 0:
            msg = ""
        return stats, msg

    def image_code_count_and_percent(self):
        """ First part of results, fill code counts and image percentages.
        Image percentages is total of coded area divided by total of Image source area. """

        # Get file area for each image
        file_ids = []
        code_names = []
        for r in self.results:
            if r['result_type'] == 'image':
                file_ids.append(r['fid'])
                code_names.append(r['codename'])
        file_ids = list(set(file_ids))
        code_names = list(set(code_names))
        code_names.sort()
        cur = self.app.conn.cursor()
        sql = "select id, name, mediapath from source where id=? order by name"
        file_areas = []
        for id_ in file_ids:
            cur.execute(sql, [id_])
            res = cur.fetchone()
            abs_path = ""
            w, h = 1, 1
            if 'images:' == res[2][0:7]:
                abs_path = res[2][7:]
            else:
                abs_path = self.app.project_path + res[2]
            try:
                image = Image.open(abs_path)
                w, h = image.size
            except FileNotFoundError:
                pass
            res_dict = {"fid": res[0], "area": w * h, "filename": res[1]}
            file_areas.append(res_dict)

        # Stats results dictionary preparation
        stats = []
        for c in code_names:
            for f in file_areas:
                stats.append({'codename': c, 'fid': f['fid'], 'filearea': f['area'],
                              'filename': f['filename'], 'codecount': 0,
                              'codedarea': 0, 'percent': 0})
        # Stats results calculated
        for st in stats:
            for r in self.results:
                if st['codename'] == r['codename'] and st['fid'] == r['fid']:
                    st['codecount'] += 1
                    st['codedarea'] += r['width'] * r['height']
                    # 2 decimal places
                    st['percent'] = round((st['codedarea'] / st['filearea']) * 100, 2)
        final_stats = []
        for st in stats:
            if st['codecount'] > 0:
                final_stats.append(st)
        msg = _("Image code statistics:")
        for st in final_stats:
            msg += f"\n{st['codename']} | {st['filename']} | {_('Count:')} {st['codecount']} | "
            msg += f"{_('Percent of file:')} {st['percent']}%"
        msg += "\n========"
        if len(final_stats) == 0:
            msg = ""
        return stats, msg

    def av_code_count_and_percent(self):
        """ First part of results, fill code counts and AV percentages.
        AV percentages is total of coded text divided by total of AV source duration. """

        # Get file lengths
        file_ids = []
        code_names = []
        for r in self.results:
            if r['result_type'] == 'av':
                file_ids.append(r['fid'])
                code_names.append(r['codename'])
        file_ids = list(set(file_ids))
        code_names = list(set(code_names))
        code_names.sort()
        cur = self.app.conn.cursor()
        sql = "select id, name, mediapath from source where id=? order by name"
        file_lengths = []
        erroneous_msecs = False
        for id_ in file_ids:
            cur.execute(sql, [id_])
            res = cur.fetchone()
            abs_path = ""
            if 'audio:' == res[2][0:6]:
                abs_path = res[2][6:]
            elif 'video:' == res[2][0:6]:
                abs_path = res[2][6:]
            else:
                abs_path = self.app.project_path + res[2]
            msecs = 1  # Default erroneous value for media duration
            if vlc:
                try:
                    instance = vlc.Instance()
                except NameError as name_err:
                    logger.error(f"vlc.Instance: {name_err}")
                    instance = None
                    erroneous_msecs = True
                if instance:
                    try:
                        media = instance.media_new(abs_path)
                        media.parse()
                        msecs = media.get_duration()
                    except FileNotFoundError:
                        erroneous_msecs = True
            else:
                erroneous_msecs = True
            res_dict = {"fid": res[0], "file_duration": msecs, "filename": res[1]}
            file_lengths.append(res_dict)
        # Stats results dictionary preparation
        stats = []
        for c in code_names:
            for f in file_lengths:
                stats.append({'codename': c, 'fid': f['fid'], 'file_duration': f['file_duration'],
                              'filename': f['filename'], 'codecount': 0,
                              'coded_duration': 0, 'percent': 0})
        # Stats results calculated
        for st in stats:
            for r in self.results:
                if st['codename'] == r['codename'] and st['fid'] == r['fid']:
                    st['codecount'] += 1
                    st['coded_duration'] += r['pos1'] - r['pos0']
                    # 2 decimal places
                    st['percent'] = round((st['coded_duration'] / st['file_duration']) * 100, 2)
        final_stats = []
        for st in stats:
            if st['codecount'] > 0:
                final_stats.append(st)
        msg = _("A/V code statistics:")
        for st in final_stats:
            msg += f"\n{st['codename']} | {st['filename']} | {_('Count:')} {st['codecount']} | "
            if not erroneous_msecs:
                msg += f"{_('Percent of file:')} {st['percent']}%"
            else:
                msg += _("Percent of file: Unknown. Either VLC not installer or file not found.")
        msg += "\n========"
        if len(final_stats) == 0:
            msg = ""
        return stats, msg

    def fill_text_edit_stats_results(self):
        """ Fill text edit with statistics for codes.
         As total counts and count and percent per file. """

        text_stats, text_msg = self.text_code_count_and_percent()
        img_stats, img_msg = self.image_code_count_and_percent()
        av_stats, av_msg = self.av_code_count_and_percent()
        counts = []
        for s in text_stats:
            counts.append(s['codename'])
        for s in img_stats:
            counts.append(s['codename'])
        for s in av_stats:
            counts.append(s['codename'])
        counts = list(set(counts))
        counts.sort()
        # Display code count totals
        msg = ""
        total_count = 0
        for c in counts:
            count = 0
            for s in text_stats:
                if s['codename'] == c:
                    count += s['codecount']
            for s in img_stats:
                if s['codename'] == c:
                    count += s['codecount']
            for s in av_stats:
                if s['codename'] == c:
                    count += s['codecount']
            msg += "\n" + c + " : " + str(count)
            total_count += count
        msg = f"{_('Code count totals')}: {total_count}\n============{msg}"
        msg += "\n============"
        self.ui.textEdit.append(msg)
        if text_msg != "":
            self.ui.textEdit.append(text_msg)
        if img_msg != "":
            self.ui.textEdit.append(img_msg)
        if av_msg != "":
            self.ui.textEdit.append(av_msg)

    def search_results_next(self):
        """ Search textedit results for text """

        search_text = self.ui.lineEdit_search_results.text()
        if search_text == "":
            return
        if self.ui.textEdit.toPlainText() == "":
            return
        if self.ui.textEdit.textCursor().position() >= len(self.ui.textEdit.toPlainText()):
            cursor = self.ui.textEdit.textCursor()
            cursor.setPosition(0, QtGui.QTextCursor.MoveMode.MoveAnchor)
            self.ui.textEdit.setTextCursor(cursor)
        te_text = self.ui.textEdit.toPlainText()
        pattern = None
        flags = 0
        try:
            pattern = re.compile(search_text, flags)
        except re.error as err:
            logger.warning('re error Bad escape ' + str(err))
        if pattern is None:
            return
        for match in pattern.finditer(te_text):
            if match.start() > self.ui.textEdit.textCursor().position():
                cursor = self.ui.textEdit.textCursor()
                cursor.setPosition(match.start(), QtGui.QTextCursor.MoveMode.MoveAnchor)
                cursor.setPosition(match.start() + len(search_text), QtGui.QTextCursor.MoveMode.KeepAnchor)
                self.ui.textEdit.setTextCursor(cursor)
                break

    def fill_text_edit_with_search_results(self):
        """ The textEdit.document is filled with the search results.
        Results are drawn from the textEdit.document to fill reports in .txt and .odt formats.
        Results are drawn from the textEdit.document and html_links variable to fill reports in html format.
        Results are drawn from self.text_results, self.image_results and self.av_results to prepare a csv file.
        The results are converted from tuples to dictionaries.
        As results are added to the textEdit, positions for the headings (code, file, codername) are recorded for
        right-click context menu to display contextualised coding in another dialog.
        """

        self.text_links = []
        self.matrix_links = []
        if self.ui.checkBox_show_stats.isChecked():
            self.fill_text_edit_stats_results()

        # Add textedit positioning for context on clicking appropriate heading in results
        # Fill text edit with heading, text, image or
        fmt_normal = QtGui.QTextCharFormat()
        fmt_normal.setFontWeight(QtGui.QFont.Weight.Normal)
        fmt_bold = QtGui.QTextCharFormat()
        fmt_bold.setFontWeight(QtGui.QFont.Weight.Bold)
        fmt_italic = QtGui.QTextCharFormat()
        fmt_italic.setFontItalic(True)
        fmt_larger = QtGui.QTextCharFormat()
        fmt_larger.setFontPointSize(self.app.settings['docfontsize'] + 2)
        # memo_choice, use current index, as other languages will not match
        memo_choice_index = self.ui.comboBox_memos.currentIndex()

        for i, row in enumerate(self.results):
            self.heading(row)
            if row['coded_memo'] != "" and memo_choice_index in (4, 5):  # Only memos, Only coded memos
                self.ui.textEdit.insertPlainText("\n")
                self.ui.textEdit.insertPlainText(row['coded_memo'] + "\n")
            if row['result_type'] == 'text' and memo_choice_index not in (4, 5):  # Only memos, Only coded memos
                cursor = self.ui.textEdit.textCursor()
                pos0 = len(self.ui.textEdit.toPlainText())
                self.ui.textEdit.insertPlainText("\n")
                self.ui.textEdit.insertPlainText(row['pretext'])
                pos1 = len(self.ui.textEdit.toPlainText())
                cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
                cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
                cursor.setCharFormat(fmt_normal)
                pos0 = len(self.ui.textEdit.toPlainText())
                self.ui.textEdit.insertPlainText(row['text'])
                pos1 = len(self.ui.textEdit.toPlainText())
                cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
                cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
                if self.ui.checkBox_text_context.isChecked() and self.app.settings['report_text_context_style'] == 'Bold':
                    cursor.setCharFormat(fmt_bold)
                if self.ui.checkBox_text_context.isChecked() and self.app.settings['report_text_context_style'] == 'Italic':
                    cursor.setCharFormat(fmt_italic)
                if self.ui.checkBox_text_context.isChecked() and self.app.settings['report_text_context_style'] == 'Bigger':
                    cursor.setCharFormat(fmt_larger)
                pos0 = len(self.ui.textEdit.toPlainText())
                self.ui.textEdit.insertPlainText(row['posttext'])
                pos1 = len(self.ui.textEdit.toPlainText())
                cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
                cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
                if self.ui.checkBox_text_context.isChecked():
                    cursor.setCharFormat(fmt_normal)
                if memo_choice_index != 5:  # Only coded memos:
                    self.ui.textEdit.insertPlainText("\n")
                if row['coded_memo'] != "" and memo_choice_index in (1, 2):  # Also all memos, Also coded memos
                    self.ui.textEdit.insertPlainText(f"{_('MEMO:')} {row['coded_memo']}\n")
            if row['result_type'] == 'image' and memo_choice_index not in (4, 5):  # Only memos, Only coded memos
                self.put_image_into_textedit(row, i, self.ui.textEdit)
            if row['result_type'] == 'av' and memo_choice_index not in (4, 5):  # Only memos, Only coded memos
                self.ui.textEdit.insertPlainText(f"\n{row['text']}\n")
            self.text_links.append(row)
        self.eventFilterTT.set_positions(self.text_links)

        # Fill matrix or clear third splitter pane.
        self.ui.tableWidget.setColumnCount(0)
        self.ui.tableWidget.setRowCount(0)
        matrix_option_index = self.ui.comboBox_matrix.currentIndex()

        if matrix_option_index == 0:
            self.ui.splitter.setSizes([200, 400, 0])
            return
        # Categories by case, Top categories by case, Codes by case
        if self.case_ids_string == "" and matrix_option_index in (1, 3, 5):
            Message(self.app, _("No case matrix"), _("Cases not selected")).exec()
            self.ui.splitter.setSizes([200, 400, 0])
            return
        if self.case_ids_string != "" and matrix_option_index == 1:  # Top categories by case
            self.matrix_by_top_categories(self.results, self.case_ids_string, "case")
        if self.case_ids_string == "" and matrix_option_index == 2:  # Top categories by file
            self.matrix_by_top_categories(self.results, self.file_ids_string)
        # Top categories BY FILE for SELECTED CASES
        if self.case_ids_string != "" and matrix_option_index == 2:  # Top categories by file
            # Need to create file ids comma separated string
            files_id_name = self.app.get_filenames()
            file_ids = []
            for r in self.results:
                file_ids.append(r['fid'])
                # Need to replace Case with File and need to replace file_or_casename
                r['file_or_case'] = 'File'
                for f in files_id_name:
                    if f['id'] == r['fid']:
                        r['file_or_casename'] = f['name']
            file_ids = str(list(set(file_ids)))[1:-1]  # Remove '[' ']'
            self.matrix_by_top_categories(self.results, file_ids)
        if self.case_ids_string != "" and matrix_option_index == 3:  # Categories by case
            self.matrix_by_categories(self.results, self.case_ids_string, "case")
        if self.case_ids_string == "" and matrix_option_index == 4:  # Categories by file
            self.matrix_by_categories(self.results, self.file_ids_string)
        # Categories BY FILE for SELECTED CASES
        if self.case_ids_string != "" and matrix_option_index == 4:  # Categories by file
            # Need to create file ids comma separated string
            files_id_name = self.app.get_filenames()
            file_ids = []
            for r in self.results:
                file_ids.append(r['fid'])
                # Need to replace Case with File and need to replace file_or_casename
                r['file_or_case'] = 'File'
                for f in files_id_name:
                    if f['id'] == r['fid']:
                        r['file_or_casename'] = f['name']
            file_ids = str(list(set(file_ids)))[1:-1]  # Remove '[' ']'
            self.matrix_by_categories(self.results, file_ids)

        if self.case_ids_string != "" and matrix_option_index == 5:  # Codes by case
            self.matrix_by_codes(self.results, self.case_ids_string, "case")
        if self.case_ids_string == "" and matrix_option_index == 6:  # Codes by file
            self.matrix_by_codes(self.results, self.file_ids_string)
        # Codes BY FILE for SELECTED CASES
        if self.case_ids_string != "" and matrix_option_index == 6:  # Codes by file
            # Need to create file ids comma separated string
            files_id_name = self.app.get_filenames()
            file_ids = []
            for r in self.results:
                file_ids.append(r['fid'])
                # Need to replace Case with File and need to replace file_or_casename
                r['file_or_case'] = 'File'
                for f in files_id_name:
                    if f['id'] == r['fid']:
                        r['file_or_casename'] = f['name']
            file_ids = str(list(set(file_ids)))[1:-1]  # Remove '[' ']'
            self.matrix_by_codes(self.results, file_ids)
        self.ui.splitter.setSizes([100, 100, 500])

    def put_image_into_textedit(self, img, counter, text_edit):
        """ Scale image, add resource to document, insert image.
        """

        text_edit.append("\n")
        path_ = self.app.project_path + img['mediapath']
        if img['mediapath'][0:7] == "images:":
            path_ = img['mediapath'][7:]
        document = text_edit.document()
        image = QtGui.QImageReader(path_).read()
        image = image.copy(int(img['x1']), int(img['y1']), int(img['width']), int(img['height']))
        # Scale to max 300 wide or high. perhaps add option to change maximum limit?
        scaler_w = 1.0
        scaler_h = 1.0
        if image.width() > 400:
            scaler_w = 400 / image.width()
        if image.height() > 400:
            scaler_h = 400 / image.height()
        if scaler_w < scaler_h:
            scaler = scaler_w
        else:
            scaler = scaler_h
        # Need unique image names or the same image from the same path is reproduced
        # Default for an image  stored in the project folder.
        imagename = str(counter) + '-' + img['mediapath']
        # Check and change path for a linked image file
        if img['mediapath'][0:7] == "images:":
            imagename = str(counter) + '-' + "/images/" + img['mediapath'].split('/')[-1]
        # imagename is now: 0-/images/filename.jpg  # where 0- is the counter 1-, 2- etc
        url = QtCore.QUrl(imagename)
        document.addResource(QtGui.QTextDocument.ResourceType.ImageResource.value, url, image)
        cursor = text_edit.textCursor()
        char_pos = cursor.position()
        image_format = QtGui.QTextImageFormat()
        image_format.setWidth(image.width() * scaler)
        image_format.setHeight(image.height() * scaler)
        image_format.setName(url.toString())
        cursor.insertImage(image_format)
        text_edit.insertHtml("<br />")
        self.html_links.append({'imagename': imagename, 'image': image, 'image_char_pos': char_pos, 'avname': None,
                                'av0': None, 'av1': None, 'avtext': None})
        if img['coded_memo'] != "":
            text_edit.insertPlainText(_("MEMO: ") + img['coded_memo'] + "\n")

    def heading(self, item):
        """ Takes a dictionary item and creates a html heading for the coded text portion.
        Inserts the heading into the main textEdit.
        Fills the textedit_start and textedit_end link positions
        param:
            item: dictionary of code, file_or_casename, positions, text, coder
        """

        cur = self.app.conn.cursor()
        cur.execute("select name from source where id=?", [item['fid']])
        filename = ""
        res = cur.fetchone()
        if res is not None:
            filename = res[0]
        head = "\n"
        if item['result_type'] == 'text':
            head += "[" + str(item['pos0']) + "-" + str(item['pos1']) + "] "
        head += item['codename'] + ", "
        memo_choice = self.ui.comboBox_memos.currentText()
        if memo_choice in (_("Also code memos"), _("Also all memos"), _("Only memos")) and item['codename_memo'] != "":
            head += _("CODE MEMO: ") + item['codename_memo'] + "<br />"
        head += _("File: ") + filename + ", "
        if memo_choice in (_("Also all memos"), _("Only memos")) and item['source_memo'] != "":
            head += _(" FILE MEMO: ") + item['source_memo']
        if item['file_or_case'] == 'Case':
            head += " " + _("Case: ") + item['file_or_casename']
            if memo_choice in (_("Also all memos"), _("Only memos")):
                cur = self.app.conn.cursor()
                cur.execute("select memo from cases where name=?", [item['file_or_casename']])
                res = cur.fetchone()
                if res is not None and res[0] != "" and res[0] is not None:
                    head += ", " + _("CASE MEMO: ") + res[0]
        head += " " + _("Coder: ") + item['coder']
        if self.app.settings['showids']:
            try:
                head += ", ctid: " + str(item['ctid'])
            except KeyError:
                pass
            try:
                head += ", imid: " + str(item['imid'])
            except KeyError:
                pass
            try:
                head += ", avid: " + str(item['avid'])
            except KeyError:
                pass

        cursor = self.ui.textEdit.textCursor()
        fmt = QtGui.QTextCharFormat()
        pos0 = len(self.ui.textEdit.toPlainText())
        item['textedit_start'] = pos0
        self.ui.textEdit.append(head)
        cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
        pos1 = len(self.ui.textEdit.toPlainText())
        cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
        brush = QBrush(QtGui.QColor(item['color']))
        fmt.setBackground(brush)
        text_brush = QBrush(QtGui.QColor(TextColor(item['color']).recommendation))
        fmt.setForeground(text_brush)
        cursor.setCharFormat(fmt)
        item['textedit_end'] = len(self.ui.textEdit.toPlainText())

    def text_edit_menu(self, position):
        """ Context menu for textEdit.
        To view coded in context. """

        if self.ui.textEdit.toPlainText() == "":
            return
        cursor_context_pos = self.ui.textEdit.cursorForPosition(position)
        # This bit to get image details for  rotation
        # https://stackoverflow.com/questions/18700945/qtextbrowser-how-to-identify-image-from-mouse-click-position
        fmt = cursor_context_pos.charFormat()
        img_fmt = None
        html_link = None
        if fmt.isImageFormat():
            img_fmt = fmt.toImageFormat()  # QtGui.QTextImageFormat
            # print("name", img_fmt.name(), img_fmt.height(), img_fmt.width())
            for h in self.html_links:
                if h['imagename'] == img_fmt.name():
                    html_link = h
                    break

        pos = cursor_context_pos.position()
        selected_text = self.ui.textEdit.textCursor().selectedText()
        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")

        # Check that there is a link to view at this location before showing menu option
        action_view = None
        action_unmark = None
        action_important = None
        action_change_code_to = None
        code_here = None
        for row in self.results:
            if row['textedit_start'] <= pos < row['textedit_end']:
                code_here = row
                break
        if code_here and code_here['result_type'] != "deleted":
            action_view = menu.addAction(_("View in context"))
            action_unmark = menu.addAction(_("Unmark"))
            action_important = menu.addAction(_("Add important mark"))
            action_change_code_to = menu.addAction(_("Change code to"))
        action_copy = None
        if selected_text != "":
            action_copy = menu.addAction(_("Copy to clipboard"))
        action_copy_all = menu.addAction(_("Copy all to clipboard"))
        action_rotate_180 = None
        if img_fmt:
            action_rotate_180 = menu.addAction(_("Rotate image 90 degrees"))
        action_hide_top_groupbox = None
        action_show_top_groupbox = None
        if self.ui.groupBox.isHidden():
            action_show_top_groupbox = menu.addAction(_("Show control panel"))
        if not self.ui.groupBox.isHidden():
            action_hide_top_groupbox = menu.addAction(_("Hide control panel"))
        action = menu.exec(self.ui.textEdit.mapToGlobal(position))
        if action is None:
            return
        if action == action_view:
            self.show_context_from_text_edit(code_here)
        if action == action_unmark:
            self.unmark(code_here)
        if action == action_important:
            self.mark_important(code_here)
        if action == action_change_code_to:
                self.change_code_to_another_code(code_here)
        if action == action_copy:
            cb = QtWidgets.QApplication.clipboard()
            cb.setText(selected_text)
        if action == action_copy_all:
            cb = QtWidgets.QApplication.clipboard()
            te_text = self.ui.textEdit.toPlainText()
            cb.setText(te_text)
        if action == action_show_top_groupbox:
            self.ui.groupBox.setVisible(True)
        if action == action_hide_top_groupbox:
            self.ui.groupBox.setVisible(False)
        if action == action_rotate_180:
            self.rotate_image(cursor_context_pos, img_fmt, html_link, 90)

    def mark_important(self, code):
        """ Add important mark to coding.
        No effect if already marked important.
        param:
            code : Dictionary of codenmae, color, file_or_casename, pos0, pos1, text, coder, fid, ctid, cid, result_type
        """

        cur = self.app.conn.cursor()
        if code['result_type'] == 'text':
            cur.execute("update code_text set important=1 where ctid=?", [code['ctid']])
        if code['result_type'] == 'image':
            cur.execute("update code_image set important=1 where imid=?", [code['imid']])
        if code['result_type'] == 'av':
            cur.execute("update code_av set important=1 where avid=?", [code['avid']])
        self.app.conn.commit()

        self.app.delete_backup = False

        # Remove widgets from coding layout
        contents = self.tab_coding.layout()
        if contents:
            for i in reversed(range(contents.count())):
                contents.itemAt(i).widget().close()
                contents.itemAt(i).widget().setParent(None)

    def change_code_to_another_code(self, existing_code):
        """ Change the selected code to another from a list. """

        # Get replacement code
        codes_list = deepcopy(self.code_names)
        to_hide = None
        for code_ in codes_list:
            if code_['cid'] == existing_code['cid']:
                to_hide = code_
        if to_hide:
            codes_list.remove(to_hide)
        ui = DialogSelectItems(self.app, codes_list, _("Select replacement code"), "single")
        ok = ui.exec()
        if not ok:
            return
        replacement_code = ui.get_selected()
        if not replacement_code:
            return
        cur = self.app.conn.cursor()
        try:
            if existing_code['result_type'] == 'text':
                cur.execute("update code_text set cid=? where ctid=?", [replacement_code['cid'], existing_code['ctid']])
            if existing_code['result_type'] == 'image':
                cur.execute("update code_image set cid=? where imid=?", [replacement_code['cid'], existing_code['imid']])
            if existing_code['result_type'] == 'av':
                cur.execute("update code_av set cid=? where avid=?", [replacement_code['cid'], existing_code['avid']])
            self.app.conn.commit()
        except sqlite3.IntegrityError:
            Message(self.app, "Cannot change code", "This is already marked with the selected code").exec()
            return
        Message(self.app, "Changed code", "Run report again to update display").exec()
        self.app.delete_backup = False
        # Remove widgets from coding layout
        contents = self.tab_coding.layout()
        if contents:
            for i in reversed(range(contents.count())):
                contents.itemAt(i).widget().close()
                contents.itemAt(i).widget().setParent(None)

    def unmark(self, code):
        """ Unmark this coding.
        param:
            code : Dictionary of codenmae, color, file_or_casename, pos0, pos1, text, coder, fid, ctid, cid, result_type"""

        coded = f"{_('Delete coded section.')} {code['codename']}. {code['coder']}"
        ui = DialogConfirmDelete(self.app, coded, _("Delete coded section"))
        ok = ui.exec()
        if not ok:
            return
        cur = self.app.conn.cursor()
        if code['result_type'] == 'text':
            cur.execute("delete from code_text where ctid=?", [code['ctid']])
        if code['result_type'] == 'image':
            cur.execute("delete from code_image where imid=?", [code['imid']])
        if code['result_type'] == 'av':
            cur.execute("delete from code_av where avid=?", [code['avid']])
        self.app.conn.commit()

        self.app.delete_backup = False
        code['result_type'] = "deleted"

        for m in self.matrix_links:
            if m['result_type'] == 'text' and m['ctid'] == code['ctid']:
                m['result_type'] = 'deleted'
                break
            if m['result_type'] == 'image' and m['imid'] == code['imid']:
                m['result_type'] = 'deleted'
                break
            if m['result_type'] == 'av' and m['avid'] == code['avid']:
                m['result_type'] = 'deleted'
                break

        # Format strike through
        cursor = self.ui.textEdit.textCursor()
        cursor.setPosition(code['textedit_start'], QtGui.QTextCursor.MoveMode.MoveAnchor)
        cursor.setPosition(code['textedit_end'], QtGui.QTextCursor.MoveMode.KeepAnchor)
        fmt = QtGui.QTextCharFormat()
        fmt.setFontStrikeOut(True)
        cursor.mergeCharFormat(fmt)

        # Remove widgets from coding layout
        contents = self.tab_coding.layout()
        if contents:
            for i in reversed(range(contents.count())):
                contents.itemAt(i).widget().close()
                contents.itemAt(i).widget().setParent(None)

    def show_context_from_text_edit(self, code):
        """ Heading (code, file, owner) in textEdit clicked so show context of coding in dialog.
        Called by: textEdit.cursorPositionChanged, after results are filled.
        Called by context menu.
        param:
            code : Dictionary of codenmae, color, file_or_casename, pos0, pos1, text, coder, fid, ctid, cid, result_type
        """

        if code['result_type'] == 'text':
            ui = DialogCodeInText(self.app, code)
            ui.exec()
        if code['result_type'] == 'image':
            ui = DialogCodeInImage(self.app, code)
            ui.exec()
        if code['result_type'] == 'av':
            ui = DialogCodeInAV(self.app, code)
            ui.exec()

    def rotate_image(self, cursor_context_pos, img_fmt, html_link, degrees):
        """  Rotate image 180 degrees.
        Tried to do 90 and 270 degree rotations but could not update the image format width and height.
        param:
            TextImage Format img_fmt
            Dictionary html_link {imagename, image:QImage, avname, av0, av1, avtext}
        """

        document = self.ui.textEdit.document()
        url = QtCore.QUrl(img_fmt.name())  # Location in document
        image = html_link['image']
        transform = QtGui.QTransform().rotate(degrees)
        image = image.transformed(transform)
        html_link['image'] = image
        document.addResource(QtGui.QTextDocument.ResourceType.ImageResource.value, url, image)
        scaler_w = 1.0
        scaler_h = 1.0
        if image.width() > 400:
            scaler_w = 400 / image.width()
        if image.height() > 400:
            scaler_h = 400 / image.height()
        if scaler_w < scaler_h:
            scaler = scaler_w
        else:
            scaler = scaler_h
        img_fmt.setWidth(image.width() * scaler)
        img_fmt.setHeight(image.height() * scaler)
        # Image is locate at a one character position, remove and replace with the new image
        cursor = self.ui.textEdit.textCursor()
        cursor.setPosition(html_link['image_char_pos'], QtGui.QTextCursor.MoveMode.MoveAnchor)
        cursor.setPosition(html_link['image_char_pos'] + 1, QtGui.QTextCursor.MoveMode.KeepAnchor)
        self.ui.textEdit.setTextCursor(cursor)
        cursor.removeSelectedText()
        cursor_context_pos.insertImage(img_fmt)

    def matrix_heading(self, item, text_edit):
        """ Takes a dictionary item and creates a heading for the coded text portion.
        Also adds the textEdit start and end character positions for this text in this text edit
        param:
            item: dictionary of code, file_or_casename, positions, text, coder
        """

        cur = self.app.conn.cursor()
        cur.execute("select name from source where id=?", [item['fid']])
        filename = ""
        res = cur.fetchone()
        if res is not None:
            filename = res[0]
        memo_choice = self.ui.comboBox_memos.currentText()
        head = "\n" + _("[VIEW] ")
        head += item['codename'] + ", "
        if memo_choice in (_("Also all memos"), _("Also code memos"), _("Only memos")) and item['codename_memo'] != "":
            head += _("CODE MEMO: All memo") + item['codename_memo'] + "<br />"
        head += f"{_('File:')} {filename}, "
        if memo_choice in (_("Also alll memos"), _("Only memos")) and item['source_memo'] != "":
            head += f" {_('FILE MEMO:')} {item['source_memo']}"
        if item['file_or_case'] == 'Case:':
            head += f" {item['file_or_case']}: {item['file_or_casename']}, "
            if memo_choice in (_("Also all memos"), _("Only memos")):
                cur = self.app.conn.cursor()
                cur.execute("select ifnull(memo,'') from cases where name=?", [item['file_or_casename']])
                res = cur.fetchone()
                if res is not None and res != "":
                    head += f", {_('CASE MEMO:')} {res[0]}"
        head += item['coder']
        cursor = text_edit.textCursor()
        fmt = QtGui.QTextCharFormat()
        pos0 = len(text_edit.toPlainText())
        item['textedit_start'] = pos0
        text_edit.append(head)
        cursor.setPosition(pos0, QtGui.QTextCursor.MoveMode.MoveAnchor)
        pos1 = len(text_edit.toPlainText())
        cursor.setPosition(pos1, QtGui.QTextCursor.MoveMode.KeepAnchor)
        brush = QBrush(QtGui.QColor(item['color']))
        fmt.setBackground(brush)
        text_brush = QBrush(QtGui.QColor(TextColor(item['color']).recommendation))
        fmt.setForeground(text_brush)
        cursor.setCharFormat(fmt)
        item['textedit_end'] = len(text_edit.toPlainText())

    def matrix_by_codes(self, results_, ids, type_="file"):
        """ Fill a tableWidget with rows of cases and columns of codes.
        First identify all codes.
        Fill tableWidget with columns of codes and rows of cases.
        Called by: fill_text_edit_with_search_results
        param:
        results_ : list of dictionary text, image, av result items
            ids : list of case ids OR file ids - as a string of integers, comma separated
            type_ : 'file' or 'case'
        """

        # Do not overwrite positions in original text_links object
        results = deepcopy(results_)
        # Need a top key
        for r in results:
            r['top'] = r['codename']
        # Get selected codes (Matrix columns)
        items = self.ui.treeWidget.selectedItems()
        horizontal_labels = [item.text(0) for item in items if item.text(1)[:3] == "cid"]

        # Get file or cases (rows)
        cur = self.app.conn.cursor()
        sql = f"select distinct id, name from source where id in ({ids}) order by name"
        if type_ == "case":
            sql = f"select caseid, name from cases where caseid in ({ids})"
        cur.execute(sql)
        id_and_name = cur.fetchall()
        vertical_labels = [c[1] for c in id_and_name]

        transpose = self.ui.checkBox_matrix_transpose.isChecked()
        if transpose:
            vertical_labels, horizontal_labels = horizontal_labels, vertical_labels
        self.fill_matrix_table(results, vertical_labels, horizontal_labels)

    def matrix_by_categories(self, results_, ids, type_="file"):
        """ Fill a tableWidget with rows of case or file name and columns of categories.
        First identify the categories. Then map all codes which are directly assigned to the categories.
        Called by: fill_text_edit_with_search_results
        param:
            results_ : list of dictionary of text, image, av result items
            ids : list of case ids OR file ids, as string of comma separated integers
            type_ : file or case ids
        """

        # Do not overwrite positions in original text_links object
        results = deepcopy(results_)
        # All categories within selection
        items = self.ui.treeWidget.selectedItems()
        top_level = []  # Categories at any level
        horizontal_labels = []
        sub_codes = []
        for item in items:
            if item.text(1)[0:3] == "cat":
                top_level.append({'name': item.text(0), 'cat': item.text(1)})
                horizontal_labels.append(item.text(0))
            # Find sub-code and traverse upwards to map to category
            if item.text(1)[0:3] == 'cid':
                sub_code = {'codename': item.text(0), 'cid': item.text(1)}
                # Maybe None of a top level code - as this will have no parent
                if item.parent() is not None:
                    sub_code['top'] = item.parent().text(0)
                    sub_codes.append(sub_code)
                    add_cat = True
                    for tl in top_level:
                        if tl['name'] == item.parent().text(0):
                            add_cat = False
                    if add_cat:
                        top_level.append({'name': item.parent().text(0), 'cat': item.parent().text(1)})
                        horizontal_labels.append(item.parent().text(0))

        # Add category name - which will match the tableWidget column category name
        res_categories = []
        for i in results:
            # Replaces the top-level name by mapping to the correct top-level category name (column)
            # Codes will not have 'top' key
            for s in sub_codes:
                if i['codename'] == s['codename']:
                    i['top'] = s['top']
            if "top" in i:
                res_categories.append(i)
        # Only show categories in the results
        results = res_categories
        cur = self.app.conn.cursor()
        sql = f"select distinct id, name from source where id in ({ids}) order by name"
        if type_ == "case":
            sql = f"select caseid, name from cases where caseid in ({ids})"
        cur.execute(sql)
        id_and_name = cur.fetchall()
        vertical_labels = []
        for c in id_and_name:
            vertical_labels.append(c[1])
        transpose = self.ui.checkBox_matrix_transpose.isChecked()
        if transpose:
            vertical_labels, horizontal_labels = horizontal_labels, vertical_labels
        self.fill_matrix_table(results, vertical_labels, horizontal_labels)

    def matrix_by_top_categories(self, results_, ids, type_="file"):
        """ Fill a tableWidget with rows of case or file name and columns of top level categories.
        First identify top-level categories. Then map all other codes to the
        top-level categories.
        Called by: fill_text_edit_with_search_results
        param:
            results_ : list of dictionary of text, image, av result items
            ids : string list of case ids or file ids, comma separated
            type_ : file or case
        """

        # Do not overwrite positions in original text_links object
        results = deepcopy(results_)
        # Get top level categories
        items = self.ui.treeWidget.selectedItems()
        top_level = []
        horizontal_labels = []
        sub_codes = []
        for item in items:
            root = self.ui.treeWidget.indexOfTopLevelItem(item)
            if root > -1 and item.text(1)[0:3] == "cat":
                top_level.append({'name': item.text(0), 'cat': item.text(1)})
                horizontal_labels.append(item.text(0))
            # Find sub-code and traverse upwards to map to top-level category
            if root == -1 and item.text(1)[0:3] == 'cid':
                not_top = True
                sub_code = {'codename': item.text(0), 'cid': item.text(1)}
                top_id = None
                while not_top:
                    item = item.parent()
                    if self.ui.treeWidget.indexOfTopLevelItem(item) > -1:
                        not_top = False
                        sub_code['top'] = item.text(0)
                        top_id = item.text(1)
                        sub_codes.append(sub_code)
                add_cat = True
                for tl in top_level:
                    if tl['name'] == sub_code['top']:
                        add_cat = False
                if add_cat and top_id is not None:
                    top_level.append({'name': sub_code['top'], 'cat': top_id})
                    horizontal_labels.append(sub_code['top'])

        # Add the top-level name - which will match the tableWidget column category name
        res_categories = []
        for i in results:
            # Replaces the top-level code name by mapping to the correct top-level category name (column)
            # Codes will not have 'top' key, so add it in.
            for s in sub_codes:
                if i['codename'] == s['codename']:
                    i['top'] = s['top']
            if "top" in i:
                res_categories.append(i)
        # Ony show top level categories
        results = res_categories

        cur = self.app.conn.cursor()
        sql = f"select distinct id, name from source where id in ({ids}) order by name"
        if type_ == "case":
            sql = f"select caseid, name from cases where caseid in ({ids})"
        cur.execute(sql)
        id_and_name = cur.fetchall()
        vertical_labels = []
        for c in id_and_name:
            vertical_labels.append(c[1])

        transpose = self.ui.checkBox_matrix_transpose.isChecked()
        if transpose:
            vertical_labels, horizontal_labels = horizontal_labels, vertical_labels
        self.fill_matrix_table(results, vertical_labels, horizontal_labels)

    def fill_matrix_table(self, results, vertical_labels, horizontal_labels):
        """ Clear then fill the table.
        Called by matrix_by_codes, matrix_by_categories, matrix_by_top_categories.
        """

        # Clear and fill tableWidget
        doc_font = 'font: ' + str(self.app.settings['docfontsize']) + 'pt '
        doc_font += '"' + self.app.settings['font'] + '";'
        self.ui.tableWidget.setStyleSheet(doc_font)
        self.ui.tableWidget.setColumnCount(len(horizontal_labels))
        # Keep horizontal labels to 80 chars per line
        horizontal_label_wrap = []
        for hl in horizontal_labels:
            segs = re.findall('.{1,80}', hl)
            horizontal_label_wrap.append("\n".join(segs))
        self.ui.tableWidget.setHorizontalHeaderLabels(horizontal_label_wrap)
        self.ui.tableWidget.setRowCount(len(vertical_labels))
        # Keep vertical labels to 30 chars per line
        vertical_labels_wrap = []
        for vl in vertical_labels:
            segs = re.findall('.{1,30}', vl)
            vertical_labels_wrap.append("\n".join(segs))
        self.ui.tableWidget.setVerticalHeaderLabels(vertical_labels_wrap)
        for i, vl in enumerate(vertical_labels):
            self.ui.tableWidget.verticalHeaderItem(i).setToolTip(vl)
        # Need to create a table of separate textEdits for reference for cursorPositionChanged event.
        self.te = []
        for vl in vertical_labels:
            column_list = []
            for hl in horizontal_labels:
                tedit = QtWidgets.QTextEdit("")
                tedit.setReadOnly(True)
                tedit.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
                tedit.customContextMenuRequested.connect(self.table_text_edit_menu)
                column_list.append(tedit)
            self.te.append(column_list)
        self.matrix_links = []
        memo_choice = self.ui.comboBox_memos.currentText()
        if self.ui.checkBox_matrix_transpose.isChecked():
            for row in range(len(vertical_labels)):
                for col in range(len(horizontal_labels)):
                    for counter, r in enumerate(results):
                        if r['file_or_casename'] == horizontal_labels[col] and r['top'] == vertical_labels[row]:
                            r['row'] = row
                            r['col'] = col
                            self.te[row][col].insertHtml(self.matrix_heading(r, self.te[row][col]))
                            if r['result_type'] == 'text' and memo_choice in (_("Only memos"), _("Only coded memos")):
                                self.te[row][col].append(r['coded_memo'])
                            if r['result_type'] == 'text' and memo_choice not in (_("Only memos"), _("Only coded memos")):
                                self.te[row][col].append(r['text'])
                                if memo_choice in (_("Also all memos"), _("Also coded memos")) and r['coded_memo'] != "":
                                    self.te[row][col].append(f"{_('MEMO:')} {r['coded_memo']}")
                                self.te[row][col].insertPlainText("\n")
                            if r['result_type'] == 'image' and memo_choice in (_("Only memos"), _("Only coded memos")):
                                self.te[row][col].append(r['coded_memo'])
                            if r['result_type'] == 'image' and memo_choice not in (_("Only memos"), _("Only coded memos")):
                                self.put_image_into_textedit(r, counter, self.te[row][col])
                            if r['result_type'] == 'av' and memo_choice not in (_("Only memos"), _("Only coded memos")):
                                self.te[row][col].insertPlainText(f"{r['text']}\n")
                            self.matrix_links.append(r)
                    self.ui.tableWidget.setCellWidget(row, col, self.te[row][col])
        else:  # Not transposed
            for row in range(len(vertical_labels)):
                for col in range(len(horizontal_labels)):
                    for counter, r in enumerate(results):
                        if r['file_or_casename'] == vertical_labels[row] and r['top'] == horizontal_labels[col]:
                            r['row'] = row
                            r['col'] = col
                            self.te[row][col].insertHtml(self.matrix_heading(r, self.te[row][col]))
                            if r['result_type'] == 'text' and memo_choice in (_("Only memos"), _("Only coded memos")):
                                self.te[row][col].append(r['coded_memo'])
                            if r['result_type'] == 'text' and memo_choice not in (_("Only memos"), _("Only coded memos")):
                                self.te[row][col].append(r['text'])
                                try:
                                    if memo_choice in (_("Also all memos"), "Also coded memos") and r['coded_memo'] != "":
                                        self.te[row][col].append(_("MEMO: ") + r['coded_memo'])
                                except TypeError as err:
                                    msg = str(err)
                                    msg += f"\nMatrix Coded Memo Error:\nchoice: {memo_choice}\n"
                                    msg += f"Result dictionary:\n{r}\n"
                                    logger.error(msg)
                                self.te[row][col].insertPlainText("\n")
                            if r['result_type'] == 'image' and memo_choice in (_("Only memos"), _("Only coded memos")):
                                self.te[row][col].append(r['coded_memo'])
                            if r['result_type'] == 'image' and memo_choice not in (_("Only memos"), _("Only coded memos")):
                                self.put_image_into_textedit(r, counter, self.te[row][col])
                            if r['result_type'] == 'av' and memo_choice not in (_("Only memos"), _("Only coded memos")):
                                self.te[row][col].insertPlainText(r['text'] + "\n")
                            self.matrix_links.append(r)
                    self.ui.tableWidget.setCellWidget(row, col, self.te[row][col])
        self.ui.tableWidget.resizeRowsToContents()
        self.ui.tableWidget.resizeColumnsToContents()
        # Maximise the space from one column or one row
        if self.ui.tableWidget.columnCount() == 1:
            self.ui.tableWidget.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        if self.ui.tableWidget.rowCount() == 1:
            self.ui.tableWidget.verticalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.ui.tableWidget.verticalHeader().setMaximumWidth(260)

    def table_text_edit_menu(self, position):
        """ Context menu for textEdit.
        To view coded in context.
        """

        x = self.ui.tableWidget.currentRow()
        y = self.ui.tableWidget.currentColumn()
        te = self.te[x][y]
        te_text = te.toPlainText()
        if te_text == "":
            return
        cursor_context_pos = te.cursorForPosition(position)
        pos = cursor_context_pos.position()
        selected_text = te.textCursor().selectedText()
        menu = QtWidgets.QMenu()
        menu.setStyleSheet("QMenu {font-size:" + str(self.app.settings['fontsize']) + "pt} ")

        # Check that there is a link to view at this location before showing menu option
        action_view = None
        for m in self.matrix_links:
            if m['row'] == x and m['col'] == y and m['textedit_start'] <= pos < m['textedit_end'] \
                    and m['result_type'] != 'deleted':
                action_view = menu.addAction(_("View in context"))
        action_copy = None
        if selected_text != "":
            action_copy = menu.addAction(_("Copy to clipboard"))
        action_copy_all = menu.addAction(_("Copy all to clipboard"))
        action = menu.exec(te.mapToGlobal(position))
        if action is None:
            return
        if action == action_copy:
            cb = QtWidgets.QApplication.clipboard()
            cb.setText(selected_text)
        if action == action_copy_all:
            cb = QtWidgets.QApplication.clipboard()
            te_text = te.toPlainText()
            cb.setText(te_text)
        if action == action_view:
            for m in self.matrix_links:
                if m['row'] == x and m['col'] == y and m['textedit_start'] <= pos < m['textedit_end']:
                    if 'mediapath' not in m:
                        ui = DialogCodeInText(self.app, m)
                        ui.exec()
                        return
                    if m['mediapath'][0:7] in ('images:', '/images'):
                        ui = DialogCodeInImage(self.app, m)
                        ui.exec()
                        return
                    if m['mediapath'][0:6] in ('audio:', 'video:', '/audio', '/video'):
                        ui = DialogCodeInAV(self.app, m)
                        ui.exec()
                        return


class ToolTipEventFilter(QtCore.QObject):
    """ Used to add a dynamic tooltip for the textBrowser.
    The tool top text is presented according to its position in the text.
    """

    media_data = None

    def set_positions(self, media_data):
        """ Code_text contains the positions for the tooltip to be displayed.

        param:
            media_data: List of dictionaries of the text contains: pos0, pos1
        """

        self.media_data = media_data

    def eventFilter(self, receiver, event):
        if event.type() == QtCore.QEvent.Type.ToolTip:
            cursor = receiver.cursorForPosition(event.pos())
            pos = cursor.position()
            receiver.setToolTip("")
            if self.media_data is None:
                return super(ToolTipEventFilter, self).eventFilter(receiver, event)
            for item in self.media_data:
                if item['textedit_start'] <= pos <= item['textedit_end']:
                    receiver.setToolTip(_("Right click to view"))
        # Call Base Class Method to Continue Normal Event Processing
        return super(ToolTipEventFilter, self).eventFilter(receiver, event)
